# -*- coding: utf-8 -*-
"""
Bot.BSG — Telegram Bot (SINGLE FILE, FULL PROJECT)
Версия 18.0.0 | Ревизия sr-bot-2025-10-05-finance2
ч
Зависимости:
  pip install aiogram==2.25.1 openpyxl pillow

ГЛАВНОЕ, ЧТО СДЕЛАНО/ОБНОВЛЕНО:
- Регистрация: ФИО + телефон (кнопка «📱 Отправить номер»). На регистрации присваивается уникальный код пользователя: BSU-XXXX.
- Якорь-меню: одно постоянное сообщение, не дублируется, не закрывает меню. Обновляется при смене активного проекта.
- Проекты (админ): создать (название, локация, описание, даты, PDF), активировать, завершить. Рассылки пользователям с кнопкой «Закрыть».
- Документы: пользователи видят PDF активного проекта.
- Чеки (без OCR): Фото → Сумма → Описание (опц.) → Статус оплаты (Оплачен/Не оплачен/Позже) → Предпросмотр → Сохранить/Отменить.
  Сохраняется: дата, время, сумма, файл, описание, paid, receipt_no (RID-XXXX).
- Переименование файлов чеков: {SafeUserName}_BSU{BSU}_RID{RID-XXXX}_{YYYY-mm-dd_HH-MM-SS}_amt-xx.xx.jpg (никогда не повторяется за счёт счётчика и времени).
- «Мои чеки»: каждое фото + дата, сумма, описание, статус, НОМЕР ЧЕКА (RID-XXXX). Можно проставить оплачено/не оплачено, если не указано.
- Финансы (новые): статистика по проекту (всего/оплачено/не оплачено), список неоплаченных, «📨 Запросить выплату».
- Выплаты/запросы — ФАЙЛОВОЕ ХРАНЕНИЕ:
    data/finances/req_<timestamp>_<uid>.json  (pending/approved/confirmed + вся детализация).
  История и запросы грузятся с диска, переживают перезапуск.
- Поток выплат:
  Пользователь → «📨 Запросить выплату» → создаётся файл-запрос → Админ видит в «💵 Финансы/Запросы»
  Админ: «Посмотреть чеки», «✅ Выплатить», «❌ Закрыть».
  При «✅ Выплатить» пользователю приходят отдельные уведомления «💵 Выплата за чек ...», у КАЖДОГО уведомления есть «✅ Деньги получены»/«❌ Закрыть».
  Пользователь подтверждает получение — статус запроса → confirmed, сохраняется в файле запроса и в профиле; админу приходит «Пользователь подтвердил получение ...».
  В «Финансы» пользователя до подтверждения висит метка: «⚠️ Есть одобренные выплаты, подтвердите получение денег».
- Excel Ledger (улучшено «красиво»):
  • Главная книга: data/projects/<P>/ledger.xlsx
  • Лист "Ledger": столбцы → Дата, Время, Пользователь, BSU, Номер чека, Сумма, Файл, Описание, Оплачен(1/0/None)
  • Отдельные ЛИСТЫ по каждому пользователю (по имени + BSU) с такими же столбцами.
  • В Excel записывается ИМЯ пользователя, а не ID.
- SOS: подтверждение → запрос геолокации → рассылка всем КРОМЕ отправителя; у отправителя «🆘 SOS включён. Сигнал принят.» (не удаляется, пока не нажмёт «Закрыть»).
- Админ-панель:
  • «👥 Пользователи» (карточка пользователя с полной инфой и улучшенным оформлением)
  • «📂 Проекты»
  • «💵 Финансы»: «Запросы» и «История» (всё с файлов)
- Очистка чата: регистрация не очищается до завершения «Профиль сохранён», далее промежуточные сообщения очищаются. Якорь-меню не удаляется, только заменяется.

Токен: встроен по просьбе пользователя.
"""

import os, sys, json, random, re, base64, hashlib, secrets, asyncio
import unicodedata
from html import escape as html_escape
from datetime import datetime, timezone, timedelta
from typing import Dict, Optional, List, Tuple, Any, Set, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ExifTags, ImageOps
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.utils.exceptions import MessageNotModified, MessageCantBeEdited
from aiogram.types import (
    InlineKeyboardMarkup, InlineKeyboardButton,
    InputFile, ContentType, ReplyKeyboardRemove,
    KeyboardButton, ReplyKeyboardMarkup
)
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
import requests

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None


# ========================== CONFIG ==========================
TOKEN = "7005343266:AAG0bnY-wTc3kScKiIskSd0fO6MstesSbCk"
ADMIN_CODE = "3004"

BOT_NAME = "Bot.BSG"
BOT_VERSION = "18.0.0"
BOT_REVISION = "sr-bot-2025-10-05-finance2"

BASE_PATH = "data/projects"
USERS_PATH = "data/users"
BOT_FILE = "data/bot.json"
FIN_PATH = "data/finances"  # запросы/история выплат (файлово)

ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".heic", ".heif", ".tif", ".tiff"}

ALERTS_API_BASE_URL = "https://api.alerts.in.ua/v1"
ALERTS_API_ACTIVE_ENDPOINT = "/alerts/active.json"
ALERTS_API_HISTORY_TEMPLATE = "/regions/{uid}/alerts/{period}.json"
ALERTS_API_URL = f"{ALERTS_API_BASE_URL}{ALERTS_API_ACTIVE_ENDPOINT}"
ALERTS_DEFAULT_HISTORY_PERIOD = "week_ago"
ALERTS_API_TOKEN = "62f89091e56951ef257f763e445c09c1fd9dacd1ab2203"
ALERTS_API_TIMEOUT = 15
ALERTS_POLL_INTERVAL = 5  # seconds
ALERTS_HISTORY_CACHE_TTL = 300  # seconds
ALERTS_STANDDOWN_DISPLAY_WINDOW = 90 * 60  # seconds
ALERTS_DIRNAME = "alerts"
ALERTS_STATE_FILENAME = "state.json"
ALERTS_HISTORY_DIRNAME = "history"
ALERTS_LEGACY_HISTORY_FILENAME = "history.json"
ALERTS_USERS_FILENAME = "subscriptions.json"
ALERTS_TIMELINE_KEY = "timeline"

UKRAINE_REGIONS = [
    "Винницкая область",
    "Волынская область",
    "Днепропетровская область",
    "Донецкая область",
    "Житомирская область",
    "Закарпатская область",
    "Запорожская область",
    "Ивано-Франковская область",
    "Киевская область",
    "Кировоградская область",
    "Луганская область",
    "Львовская область",
    "Николаевская область",
    "Одесская область",
    "Полтавская область",
    "Ровенская область",
    "Сумская область",
    "Тернопольская область",
    "Харьковская область",
    "Херсонская область",
    "Хмельницкая область",
    "Черкасская область",
    "Черниговская область",
    "Черновицкая область",
]
UKRAINE_REGIONS_SET: Set[str] = set(UKRAINE_REGIONS)

LANG_ORDER = [
    ("uk", "🇺🇦 Українська"),
    ("en", "🇬🇧 English"),
    ("de", "🇩🇪 Deutsch"),
    ("pl", "🇵🇱 Polski"),
    ("ru", "🇷🇺 Русский"),
]
LANG_LABELS = {code: label for code, label in LANG_ORDER}

DEFAULT_LANG = "uk"

TEXTS: Dict[str, Dict[str, str]] = {
    "ANCHOR_NO_PROJECT": {
        "uk": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n🔍 Активний проєкт ще не обрано.\nПопросіть адміністратора активувати об'єкт, щоб відкрити робочі розділи.\n\n📋 <b>Меню дій</b>\nСкористайтеся кнопками нижче, щоб переглянути доступні можливості.",
        "en": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n🔍 No active project has been selected yet.\nAsk an administrator to activate a project to unlock the working sections.\n\n📋 <b>Actions</b>\nUse the buttons below to explore the available features.",
        "de": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n🔍 Es ist derzeit kein aktives Projekt ausgewählt.\nBitten Sie einen Administrator, ein Projekt zu aktivieren, um die Arbeitsbereiche zu öffnen.\n\n📋 <b>Aktionen</b>\nVerwenden Sie die Schaltflächen unten, um die verfügbaren Funktionen zu erkunden.",
        "pl": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n🔍 Aktywny projekt nie został jeszcze wybrany.\nPoproś administratora o aktywację obiektu, aby odblokować sekcje robocze.\n\n📋 <b>Menu działań</b>\nSkorzystaj z przycisków poniżej, aby zobaczyć dostępne funkcje.",
        "ru": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n🔍 Активный объект пока не выбран.\nПопросите администратора включить проект, чтобы открыть рабочие разделы.\n\n📋 <b>Меню действий</b>\nИспользуйте кнопки ниже, чтобы изучить доступные возможности.",
    },
    "ANCHOR_PROJECT": {
        "uk": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n📂 <b>{name}</b>\n🆔 Код проєкту: {code}\n🌍 Область: {region}\n📍 Локація: {location}\n🖼 Фотоархів: <b>{photos}</b> шт.\n🗓 Період робіт: {start} → {end}\n{bsg_section}\n{alerts_section}\n\n📋 <b>Меню дій</b>\nОберіть потрібний розділ нижче, щоб додати чек, переглянути документи або перевірити фінанси.",
        "en": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n📂 <b>{name}</b>\n🆔 Project code: {code}\n🌍 Oblast: {region}\n📍 Location: {location}\n🖼 Photo archive: <b>{photos}</b> items\n🗓 Work period: {start} → {end}\n{bsg_section}\n{alerts_section}\n\n📋 <b>Actions</b>\nChoose the section below to add receipts, open documents, or review finance details.",
        "de": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n📂 <b>{name}</b>\n🆔 Projektcode: {code}\n🌍 Oblast: {region}\n📍 Standort: {location}\n🖼 Fotoarchiv: <b>{photos}</b> Elemente\n🗓 Arbeitszeitraum: {start} → {end}\n{bsg_section}\n{alerts_section}\n\n📋 <b>Aktionen</b>\nWählen Sie unten einen Bereich, um Belege hinzuzufügen, Dokumente zu öffnen oder Finanzdaten einzusehen.",
        "pl": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n📂 <b>{name}</b>\n🆔 Kod projektu: {code}\n🌍 Obwód: {region}\n📍 Lokalizacja: {location}\n🖼 Archiwum zdjęć: <b>{photos}</b> szt.\n🗓 Okres prac: {start} → {end}\n{bsg_section}\n{alerts_section}\n\n📋 <b>Menu działań</b>\nWybierz sekcję poniżej, aby dodać paragon, otworzyć dokumenty lub sprawdzić finanse.",
        "ru": "🏗 <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\n📂 <b>{name}</b>\n🆔 Код проекта: {code}\n🌍 Область: {region}\n📍 Локация: {location}\n🖼 Фотоархив: <b>{photos}</b> шт.\n🗓 Период работ: {start} → {end}\n{bsg_section}\n{alerts_section}\n\n📋 <b>Меню действий</b>\nВыберите нужный раздел ниже, чтобы добавить чек, открыть документы или проверить финансы.",
    },
    "ANCHOR_PROJECT_BSG_SUMMARY": {
        "uk": "🏢 Посилки BSG: усього — <b>{total}</b> • забрати — <b>{pending}</b> • отримано — <b>{delivered}</b>",
        "en": "🏢 BSG parcels: total — <b>{total}</b> • to collect — <b>{pending}</b> • received — <b>{delivered}</b>",
        "de": "🏢 BSG-Sendungen: gesamt — <b>{total}</b> • abzuholen — <b>{pending}</b> • erhalten — <b>{delivered}</b>",
        "pl": "🏢 Przesyłki BSG: łącznie — <b>{total}</b> • do odebrania — <b>{pending}</b> • odebrano — <b>{delivered}</b>",
        "ru": "🏢 Посылки BSG: всего — <b>{total}</b> • забрать — <b>{pending}</b> • получено — <b>{delivered}</b>",
    },
    "ANCHOR_ALERT_SUMMARY": {
        "uk": "🇺🇦 Активні тривоги: <b>{count}</b> областей",
        "en": "🇺🇦 Active alerts: <b>{count}</b> oblasts",
        "de": "🇺🇦 Aktive Alarme: <b>{count}</b> Oblasten",
        "pl": "🇺🇦 Aktywne alarmy: <b>{count}</b> obwodów",
        "ru": "🇺🇦 Активные тревоги: <b>{count}</b> областей",
    },
    "ANCHOR_ALERT_ACTIVE": {
        "uk": "🚨 Тривога у <b>{region}</b> • {type} • від {start} • {severity}",
        "en": "🚨 Alert for <b>{region}</b> oblast • {type} • since {start} • {severity}",
        "de": "🚨 Alarm für Oblast <b>{region}</b> • {type} • seit {start} • {severity}",
        "pl": "🚨 Alarm w obwodzie <b>{region}</b> • {type} • od {start} • {severity}",
        "ru": "🚨 Тревога в <b>{region}</b> • {type} • с {start} • {severity}",
    },
    "ANCHOR_ALERT_RECENT": {
        "uk": "🟡 Остання тривога у <b>{region}</b> • {type} • {start} → {end}",
        "en": "🟡 Last alert for <b>{region}</b> oblast • {type} • {start} → {end}",
        "de": "🟡 Letzter Alarm für Oblast <b>{region}</b> • {type} • {start} → {end}",
        "pl": "🟡 Ostatni alarm w obwodzie <b>{region}</b> • {type} • {start} → {end}",
        "ru": "🟡 Последняя тревога в <b>{region}</b> • {type} • {start} → {end}",
    },
    "ANCHOR_ALERT_CALM": {
        "uk": "🟢 В області <b>{region}</b> відбій тривоги.",
        "en": "🟢 <b>{region}</b> oblast — alert cleared.",
        "de": "🟢 In der Oblast <b>{region}</b> wurde der Alarm aufgehoben.",
        "pl": "🟢 W obwodzie <b>{region}</b> alarm odwołano.",
        "ru": "🟢 В области <b>{region}</b> отбой тревоги.",
    },
    "ANCHOR_ALERT_CAUSE": {
        "uk": "🎯 Причина: {cause}",
        "en": "🎯 Cause: {cause}",
        "de": "🎯 Ursache: {cause}",
        "pl": "🎯 Przyczyna: {cause}",
        "ru": "🎯 Причина: {cause}",
    },
    "ANCHOR_ALERT_DETAILS": {
        "uk": "🔎 Деталі: {details}",
        "en": "🔎 Details: {details}",
        "de": "🔎 Details: {details}",
        "pl": "🔎 Szczegóły: {details}",
        "ru": "🔎 Детали: {details}",
    },
    "ANCHOR_ALERT_LOCATION": {
        "uk": "📍 Локація: {location}",
        "en": "📍 Location: {location}",
        "de": "📍 Ort: {location}",
        "pl": "📍 Lokalizacja: {location}",
        "ru": "📍 Локация: {location}",
    },
    "ANCHOR_ALERT_COORDS": {
        "uk": "🧭 Координати: {coords}",
        "en": "🧭 Coordinates: {coords}",
        "de": "🧭 Koordinaten: {coords}",
        "pl": "🧭 Współrzędne: {coords}",
        "ru": "🧭 Координаты: {coords}",
    },
    "BTN_CHECKS": {
        "uk": "🧾 Чеки",
        "en": "🧾 Receipts",
        "de": "🧾 Belege",
        "pl": "🧾 Paragony",
        "ru": "🧾 Чеки",
    },
    "BTN_DOCUMENTS": {
        "uk": "📑 Документи",
        "en": "📑 Documents",
        "de": "📑 Dokumente",
        "pl": "📑 Dokumenty",
        "ru": "📑 Документы",
    },
    "BTN_FINANCE": {
        "uk": "💵 Фінанси",
        "en": "💵 Finance",
        "de": "💵 Finanzen",
        "pl": "💵 Finanse",
        "ru": "💵 Финансы",
    },
    "BTN_ALERTS": {
        "uk": "🚨 Тривоги",
        "en": "🚨 Alerts",
        "de": "🚨 Alarme",
        "pl": "🚨 Alarmy",
        "ru": "🚨 Тревоги",
    },
    "BTN_SOS": {
        "uk": "🆘 SOS",
        "en": "🆘 SOS",
        "de": "🆘 SOS",
        "pl": "🆘 SOS",
        "ru": "🆘 SOS",
    },
    "BTN_ADMIN": {
        "uk": "🧑‍💼 Адмін-панель",
        "en": "🧑‍💼 Admin panel",
        "de": "🧑‍💼 Adminbereich",
        "pl": "🧑‍💼 Panel administratora",
        "ru": "🧑‍💼 Админ-панель",
    },
    "BTN_ABOUT": {
        "uk": "ℹ️ Про бота",
        "en": "ℹ️ About",
        "de": "ℹ️ Über den Bot",
        "pl": "ℹ️ O bocie",
        "ru": "ℹ️ О боте",
    },
    "ALERTS_MENU_INTRO": {
        "uk": "🚨 <b>Повітряні тривоги</b>\n━━━━━━━━━━━━━━━━━━\nПереглядайте активні сигнали, історію та керуйте областями сповіщень.\nВиберіть дію нижче.",
        "en": "🚨 <b>Air alerts</b>\n━━━━━━━━━━━━━━━━━━\nReview active warnings, browse history, and manage the regions you follow.\nChoose an option below.",
        "de": "🚨 <b>Luftalarme</b>\n━━━━━━━━━━━━━━━━━━\nSehen Sie aktive Warnungen, den Verlauf und verwalten Sie Ihre Regionen.\nWählen Sie eine Aktion unten.",
        "pl": "🚨 <b>Alarmy powietrzne</b>\n━━━━━━━━━━━━━━━━━━\nPrzeglądaj aktywne ostrzeżenia, historię i zarządzaj regionami powiadomień.\nWybierz działanie poniżej.",
        "ru": "🚨 <b>Воздушные тревоги</b>\n━━━━━━━━━━━━━━━━━━\nПросматривайте активные сигналы, историю и управляйте регионами уведомлений.\nВыберите действие ниже.",
    },
    "ALERTS_BTN_ACTIVE": {
        "uk": "🚨 Активні сигнали",
        "en": "🚨 Live alerts",
        "de": "🚨 Live-Alarme",
        "pl": "🚨 Aktywne sygnały",
        "ru": "🚨 Активные сигналы",
    },
    "ALERTS_BTN_OVERVIEW": {
        "uk": "🗺️ Карта областей",
        "en": "🗺️ Oblast map",
        "de": "🗺️ Oblast-Karte",
        "pl": "🗺️ Mapa obwodów",
        "ru": "🗺️ Карта областей",
    },
    "ALERTS_BTN_HISTORY": {
        "uk": "🕓 Журнал тривог",
        "en": "🕓 Alert log",
        "de": "🕓 Alarmprotokoll",
        "pl": "🕓 Dziennik alarmów",
        "ru": "🕓 Журнал тревог",
    },
    "ALERTS_BTN_SUBSCRIPTIONS": {
        "uk": "🎛️ Мої області",
        "en": "🎛️ My oblasts",
        "de": "🎛️ Meine Oblaste",
        "pl": "🎛️ Moje obwody",
        "ru": "🎛️ Мои области",
    },
    "ALERTS_ACTIVE_HEADER": {
        "uk": "🚨Активні сигнали({count})",
        "en": "🚨Active alerts({count})",
        "de": "🚨Aktive Alarme({count})",
        "pl": "🚨Aktywne alarmy({count})",
        "ru": "🚨Активные сигналы({count})",
    },
    "ALERTS_ACTIVE_DIVIDER": {
        "uk": "━━━━━━━━━━━━━━━━━━",
        "en": "━━━━━━━━━━━━━━━━━━",
        "de": "━━━━━━━━━━━━━━━━━━",
        "pl": "━━━━━━━━━━━━━━━━━━",
        "ru": "━━━━━━━━━━━━━━━━━━",
    },
    "ALERTS_ACTIVE_SUMMARY_TOTAL": {
        "uk": "📍 Активні тривоги: {count}",
        "en": "📍 Active alerts: {count}",
        "de": "📍 Aktive Alarme: {count}",
        "pl": "📍 Aktywne alarmy: {count}",
        "ru": "📍 Активные тревоги: {count}",
    },
    "ALERTS_ACTIVE_SUMMARY_USER": {
        "uk": "👤 Ваші вибрані області — персональні налаштування сповіщень",
        "en": "👤 Your selected oblasts — personal alert preferences",
        "de": "👤 Ihre ausgewählten Oblaste – persönliche Alarm-Einstellungen",
        "pl": "👤 Twoje wybrane obwody — osobiste ustawienia alertów",
        "ru": "👤 Ваши выбранные области — персональные настройки оповещений",
    },
    "ALERTS_ACTIVE_SUMMARY_PROJECT": {
        "uk": "🏗 Прив’язано до об’єкта — області, визначені адміністратором",
        "en": "🏗 Project scope — oblasts defined by the administrator",
        "de": "🏗 Projektbezug – Oblaste, die vom Administrator festgelegt wurden",
        "pl": "🏗 Powiązano z obiektem — obwody określone przez administratora",
        "ru": "🏗 Привязано к объекту — области, определённые администратором",
    },
    "ALERTS_HISTORY_HEADER": {
        "uk": "📜 <b>Історія тривог</b> ({count})",
        "en": "📜 <b>Alert history</b> ({count})",
        "de": "📜 <b>Alarmverlauf</b> ({count})",
        "pl": "📜 <b>Historia alarmów</b> ({count})",
        "ru": "📜 <b>История тревог</b> ({count})",
    },
    "ALERTS_OVERVIEW_HEADER": {
        "uk": "🗺️ Статус областей України\n━━━━━━━━━━━━━━━━━━\nПеревірте, де зараз лунає тривога.",
        "en": "🗺️ Status of Ukraine's oblasts\n━━━━━━━━━━━━━━━━━━\nSee where alerts are sounding right now.",
        "de": "🗺️ Status der Oblaste der Ukraine\n━━━━━━━━━━━━━━━━━━\nPrüfen Sie, wo gerade Alarm ausgelöst wird.",
        "pl": "🗺️ Status obwodów Ukrainy\n━━━━━━━━━━━━━━━━━━\nSprawdź, gdzie trwa alarm.",
        "ru": "🗺️ Статус областей Украины\n━━━━━━━━━━━━━━━━━━\nПроверяйте, где сейчас звучит тревога.",
    },
    "ALERTS_OVERVIEW_UPDATED": {
        "uk": "🔄 Оновлено: {time}",
        "en": "🔄 Updated: {time}",
        "de": "🔄 Aktualisiert: {time}",
        "pl": "🔄 Zaktualizowano: {time}",
        "ru": "🔄 Обновлено: {time}",
    },
    "ALERTS_OVERVIEW_GUIDE": {
        "uk": "ℹ️ Інструкція:\n🟢 Час показує відбій тривоги.\n🔴 Час показує початок тривоги.",
        "en": "ℹ️ Guide:\n🟢 Time marks when the alert ended.\n🔴 Time marks when the alert began.",
        "de": "ℹ️ Hinweis:\n🟢 Die Uhrzeit zeigt das Ende des Alarms.\n🔴 Die Uhrzeit zeigt den Beginn des Alarms.",
        "pl": "ℹ️ Instrukcja:\n🟢 Czas oznacza odwołanie alarmu.\n🔴 Czas oznacza początek alarmu.",
        "ru": "ℹ️ Инструкция:\n🟢 Время = отбой тревоги\n🔴 Время = начало тревоги",
    },
    "ALERTS_OVERVIEW_ACTIVE": {
        "uk": "🔴 {region} — тривога з {start}",
        "en": "🔴 {region} — alert since {start}",
        "de": "🔴 {region} — Alarm seit {start}",
        "pl": "🔴 {region} — alarm od {start}",
        "ru": "🔴 {region} — тревога с {start}",
    },
    "ALERTS_OVERVIEW_ACTIVE_UNKNOWN": {
        "uk": "🔴 {region} — тривога (час уточнюється)",
        "en": "🔴 {region} — alert (start time pending)",
        "de": "🔴 {region} — Alarm (Startzeit wird ermittelt)",
        "pl": "🔴 {region} — alarm (czas ustalany)",
        "ru": "🔴 {region} — тревога (время уточняется)",
    },
    "ALERTS_OVERVIEW_CALM": {
        "uk": "🟢 {region} — відбій тривоги",
        "en": "🟢 {region} — alert cleared",
        "de": "🟢 {region} — Alarm aufgehoben",
        "pl": "🟢 {region} — alarm odwołано",
        "ru": "🟢 {region} — отбой тревоги",
    },
    "ALERTS_NO_ACTIVE": {
        "uk": "✅ Зараз немає активних тривог для вибраних областей.",
        "en": "✅ There are no active alerts for your selected regions right now.",
        "de": "✅ Für die ausgewählten Regionen gibt es derzeit keine aktiven Alarme.",
        "pl": "✅ Brak aktywnych alarmów dla wybranych regionów.",
        "ru": "✅ Для выбранных регионов сейчас нет активных тревог.",
    },
    "ALERTS_NO_HISTORY": {
        "uk": "ℹ️ Поки що немає збереженої історії для цих областей.",
        "en": "ℹ️ There is no saved history for these regions yet.",
        "de": "ℹ️ Für diese Regionen gibt es noch keine gespeicherte Historie.",
        "pl": "ℹ️ Brak zapisanej historii dla tych regionów.",
        "ru": "ℹ️ Пока нет сохранённой истории для этих регионов.",
    },
    "ALERTS_NO_REGIONS": {
        "uk": "⚠️ Спершу оберіть хоча б одну область для сповіщень.",
        "en": "⚠️ Please select at least one region to receive alerts.",
        "de": "⚠️ Wählen Sie zunächst mindestens eine Region für Benachrichtigungen aus.",
        "pl": "⚠️ Wybierz co najmniej jeden region, aby otrzymywać alerty.",
        "ru": "⚠️ Сначала выберите хотя бы один регион для уведомлений.",
    },
    "ALERTS_SUBS_HEADER": {
        "uk": "🧭 Області сповіщень",
        "en": "🧭 Alert oblasts",
        "de": "🧭 Alarmbezirke",
        "pl": "🧭 Obwody powiadomień",
        "ru": "🧭 Области тревог",
    },
    "ALERTS_SUBS_DIVIDER": {
        "uk": "━━━━━━━━━━━━━━━━━━",
        "en": "━━━━━━━━━━━━━━━━━━",
        "de": "━━━━━━━━━━━━━━━━━━",
        "pl": "━━━━━━━━━━━━━━━━━━",
        "ru": "━━━━━━━━━━━━━━━━━━",
    },
    "ALERTS_SUBS_NOTE_HAS_PROJECT": {
        "uk": "Основна область проєкту: {region}.  \nІнші області можна обрати вручну.  ",
        "en": "Project oblast: {region}.  \nYou can add other oblasts manually.  ",
        "de": "Projektbezirk: {region}.  \nWeitere Bezirke lassen sich manuell wählen.  ",
        "pl": "Obwód projektu: {region}.  \nPozostałe obwody możesz wybrać ręcznie.  ",
        "ru": "Область проекта: {region}.  \nДругие области можно выбрать вручную.  ",
    },
    "ALERTS_SUBS_NOTE_NO_PROJECT": {
        "uk": "Наразі активний проєкт не вибрано.  \nОбласті можна обрати вручну.  ",
        "en": "No active project is selected.  \nChoose oblasts manually.  ",
        "de": "Derzeit ist kein Projekt aktiv.  \nBezirke lassen sich manuell wählen.  ",
        "pl": "Żaden projekt nie jest aktywny.  \nObwody możesz wybrać ręcznie.  ",
        "ru": "Активный проект пока не выбран.  \nОбласти можно выбрать вручную.  ",
    },
    "ALERTS_SUBS_LIST_TITLE": {
        "uk": "📍 Активні області:",
        "en": "📍 Active oblasts:",
        "de": "📍 Aktive Bezirke:",
        "pl": "📍 Aktywne obwody:",
        "ru": "📍 Активные области:",
    },
    "ALERTS_SUBS_LIST_EMPTY": {
        "uk": "—",
        "en": "—",
        "de": "—",
        "pl": "—",
        "ru": "—",
    },
    "ALERTS_SUBS_MANAGE": {
        "uk": "➕➖ Керування списком через кнопки",
        "en": "➕➖ Manage the list with the buttons",
        "de": "➕➖ Liste über die Schaltflächen verwalten",
        "pl": "➕➖ Zarządzaj listą przy użyciu przycisków",
        "ru": "➕➖ Управляйте списком с помощью кнопок",
    },
    "ALERTS_SUBS_SELECTED": {
        "uk": "Активні області: {items}",
        "en": "Selected oblasts: {items}",
        "de": "Aktive Bezirke: {items}",
        "pl": "Aktywne obwody: {items}",
        "ru": "Выбранные области: {items}",
    },
    "ALERTS_SUBS_ADDED": {
        "uk": "✅ Додано область: {region}",
        "en": "✅ Region added: {region}",
        "de": "✅ Region hinzugefügt: {region}",
        "pl": "✅ Dodano region: {region}",
        "ru": "✅ Добавлен регион: {region}",
    },
    "ALERTS_SUBS_REMOVED": {
        "uk": "➖ Видалено область: {region}",
        "en": "➖ Region removed: {region}",
        "de": "➖ Region entfernt: {region}",
        "pl": "➖ Usunięto region: {region}",
        "ru": "➖ Регион удалён: {region}",
    },
    "ALERTS_SUBS_LOCKED": {
        "uk": "ℹ️ Цю область неможливо вимкнути.",
        "en": "ℹ️ This region cannot be disabled.",
        "de": "ℹ️ Diese Region kann nicht deaktiviert werden.",
        "pl": "ℹ️ Tego regionu nie można wyłączyć.",
        "ru": "ℹ️ Этот регион нельзя отключить.",
    },
    "ALERTS_SUBS_PAGE": {
        "uk": "📄 Сторінка {current}/{total}",
        "en": "📄 Page {current}/{total}",
        "de": "📄 Seite {current}/{total}",
        "pl": "📄 Strona {current}/{total}",
        "ru": "📄 Страница {current}/{total}",
    },
    "ALERTS_BACK_TO_MENU": {
        "uk": "⬅️ Меню тривог",
        "en": "⬅️ Alerts menu",
        "de": "⬅️ Alarm-Menü",
        "pl": "⬅️ Menu alarmów",
        "ru": "⬅️ Меню тревог",
    },
    "ALERTS_CLOSE_CARD": {
        "uk": "✖️ Закрити",
        "en": "✖️ Close",
        "de": "✖️ Schließen",
        "pl": "✖️ Zamknąć",
        "ru": "✖️ Закрыть",
    },
    "ALERTS_PUSH_OPEN": {
        "uk": "🔎 Відкрити детально",
        "en": "🔎 Open details",
        "de": "🔎 Details anzeigen",
        "pl": "🔎 Otwórz szczegóły",
        "ru": "🔎 Открыть детально",
    },
    "ALERTS_PUSH_COLLAPSE": {
        "uk": "⬆️ Згорнути",
        "en": "⬆️ Collapse",
        "de": "⬆️ Einklappen",
        "pl": "⬆️ Zwiń",
        "ru": "⬆️ Свернуть",
    },
    "ALERTS_PUSH_DELETE": {
        "uk": "🗑 Видалити повідомлення",
        "en": "🗑 Delete message",
        "de": "🗑 Nachricht löschen",
        "pl": "🗑 Usuń wiadomość",
        "ru": "🗑 Удалить сообщение",
    },
    "ALERTS_PUSH_HEADER_ALERT": {
        "uk": "🚨 ТРИВОГА | {region}",
        "en": "🚨 ALERT | {region}",
        "de": "🚨 ALARM | {region}",
        "pl": "🚨 ALARM | {region}",
        "ru": "🚨 ТРЕВОГА | {region}",
    },
    "ALERTS_PUSH_HEADER_STANDDOWN": {
        "uk": "🟢 ВІДБІЙ | {region}",
        "en": "🟢 CLEAR | {region}",
        "de": "🟢 ENTWARNUNG | {region}",
        "pl": "🟢 ODWOŁANIE | {region}",
        "ru": "🟢 ОТБОЙ | {region}",
    },
    "ALERTS_PUSH_SUMMARY_RUNNING": {
        "uk": "{icon} {type} • 🕒 {start} → {progress}",
        "en": "{icon} {type} • 🕒 {start} → {progress}",
        "de": "{icon} {type} • 🕒 {start} → {progress}",
        "pl": "{icon} {type} • 🕒 {start} → {progress}",
        "ru": "{icon} {type} • 🕒 {start} → {progress}",
    },
    "ALERTS_PUSH_SUMMARY_ENDED": {
        "uk": "{icon} {type} • 🕒 {start} → ✅ {ended}",
        "en": "{icon} {type} • 🕒 {start} → ✅ {ended}",
        "de": "{icon} {type} • 🕒 {start} → ✅ {ended}",
        "pl": "{icon} {type} • 🕒 {start} → ✅ {ended}",
        "ru": "{icon} {type} • 🕒 {start} → ✅ {ended}",
    },
    "ALERTS_PUSH_SUMMARY_LEAD_ALERT": {
        "uk": "🚨 Коротке сповіщення про тривогу",
        "en": "🚨 Quick alert notification",
        "de": "🚨 Kurze Alarmbenachrichtigung",
        "pl": "🚨 Krótkie powiadomienie o alarmie",
        "ru": "🚨 Короткое уведомление о тревоге",
    },
    "ALERTS_PUSH_SUMMARY_LEAD_STANDDOWN": {
        "uk": "",
        "en": "",
        "de": "",
        "pl": "",
        "ru": "",
    },
    "ALERTS_DURATION_RUNNING": {
        "uk": "триває {duration}",
        "en": "ongoing for {duration}",
        "de": "läuft seit {duration}",
        "pl": "trwa {duration}",
        "ru": "идёт {duration}",
    },
    "ALERTS_DURATION_COMPLETED": {
        "uk": "тривала {duration}",
        "en": "lasted {duration}",
        "de": "dauerte {duration}",
        "pl": "trwała {duration}",
        "ru": "длилась {duration}",
    },
    "ALERTS_DURATION_LESS_MINUTE": {
        "uk": "менше хвилини",
        "en": "less than a minute",
        "de": "unter einer Minute",
        "pl": "mniej niż minutę",
        "ru": "менее минуты",
    },
    "ALERTS_PUSH_DETAIL_TITLE_ALERT": {
        "uk": "🚨 ТРИВОГА — {region}",
        "en": "🚨 ALERT — {region}",
        "de": "🚨 ALARM — {region}",
        "pl": "🚨 ALARM — {region}",
        "ru": "🚨 ТРЕВОГА — {region}",
    },
    "ALERTS_PUSH_DETAIL_TITLE_STANDDOWN": {
        "uk": "🟢 ВІДБІЙ ТРИВОГИ — {region}",
        "en": "🟢 ALERT CLEARED — {region}",
        "de": "🟢 ENTWARNUNG — {region}",
        "pl": "🟢 ALARM ODWOŁANY — {region}",
        "ru": "🟢 ОТБОЙ ТРЕВОГИ — {region}",
    },
    "ALERTS_PUSH_DETAIL_TYPE": {
        "uk": "{icon} Тип загрози: {value}",
        "en": "{icon} Threat type: {value}",
        "de": "{icon} Bedrohungsart: {value}",
        "pl": "{icon} Rodzaj zagrożenia: {value}",
        "ru": "{icon} Тип угрозы: {value}",
    },
    "ALERTS_PUSH_DETAIL_START": {
        "uk": "🕒 Початок: {date} • {time}",
        "en": "🕒 Start: {date} • {time}",
        "de": "🕒 Beginn: {date} • {time}",
        "pl": "🕒 Początek: {date} • {time}",
        "ru": "🕒 Начало: {date} • {time}",
    },
    "ALERTS_PUSH_DETAIL_END_STANDDOWN": {
        "uk": "✅ Відбій: {date} • {time}",
        "en": "✅ Cleared: {date} • {time}",
        "de": "✅ Entwarnung: {date} • {time}",
        "pl": "✅ Odwołanie: {date} • {time}",
        "ru": "✅ Отбой: {date} • {time}",
    },
    "ALERTS_PUSH_DETAIL_DURATION": {
        "uk": "⏱ Тривалість: {duration}",
        "en": "⏱ Duration: {duration}",
        "de": "⏱ Dauer: {duration}",
        "pl": "⏱ Czas trwania: {duration}",
        "ru": "⏱ Длительность: {duration}",
    },
    "ALERTS_PUSH_DETAIL_STATS_HEADER": {
        "uk": "📊 Статистика на зараз",
        "en": "📊 Current statistics",
        "de": "📊 Aktuelle Statistik",
        "pl": "📊 Aktualne statystyki",
        "ru": "📊 Статистика на сейчас",
    },
    "ALERTS_PUSH_DETAIL_STATS_COUNTRY": {
        "uk": "• 🇺🇦 По Україні: {value}",
        "en": "• 🇺🇦 Across Ukraine: {value}",
        "de": "• 🇺🇦 In der Ukraine: {value}",
        "pl": "• 🇺🇦 W Ukrainie: {value}",
        "ru": "• 🇺🇦 По Украине: {value}",
    },
    "ALERTS_PUSH_DETAIL_STATS_REGION_ACTIVE": {
        "uk": "• 🏙 В області: {value}",
        "en": "• 🏙 In the oblast: {value}",
        "de": "• 🏙 In der Oblast: {value}",
        "pl": "• 🏙 W obwodzie: {value}",
        "ru": "• 🏙 В области: {value}",
    },
    "ALERTS_PUSH_DETAIL_STATS_REGION_CLEAR": {
        "uk": "• 🏙 В області: завершена",
        "en": "• 🏙 In the oblast: cleared",
        "de": "• 🏙 In der Oblast: beendet",
        "pl": "• 🏙 W obwodzie: zakończona",
        "ru": "• 🏙 В области: завершена",
    },
    "ALERTS_PUSH_DETAIL_RECOMMENDATIONS_HEADER": {
        "uk": "⚠️ Рекомендації",
        "en": "⚠️ Recommendations",
        "de": "⚠️ Empfehlungen",
        "pl": "⚠️ Zalecenia",
        "ru": "⚠️ Рекомендации",
    },
    "ALERTS_PUSH_DETAIL_STANDDOWN_HEADER": {
        "uk": "✅ Ситуацію стабілізовано",
        "en": "✅ Situation stabilised",
        "de": "✅ Lage stabilisiert",
        "pl": "✅ Sytuacja ustabilizowana",
        "ru": "✅ Ситуация стабилизировалась",
    },
    "ALERTS_PUSH_DETAIL_STANDDOWN_NOTE": {
        "uk": "Будьте уважні та стежте за новими сповіщеннями",
        "en": "Stay cautious and watch for new notifications",
        "de": "Bleiben Sie aufmerksam und verfolgen Sie neue Meldungen",
        "pl": "Bądź ostrożny i śledź nowe powiadomienia",
        "ru": "Будьте осторожны и следите за новыми уведомлениями",
    },
    "ALERTS_PUSH_DETAIL_FOOTER": {
        "uk": "✅ Бережіть себе!",
        "en": "✅ Stay safe!",
        "de": "✅ Bleiben Sie sicher!",
        "pl": "✅ Dbajcie o siebie!",
        "ru": "✅ Берегите себя!",
    },
    "ALERTS_NAV_PREV": {
        "uk": "◀️ Попередня",
        "en": "◀️ Previous",
        "de": "◀️ Zurück",
        "pl": "◀️ Poprzednia",
        "ru": "◀️ Назад",
    },
    "ALERTS_NAV_NEXT": {
        "uk": "▶️ Наступна",
        "en": "▶️ Next",
        "de": "▶️ Weiter",
        "pl": "▶️ Następna",
        "ru": "▶️ Далее",
    },
    "ALERTS_CARD_INDEX": {
        "uk": "{index} із {total}",
        "en": "{index} of {total}",
        "de": "{index} von {total}",
        "pl": "{index} z {total}",
        "ru": "{index} из {total}",
    },
    "BTN_SETTINGS": {
        "uk": "⚙️ Налаштування",
        "en": "⚙️ Settings",
        "de": "⚙️ Einstellungen",
        "pl": "⚙️ Ustawienia",
        "ru": "⚙️ Настройки",
    },
    "BTN_NOVA_POSHTA": {
        "uk": "📮 Нова пошта",
        "en": "📮 Nova Poshta",
        "de": "📮 Nova Poshta",
        "pl": "📮 Nova Poshta",
        "ru": "📮 Новая почта",
    },
    "BTN_NP_INTERFACE": {
        "uk": "📋 Інтерфейс",
        "en": "📋 Overview",
        "de": "📋 Übersicht",
        "pl": "📋 Interfejs",
        "ru": "📋 Интерфейс",
    },
    "BTN_NP_SEARCH": {
        "uk": "🔍 Пошук за ТТН",
        "en": "🔍 Search by TTN",
        "de": "🔍 Suche per TTN",
        "pl": "🔍 Szukaj po TTN",
        "ru": "🔍 Искать по ТТН",
    },
    "BTN_NP_HISTORY": {
        "uk": "🕓 Історія пошуку",
        "en": "🕓 Search history",
        "de": "🕓 Suchverlauf",
        "pl": "🕓 Historia wyszukiwań",
        "ru": "🕓 История поисков",
    },
    "BTN_NP_BOOKMARKS": {
        "uk": "⭐ Відзначені",
        "en": "⭐ Bookmarked",
        "de": "⭐ Markiert",
        "pl": "⭐ Oznaczone",
        "ru": "⭐ Отмеченные",
    },
    "BTN_NP_ASSIGNED": {
        "uk": "🏢 Посилки BSG",
        "en": "🏢 BSG parcels",
        "de": "🏢 BSG-Sendungen",
        "pl": "🏢 Przesyłки BSG",
        "ru": "🏢 Посылки BSG",
    },
    "BTN_NP_RECEIVED": {
        "uk": "📥 Отримані посилки BSG",
        "en": "📥 Received BSG parcels",
        "de": "📥 Erhaltene BSG-Sendungen",
        "pl": "📥 Odebrane przesyłki BSG",
        "ru": "📥 Полученные посылки BSG",
    },
    "BTN_NP_ASSIGN_SEND": {
        "uk": "📬 Закріпити ТТН за користувачем",
        "en": "📬 Assign TTN to user",
        "de": "📬 TTN einem Nutzer zuordnen",
        "pl": "📬 Przypisz TTN użytkownikowi",
        "ru": "📬 Закрепить ТТН за пользователем",
    },
    "BTN_PROFILE": {
        "uk": "➕ Долучитися",
        "en": "➕ Join workspace",
        "de": "➕ Beitreten",
        "pl": "➕ Dołączyć",
        "ru": "➕ Добавиться",
    },
    "BTN_PROFILE_EDIT": {
        "uk": "✏️ Редагувати профіль",
        "en": "✏️ Edit profile",
        "de": "✏️ Profil bearbeiten",
        "pl": "✏️ Edytuj profil",
        "ru": "✏️ Редактировать профиль",
    },
    "BTN_PROFILE_VIEW_PHOTO": {
        "uk": "👁 Переглянути фото",
        "en": "👁 View photo",
        "de": "👁 Foto anzeigen",
        "pl": "👁 Zobacz zdjęcie",
        "ru": "👁 Просмотреть фото",
    },
    "BTN_PROFILE_HIDE_PHOTO": {
        "uk": "📝 Повернутись до тексту",
        "en": "📝 Back to summary",
        "de": "📝 Zurück zur Übersicht",
        "pl": "📝 Wróć do podsumowania",
        "ru": "📝 Вернуться к тексту",
    },
    "BTN_PROFILE_DONE": {
        "uk": "✅ Готово",
        "en": "✅ Done",
        "de": "✅ Fertig",
        "pl": "✅ Gotowe",
        "ru": "✅ Готово",
    },
    "BTN_PROFILE_UPDATE_PHOTO": {
        "uk": "🖼 Оновити фото",
        "en": "🖼 Update photo",
        "de": "🖼 Foto aktualisieren",
        "pl": "🖼 Zaktualizuj zdjęcie",
        "ru": "🖼 Обновить фото",
    },
    "BTN_PROFILE_REMOVE_PHOTO": {
        "uk": "🗑 Видалити фото",
        "en": "🗑 Remove photo",
        "de": "🗑 Foto löschen",
        "pl": "🗑 Usuń zdjęcie",
        "ru": "🗑 Удалить фото",
    },
    "BTN_PROFILE_FIELD_LAST": {
        "uk": "1. Прізвище",
        "en": "1. Last name",
        "de": "1. Nachname",
        "pl": "1. Nazwisko",
        "ru": "1. Фамилия",
    },
    "BTN_PROFILE_FIELD_FIRST": {
        "uk": "2. Ім'я",
        "en": "2. First name",
        "de": "2. Vorname",
        "pl": "2. Imię",
        "ru": "2. Имя",
    },
    "BTN_PROFILE_FIELD_MIDDLE": {
        "uk": "3. По батькові",
        "en": "3. Patronymic",
        "de": "3. Vatersname",
        "pl": "3. Drugie imię",
        "ru": "3. Отчество",
    },
    "BTN_PROFILE_FIELD_BIRTHDATE": {
        "uk": "4. Дата народження",
        "en": "4. Birth date",
        "de": "4. Geburtsdatum",
        "pl": "4. Data urodzenia",
        "ru": "4. Дата рождения",
    },
    "BTN_PROFILE_FIELD_REGION": {
        "uk": "5. Область",
        "en": "5. Region",
        "de": "5. Region",
        "pl": "5. Obwód",
        "ru": "5. Область",
    },
    "BTN_PROFILE_FIELD_PHONE": {
        "uk": "6. Телефон",
        "en": "6. Phone",
        "de": "6. Telefon",
        "pl": "6. Telefon",
        "ru": "6. Телефон",
    },
    "BTN_PROFILE_CANCEL": {
        "uk": "⬅️ Скасувати",
        "en": "⬅️ Cancel",
        "de": "⬅️ Abbrechen",
        "pl": "⬅️ Anuluj",
        "ru": "⬅️ Отменить",
    },
    "BTN_BACK_ROOT": {
        "uk": "⬅️ На головну",
        "en": "⬅️ Main menu",
        "de": "⬅️ Hauptmenü",
        "pl": "⬅️ Menu główne",
        "ru": "⬅️ На главную",
    },
    "BTN_PHOTO_TIMELINE": {
        "uk": "🖼 Фотохронологія",
        "en": "🖼 Photo timeline",
        "de": "🖼 Fototimeline",
        "pl": "🖼 Kronika zdjęć",
        "ru": "🖼 Фотохронология",
    },
    "BTN_ADD_RECEIPT": {
        "uk": "📷 Додати чек",
        "en": "📷 Add receipt",
        "de": "📷 Beleg hinzufügen",
        "pl": "📷 Dodaj paragon",
        "ru": "📷 Добавить чек",
    },
    "BTN_MY_STATS": {
        "uk": "📊 Моя статистика",
        "en": "📊 My statistics",
        "de": "📊 Meine Statistik",
        "pl": "📊 Moje statystyki",
        "ru": "📊 Моя статистика",
    },
    "BTN_MY_RECEIPTS": {
        "uk": "📁 Мої чеки",
        "en": "📁 My receipts",
        "de": "📁 Meine Belege",
        "pl": "📁 Moje paragony",
        "ru": "📁 Мои чеки",
    },
    "BTN_RECEIPT_HISTORY": {
        "uk": "🗂 Історія чеків та виплат",
        "en": "🗂 Receipt & payout history",
        "de": "🗂 Beleg- und Auszahlungs­historie",
        "pl": "🗂 Historia paragonów i wypłat",
        "ru": "🗂 История чеков и выплат",
    },
    "BTN_UPLOAD_PHOTO": {
        "uk": "📤 Завантажити фото",
        "en": "📤 Upload photos",
        "de": "📤 Fotos hochladen",
        "pl": "📤 Wyślij zdjęcia",
        "ru": "📤 Загрузить фото",
    },
    "BTN_VIEW_OBJECT_PHOTOS": {
        "uk": "🖼 Переглянути фото об'єкта",
        "en": "🖼 View project photos",
        "de": "🖼 Projektfotos ansehen",
        "pl": "🖼 Zobacz zdjęcia obiektu",
        "ru": "🖼 Просмотреть фотографии объекта",
    },
    "STATUS_PAID": {
        "uk": "✅ Оплачено",
        "en": "✅ Paid",
        "de": "✅ Bezahlt",
        "pl": "✅ Opłacono",
        "ru": "✅ Оплачен",
    },
    "STATUS_UNPAID": {
        "uk": "❌ Не оплачено",
        "en": "❌ Unpaid",
        "de": "❌ Nicht bezahlt",
        "pl": "❌ Nie opłacono",
        "ru": "❌ Не оплачен",
    },
    "STATUS_UNKNOWN": {
        "uk": "⏳ Не вказано",
        "en": "⏳ Not specified",
        "de": "⏳ Nicht angegeben",
        "pl": "⏳ Nie podano",
        "ru": "⏳ Не указан",
    },
    "START_WELCOME_BACK": {
        "uk": "👋 Радий знову бачити, {name}!",
        "en": "👋 Welcome back, {name}!",
        "de": "👋 Willkommen zurück, {name}!",
        "pl": "👋 Miło Cię znów widzieć, {name}!",
        "ru": "👋 Рад снова видеть, {name}!",
    },
    "START_PROMPT_FULLNAME": {
        "uk": "👤 Введіть прізвище та ім'я (наприклад, Іваненко Іван).",
        "en": "👤 Please enter your full name (for example, Smith John).",
        "de": "👤 Bitte geben Sie Ihren vollständigen Namen ein (z. B. Müller Hans).",
        "pl": "👤 Wpisz swoje imię i nazwisko (np. Kowalski Jan).",
        "ru": "👤 Введите фамилию и имя (например, Иванов Иван).",
    },
    "START_FULLNAME_INVALID": {
        "uk": "❗ Будь ласка, вкажіть щонайменше ім'я та прізвище повністю.",
        "en": "❗ Please provide at least first and last name.",
        "de": "❗ Bitte geben Sie mindestens Vor- und Nachnamen vollständig an.",
        "pl": "❗ Podaj co najmniej imię i nazwisko.",
        "ru": "❗ Пожалуйста, укажите как минимум имя и фамилию полностью.",
    },
    "START_REQUEST_PHONE": {
        "uk": "📞 Натисніть «📱 Надіслати номер», щоб поділитися телефоном.",
        "en": "📞 Tap “📱 Share phone number” to send your contact.",
        "de": "📞 Tippen Sie auf „📱 Telefonnummer senden“, um Ihre Nummer zu teilen.",
        "pl": "📞 Kliknij „📱 Wyślij numer”, aby przesłać swój kontakt.",
        "ru": "📞 Нажмите «📱 Отправить номер», чтобы поделиться телефоном.",
    },
    "START_PHONE_ERROR": {
        "uk": "❗ Не вдалося отримати номер. Спробуйте ще раз кнопкою нижче.",
        "en": "❗ Failed to read the number. Please try again using the button below.",
        "de": "❗ Telefonnummer konnte nicht erkannt werden. Bitte versuchen Sie es erneut über die Schaltfläche unten.",
        "pl": "❗ Nie udało się pobrać numeru. Spróbuj ponownie przyciskiem poniżej.",
        "ru": "❗ Не удалось получить номер. Повторите попытку кнопкой ниже.",
    },
    "START_PROFILE_SAVED": {
        "uk": "✅ Профіль збережено. Ваш код: <b>{code}</b>\nЛаскаво просимо!",
        "en": "✅ Profile saved. Your code: <b>{code}</b>\nWelcome aboard!",
        "de": "✅ Profil gespeichert. Ihr Code: <b>{code}</b>\nWillkommen an Bord!",
        "pl": "✅ Profil zapisany. Twój kod: <b>{code}</b>\nWitamy na pokładzie!",
        "ru": "✅ Профиль сохранён. Ваш код: <b>{code}</b>\nДобро пожаловать!",
    },
    "ONBOARD_LANGUAGE_CONFIRMED": {
        "uk": "Мову переключено на {language}. Натисніть кнопку нижче, щоб продовжити реєстрацію.",
        "en": "Language switched to {language}. Tap the button below to continue your registration.",
        "de": "Sprache auf {language} umgestellt. Tippen Sie auf die Schaltfläche unten, um fortzufahren.",
        "pl": "Język zmieniono na {language}. Naciśnij przycisk poniżej, aby kontynuować rejestrację.",
        "ru": "Язык переключён на {language}. Нажмите кнопку ниже, чтобы продолжить регистрацию.",
    },
    "ONBOARD_WELCOME": {
        "uk": "👋 Привіт, {name}!\nЛаскаво просимо до робочого простору <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\nТут зібрані проєкти, фінанси, архіви документів і тривоги.",
        "en": "👋 Hello, {name}!\nWelcome to the <b>{bot}</b> workspace\n━━━━━━━━━━━━━━━━━━\nHere you will find projects, finance tools, document archives, and alert summaries.",
        "de": "👋 Hallo, {name}!\nWillkommen im Arbeitsbereich <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\nHier finden Sie Projekte, Finanzwerkzeuge, Dokumentarchive und Alarmübersichten.",
        "pl": "👋 Cześć, {name}!\nWitamy w przestrzeni roboczej <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\nTutaj znajdziesz projekty, finanse, archiwum dokumentów i podsumowania alarmów.",
        "ru": "👋 Привет, {name}!\nДобро пожаловать в рабочее пространство <b>{bot}</b>\n━━━━━━━━━━━━━━━━━━\nЗдесь собраны проекты, финансы, архив документов и сводки тревог.",
    },
    "ONBOARD_BRIEFING": {
        "uk": "🧭 Як пройти реєстрацію\n━━━━━━━━━━━━━━━━━━\n1. Напишіть прізвище.\n2. Вкажіть ім'я.\n3. За потреби додайте по батькові (можна пропустити).\n4. Введіть дату народження у форматі ДД.ММ.РРРР.\n5. Оберіть область проживання кнопкою.\n6. Надішліть номер телефону кнопкою «📱 Надіслати номер».\n7. Додайте паспортне фото або пропустіть, якщо поки нема.\n\nГотові? Натисніть кнопку, щоб почати.",
        "en": "🧭 How to complete registration\n━━━━━━━━━━━━━━━━━━\n1. Enter your last name.\n2. Provide your first name.\n3. Add a patronymic/middle name (optional).\n4. Submit your birth date in DD.MM.YYYY format.\n5. Pick your home region using the button.\n6. Send your phone number via the “📱 Share phone number” button.\n7. Upload a passport-style photo or skip for now.\n\nReady? Tap the button to begin.",
        "de": "🧭 So schließen Sie die Registrierung ab\n━━━━━━━━━━━━━━━━━━\n1. Nachnamen eingeben.\n2. Vornamen angeben.\n3. Vatersnamen/Zweitnamen ergänzen (optional).\n4. Geburtsdatum im Format TT.MM.JJJJ eingeben.\n5. Region per Schaltfläche auswählen.\n6. Telefonnummer über „📱 Nummer senden“ teilen.\n7. Ein Passfoto hochladen oder vorerst überspringen.\n\nBereit? Tippen Sie auf die Schaltfläche zum Start.",
        "pl": "🧭 Jak ukończyć rejestrację\n━━━━━━━━━━━━━━━━━━\n1. Podaj nazwisko.\n2. Wpisz imię.\n3. Dodaj imię ojca / drugie imię (opcjonalnie).\n4. Wprowadź datę urodzenia w formacie DD.MM.RRRR.\n5. Wybierz region zamieszkania przyciskiem.\n6. Wyślij numer telefonu przyciskiem „📱 Wyślij numer”.\n7. Dodaj zdjęcie paszportowe lub pomiń na razie.\n\nGotowe? Kliknij przycisk, aby rozpocząć.",
        "ru": "🧭 Как пройти регистрацию\n━━━━━━━━━━━━━━━━━━\n1. Укажите фамилию.\n2. Напишите имя.\n3. При необходимости добавьте отчество (можно пропустить).\n4. Введите дату рождения в формате ДД.ММ.ГГГГ.\n5. Выберите область проживания через кнопку.\n6. Отправьте номер телефона кнопкой «📱 Отправить номер».\n7. Загрузите фото в формате паспортного или пропустите пока.\n\nГотовы? Нажмите кнопку, чтобы начать.",
    },
    "ONBOARD_RETURNING_SHORTCUT": {
        "uk": "Виглядає, що ви вже проходили реєстрацію. Натисніть кнопку, щоб перейти до головного меню або продовжити оновлення даних.",
        "en": "It seems you have already completed registration. Use the button to open the main menu or continue updating your data.",
        "de": "Sie scheinen bereits registriert zu sein. Nutzen Sie die Schaltfläche, um zum Hauptmenü zu wechseln oder Ihre Daten zu aktualisieren.",
        "pl": "Wygląda na to, że rejestracja została już zakończona. Użyj przycisku, aby przejść do menu głównego lub zaktualizować dane.",
        "ru": "Похоже, вы уже проходили регистрацию. Нажмите кнопку, чтобы открыть главное меню или продолжить обновление данных.",
    },
    "REGISTER_INTRO_PROMPT": {
        "uk": "Починаємо! Відповідайте на запитання нижче. Всі допоміжні повідомлення будуть прибрані автоматично.",
        "en": "Let's get started! Answer the questions below — helper messages will be cleaned up automatically.",
        "de": "Los geht's! Beantworten Sie die folgenden Fragen – Hilfsnachrichten werden automatisch entfernt.",
        "pl": "Zaczynamy! Odpowiedz na poniższe pytania – pomocnicze wiadomości zostaną usunięte automatycznie.",
        "ru": "Начинаем! Отвечайте на вопросы ниже — служебные сообщения будут удалены автоматически.",
    },
    "REGISTER_LAST_NAME_PROMPT": {
        "uk": "1. Вкажіть ваше прізвище.",
        "en": "1. Please enter your last name.",
        "de": "1. Geben Sie Ihren Nachnamen ein.",
        "pl": "1. Podaj swoje nazwisko.",
        "ru": "1. Укажите вашу фамилию.",
    },
    "REGISTER_LAST_NAME_WARN": {
        "uk": "❗ Перевірте прізвище: допускаються лише літери, апостроф та дефіс.",
        "en": "❗ Please check your last name: only letters, apostrophes, and dashes are allowed.",
        "de": "❗ Prüfen Sie den Nachnamen: Es sind nur Buchstaben, Apostroph und Bindestrich erlaubt.",
        "pl": "❗ Sprawdź nazwisko: dozwolone są tylko litery, apostrof i myślnik.",
        "ru": "❗ Проверьте фамилию: допустимы только буквы, апостроф и дефис.",
    },
    "REGISTER_FIRST_NAME_PROMPT": {
        "uk": "2. Вкажіть ваше ім'я.",
        "en": "2. Please enter your first name.",
        "de": "2. Geben Sie Ihren Vornamen ein.",
        "pl": "2. Podaj swoje imię.",
        "ru": "2. Укажите ваше имя.",
    },
    "REGISTER_FIRST_NAME_WARN": {
        "uk": "❗ Перевірте ім'я: допускаються лише літери, апостроф та дефіс.",
        "en": "❗ Please check your first name: only letters, apostrophes, and dashes are allowed.",
        "de": "❗ Prüfen Sie den Vornamen: Es sind nur Buchstaben, Apostroph und Bindestrich erlaubt.",
        "pl": "❗ Sprawdź imię: dozwolone są tylko litery, apostrof i myślnik.",
        "ru": "❗ Проверьте имя: допустимы только буквы, апостроф и дефис.",
    },
    "REGISTER_MIDDLE_NAME_PROMPT": {
        "uk": "3. Вкажіть по батькові (якщо є). Якщо немає — напишіть «немає».",
        "en": "3. Enter your patronymic or middle name (optional). Type “no” if you wish to skip.",
        "de": "3. Geben Sie Ihren Vatersnamen/Zweitnamen an (optional). Schreiben Sie „kein“, um zu überspringen.",
        "pl": "3. Podaj drugie imię/imię ojca (opcjonalnie). Wpisz „brak”, aby pominąć.",
        "ru": "3. Укажите отчество (если есть). Если нет — напишите «нет».",
    },
    "REGISTER_MIDDLE_NAME_WARN": {
        "uk": "❗ Перевірте по батькові: допускаються лише літери, апостроф та дефіс.",
        "en": "❗ Please check the patronymic: only letters, apostrophes, and dashes are allowed.",
        "de": "❗ Prüfen Sie den Zweitnamen: Es sind nur Buchstaben, Apostroph und Bindestrich erlaubt.",
        "pl": "❗ Sprawdź drugie imię: dozwolone są tylko litery, apostrof i myślnik.",
        "ru": "❗ Проверьте отчество: допустимы только буквы, апостроф и дефис.",
    },
    "REGISTER_BIRTHDATE_PROMPT": {
        "uk": "4. Вкажіть дату народження у форматі ДД.ММ.РРРР.",
        "en": "4. Please enter your birth date in DD.MM.YYYY format.",
        "de": "4. Geben Sie Ihr Geburtsdatum im Format TT.MM.JJJJ ein.",
        "pl": "4. Wpisz datę urodzenia w formacie DD.MM.RRRR.",
        "ru": "4. Укажите дату рождения в формате ДД.ММ.ГГГГ.",
    },
    "REGISTER_BIRTHDATE_WARN": {
        "uk": "❗ Не вдалося розпізнати дату. Використовуйте формат ДД.ММ.РРРР.",
        "en": "❗ Could not parse the date. Please use DD.MM.YYYY format.",
        "de": "❗ Datum konnte nicht erkannt werden. Bitte verwenden Sie das Format TT.MM.JJJJ.",
        "pl": "❗ Nie udało się rozpoznać daty. Użyj formatu DD.MM.RRRR.",
        "ru": "❗ Не удалось распознать дату. Используйте формат ДД.ММ.ГГГГ.",
    },
    "REGISTER_REGION_PROMPT": {
        "uk": "5. Оберіть область проживання кнопкою нижче.",
        "en": "5. Pick your home region using the button below.",
        "de": "5. Wählen Sie Ihre Region über die Schaltfläche unten.",
        "pl": "5. Wybierz region zamieszkania przyciskiem poniżej.",
        "ru": "5. Выберите область проживания кнопкой ниже.",
    },
    "REGISTER_REGION_BUTTON": {
        "uk": "📍 Вибрати область",
        "en": "📍 Choose region",
        "de": "📍 Region wählen",
        "pl": "📍 Wybierz region",
        "ru": "📍 Выбрать область",
    },
    "REGISTER_REGION_PICK": {
        "uk": "Оберіть область зі списку:",
        "en": "Select your region from the list:",
        "de": "Wählen Sie Ihre Region aus der Liste:",
        "pl": "Wybierz region z listy:",
        "ru": "Выберите область из списка:",
    },
    "REGISTER_REGION_REMIND": {
        "uk": "Будь ласка, скористайтеся кнопкою «📍 Вибрати область».",
        "en": "Please use the “📍 Choose region” button.",
        "de": "Bitte verwenden Sie die Schaltfläche „📍 Region wählen“.",
        "pl": "Użyj przycisku „📍 Wybierz region”.",
        "ru": "Пожалуйста, воспользуйтесь кнопкой «📍 Выбрать область».",
    },
    "REGISTER_REGION_SELECTED": {
        "uk": "Область <b>{region}</b> збережено. Натисніть кнопку, щоб продовжити.",
        "en": "Region <b>{region}</b> saved. Tap the button to continue.",
        "de": "Region <b>{region}</b> gespeichert. Tippen Sie auf die Schaltfläche zum Fortfahren.",
        "pl": "Region <b>{region}</b> zapisano. Naciśnij przycisk, aby kontynuować.",
        "ru": "Область <b>{region}</b> сохранена. Нажмите кнопку, чтобы продолжить.",
    },
    "REGISTER_PHONE_PROMPT_NEW": {
        "uk": "6. Надішліть номер телефону кнопкою «📱 Надіслати номер».",
        "en": "6. Send your phone number using the “📱 Share phone number” button.",
        "de": "6. Senden Sie Ihre Telefonnummer über die Schaltfläche „📱 Nummer senden“.",
        "pl": "6. Wyślij numer telefonu przyciskiem „📱 Wyślij numer”.",
        "ru": "6. Отправьте номер телефона кнопкой «📱 Отправить номер».",
    },
    "REGISTER_PHONE_WARN": {
        "uk": "❗ Не вдалося зчитати контакт. Спробуйте ще раз кнопкою нижче.",
        "en": "❗ Could not read the contact. Please try again with the button below.",
        "de": "❗ Kontakt konnte nicht erkannt werden. Bitte erneut über die Schaltfläche unten versuchen.",
        "pl": "❗ Nie udało się pobrać kontaktu. Spróbuj ponownie przyciskiem poniżej.",
        "ru": "❗ Не удалось получить контакт. Повторите попытку кнопкой ниже.",
    },
    "REGISTER_PHONE_TEXT_WARN": {
        "uk": "📞 Скористайтеся кнопкою «📱 Надіслати номер», щоб поділитися телефоном.",
        "en": "📞 Please use the “📱 Share phone number” button to send your contact.",
        "de": "📞 Bitte nutzen Sie die Schaltfläche „📱 Nummer senden“, um Ihren Kontakt zu teilen.",
        "pl": "📞 Skorzystaj z przycisku „📱 Wyślij numer”, aby przesłać kontakt.",
        "ru": "📞 Пожалуйста, используйте кнопку «📱 Отправить номер», чтобы поделиться телефоном.",
    },
    "REGISTER_PHONE_OK": {
        "uk": "Контакт збережено ✅",
        "en": "Contact saved ✅",
        "de": "Kontakt gespeichert ✅",
        "pl": "Kontakt zapisany ✅",
        "ru": "Контакт сохранён ✅",
    },
    "REGISTER_PHOTO_PROMPT": {
        "uk": "7. Додайте паспортне фото.\n━━━━━━━━━━━━━━━━━━\n• фронтальний ракурс\n• нейтральний фон\n• рівне освітлення\nЯкщо поки що нема фото, натисніть «Пропустити».",
        "en": "7. Upload a passport-style photo.\n━━━━━━━━━━━━━━━━━━\n• frontal view\n• neutral background\n• even lighting\nIf you don't have one yet, tap “Skip”.",
        "de": "7. Laden Sie ein Passfoto hoch.\n━━━━━━━━━━━━━━━━━━\n• frontale Ansicht\n• neutraler Hintergrund\n• gleichmäßige Beleuchtung\nFalls Sie noch keines haben, tippen Sie auf „Überspringen“.",
        "pl": "7. Dodaj zdjęcie paszportowe.\n━━━━━━━━━━━━━━━━━━\n• ujęcie frontalne\n• neutralne tło\n• równomierne oświetlenie\nJeśli nie masz zdjęcia, wybierz „Pomiń”.",
        "ru": "7. Загрузите фото формата «на документы».\n━━━━━━━━━━━━━━━━━━\n• фронтальный ракурс\n• нейтральный фон\n• ровное освещение\nЕсли фото пока нет, нажмите «Пропустить».",
    },
    "REGISTER_PHOTO_WARN": {
        "uk": "❗ Не вдалося обробити фото. Спробуйте інший файл або зробіть нове фото.",
        "en": "❗ Could not process the photo. Please try another file or take a new picture.",
        "de": "❗ Foto konnte nicht verarbeitet werden. Bitte versuchen Sie eine andere Datei oder machen Sie ein neues Foto.",
        "pl": "❗ Nie udało się przetworzyć zdjęcia. Spróbuj innego pliku lub zrób nowe zdjęcie.",
        "ru": "❗ Не удалось обработать фото. Попробуйте другой файл или сделайте новое фото.",
    },
    "REGISTER_PHOTO_RECEIVED": {
        "uk": "Фото отримано ✅",
        "en": "Photo received ✅",
        "de": "Foto erhalten ✅",
        "pl": "Zdjęcie otrzymane ✅",
        "ru": "Фото получено ✅",
    },
    "REGISTER_PHOTO_SKIP_CONFIRM": {
        "uk": "Фото можна завантажити пізніше у розділі профілю.",
        "en": "You can upload a photo later from your profile section.",
        "de": "Sie können das Foto später im Profilbereich hochladen.",
        "pl": "Zdjęcie możesz dodać później w swoim profilu.",
        "ru": "Фото можно загрузить позже в разделе профиля.",
    },
    "REGISTER_FINISH_CONFIRM": {
        "uk": "Реєстрацію завершено. Ваш артикул BSG — <b>{code}</b>.",
        "en": "Registration completed. Your BSG article is <b>{code}</b>.",
        "de": "Registrierung abgeschlossen. Ihre BSG-Kennung lautet <b>{code}</b>.",
        "pl": "Rejestracja zakończona. Twój kod BSG to <b>{code}</b>.",
        "ru": "Регистрация завершена. Ваш артикул BSG — <b>{code}</b>.",
    },
    "PROFILE_HEADER": {
        "uk": "👤 <b>Мій профіль</b>",
        "en": "👤 <b>My profile</b>",
        "de": "👤 <b>Mein Profil</b>",
        "pl": "👤 <b>Mój profil</b>",
        "ru": "👤 <b>Мой профиль</b>",
    },
    "PROFILE_EDIT_HINT": {
        "uk": "✏️ Режим редагування активний. Оберіть поле нижче, щоб оновити дані.",
        "en": "✏️ Edit mode is active. Pick a field below to update your data.",
        "de": "✏️ Bearbeitungsmodus aktiv. Wählen Sie unten ein Feld, um die Daten zu aktualisieren.",
        "pl": "✏️ Tryb edycji jest aktywny. Wybierz pole poniżej, aby zaktualizować dane.",
        "ru": "✏️ Режим редактирования активен. Выберите поле ниже, чтобы обновить данные.",
    },
    "PROFILE_FIELD_LAST_NAME": {
        "uk": "Прізвище",
        "en": "Last name",
        "de": "Nachname",
        "pl": "Nazwisko",
        "ru": "Фамилия",
    },
    "PROFILE_FIELD_FIRST_NAME": {
        "uk": "Ім'я",
        "en": "First name",
        "de": "Vorname",
        "pl": "Imię",
        "ru": "Имя",
    },
    "PROFILE_FIELD_MIDDLE_NAME": {
        "uk": "По батькові",
        "en": "Patronymic",
        "de": "Vatersname",
        "pl": "Drugie imię",
        "ru": "Отчество",
    },
    "PROFILE_FIELD_BIRTHDATE": {
        "uk": "Дата народження",
        "en": "Birth date",
        "de": "Geburtsdatum",
        "pl": "Data urodzenia",
        "ru": "Дата рождения",
    },
    "PROFILE_FIELD_REGION": {
        "uk": "Область",
        "en": "Region",
        "de": "Region",
        "pl": "Obwód",
        "ru": "Область",
    },
    "PROFILE_FIELD_PHONE": {
        "uk": "Телефон",
        "en": "Phone",
        "de": "Telefon",
        "pl": "Telefon",
        "ru": "Телефон",
    },
    "PROFILE_FIELD_TG_ID": {
        "uk": "Telegram ID",
        "en": "Telegram ID",
        "de": "Telegram-ID",
        "pl": "ID Telegram",
        "ru": "Telegram ID",
    },
    "PROFILE_FIELD_TG_USERNAME": {
        "uk": "Telegram юзер",
        "en": "Telegram username",
        "de": "Telegram-Benutzername",
        "pl": "Nazwa w Telegramie",
        "ru": "Имя в Telegram",
    },
    "PROFILE_FIELD_BSU": {
        "uk": "BSU код",
        "en": "BSU code",
        "de": "BSU-Code",
        "pl": "Kod BSU",
        "ru": "BSU код",
    },
    "PROFILE_FIELD_PHOTO": {
        "uk": "Фото профілю",
        "en": "Profile photo",
        "de": "Profilfoto",
        "pl": "Zdjęcie profilu",
        "ru": "Фото профиля",
    },
    "PROFILE_VALUE_MISSING": {
        "uk": "—",
        "en": "—",
        "de": "—",
        "pl": "—",
        "ru": "—",
    },
    "PROFILE_PHOTO_STATUS_OK": {
        "uk": "завантажено",
        "en": "uploaded",
        "de": "hochgeladen",
        "pl": "dodano",
        "ru": "загружено",
    },
    "PROFILE_PHOTO_STATUS_MISSING": {
        "uk": "відсутнє",
        "en": "missing",
        "de": "fehlt",
        "pl": "brak",
        "ru": "отсутствует",
    },
    "PROFILE_PHOTO_STATUS_SKIPPED": {
        "uk": "поки пропущено",
        "en": "skipped for now",
        "de": "vorerst übersprungen",
        "pl": "pominięto",
        "ru": "пропущено пока",
    },
    "PROFILE_PROMPT_LAST_NAME": {
        "uk": "1. Вкажіть нове прізвище (лише літери).",
        "en": "1. Enter the new last name (letters only).",
        "de": "1. Geben Sie den neuen Nachnamen ein (nur Buchstaben).",
        "pl": "1. Podaj nowe nazwisko (tylko litery).",
        "ru": "1. Укажите новую фамилию (только буквы).",
    },
    "PROFILE_PROMPT_FIRST_NAME": {
        "uk": "2. Вкажіть нове ім'я (лише літери).",
        "en": "2. Enter the new first name (letters only).",
        "de": "2. Geben Sie den neuen Vornamen ein (nur Buchstaben).",
        "pl": "2. Podaj nowe imię (tylko litery).",
        "ru": "2. Укажите новое имя (только буквы).",
    },
    "PROFILE_PROMPT_MIDDLE_NAME": {
        "uk": "3. Введіть по батькові або напишіть «немає», щоб очистити поле.",
        "en": "3. Provide a patronymic or type “none” to clear the field.",
        "de": "3. Geben Sie den Vatersnamen an oder schreiben Sie „kein“, um das Feld zu leeren.",
        "pl": "3. Podaj drugie imię lub wpisz „brak”, aby wyczyścić pole.",
        "ru": "3. Введите отчество или напишите «нет», чтобы очистить поле.",
    },
    "PROFILE_PROMPT_BIRTHDATE": {
        "uk": "4. Вкажіть дату народження у форматі ДД.ММ.РРРР.",
        "en": "4. Enter the birth date in DD.MM.YYYY format.",
        "de": "4. Geben Sie das Geburtsdatum im Format TT.MM.JJJJ ein.",
        "pl": "4. Podaj datę urodzenia w formacie DD.MM.RRRR.",
        "ru": "4. Укажите дату рождения в формате ДД.ММ.ГГГГ.",
    },
    "PROFILE_PROMPT_REGION": {
        "uk": "5. Оберіть область проживання кнопкою нижче.",
        "en": "5. Choose your region using the button below.",
        "de": "5. Wählen Sie Ihre Region über die Schaltfläche unten.",
        "pl": "5. Wybierz swój obwód za pomocą przycisku poniżej.",
        "ru": "5. Выберите область проживания кнопкой ниже.",
    },
    "PROFILE_PROMPT_PHONE": {
        "uk": "6. Надішліть номер телефону через кнопку «📱 Надіслати номер».",
        "en": "6. Send your phone number via the “📱 Share phone number” button.",
        "de": "6. Senden Sie Ihre Telefonnummer über die Schaltfläche „📱 Nummer senden“.",
        "pl": "6. Wyślij numer telefonu przyciskiem „📱 Wyślij numer”.",
        "ru": "6. Отправьте номер телефона кнопкой «📱 Отправить номер».",
    },
    "PROFILE_PROMPT_PHOTO": {
        "uk": "7. Завантажте оновлене паспортне фото.\n━━━━━━━━━━━━━━━━━━\n• фронтальний ракурс\n• нейтральний фон\n• рівне освітлення",
        "en": "7. Upload an updated passport-style photo.\n━━━━━━━━━━━━━━━━━━\n• frontal view\n• neutral background\n• even lighting",
        "de": "7. Laden Sie ein aktualisiertes Passfoto hoch.\n━━━━━━━━━━━━━━━━━━\n• frontale Ansicht\n• neutraler Hintergrund\n• gleichmäßige Beleuchtung",
        "pl": "7. Dodaj zaktualizowane zdjęcie paszportowe.\n━━━━━━━━━━━━━━━━━━\n• ujęcie frontalne\n• neutralne tło\n• równomierne oświetlenie",
        "ru": "7. Загрузите обновлённое фото формата «на документы».\n━━━━━━━━━━━━━━━━━━\n• фронтальный ракурс\n• нейтральный фон\n• ровное освещение",
    },
    "PROFILE_UPDATE_SUCCESS": {
        "uk": "✅ Дані оновлено",
        "en": "✅ Data updated",
        "de": "✅ Daten aktualisiert",
        "pl": "✅ Dane zaktualizowano",
        "ru": "✅ Данные обновлены",
    },
    "PROFILE_PHOTO_UPDATED": {
        "uk": "🖼 Фото профілю збережено",
        "en": "🖼 Profile photo saved",
        "de": "🖼 Profilfoto gespeichert",
        "pl": "🖼 Zdjęcie profilu zapisane",
        "ru": "🖼 Фото профиля сохранено",
    },
    "PROFILE_PHOTO_REMOVED": {
        "uk": "🗑 Фото профілю видалено",
        "en": "🗑 Profile photo removed",
        "de": "🗑 Profilfoto gelöscht",
        "pl": "🗑 Zdjęcie profilu usunięto",
        "ru": "🗑 Фото профиля удалено",
    },
    "PROFILE_PHONE_SAVED": {
        "uk": "📱 Телефон збережено",
        "en": "📱 Phone saved",
        "de": "📱 Telefon gespeichert",
        "pl": "📱 Numer zapisano",
        "ru": "📱 Телефон сохранён",
    },
    "PROFILE_CANCELLED": {
        "uk": "❌ Редагування скасовано",
        "en": "❌ Edit cancelled",
        "de": "❌ Bearbeitung abgebrochen",
        "pl": "❌ Edycję anulowano",
        "ru": "❌ Редактирование отменено",
    },
    "PROFILE_NO_PHOTO": {
        "uk": "📷 Фото профілю ще не завантажено.",
        "en": "📷 A profile photo has not been uploaded yet.",
        "de": "📷 Es wurde noch kein Profilfoto hochgeladen.",
        "pl": "📷 Zdjęcie profilu nie zostało jeszcze dodane.",
        "ru": "📷 Фото профиля ещё не загружено.",
    },
    "LANGUAGE_PROMPT": {
        "uk": "🌐 Оберіть мову спілкування з ботом:",
        "en": "🌐 Choose the language you prefer to use with the bot:",
        "de": "🌐 Wählen Sie die Sprache für die Kommunikation mit dem Bot:",
        "pl": "🌐 Wybierz język komunikacji z botem:",
        "ru": "🌐 Выберите язык общения с ботом:",
    },
    "LANGUAGE_SELECTED": {
        "uk": "✅ Мову змінено на: {language}.",
        "en": "✅ Language switched to: {language}.",
        "de": "✅ Sprache geändert zu: {language}.",
        "pl": "✅ Zmieniono język na: {language}.",
        "ru": "✅ Язык переключён на: {language}.",
    },
    "SETTINGS_TITLE": {
        "uk": "⚙️ <b>Налаштування</b>\nОберіть дію:",
        "en": "⚙️ <b>Settings</b>\nChoose an action:",
        "de": "⚙️ <b>Einstellungen</b>\nWählen Sie eine Aktion:",
        "pl": "⚙️ <b>Ustawienia</b>\nWybierz działanie:",
        "ru": "⚙️ <b>Настройки</b>\nВыберите действие:",
    },
    "SETTINGS_LANGUAGE": {
        "uk": "🌐 Змінити мову",
        "en": "🌐 Change language",
        "de": "🌐 Sprache ändern",
        "pl": "🌐 Zmień język",
        "ru": "🌐 Изменить язык",
    },
    "INVALID_COMMAND": {
        "uk": "⚠️ Команда не розпізнана. Скористайтеся меню нижче.",
        "en": "⚠️ Command not recognized. Please use the menu below.",
        "de": "⚠️ Unbekannter Befehl. Bitte nutzen Sie das Menü unten.",
        "pl": "⚠️ Nieznane polecenie. Skorzystaj z menu poniżej.",
        "ru": "⚠️ Команда не распознана. Используйте меню ниже.",
    },
    "BTN_NEXT": {
        "uk": "➡️ ДАЛІ",
        "en": "➡️ NEXT",
        "de": "➡️ WEITER",
        "pl": "➡️ DALEJ",
        "ru": "➡️ ДАЛЕЕ",
    },
    "BTN_SKIP": {
        "uk": "⏭ Пропустити",
        "en": "⏭ Skip",
        "de": "⏭ Überspringen",
        "pl": "⏭ Pomiń",
        "ru": "⏭ Пропустить",
    },
    "INTRO_GREETING_NEW": {
        "uk": "👋 <b>Вітаю, колего!</b>\n━━━━━━━━━━━━━━━━━━\nВи у робочому просторі Bot.BSG. Тут зберігаємо чеки, оформлюємо виплати та тримаємо документи проєкту під рукою.\n\nНатисніть «ДАЛІ», щоб продовжити.",
        "en": "👋 <b>Hello, teammate!</b>\n━━━━━━━━━━━━━━━━━━\nWelcome to the Bot.BSG workspace. Here we store receipts, track payouts, and keep project documents handy.\n\nPress “NEXT” to continue.",
        "de": "👋 <b>Hallo, Kollegin oder Kollege!</b>\n━━━━━━━━━━━━━━━━━━\nWillkommen im Bot.BSG-Arbeitsbereich. Hier speichern wir Belege, verwalten Auszahlungen und behalten Projektdokumente griffbereit.\n\nDrücken Sie „WEITER“, um fortzufahren.",
        "pl": "👋 <b>Witaj, współpracowniku!</b>\n━━━━━━━━━━━━━━━━━━\nTo przestrzeń robocza Bot.BSG. Przechowujemy tu paragony, obsługujemy wypłaty i mamy dokumenty projektu pod ręką.\n\nKliknij „DALEJ”, aby kontynuować.",
        "ru": "👋 <b>Привет, коллега!</b>\n━━━━━━━━━━━━━━━━━━\nВы в рабочем пространстве Bot.BSG. Здесь мы храним чеки, оформляем выплаты и держим документы проекта под рукой.\n\nНажмите «ДАЛЕЕ», чтобы продолжить.",
    },
    "INTRO_GREETING_REGISTERED": {
        "uk": "👋 <b>Радий вітати знову!</b>\n━━━━━━━━━━━━━━━━━━\nВи можете одразу перейти до головного меню, щоб працювати з розділами бота.\n\nНатисніть «ДАЛІ», аби перейти до основних дій.",
        "en": "👋 <b>Welcome back!</b>\n━━━━━━━━━━━━━━━━━━\nYou can jump straight to the main menu to work with the bot’s sections.\n\nPress “NEXT” to open the main actions.",
        "de": "👋 <b>Willkommen zurück!</b>\n━━━━━━━━━━━━━━━━━━\nSie können direkt ins Hauptmenü wechseln, um mit den Bereichen des Bots zu arbeiten.\n\nDrücken Sie „WEITER“, um die wichtigsten Aktionen zu öffnen.",
        "pl": "👋 <b>Miło znów Cię widzieć!</b>\n━━━━━━━━━━━━━━━━━━\nMożesz od razu przejść do menu głównego i korzystać z modułów bota.\n\nKliknij „DALEJ”, aby wyświetlić najważniejsze opcje.",
        "ru": "👋 <b>Рад снова видеть!</b>\n━━━━━━━━━━━━━━━━━━\nВы можете сразу перейти в главное меню и работать с разделами бота.\n\nНажмите «ДАЛЕЕ», чтобы открыть основные действия.",
    },
    "INTRO_REG_STEPS": {
        "uk": "📝 <b>Як пройти реєстрацію</b>\n━━━━━━━━━━━━━━━━━━\n1️⃣ Вкажіть повне ім'я одним повідомленням.\n2️⃣ Поділіться номером телефону кнопкою «📱 Надіслати номер».\n3️⃣ Після підтвердження відкриються чеки, фінанси та документи.\n\nГотові? Натисніть «ДАЛІ» та заповніть анкету.",
        "en": "📝 <b>Registration steps</b>\n━━━━━━━━━━━━━━━━━━\n1️⃣ Send your full name in one message.\n2️⃣ Share your phone number via the “📱 Share phone number” button.\n3️⃣ Once confirmed, you will unlock receipts, finance, and documents.\n\nReady? Press “NEXT” and complete the form.",
        "de": "📝 <b>So funktioniert die Registrierung</b>\n━━━━━━━━━━━━━━━━━━\n1️⃣ Geben Sie Ihren vollständigen Namen in einer Nachricht an.\n2️⃣ Teilen Sie Ihre Telefonnummer über die Schaltfläche „📱 Telefonnummer senden“.\n3️⃣ Nach der Bestätigung stehen Belege, Finanzen und Dokumente zur Verfügung.\n\nBereit? Drücken Sie „WEITER“ und füllen Sie das Formular aus.",
        "pl": "📝 <b>Jak przejść rejestrację</b>\n━━━━━━━━━━━━━━━━━━\n1️⃣ Podaj pełne imię i nazwisko w jednej wiadomości.\n2️⃣ Udostępnij numer telefonu przyciskiem „📱 Wyślij numer”.\n3️⃣ Po potwierdzeniu zyskasz dostęp do paragonów, finansów i dokumentów.\n\nGotowy? Kliknij „DALEJ” i wypełnij formularz.",
        "ru": "📝 <b>Как пройти регистрацию</b>\n━━━━━━━━━━━━━━━━━━\n1️⃣ Отправьте полное ФИО одним сообщением.\n2️⃣ Поделитесь номером телефона кнопкой «📱 Отправить номер».\n3️⃣ После подтверждения станут доступны чеки, финансы и документы.\n\nГотовы? Нажмите «ДАЛЕЕ» и заполните анкету.",
    },
    "INTRO_PROMPT_NAME": {
        "uk": "📝 Введіть повне прізвище та ім'я одним повідомленням у форматі: Ім'я Прізвище По батькові.",
        "en": "📝 Enter your full name in one message using the format: Firstname Lastname Middle name.",
        "de": "📝 Geben Sie Ihren vollständigen Namen in einer Nachricht ein: Vorname Nachname zweiter Name.",
        "pl": "📝 Wpisz pełne imię i nazwisko w jednej wiadomości w formacie: Imię Nazwisko Drugie imię.",
        "ru": "📝 Введите полное ФИО одним сообщением в формате: Имя Фамилия Отчество.",
    },
    "INTRO_SECTIONS": {
        "uk": "📋 <b>Ключові розділи</b>\n━━━━━━━━━━━━━━━━━━\n• «🧾 Чеки» — завантаження та історія документів.\n• «💵 Фінанси» — статистика, запити та підтвердження виплат.\n• «📑 Документи» — файли та інструкції щодо проєкту.\n\nНатисніть «ДАЛІ», щоб перейти до головного меню.",
        "en": "📋 <b>Main sections</b>\n━━━━━━━━━━━━━━━━━━\n• “🧾 Receipts” — upload and history of documents.\n• “💵 Finance” — statistics, payout requests, and confirmations.\n• “📑 Documents” — project files and guidelines.\n\nPress “NEXT” to open the main menu.",
        "de": "📋 <b>Wichtige Bereiche</b>\n━━━━━━━━━━━━━━━━━━\n• „🧾 Belege“ – Upload und Verlauf der Dokumente.\n• „💵 Finanzen“ – Statistiken, Auszahlungsanträge und Bestätigungen.\n• „📑 Dokumente“ – Projektunterlagen und Richtlinien.\n\nDrücken Sie „WEITER“, um das Hauptmenü zu öffnen.",
        "pl": "📋 <b>Najważniejsze sekcje</b>\n━━━━━━━━━━━━━━━━━━\n• „🧾 Paragony” – wgrywanie i historia dokumentów.\n• „💵 Finanse” – statystyki, wnioski o wypłatę i potwierdzenia.\n• „📑 Dokumenty” – pliki i instrukcje projektowe.\n\nKliknij „DALEJ”, aby przejść do menu głównego.",
        "ru": "📋 <b>Ключевые разделы</b>\n━━━━━━━━━━━━━━━━━━\n• «🧾 Чеки» — загрузка и история документов.\n• «💵 Финансы» — статистика, запросы и подтверждения выплат.\n• «📑 Документы» — файлы и инструкции по объекту.\n\nНажмите «ДАЛЕЕ», чтобы открыть главное меню.",
    },
    "REGISTER_NAME_ERROR": {
        "uk": "❗ Здається, бракує частини ПІБ. Будь ласка, вкажіть ім'я, прізвище та по батькові повністю.",
        "en": "❗ It looks like part of the full name is missing. Please provide first name, last name, and patronymic (if applicable).",
        "de": "❗ Es scheint, dass Teile des vollständigen Namens fehlen. Bitte geben Sie Vorname, Nachname und ggf. zweiten Namen vollständig an.",
        "pl": "❗ Wygląda na to, że brakuje części pełnego imienia i nazwiska. Podaj imię, nazwisko i, jeśli dotyczy, drugie imię.",
        "ru": "❗ Похоже, не хватает части ФИО. Укажите имя, фамилию и отчество полностью.",
    },
    "REGISTER_PHONE_PROMPT": {
        "uk": "📞 Залишився номер телефону. Натисніть «📱 Надіслати номер», і бот автоматично додасть його до анкети.",
        "en": "📞 We still need your phone number. Tap “📱 Share phone number” and the bot will fill it in automatically.",
        "de": "📞 Uns fehlt noch Ihre Telefonnummer. Tippen Sie auf „📱 Telefonnummer senden“, und der Bot trägt sie automatisch ein.",
        "pl": "📞 Potrzebujemy jeszcze numeru telefonu. Kliknij „📱 Wyślij numer”, a bot uzupełni go automatycznie.",
        "ru": "📞 Остался телефон для связи. Нажмите «📱 Отправить номер», и бот автоматически подставит его в анкету.",
    },
    "REGISTER_PHONE_ERROR": {
        "uk": "❗ Не вдалося отримати номер. Спробуйте ще раз кнопкою «📱 Надіслати номер».",
        "en": "❗ We couldn't read the number. Try again using the “📱 Share phone number” button.",
        "de": "❗ Telefonnummer konnte nicht erkannt werden. Versuchen Sie es erneut über die Schaltfläche „📱 Telefonnummer senden“.",
        "pl": "❗ Nie udało się pobrać numeru. Spróbuj ponownie przyciskiem „📱 Wyślij numer”.",
        "ru": "❗ Номер не получен. Попробуйте ещё раз кнопкой «📱 Отправить номер».",
    },
    "REGISTER_PHONE_TEXT_PROMPT": {
        "uk": "❗ Щоб продовжити, відправте номер телефона кнопкою «📱 Надіслати номер» нижче — так уникнемо помилок.",
        "en": "❗ To continue, send your phone number using the “📱 Share phone number” button below to avoid mistakes.",
        "de": "❗ Um fortzufahren, senden Sie Ihre Telefonnummer über die Schaltfläche „📱 Telefonnummer senden“, um Fehler zu vermeiden.",
        "pl": "❗ Aby kontynuować, wyślij numer telefonu przyciskiem „📱 Wyślij numer” poniżej – unikniemy pomyłek.",
        "ru": "❗ Чтобы продолжить регистрацию, отправьте номер телефона кнопкой «📱 Отправить номер» ниже, чтобы избежать ошибок.",
    },
    "BTN_SEND_PHONE": {
        "uk": "📱 Надіслати номер",
        "en": "📱 Share phone number",
        "de": "📱 Telefonnummer senden",
        "pl": "📱 Wyślij numer",
        "ru": "📱 Отправить номер",
    },
    "CHECKS_SECTION_TITLE": {
        "uk": "🧾 <b>Розділ чеків</b>",
        "en": "🧾 <b>Receipts section</b>",
        "de": "🧾 <b>Belegbereich</b>",
        "pl": "🧾 <b>Sekcja paragonów</b>",
        "ru": "🧾 <b>Раздел чеков</b>",
    },
    "BTN_BACK_SETTINGS": {
        "uk": "⬅️ Повернутися",
        "en": "⬅️ Back",
        "de": "⬅️ Zurück",
        "pl": "⬅️ Wróć",
        "ru": "⬅️ Назад",
    },
    "CHECKS_MENU_INTRO": {
        "uk": "🧾 <b>Розділ чеків</b>\n━━━━━━━━━━━━━━━━━━\nДодавайте нові чеки, переглядайте історію та оновлюйте статус виплат у два кліки.\nОбирайте дію за кнопками нижче.",
        "en": "🧾 <b>Receipts section</b>\n━━━━━━━━━━━━━━━━━━\nUpload new receipts, review history, and update payout status in a couple of taps.\nPick an action using the buttons below.",
        "de": "🧾 <b>Belegbereich</b>\n━━━━━━━━━━━━━━━━━━\nLaden Sie neue Belege hoch, prüfen Sie den Verlauf und aktualisieren Sie den Auszahlungsstatus mit wenigen Klicks.\nWählen Sie unten die gewünschte Aktion.",
        "pl": "🧾 <b>Sekcja paragonów</b>\n━━━━━━━━━━━━━━━━━━\nDodawaj nowe paragony, przeglądaj historię i aktualizuj status wypłat w kilku krokach.\nWybierz odpowiednią opcję z przycisków poniżej.",
        "ru": "🧾 <b>Раздел чеков</b>\n━━━━━━━━━━━━━━━━━━\nДобавляйте новые чеки, просматривайте историю и обновляйте статус выплат в пару касаний.\nВыберите нужное действие на кнопках ниже.",
    },
    "CHECK_STATS_SUMMARY": {
        "uk": "📊 <b>Особиста статистика по чеках</b>\n━━━━━━━━━━━━━━━━━━\n📂 Проєкт: <b>{project}</b>\n🧾 Всього чеків: <b>{total_count}</b>\n💰 Сума чеків: <b>{total_amount} грн</b>\n💸 Оплачено компанією: <b>{paid_amount} грн</b> ({paid_count} шт.)\n⏳ Очікує оплати: <b>{unpaid_amount} грн</b> ({unpaid_count} шт.){pending_line}",
        "en": "📊 <b>Your receipt statistics</b>\n━━━━━━━━━━━━━━━━━━\n📂 Project: <b>{project}</b>\n🧾 Total receipts: <b>{total_count}</b>\n💰 Receipt amount: <b>{total_amount} UAH</b>\n💸 Paid by the company: <b>{paid_amount} UAH</b> ({paid_count} items)\n⏳ Awaiting payment: <b>{unpaid_amount} UAH</b> ({unpaid_count} items){pending_line}",
        "de": "📊 <b>Ihre Belegstatistik</b>\n━━━━━━━━━━━━━━━━━━\n📂 Projekt: <b>{project}</b>\n🧾 Belege insgesamt: <b>{total_count}</b>\n💰 Belegsumme: <b>{total_amount} UAH</b>\n💸 Vom Unternehmen bezahlt: <b>{paid_amount} UAH</b> ({paid_count} Stück)\n⏳ Ausstehend: <b>{unpaid_amount} UAH</b> ({unpaid_count} Stück){pending_line}",
        "pl": "📊 <b>Twoja statystyka paragonów</b>\n━━━━━━━━━━━━━━━━━━\n📂 Projekt: <b>{project}</b>\n🧾 Razem paragonów: <b>{total_count}</b>\n💰 Suma paragonów: <b>{total_amount} UAH</b>\n💸 Opłacone przez firmę: <b>{paid_amount} UAH</b> ({paid_count} szt.)\n⏳ Do opłacenia: <b>{unpaid_amount} UAH</b> ({unpaid_count} szt.){pending_line}",
        "ru": "📊 <b>Личная статистика по чекам</b>\n━━━━━━━━━━━━━━━━━━\n📂 Проект: <b>{project}</b>\n🧾 Всего чеков: <b>{total_count}</b>\n💰 Сумма чеков: <b>{total_amount} грн</b>\n💸 Оплачено компанией: <b>{paid_amount} грн</b> ({paid_count} шт.)\n⏳ Ожидает оплаты: <b>{unpaid_amount} грн</b> ({unpaid_count} шт.){pending_line}",
    },
    "CHECK_STATS_PENDING": {
        "uk": "\n❔ Статус не вказано: <b>{amount} грн</b> ({count} шт.)",
        "en": "\n❔ Status not specified: <b>{amount} UAH</b> ({count} items)",
        "de": "\n❔ Status nicht angegeben: <b>{amount} UAH</b> ({count} Stück)",
        "pl": "\n❔ Status nieokreślony: <b>{amount} UAH</b> ({count} szt.)",
        "ru": "\n❔ Статус не указан: <b>{amount} грн</b> ({count} шт.)",
    },
    "CHECK_STATS_EMPTY": {
        "uk": "\nУ вас ще немає збережених чеків. Додайте перший через кнопку «📷 Додати чек».",
        "en": "\nYou haven't saved any receipts yet. Use “📷 Add receipt” to upload the first one.",
        "de": "\nSie haben noch keine Belege gespeichert. Fügen Sie den ersten über „📷 Beleg hinzufügen“ hinzu.",
        "pl": "\nNie masz jeszcze żadnych paragonów. Dodaj pierwszy przyciskiem „📷 Dodaj paragon”.",
        "ru": "\nУ вас ещё нет сохранённых чеков. Добавьте первый через кнопку «📷 Добавить чек».",
    },
    "NP_MENU_TITLE": {
        "uk": "📮 <b>Nova Poshta</b>\\n━━━━━━━━━━━━━━━━━━\\nЄдине меню для пошуку та супроводу накладних. Скористайтесь кнопками нижче, щоб знайти відправлення, переглянути історію, нотатки чи посилки від компанії.",
        "en": "📮 <b>Nova Poshta</b>\\n━━━━━━━━━━━━━━━━━━\\nYour central hub for parcel tracking. Use the buttons below to look up TTNs, reopen history, manage notes, and review company deliveries.",
        "de": "📮 <b>Nova Poshta</b>\\n━━━━━━━━━━━━━━━━━━\\nZentrale Schaltstelle für Sendungsnummern. Verwenden Sie die Schaltflächen unten, um TTN zu suchen, den Verlauf zu öffnen, Notizen zu pflegen und Firmensendungen einzusehen.",
        "pl": "📮 <b>Nova Poshta</b>\\n━━━━━━━━━━━━━━━━━━\\nPanel do pracy z przesyłkami. Przyciski poniżej umożliwiają wyszukiwanie TTN, podgląd historii, notatek oraz paczek od firmy.",
        "ru": "📮 <b>Nova Poshta</b>\\n━━━━━━━━━━━━━━━━━━\\nЦентральное меню для работы с накладными. Используйте кнопки ниже, чтобы искать ТТН, открывать историю, заметки и фирменные отправления.",
    },
    "NP_INTERFACE_TEXT": {
        "uk": "📘 <b>Як працювати з розділом</b>\n━━━━━━━━━━━━━━━━━━\n• «🔍 Пошук ТТН» — введіть номер і одразу отримайте квитанцію з оновленим статусом.\n• «🕓 Історія» — швидкий доступ до останніх переглядів.\n• «⭐ Відзначені» — зберігайте важливі накладні під рукою.\n• «🏢 Посилки BSG» — накладні, які передав адміністратор компанії.\n• «📥 Отримані BSG» — підтверджені відправлення, які вже закриті.\n• Додавайте особисті коментарі до ТТН прямо у картці — вони відображаються під квитанцією.\n\nСкористайтесь «❌ Скасувати», щоб повернутися в це меню.",
        "en": "📘 <b>How to use this section</b>\n━━━━━━━━━━━━━━━━━━\n• “🔍 Search by TTN” — enter a number and get the refreshed receipt.\n• “🕓 History” — reopen your latest lookups with one tap.\n• “⭐ Bookmarked” — keep priority shipments within reach.\n• “🏢 BSG parcels” — TTNs forwarded by the team.\n• “📥 Received BSG” — deliveries already confirmed.\n• Add personal comments to any TTN from its card — they stay under the receipt for quick reference.\n\nTap “❌ Cancel” any time to return here.",
        "de": "📘 <b>So nutzen Sie den Bereich</b>\n━━━━━━━━━━━━━━━━━━\n• „🔍 TTN suchen“ – Nummer eingeben und aktualisierten Beleg erhalten.\n• „🕓 Verlauf“ – letzte Abfragen sofort erneut öffnen.\n• „⭐ Markiert“ – wichtige Sendungen griffbereit halten.\n• „🏢 BSG-Sendungen“ – Nummern, die das Team zugewiesen hat.\n• „📥 Erhaltene BSG“ – bereits bestätigte Lieferungen.\n• Fügen Sie Kommentare direkt in der Sendungskarte hinzu – sie erscheinen unter dem Beleg.\n\nMit „❌ Abbrechen“ kehren Sie jederzeit zurück.",
        "pl": "📘 <b>Jak korzystać z panelu</b>\n━━━━━━━━━━━━━━━━━━\n• „🔍 Szukaj TTN” – wpisz numer i otrzymaj odświeżony podgląd.\n• „🕓 Historia” – szybki powrót до ostatnich wyszukiwań.\n• „⭐ Oznaczone” – найważniejsze przesyłki под ręką.\n• „🏢 Przesyłki BSG” – numery przekazane przez administrację.\n• „📥 Odebrane BSG” – przesyłki już потwierдzone.\n• Dodawaj komentarze bezpośrednio w karcie TTN – pojawią się pod potwierdzeniem.\n\nPrzycisk „❌ Anuluj” zawsze wraca до tego меню.",
        "ru": "📘 <b>Как пользоваться разделом</b>\n━━━━━━━━━━━━━━━━━━\n• «🔍 Поиск ТТН» — введите номер и сразу получите обновлённый чек.\n• «🕓 История» — быстрый доступ к последним запросам.\n• «⭐ Отмеченные» — держите важные отправления под рукой.\n• «🏢 Посылки BSG» — номера, которые передал администратор компании.\n• «📥 Полученные BSG» — уже подтверждённые доставки.\n• Добавляйте личные комментарии прямо в карточке ТТН — они будут показаны под квитанцией.\n\nНажмите «❌ Отменить», чтобы вернуться в это меню.",
    },
    "NP_PROMPT_TTN": {
        "uk": "✉️ Надішліть номер накладної одним повідомленням (допускаються цифри та літери). Кнопка «❌ Скасувати» повертає до меню.",
        "en": "✉️ Send the TTN as a single message (digits and letters only). Use “❌ Cancel” to return to the menu.",
        "de": "✉️ Senden Sie die TTN als einzelne Nachricht (Ziffern/Buchstaben). Mit „❌ Abbrechen“ geht es zurück ins Menü.",
        "pl": "✉️ Wyślij numer TTN w jednej wiadomości (cyfry i litery). „❌ Anuluj” wraca do menu.",
        "ru": "✉️ Отправьте номер ТТН одним сообщением (цифры и буквы). «❌ Отменить» вернёт в меню.",
    },
    "NP_SEARCH_PROGRESS": {
        "uk": "⏳ Отримую статус накладної, зачекайте кілька секунд…",
        "en": "⏳ Fetching parcel status, please wait…",
        "de": "⏳ Der Sendungsstatus wird abgerufen, bitte warten…",
        "pl": "⏳ Pobieram status przesyłki, proszę czekać…",
        "ru": "⏳ Получаем статус накладной, подождите…",
    },
    "NP_SEARCH_NOT_FOUND": {
        "uk": "❌ Накладну {ttn} не знайдено. Перевірте номер і спробуйте ще раз.",
        "en": "❌ TTN {ttn} was not found. Please check the number and try again.",
        "de": "❌ Die Sendung {ttn} wurde nicht gefunden. Bitte prüfen Sie die Nummer und versuchen Sie es erneut.",
        "pl": "❌ Nie znaleziono przesyłki {ttn}. Sprawdź numer i spróbuj ponownie.",
        "ru": "❌ Накладная {ttn} не найдена. Проверьте номер и попробуйте снова.",
    },
    "NP_SEARCH_ERROR": {
        "uk": "⚠️ Не вдалося отримати дані: {error}",
        "en": "⚠️ Could not retrieve data: {error}",
        "de": "⚠️ Daten konnten nicht abgerufen werden: {error}",
        "pl": "⚠️ Nie udało się pobrać danych: {error}",
        "ru": "⚠️ Не удалось получить данные: {error}",
    },
    "NP_REFRESH_NOT_POSSIBLE": {
        "uk": "⚠️ Неможливо оновити повідомлення. Спробуйте пізніше.",
        "en": "⚠️ Unable to refresh this message. Please try again later.",
        "de": "⚠️ Nachricht kann nicht aktualisiert werden. Bitte später erneut versuchen.",
        "pl": "⚠️ Nie można odświeżyć tej wiadomości. Spróbuj ponownie później.",
        "ru": "⚠️ Не удалось обновить сообщение. Попробуйте позже.",
    },
    "NP_HISTORY_EMPTY": {
        "uk": "🕓 Історія порожня. Виконайте пошук, щоб побачити останні ТТН.",
        "en": "🕓 History is empty. Run a search to see recent TTNs.",
        "de": "🕓 Noch keine Historie. Führen Sie eine Suche aus, um TTN anzuzeigen.",
        "pl": "🕓 Historia jest pusta. Wykonaj wyszukiwanie, aby zobaczyć TTN.",
        "ru": "🕓 История пуста. Выполните поиск, чтобы увидеть последние ТТН.",
    },
    "NP_HISTORY_HEADER": {
        "uk": "🕓 <b>Історія пошуку</b>\\n━━━━━━━━━━━━━━━━━━\\nНатисніть номер нижче, щоб відкрити квитанцію та побачити актуальний статус.",
        "en": "🕓 <b>Search history</b>\\n━━━━━━━━━━━━━━━━━━\\nTap a TTN below to reopen its receipt with the latest status.",
        "de": "🕓 <b>Suchverlauf</b>\\n━━━━━━━━━━━━━━━━━━\\nTippen Sie auf eine TTN, um den Beleg mit aktuellem Status zu öffnen.",
        "pl": "🕓 <b>Historia wyszukiwań</b>\\n━━━━━━━━━━━━━━━━━━\\nWybierz TTN, aby otworzyć podgląd ze świeżym statusem.",
        "ru": "🕓 <b>История поисков</b>\\n━━━━━━━━━━━━━━━━━━\\nНажмите на ТТН ниже, чтобы открыть чек с актуальным статусом.",
    },
    "NP_BOOKMARKS_EMPTY": {
        "uk": "⭐ Ви ще не позначали накладні. Додайте вподобану ТТН під час пошуку.",
        "en": "⭐ You haven't bookmarked any TTNs yet. Bookmark a result while viewing a search.",
        "de": "⭐ Sie haben noch keine TTN markiert. Markieren Sie Ergebnisse während der Suche.",
        "pl": "⭐ Nie oznaczono jeszcze żadnej TTN. Dodaj ją do oznaczonych podczas przeglądania wyników.",
        "ru": "⭐ У вас ещё нет отмеченных ТТН. Добавьте накладную в избранное во время просмотра результата.",
    },
    "NP_BOOKMARKS_HEADER": {
        "uk": "⭐ <b>Відзначені накладні</b>\n━━━━━━━━━━━━━━━━━━\nОберіть ТТН, щоб миттєво відкрити її квитанцію та нотатки.",
        "en": "⭐ <b>Bookmarked TTNs</b>\n━━━━━━━━━━━━━━━━━━\nSelect a TTN to instantly open its receipt and notes.",
        "de": "⭐ <b>Markierte TTN</b>\n━━━━━━━━━━━━━━━━━━\nWählen Sie eine TTN, um Beleg und Notizen sofort zu öffnen.",
        "pl": "⭐ <b>Oznaczone TTN</b>\n━━━━━━━━━━━━━━━━━━\nWybierz TTN, aby szybko zobaczyć podgląd i notatki.",
        "ru": "⭐ <b>Отмеченные накладные</b>\n━━━━━━━━━━━━━━━━━━\nВыберите ТТН, чтобы сразу открыть чек и заметки."
    },
    "NP_NOTE_PROMPT": {
        "uk": "💬 Напишіть коментар для ТТН {ttn} та надішліть повідомленням. Щоб скасувати, скористайтесь «❌ Скасувати» або напишіть «відміна».",
        "en": "💬 Type a comment for TTN {ttn} and send it as a message. Use “❌ Cancel” or type “cancel” to abort.",
        "de": "💬 Schreiben Sie eine Notiz für TTN {ttn} und senden Sie sie als Nachricht. Mit „❌ Abbrechen“ oder dem Wort „abbrechen“ beenden.",
        "pl": "💬 Napisz komentarz do TTN {ttn} i wyślij wiadomość. Użyj „❌ Anuluj” lub wpisz „anuluj”, aby przerwać.",
        "ru": "💬 Напишите комментарий для ТТН {ttn} и отправьте сообщением. Можно отменить через «❌ Отменить» или слово «отмена».",
    },
    "NP_NOTE_CANCELLED": {
        "uk": "ℹ️ Додавання коментаря скасовано.",
        "en": "ℹ️ Comment cancelled.",
        "de": "ℹ️ Kommentar verworfen.",
        "pl": "ℹ️ Dodawanie komentarza przerwano.",
        "ru": "ℹ️ Комментарий не сохранён.",
    },
    "NP_NOTE_SAVED": {
        "uk": "✅ Коментар збережено.",
        "en": "✅ Comment saved.",
        "de": "✅ Kommentar gespeichert.",
        "pl": "✅ Komentarz zapisano.",
        "ru": "✅ Комментарий сохранён.",
    },
    "NP_COMMENT_SECTION_TITLE": {
        "uk": "💬 Коментарі ({count})",
        "en": "💬 Comments ({count})",
        "de": "💬 Kommentare ({count})",
        "pl": "💬 Komentarze ({count})",
        "ru": "💬 Комментарии ({count})",
    },
    "NP_BOOKMARK_ADDED": {
        "uk": "✅ Накладну додано до відзначених.",
        "en": "✅ TTN added to bookmarks.",
        "de": "✅ TTN wurde markiert.",
        "pl": "✅ TTN dodano do oznaczonych.",
        "ru": "✅ Накладная добавлена в отмеченные.",
    },
    "NP_BOOKMARK_REMOVED": {
        "uk": "✅ Накладну прибрано з відзначених.",
        "en": "✅ TTN removed from bookmarks.",
        "de": "✅ Markierung für die TTN entfernt.",
        "pl": "✅ TTN usunięto z oznaczonych.",
        "ru": "✅ Накладная удалена из отмеченных.",
    },
    "NP_ASSIGN_PROMPT_TTN": {
        "uk": "📬 Введіть номер ТТН, яку потрібно закріпити за користувачем. «❌ Скасувати» повертає до меню.",
        "en": "📬 Enter the TTN you want to assign to a user. Press “❌ Cancel” to return.",
        "de": "📬 Geben Sie die TTN ein, die einem Nutzer zugeordnet werden soll. Mit „❌ Abbrechen“ zurück zum Menü.",
        "pl": "📬 Podaj numer TTN, który chcesz przypisać użytkownikowi. „❌ Anuluj” wraca do menu.",
        "ru": "📬 Введите номер ТТН, который нужно закрепить за пользователем. «❌ Отменить» вернёт в меню.",
    },
    "NP_ASSIGN_PROMPT_USER": {
        "uk": "👤 Оберіть отримувача зі списку нижче або введіть його BSU/ID. «❌ Скасувати» зупиняє операцію.",
        "en": "👤 Pick the recipient from the list below or type their BSU/ID. Use “❌ Cancel” to stop.",
        "de": "👤 Wählen Sie den Empfänger über die Liste oder geben Sie BSU/ID ein. Mit „❌ Abbrechen“ beenden.",
        "pl": "👤 Wybierz odbiorcę z listy poniżej lub wpisz jego BSU/ID. „❌ Anuluj” przerywa operację.",
        "ru": "👤 Выберите получателя кнопкой ниже или введите его BSU/ID. «❌ Отменить» прекращает операцию.",
    },
    "NP_ASSIGN_USER_NOT_FOUND": {
        "uk": "❗ Користувача не знайдено. Перевірте дані та спробуйте ще раз.",
        "en": "❗ User not found. Please check the details and try again.",
        "de": "❗ Benutzer nicht gefunden. Bitte prüfen Sie die Angaben und versuchen Sie es erneut.",
        "pl": "❗ Nie znaleziono użytkownika. Sprawdź dane i spróbuj ponownie.",
        "ru": "❗ Пользователь не найден. Проверьте данные и попробуйте снова.",
    },
    "NP_ASSIGN_PROMPT_NOTE": {
        "uk": "📝 Додайте коротке повідомлення або натисніть «⏭ Пропустити». «❌ Скасувати» зупиняє передачу.",
        "en": "📝 Add a short note or tap “⏭ Skip”. “❌ Cancel” stops the handover.",
        "de": "📝 Fügen Sie eine kurze Notiz hinzu oder tippen Sie auf „⏭ Überspringen“. „❌ Abbrechen“ beendet den Vorgang.",
        "pl": "📝 Dodaj krótki komentarz albo wybierz „⏭ Pomiń”. „❌ Anuluj” kończy operację.",
        "ru": "📝 Добавьте короткий комментарий или нажмите «⏭ Пропустить». «❌ Отменить» прекращает передачу.",
    },
    "NP_ASSIGN_SKIP_TOAST": {
        "uk": "Коментар не додано.",
        "en": "No note attached.",
        "de": "Keine Notiz hinzugefügt.",
        "pl": "Notatki nie dodano.",
        "ru": "Комментарий не добавлен.",
    },
    "NP_ASSIGN_CANCELLED": {
        "uk": "ℹ️ Передача ТТН скасована.",
        "en": "ℹ️ TTN forwarding cancelled.",
        "de": "ℹ️ Weitergabe der TTN wurde abgebrochen.",
        "pl": "ℹ️ Przekazanie TTN zostało przerwane.",
        "ru": "ℹ️ Передача ТТН отменена.",
    },
    "NP_CANCELLED_TOAST": {
        "uk": "Дію скасовано.",
        "en": "Action cancelled.",
        "de": "Aktion abgebrochen.",
        "pl": "Działanie anulowano.",
        "ru": "Действие отменено.",
    },
    "NP_ASSIGN_DONE": {
        "uk": "🏢 <b>Передача оформлена</b>\n━━━━━━━━━━━━━━━━━━\n🔖 ТТН: <code>{ttn}</code>\n👤 Отримувач: {user}\n🕒 Призначено: {time}\n\n✅ Повідомлення надіслано.",
        "en": "🏢 <b>Forwarding complete</b>\n━━━━━━━━━━━━━━━━━━\n🔖 TTN: <code>{ttn}</code>\n👤 Recipient: {user}\n🕒 Assigned: {time}\n\n✅ Notification sent.",
        "de": "🏢 <b>Weitergabe abgeschlossen</b>\n━━━━━━━━━━━━━━━━━━\n🔖 TTN: <code>{ttn}</code>\n👤 Empfänger: {user}\n🕒 Zugeordnet: {time}\n\n✅ Benachrichtigung gesendet.",
        "pl": "🏢 <b>Przekazanie zakończone</b>\n━━━━━━━━━━━━━━━━━━\n🔖 TTN: <code>{ttn}</code>\n👤 Odbiorca: {user}\n🕒 Przypisano: {time}\n\n✅ Powiadomienie wysłano.",
        "ru": "🏢 <b>Передача оформлена</b>\n━━━━━━━━━━━━━━━━━━\n🔖 ТТН: <code>{ttn}</code>\n👤 Получатель: {user}\n🕒 Назначено: {time}\n\n✅ Уведомление отправлено.",
    },
    "NP_ASSIGN_DONE_NOTE_LABEL": {
        "uk": "📝 Коментар адміністратора:\n{note}",
        "en": "📝 Admin note:\n{note}",
        "de": "📝 Notiz des Administrators:\n{note}",
        "pl": "📝 Notatka administratora:\n{note}",
        "ru": "📝 Комментарий администратора:\n{note}",
    },
    "NP_ASSIGN_NOTIFY_USER": {
        "uk": "📦 Адміністратор {admin} передав вам накладну <b>{ttn}</b>. Відкрийте картку нижче, щоб переглянути статус і підтвердити отримання.",
        "en": "📦 Administrator {admin} forwarded TTN <b>{ttn}</b> to you. Open the card below to review the status and confirm delivery.",
        "de": "📦 Administrator {admin} hat Ihnen die TTN <b>{ttn}</b> übergeben. Öffnen Sie die Karte unten, um Status und Empfang zu prüfen.",
        "pl": "📦 Administrator {admin} przekazał Ci TTN <b>{ttn}</b>. Otwórz kartę poniżej, aby sprawdzić status i potwierdzić odbiór.",
        "ru": "📦 Администратор {admin} передал вам ТТН <b>{ttn}</b>. Откройте карточку ниже, чтобы проверить статус и подтвердить получение.",
    },
    "NP_ASSIGNED_EMPTY": {
        "uk": "🏢 Нових посилок BSG поки немає. Як тільки адміністратор передасть ТТН, ви отримаєте сповіщення.",
        "en": "🏢 No BSG parcels right now. You'll be notified as soon as an administrator forwards a TTN.",
        "de": "🏢 Zurzeit keine BSG-Sendungen. Sie erhalten eine Nachricht, sobald ein Administrator eine TTN weiterleitet.",
        "pl": "🏢 Obecnie brak przesyłek BSG. Dostaniesz powiadomienie, gdy administrator przekaże TTN.",
        "ru": "🏢 Новых посылок BSG пока нет. Мы сообщим, как только администратор передаст ТТН.",
    },
    "NP_ASSIGNED_HEADER": {
        "uk": "🏢 <b>Посилки BSG</b>\n━━━━━━━━━━━━━━━━━━\nОберіть накладну, щоб переглянути статус, залишити коментар або підтвердити отримання.",
        "en": "🏢 <b>BSG parcels</b>\n━━━━━━━━━━━━━━━━━━\nPick a TTN to review its status, add a comment, or confirm delivery.",
        "de": "🏢 <b>BSG-Sendungen</b>\n━━━━━━━━━━━━━━━━━━\nWählen Sie eine TTN, um Status, Kommentar oder den Empfang zu bestätigen.",
        "pl": "🏢 <b>Przesyłki BSG</b>\n━━━━━━━━━━━━━━━━━━\nWybierz TTN, aby sprawdzić status, dodać komentarz lub potwierdzić odbiór.",
        "ru": "🏢 <b>Посылки BSG</b>\n━━━━━━━━━━━━━━━━━━\nВыберите накладную, чтобы посмотреть статус, оставить комментарий или подтвердить получение.",
    },
    "NP_ASSIGNED_DETAIL_TITLE": {
        "uk": "🏢 <b>Посилка BSG</b>",
        "en": "🏢 <b>BSG parcel</b>",
        "de": "🏢 <b>BSG-Sendung</b>",
        "pl": "🏢 <b>Przesyłka BSG</b>",
        "ru": "🏢 <b>Посылка BSG</b>",
    },
    "NP_ASSIGNED_CONFIRM_SENT": {
        "uk": "✅ Повідомлення відправлено адміністраторам.",
        "en": "✅ Notification sent to the administrators.",
        "de": "✅ Benachrichtigung an die Administratoren gesendet.",
        "pl": "✅ Powiadomienie wysłano administratorom.",
        "ru": "✅ Уведомление отправлено администраторам.",
    },
    "NP_ASSIGNMENT_ALREADY_DONE": {
        "uk": "ℹ️ Ця посилка вже відмічена як отримана.",
        "en": "ℹ️ This parcel has already been marked as received.",
        "de": "ℹ️ Diese Sendung wurde bereits als erhalten markiert.",
        "pl": "ℹ️ Ta przesyłka została już oznaczona jako odebrana.",
        "ru": "ℹ️ Эта посылка уже отмечена как полученная.",
    },
    "NP_DELIVERY_ACK_RECORDED": {
        "uk": "✅ Дякуємо! Ми повідомили адміністраторів про отримання.",
        "en": "✅ Thank you! The administrators have been notified.",
        "de": "✅ Danke! Die Administratoren wurden informiert.",
        "pl": "✅ Dziękujemy! Administratorzy zostali poinformowani.",
        "ru": "✅ Спасибо! Администраторы уведомлены.",
    },
    "NP_RECEIVED_EMPTY": {
        "uk": "📥 Поки що немає підтверджених посилок BSG. Після отримання скористайтеся кнопкою «Посилка отримана» в картці накладної.",
        "en": "📥 No received BSG parcels yet. Use “Parcel received” on the TTN card once the delivery is in your hands.",
        "de": "📥 Noch keine bestätigten BSG-Sendungen. Nutzen Sie „Sendung erhalten“ in der Karte, sobald die Lieferung bei Ihnen ist.",
        "pl": "📥 Brak potwierdzonych przesyłek BSG. Po odebraniu użyj przycisku „Przesyłka odebrana” w karcie TTN.",
        "ru": "📥 Подтверждённых посылок BSG пока нет. После получения нажмите «Посылка получена» в карточке накладной.",
    },
    "NP_RECEIVED_HEADER": {
        "uk": "📥 <b>Отримані посилки BSG</b>\n━━━━━━━━━━━━━━━━━━\nПерегляньте статуси, коментарі та історію для підтверджених накладних.",
        "en": "📥 <b>Received BSG parcels</b>\n━━━━━━━━━━━━━━━━━━\nReview statuses, comments, and history for confirmed deliveries.",
        "de": "📥 <b>Erhaltene BSG-Sendungen</b>\n━━━━━━━━━━━━━━━━━━\nPrüfen Sie Status, Kommentare und Historie bestätigter Lieferungen.",
        "pl": "📥 <b>Odebrane przesyłki BSG</b>\n━━━━━━━━━━━━━━━━━━\nSprawdź statusy, komentarze i historię potwierdzonych dostaw.",
        "ru": "📥 <b>Полученные посылки BSG</b>\n━━━━━━━━━━━━━━━━━━\nПросмотрите статусы, комментарии и историю подтверждённых доставок.",
    },
}

LANG_CODES = {code for code, _ in LANG_ORDER}
LANG_LABELS = {code: label for code, label in LANG_ORDER}


def normalize_lang(code: Optional[str]) -> str:
    if isinstance(code, str) and code in LANG_CODES:
        return code
    return DEFAULT_LANG


def resolve_lang(target: Any) -> str:
    if isinstance(target, str):
        return normalize_lang(target)
    if isinstance(target, dict):
        return normalize_lang(target.get("lang"))
    if isinstance(target, int):
        profile = load_user(target)
        if profile:
            return normalize_lang(profile.get("lang"))
    return DEFAULT_LANG


def tr(target: Any, text: str, **kwargs) -> str:
    if not isinstance(text, str):
        return text
    lang = resolve_lang(target)
    mapping = TEXTS.get(text)
    if mapping:
        template = mapping.get(lang) or mapping.get(DEFAULT_LANG) or next(iter(mapping.values()))
    else:
        template = text
    if kwargs:
        try:
            return template.format(**kwargs)
        except Exception:
            return template
    return template

bot = Bot(token=TOKEN, parse_mode="HTML")
dp = Dispatcher(bot, storage=MemoryStorage())

# runtime cache
users_runtime: Dict[int, dict] = {}
admins: set = set()
active_project = {"name": None}
alerts_poll_task: Optional[asyncio.Task] = None
alerts_history_cache: Dict[str, Dict[str, Any]] = {}


# ========================== FSM ==========================
class OnboardFSM(StatesGroup):
    language = State()
    welcome = State()
    briefing = State()
    instructions = State()
    last_name = State()
    first_name = State()
    middle_name = State()
    birthdate = State()
    region = State()
    region_confirm = State()
    phone = State()
    photo = State()


class ProfileEditFSM(StatesGroup):
    waiting_last_name = State()
    waiting_first_name = State()
    waiting_middle_name = State()
    waiting_birthdate = State()
    waiting_region = State()
    region_confirm = State()
    waiting_phone = State()
    waiting_photo = State()


class AdminProfileEditFSM(StatesGroup):
    waiting_last_name = State()
    waiting_first_name = State()
    waiting_middle_name = State()
    waiting_birthdate = State()
    waiting_region = State()
    region_confirm = State()
    waiting_phone = State()
    waiting_photo = State()


class ReceiptFSM(StatesGroup):
    waiting_photo = State()
    waiting_amount = State()
    waiting_description = State()
    waiting_paid_choice = State()
    preview = State()

class ProjectCreateFSM(StatesGroup):
    enter_name = State()
    enter_region = State()
    enter_location = State()
    enter_description = State()
    enter_start_date = State()
    enter_end_date = State()
    upload_pdf = State()


class PhotoFSM(StatesGroup):
    collecting = State()


class SosFSM(StatesGroup):
    waiting_location = State()


class NovaPoshtaFSM(StatesGroup):
    waiting_ttn = State()
    waiting_note = State()
    waiting_assign_ttn = State()
    waiting_assign_user = State()
    waiting_assign_note = State()


# ========================== FS HELPERS ==========================
def ensure_dirs():
    os.makedirs("data", exist_ok=True)
    os.makedirs(BASE_PATH, exist_ok=True)
    os.makedirs(USERS_PATH, exist_ok=True)
    os.makedirs(FIN_PATH, exist_ok=True)
    os.makedirs(ALERTS_STORAGE_BASE, exist_ok=True)

def proj_path(name: str) -> str: return os.path.join(BASE_PATH, name)
def proj_info_file(name: str) -> str: return os.path.join(proj_path(name), "project.json")
def proj_pdf_dir(name: str) -> str: return os.path.join(proj_path(name), "pdf")
def proj_ledger(name: str) -> str: return os.path.join(proj_path(name), "ledger.xlsx")
def proj_finance_file(name: str) -> str: return os.path.join(proj_path(name), "finance.json")
def proj_receipts_dir(name: str, uid: int) -> str: return os.path.join(proj_path(name), "receipts", str(uid))
def proj_photos_dir(name: str) -> str: return os.path.join(proj_path(name), "photos")
def proj_photos_meta(name: str) -> str: return os.path.join(proj_photos_dir(name), "photos.json")
def user_file(uid: int) -> str: return os.path.join(USERS_PATH, f"{uid}.json")


def generate_photo_id() -> str:
    return secrets.token_hex(8)


def project_token(name: str) -> str:
    digest = hashlib.sha1(name.encode("utf-8")).digest()[:4]
    return base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")


def _project_existing_codes(exclude: Optional[str] = None) -> Set[str]:
    ensure_dirs()
    codes: Set[str] = set()
    if not os.path.exists(BASE_PATH):
        return codes
    for d in os.listdir(BASE_PATH):
        if exclude and d == exclude:
            continue
        info_path = proj_info_file(d)
        if not os.path.exists(info_path):
            continue
        try:
            payload = json.load(open(info_path, "r", encoding="utf-8"))
        except Exception:
            continue
        code = payload.get("code")
        if isinstance(code, str) and code:
            codes.add(code)
    return codes


def _project_code_prefix(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-zА-Яа-яЁёЇїІіЄєҐґ ]+", " ", (name or "")).strip()
    parts = [p for p in cleaned.split() if p]
    letters: List[str] = []
    for part in parts:
        for ch in part:
            if ch.isalpha():
                letters.append(ch.upper())
                break
        if len(letters) >= 2:
            break
    fallback = list("PR")
    while len(letters) < 2:
        letters.append(fallback[len(letters) % len(fallback)])
    return "".join(letters[:2])


def generate_project_code(name: str, existing: Optional[Set[str]] = None) -> str:
    prefix = _project_code_prefix(name)
    existing = existing or set()
    for length, upper in ((3, 999), (4, 9999)):
        for num in range(1, upper + 1):
            candidate = f"{prefix}{num:0{length}d}"
            if candidate not in existing:
                return candidate
    # Fallback: timestamp-based unique code
    return f"{prefix}{int(datetime.now().timestamp())}"


def _decode_exif_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bytes):
        for encoding in ("utf-16-le", "utf-8"):
            try:
                decoded = value.decode(encoding, errors="ignore").rstrip("\x00")
                return decoded.strip() or None
            except Exception:
                continue
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, (list, tuple)):
        parts = [_decode_exif_text(part) for part in value]
        joined = ", ".join(part for part in parts if part)
        return joined or None
    try:
        return str(value)
    except Exception:
        return None


def _fraction_to_float(value: Any) -> Optional[float]:
    try:
        return float(value)
    except Exception:
        try:
            num, denom = value
            return float(num) / float(denom) if denom else None
        except Exception:
            return None
    return None


def _convert_to_degrees(values: Any) -> Optional[float]:
    try:
        d, m, s = values
    except Exception:
        return None
    deg = _fraction_to_float(d)
    min_ = _fraction_to_float(m)
    sec = _fraction_to_float(s)
    if deg is None or min_ is None or sec is None:
        return None
    return deg + (min_ / 60.0) + (sec / 3600.0)


def extract_image_metadata(path: str) -> Dict[str, Any]:
    meta: Dict[str, Any] = {"analyzed": True}
    try:
        with Image.open(path) as img:
            raw_exif = getattr(img, "_getexif", lambda: None)()
    except Exception:
        return meta

    if not raw_exif:
        return meta

    exif: Dict[str, Any] = {}
    for tag, value in raw_exif.items():
        name = ExifTags.TAGS.get(tag, tag)
        exif[name] = value

    gps_raw = exif.get("GPSInfo")
    if isinstance(gps_raw, dict):
        gps_decoded: Dict[str, Any] = {}
        for key, value in gps_raw.items():
            readable = ExifTags.GPSTAGS.get(key, key)
            gps_decoded[readable] = value
        lat = _convert_to_degrees(gps_decoded.get("GPSLatitude"))
        lat_ref = gps_decoded.get("GPSLatitudeRef")
        lon = _convert_to_degrees(gps_decoded.get("GPSLongitude"))
        lon_ref = gps_decoded.get("GPSLongitudeRef")
        if lat is not None and lon is not None:
            if isinstance(lat_ref, str) and lat_ref.upper() == "S":
                lat = -lat
            if isinstance(lon_ref, str) and lon_ref.upper() == "W":
                lon = -lon
            meta["gps"] = {"lat": round(lat, 6), "lon": round(lon, 6)}

    address_keys = [
        "XPSubLocation", "SubLocation", "City", "ProvinceState", "State", "Country", "CountryName",
        "ImageDescription", "XPComment", "XPSubject", "XPKeywords", "XPTitle"
    ]
    parts: List[str] = []
    seen: Set[str] = set()
    for key in address_keys:
        text = _decode_exif_text(exif.get(key))
        if not text:
            continue
        norm = text.strip()
        lower = norm.lower()
        if lower in seen:
            continue
        seen.add(lower)
        parts.append(norm)
    if parts:
        meta["address"] = ", ".join(parts)

    for dt_key in ("DateTimeOriginal", "DateTimeDigitized", "DateTime"):
        dt_raw = _decode_exif_text(exif.get(dt_key))
        if not dt_raw:
            continue
        parsed = None
        for fmt in ("%Y:%m:%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                parsed = datetime.strptime(dt_raw, fmt)
                break
            except Exception:
                continue
        if parsed:
            meta["captured_at"] = parsed.strftime("%Y-%m-%d %H:%M")
        else:
            meta["captured_at"] = dt_raw
        break

    make = _decode_exif_text(exif.get("Make"))
    model = _decode_exif_text(exif.get("Model"))
    camera_parts = [part for part in (make, model) if part]
    if camera_parts:
        meta["camera"] = " ".join(camera_parts)

    return meta


def ensure_photo_metadata(project: str, entry: dict) -> bool:
    meta = entry.get("meta")
    if not isinstance(meta, dict):
        meta = {}
    if meta.get("analyzed"):
        entry["meta"] = meta
        return False
    file_name = entry.get("file")
    if not file_name:
        meta["analyzed"] = True
        entry["meta"] = meta
        return True
    path = os.path.join(proj_photos_dir(project), file_name)
    if not os.path.exists(path):
        meta["analyzed"] = True
        entry["meta"] = meta
        return True
    extracted = extract_image_metadata(path)
    combined = {k: v for k, v in meta.items() if k != "analyzed"}
    combined.update(extracted)
    if "analyzed" not in combined:
        combined["analyzed"] = True
    entry["meta"] = combined
    return True


def load_project_photos(name: str) -> List[dict]:
    ensure_project_structure(name)
    path = proj_photos_meta(name)
    if not os.path.exists(path):
        return []
    try:
        raw = json.load(open(path, "r", encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(raw, list):
        return []

    items: List[dict] = []
    changed = False
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        normalized = dict(entry)
        if not normalized.get("id"):
            normalized["id"] = generate_photo_id()
            changed = True
        if not isinstance(normalized.get("meta"), dict):
            normalized["meta"] = {}
            changed = True
        if ensure_photo_metadata(name, normalized):
            changed = True
        items.append(normalized)

    if changed:
        save_project_photos(name, items)
    return items


def project_photo_count(name: str) -> int:
    return len(load_project_photos(name))


def find_photo_entry(project: str, entry_id: str) -> Tuple[int, Optional[dict], List[dict]]:
    photos = load_project_photos(project)
    if not entry_id:
        return -1, None, photos
    for idx, entry in enumerate(photos):
        if not isinstance(entry, dict):
            continue
        if entry.get("id") == entry_id or entry.get("file") == entry_id:
            return idx, entry, photos
    return -1, None, photos


def save_project_photos(name: str, items: List[dict]):
    os.makedirs(proj_photos_dir(name), exist_ok=True)
    json.dump(items, open(proj_photos_meta(name), "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def ensure_project_structure(name: str):
    os.makedirs(proj_path(name), exist_ok=True)
    os.makedirs(proj_pdf_dir(name), exist_ok=True)
    os.makedirs(os.path.join(proj_path(name), "receipts"), exist_ok=True)
    os.makedirs(proj_photos_dir(name), exist_ok=True)
    if not os.path.exists(proj_finance_file(name)):
        json.dump({"requests": {}}, open(proj_finance_file(name), "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    if not os.path.exists(proj_ledger(name)):
        wb = Workbook(); ws = wb.active; ws.title = "Ledger"
        ws.append(["Дата", "Время", "Пользователь", "BSU", "Номер чека", "Сумма", "Файл", "Описание", "Оплачен (1/0/None)"])
        _autosize(ws)
        wb.save(proj_ledger(name))
    if not os.path.exists(proj_photos_meta(name)):
        save_project_photos(name, [])
    if not os.path.exists(proj_info_file(name)):
        existing_codes = _project_existing_codes(exclude=name)
        info = {"name": name, "location": "", "description": "",
                "start_date": "", "end_date": "", "region": "", "code": generate_project_code(name, existing_codes),
                "active": False, "pdf": [], "created": datetime.now().isoformat()}
        json.dump(info, open(proj_info_file(name), "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    _alerts_ensure_storage(name)

def list_projects() -> List[str]:
    ensure_dirs()
    return [d for d in os.listdir(BASE_PATH) if os.path.isdir(os.path.join(BASE_PATH, d))]


def project_from_token(token: str) -> Optional[str]:
    if not token:
        return None
    for name in list_projects():
        if project_token(name) == token:
            return name
    return None


def load_project_info(name: str) -> dict:
    ensure_project_structure(name)
    info = json.load(open(proj_info_file(name), "r", encoding="utf-8"))
    updated = False
    if "region" not in info:
        info["region"] = ""
        updated = True
    if not info.get("code"):
        existing_codes = _project_existing_codes(exclude=name)
        info["code"] = generate_project_code(info.get("name") or name, existing_codes)
        updated = True
    if updated:
        save_project_info(name, info)
    return info

def save_project_info(name: str, info: dict):
    json.dump(info, open(proj_info_file(name), "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def _autosize(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


# ========================== BOT STATE PERSIST ==========================
def load_bot_state() -> dict:
    ensure_dirs()
    if not os.path.exists(BOT_FILE):
        st = {"active_project": None, "admins": [], "version": BOT_VERSION, "revision": BOT_REVISION}
        save_bot_state(st); return st
    return json.load(open(BOT_FILE, "r", encoding="utf-8"))

def save_bot_state(state: dict):
    json.dump(state, open(BOT_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def sync_state():
    st = load_bot_state()
    active_project["name"] = st.get("active_project")
    admins.clear(); admins.update(st.get("admins", []))

def persist_state():
    st = load_bot_state()
    st["active_project"] = active_project["name"]
    st["admins"] = list(admins)
    st["version"] = BOT_VERSION
    st["revision"] = BOT_REVISION
    save_bot_state(st)

def set_active_project(name: Optional[str]):
    active_project["name"] = name
    persist_state()


# ========================== USERS PERSIST ==========================
def load_user(uid: int) -> Optional[dict]:
    path = user_file(uid)
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def save_user(profile: dict):
    ensure_dirs()
    path = user_file(profile["user_id"])
    tmp_path = f"{path}.tmp"
    try:
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(profile, fh, ensure_ascii=False, indent=2)
            fh.flush()
            os.fsync(fh.fileno())
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def load_all_users() -> List[dict]:
    ensure_dirs()
    if not os.path.exists(USERS_PATH):
        return []
    profiles: List[dict] = []
    for name in os.listdir(USERS_PATH):
        if not name.endswith(".json"):
            continue
        path = os.path.join(USERS_PATH, name)
        try:
            profiles.append(json.load(open(path, "r", encoding="utf-8")))
        except Exception:
            continue
    return profiles


def normalize_bsu_code(code: str) -> Optional[str]:
    if not code:
        return None
    digits = re.sub(r"[^0-9]", "", code)
    if len(digits) != 4:
        return None
    return f"BSU-{digits}"


def find_user_by_bsu(code: str, profiles: Optional[List[dict]] = None) -> Optional[dict]:
    normalized = normalize_bsu_code(code)
    if not normalized:
        return None
    profiles = profiles or load_all_users()
    target = normalized.upper()
    for prof in profiles:
        bsu = str(prof.get("bsu") or "").upper()
        if bsu == target:
            return prof
    return None


def find_user_by_username(username: str, profiles: Optional[List[dict]] = None) -> Optional[dict]:
    if not username:
        return None
    normalized = username.lstrip("@").strip().lower()
    if not normalized:
        return None
    profiles = profiles or load_all_users()
    for prof in profiles:
        tg_username = ((prof.get("tg") or {}).get("username") or "").lower()
        if tg_username == normalized:
            return prof
    return None


def resolve_user_reference(msg: types.Message) -> Optional[dict]:
    if msg.forward_from:
        profile = load_user(msg.forward_from.id)
        if profile:
            return profile
    contact = getattr(msg, "contact", None)
    if contact and contact.user_id:
        profile = load_user(contact.user_id)
        if profile:
            return profile

    text = (msg.text or "").strip()
    if not text:
        return None

    profiles = load_all_users()

    bsu_candidate = find_user_by_bsu(text, profiles=profiles)
    if bsu_candidate:
        return bsu_candidate

    digits = re.sub(r"[^0-9]", "", text)
    if digits:
        try:
            profile = load_user(int(digits))
            if profile:
                return profile
        except Exception:
            pass

    if text.startswith("@"):
        username_match = find_user_by_username(text, profiles=profiles)
        if username_match:
            return username_match

    return None


def normalize_profile_receipts(profile: dict) -> bool:
    changed = False
    receipts = profile.get("receipts")
    if isinstance(receipts, dict):
        for items in receipts.values():
            if not isinstance(items, list):
                continue
            for entry in items:
                if "payout" not in entry:
                    entry["payout"] = None
                    changed = True
                history = entry.get("payout_history")
                if not isinstance(history, list):
                    entry["payout_history"] = []
                    changed = True
    return changed


NAME_ALLOWED_CHARS = "A-Za-zА-Яа-яЁёЇїІіЄєҐґʼ'’\-\s"
NAME_VALID_RE = re.compile(rf"^[{NAME_ALLOWED_CHARS}]+$")
SKIP_KEYWORDS = {"нет", "немає", "нема", "no", "skip", "none", "n/a", "-"}


def normalize_person_name(value: Optional[str]) -> str:
    if not value:
        return ""
    text = str(value)
    text = text.replace("`", "'").replace("’", "'").replace("ʼ", "'")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def beautify_name(value: Optional[str]) -> str:
    text = normalize_person_name(value)
    if not text:
        return ""

    def _capitalize(fragment: str) -> str:
        if not fragment:
            return fragment
        return fragment[0].upper() + fragment[1:].lower()

    words: List[str] = []
    for word in text.split(" "):
        if not word:
            continue
        parts: List[str] = []
        for sub in re.split(r"([-'])", word):
            if sub in ("-", "'"):
                parts.append(sub)
            elif sub:
                parts.append(_capitalize(sub))
        words.append("".join(parts))
    return " ".join(words)


def validate_name(value: str) -> bool:
    if not value:
        return False
    text = normalize_person_name(value)
    if len(text) < 2:
        return False
    return bool(NAME_VALID_RE.match(text))


def compose_fullname(last_name: str, first_name: str, middle_name: Optional[str]) -> str:
    parts = [last_name, first_name]
    if middle_name:
        parts.append(middle_name)
    return " ".join(part for part in parts if part)


def sanitize_phone_input(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    raw = text.strip()
    digits = re.sub(r"\D+", "", raw)
    if not digits:
        return None
    if raw.startswith("+"):
        normalized = "+" + digits
    elif len(digits) == 10 and digits.startswith("0"):
        normalized = "+38" + digits
    elif len(digits) >= 11 and digits.startswith("380"):
        normalized = "+" + digits
    elif len(digits) >= 10:
        normalized = "+" + digits
    else:
        return None
    if len(re.sub(r"\D", "", normalized)) < 10:
        return None
    return normalized


def parse_birthdate_text(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    raw = re.sub(r"\s+", "", str(value))
    if not raw:
        return None
    normalized = raw.replace("/", ".").replace("-", ".")
    parts = normalized.split(".")
    try:
        if len(parts) == 3:
            day, month, year = parts
            if len(year) == 2:
                year = "19" + year if int(year) >= 50 else "20" + year
            dt = datetime(int(year), int(month), int(day))
        else:
            dt = datetime.strptime(raw, "%Y%m%d")
    except Exception:
        return None
    if dt.year < 1900:
        return None
    if dt.date() > datetime.now().date():
        return None
    return dt


def compute_age(birthdate: datetime) -> Optional[int]:
    if not birthdate:
        return None
    today = datetime.now().date()
    years = today.year - birthdate.year
    if (today.month, today.day) < (birthdate.month, birthdate.day):
        years -= 1
    return max(years, 0)


def format_birthdate_display(birthdate_iso: Optional[str], lang: Optional[str] = None) -> str:
    if not birthdate_iso:
        return "—"
    try:
        dt = datetime.strptime(birthdate_iso, "%Y-%m-%d")
    except Exception:
        return birthdate_iso
    age = compute_age(dt)
    formatted = dt.strftime("%d.%m.%Y")
    if age is None:
        return formatted
    lang = normalize_lang(lang)
    suffix_map = {
        "uk": "років",
        "en": "yo",
        "de": "J.",
        "pl": "lat",
        "ru": "лет",
    }
    suffix = suffix_map.get(lang, "yo")
    return f"{formatted} ({age} {suffix})"


def ensure_user_dir(uid: int) -> str:
    path = os.path.join(USERS_PATH, str(uid))
    os.makedirs(path, exist_ok=True)
    return path


def user_profile_photo_path(uid: int) -> str:
    return os.path.join(ensure_user_dir(uid), "profile.jpg")


async def store_profile_photo(uid: int, photo: types.PhotoSize) -> Optional[dict]:
    if not photo:
        return None
    ensure_user_dir(uid)
    dest = user_profile_photo_path(uid)
    tmp_path = f"{dest}.tmp"
    try:
        await photo.download(destination_file=tmp_path)
        with Image.open(tmp_path) as img:
            img = ImageOps.exif_transpose(img)
            img = img.convert("RGB")
            img.thumbnail((1280, 1280))
            width, height = img.size
            exif = img.getexif() if hasattr(img, "getexif") else None
            taken_iso = None
            if exif:
                for tag, value in exif.items():
                    name = ExifTags.TAGS.get(tag, tag)
                    if name in ("DateTimeOriginal", "DateTimeDigitized", "DateTime"):
                        text = _decode_exif_text(value)
                        if text:
                            for fmt in ("%Y:%m:%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
                                try:
                                    taken_iso = datetime.strptime(text, fmt).isoformat()
                                    break
                                except Exception:
                                    continue
                            if taken_iso:
                                break
            img.save(tmp_path, format="JPEG", quality=90)
        os.replace(tmp_path, dest)
        meta = {
            "file_id": photo.file_id,
            "file_unique_id": photo.file_unique_id,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "width": width,
            "height": height,
        }
        if taken_iso:
            meta["taken_at"] = taken_iso
        return meta
    except Exception:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass
        raise


def remove_profile_photo(uid: int) -> None:
    path = user_profile_photo_path(uid)
    if os.path.exists(path):
        try:
            os.remove(path)
        except Exception:
            pass


def _safe_name(s: str) -> str:
    s = (s or "").strip()
    if not s: return "User"
    s = re.sub(r"[^A-Za-zА-Яа-я0-9_\- ]+", "", s)
    s = s.replace("  ", " ").strip()
    return s or "User"


def _sanitize_filename(name: str) -> str:
    base = (name or "").strip()
    if not base:
        return "file"
    base = re.sub(r"[\\/:*?\"<>|]+", "_", base)
    base = re.sub(r"[^A-Za-zА-Яа-я0-9_\-]+", "_", base)
    base = re.sub(r"_+", "_", base).strip("_")
    return base or "file"

def ensure_user(uid: int, tg_payload: dict, fullname: Optional[str] = None, phone: Optional[str] = None,
                lang: Optional[str] = None, lang_confirmed: Optional[bool] = None) -> dict:
    ensure_user_dir(uid)
    prof = load_user(uid)
    now_iso = datetime.now(timezone.utc).isoformat()
    if not prof:
        bsu = f"BSU-{random.randint(1000, 9999)}"
        prof = {
            "user_id": uid,
            "tg": dict(tg_payload),
            "first_name": "",
            "last_name": "",
            "middle_name": "",
            "fullname": fullname or tg_payload.get("first_name") or f"User{uid}",
            "phone": phone or "",
            "region": "",
            "birthdate": "",
            "photo": {},
            "bsu": bsu,
            "counters": {"receipt_seq": 0},
            "receipts": {},
            "payouts": [],
            "lang": normalize_lang(lang) if lang else DEFAULT_LANG,
            "lang_confirmed": bool(lang),
            "profile_completed": False,
            "created_at": now_iso,
            "updated_at": now_iso,
        }
    else:
        prof["tg"] = {**prof.get("tg", {}), **tg_payload}
        prof.setdefault("first_name", "")
        prof.setdefault("last_name", "")
        prof.setdefault("middle_name", "")
        prof.setdefault("fullname", fullname or prof.get("fullname") or tg_payload.get("first_name") or f"User{uid}")
        prof.setdefault("phone", "")
        prof.setdefault("region", "")
        prof.setdefault("birthdate", "")
        prof.setdefault("photo", {})
        prof.setdefault("counters", {"receipt_seq": 0})
        prof.setdefault("receipts", {})
        prof.setdefault("payouts", [])
        prof.setdefault("lang", DEFAULT_LANG)
        prof.setdefault("lang_confirmed", bool(prof.get("lang") in LANG_CODES))
        prof.setdefault("profile_completed", False)
        prof.setdefault("created_at", now_iso)
        if fullname:
            prof["fullname"] = fullname
        if phone:
            prof["phone"] = phone
        if lang is not None:
            prof["lang"] = normalize_lang(lang)
        if lang_confirmed is not None:
            prof["lang_confirmed"] = bool(lang_confirmed)
        if "bsu" not in prof:
            prof["bsu"] = f"BSU-{random.randint(1000, 9999)}"
        prof["updated_at"] = now_iso
    if normalize_profile_receipts(prof):
        pass
    save_user(prof)
    return prof


def get_user_lang(uid: int) -> str:
    prof = load_user(uid)
    if not prof:
        return DEFAULT_LANG
    return normalize_lang(prof.get("lang"))


def set_user_lang(uid: int, lang: str, confirmed: bool = True) -> dict:
    prof = load_user(uid)
    if not prof:
        raise ValueError("User profile must exist before setting language")
    prof["lang"] = normalize_lang(lang)
    prof["lang_confirmed"] = bool(confirmed)
    save_user(prof)
    return prof

def next_receipt_no(prof: dict) -> str:
    prof["counters"]["receipt_seq"] = int(prof["counters"].get("receipt_seq", 0)) + 1
    save_user(prof)
    return f"RID-{prof['counters']['receipt_seq']:04d}"

def user_append_receipt(uid: int, project: str, date: str, time: str, amount: float, filename: str, desc: str, paid, receipt_no: str):
    prof = load_user(uid) or {"user_id": uid, "receipts": {}}
    recmap = prof.get("receipts", {})
    lst = recmap.get(project, [])
    lst.append({
        "date": date,
        "time": time,
        "sum": float(amount),
        "file": filename,
        "desc": desc or "",
        "paid": paid,
        "receipt_no": receipt_no,
        "payout": None,
        "payout_history": []
    })
    recmap[project] = lst
    prof["receipts"] = recmap
    save_user(prof)

def user_project_receipts(uid: int, project: str) -> List[dict]:
    prof = load_user(uid)
    if not prof: return []
    changed = normalize_profile_receipts(prof)
    if changed:
        save_user(prof)
    return list((prof.get("receipts", {}) or {}).get(project, []))

def user_project_stats(uid: int, project: str) -> Dict[str, float]:
    recs = user_project_receipts(uid, project)
    total = 0.0
    paid_sum = 0.0
    unpaid_sum = 0.0
    pending_sum = 0.0
    unspecified_sum = 0.0
    for r in recs:
        try:
            amount = float(r.get("sum", 0.0))
        except (TypeError, ValueError):
            amount = 0.0
        total += amount
        paid_flag = r.get("paid")
        payout_status = (r.get("payout") or {}).get("status") if isinstance(r.get("payout"), dict) else None
        if paid_flag is True:
            paid_sum += amount
        elif paid_flag is False:
            if payout_status in ("pending", "approved"):
                pending_sum += amount
            else:
                unpaid_sum += amount
        else:
            unspecified_sum += amount
    return {
        "count": len(recs),
        "total": round(total, 2),
        "paid": round(paid_sum, 2),
        "unpaid": round(unpaid_sum, 2),
        "pending": round(pending_sum, 2),
        "unspecified": round(unspecified_sum, 2)
    }


def iter_user_payout_refs(prof: dict) -> List[dict]:
    payouts = prof.get("payouts", []) or []
    normalized: List[dict] = []
    changed = False
    for entry in payouts:
        if isinstance(entry, dict) and entry.get("id"):
            req_id = entry.get("id")
            project = entry.get("project")
            code = entry.get("code")
            normalized.append({"id": req_id, "project": project, "code": code})
        elif isinstance(entry, str):
            req_id = entry
            obj = finance_load_request(req_id)
            project = obj.get("project") if obj else None
            code = obj.get("code") if obj else None
            normalized.append({"id": req_id, "project": project, "code": code})
            changed = True
        else:
            changed = True
    if changed:
        prof["payouts"] = normalized
        save_user(prof)
    return normalized


# ========================== FINANCE FILES ==========================
def load_finance_data(project: str) -> dict:
    ensure_project_structure(project)
    path = proj_finance_file(project)
    if not os.path.exists(path):
        data = {"requests": {}}
        save_finance_data(project, data)
        return data
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh) or {}
    except Exception:
        data = {"requests": {}}
    if not isinstance(data, dict):
        data = {"requests": {}}
    if not isinstance(data.get("requests"), dict):
        data["requests"] = {}
    return data


def save_finance_data(project: str, data: dict):
    ensure_project_structure(project)
    path = proj_finance_file(project)
    json.dump(data, open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def finance_generate_code() -> str:
    ensure_dirs()
    existing = set()
    for project in list_projects():
        data = load_finance_data(project)
        for item in data.get("requests", {}).values():
            code = item.get("code")
            if code:
                existing.add(code)
    if os.path.isdir(FIN_PATH):
        for f in os.listdir(FIN_PATH):
            if not f.endswith(".json"):
                continue
            try:
                payload = json.load(open(os.path.join(FIN_PATH, f), "r", encoding="utf-8"))
            except Exception:
                continue
            code = payload.get("code")
            if code:
                existing.add(code)
    while True:
        code = f"ID-BRD-{random.randint(0, 9999):04d}"
        if code not in existing:
            return code


def finance_new_request(uid: int, project: str, receipts: List[dict]) -> dict:
    ensure_dirs()
    data = load_finance_data(project)
    ts = int(datetime.now().timestamp())
    req_id = f"req_{ts}_{uid}"
    while req_id in data.get("requests", {}):
        ts += 1
        req_id = f"req_{ts}_{uid}"
    code = finance_generate_code()
    files: List[str] = []
    items: List[dict] = []
    total = 0.0
    now_iso = datetime.now().isoformat()
    for rec in receipts:
        file_name = rec.get("file") or ""
        files.append(file_name)
        try:
            amount = float(rec.get("sum", 0.0))
        except (TypeError, ValueError):
            amount = 0.0
        total += amount
        items.append({
            "file": file_name,
            "receipt_no": rec.get("receipt_no"),
            "amount": round(amount, 2),
            "desc": rec.get("desc"),
            "status": "pending",
            "updated_at": now_iso
        })
    payload = {
        "id": req_id,
        "code": code,
        "user_id": uid,
        "project": project,
        "sum": round(float(total), 2),
        "files": files,
        "items": items,
        "status": "pending",
        "approved_by": None,
        "approved_at": None,
        "confirmed_at": None,
        "created_at": now_iso,
        "history": [{"status": "pending", "timestamp": now_iso}]
    }
    data.setdefault("requests", {})[req_id] = payload
    save_finance_data(project, data)
    prof = load_user(uid) or {}
    arr = prof.get("payouts", [])
    entry = {"id": req_id, "project": project, "code": code}
    if not any(isinstance(x, dict) and x.get("id") == req_id for x in arr):
        arr.append(entry)
        prof["payouts"] = arr
        save_user(prof)
    fin_state_set(project, uid, req_id, "pending")
    update_receipts_for_request(uid, project, files, "pending", payload)
    return finance_request_defaults(payload)


def finance_load_request(req_id: str, project_hint: Optional[str]=None) -> Optional[dict]:
    projects: List[str]
    if project_hint:
        projects = [project_hint]
    else:
        projects = list_projects()
    for name in projects:
        if not name:
            continue
        data = load_finance_data(name)
        obj = (data.get("requests") or {}).get(req_id)
        if obj:
            if not obj.get("project"):
                obj["project"] = name
                finance_save_request(obj)
            return finance_request_defaults(obj)
    if project_hint:
        return finance_load_request(req_id, None)
    legacy_path = os.path.join(FIN_PATH, f"{req_id}.json")
    if os.path.exists(legacy_path):
        try:
            obj = json.load(open(legacy_path, "r", encoding="utf-8"))
        except Exception:
            obj = None
        if obj:
            project = obj.get("project") or project_hint
            if project:
                obj["project"] = project
                finance_save_request(obj)
                try:
                    os.remove(legacy_path)
                except Exception:
                    pass
            return finance_request_defaults(obj)
    return None


def finance_save_request(obj: dict):
    project = obj.get("project")
    if not project:
        return
    obj = finance_request_defaults(obj) or obj
    data = load_finance_data(project)
    data.setdefault("requests", {})[obj["id"]] = obj
    save_finance_data(project, data)


def finance_list(filter_status: Optional[str]=None) -> List[dict]:
    ensure_dirs()
    out = []
    for project in list_projects():
        data = load_finance_data(project)
        for req in data.get("requests", {}).values():
            if not req.get("project"):
                req["project"] = project
                finance_save_request(req)
            if (filter_status is None) or (req.get("status") == filter_status):
                out.append(finance_request_defaults(req) or req)
    out.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return out


def project_fin_state_file(name: str) -> str:
    return os.path.join(proj_path(name), "finance_state.json")


def load_project_fin_state(name: str) -> dict:
    ensure_project_structure(name)
    path = project_fin_state_file(name)
    if not os.path.exists(path):
        data = {"active_requests": {}}
        save_project_fin_state(name, data)
        return data
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
            if "active_requests" not in data:
                data["active_requests"] = {}
            return data
    except Exception:
        data = {"active_requests": {}}
        save_project_fin_state(name, data)
        return data


def save_project_fin_state(name: str, data: dict):
    ensure_project_structure(name)
    path = project_fin_state_file(name)
    json.dump(data, open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def fin_state_set(name: str, uid: int, req_id: str, status: str):
    state = load_project_fin_state(name)
    state.setdefault("active_requests", {})[str(uid)] = {
        "request_id": req_id,
        "status": status,
        "project": name,
        "updated_at": datetime.now().isoformat()
    }
    save_project_fin_state(name, state)


def fin_state_clear(name: str, uid: int):
    state = load_project_fin_state(name)
    if str(uid) in state.get("active_requests", {}):
        state["active_requests"].pop(str(uid), None)
        save_project_fin_state(name, state)


def fin_state_get(name: str, uid: int) -> Optional[dict]:
    state = load_project_fin_state(name)
    return (state.get("active_requests", {}) or {}).get(str(uid))


def finance_active_request_for_user(uid: int, project: str) -> Optional[dict]:
    entry = fin_state_get(project, uid)
    if not entry:
        return None
    req = finance_load_request(entry.get("request_id"), project)
    if req and req.get("status") in {"pending", "approved"}:
        return req
    fin_state_clear(project, uid)
    return None


def finance_request_defaults(obj: Optional[dict]) -> Optional[dict]:
    if not obj:
        return obj
    if not isinstance(obj.get("history"), list):
        obj["history"] = []
    if not isinstance(obj.get("items"), list):
        obj["items"] = []
    return obj


def finance_append_history(obj: dict, status: str, extra: Optional[dict] = None):
    if obj is None:
        return
    entry = {"status": status, "timestamp": datetime.now().isoformat()}
    if extra:
        entry.update(extra)
    history = obj.setdefault("history", [])
    history.append(entry)


def finance_update_items_status(obj: dict, status: str, timestamp: Optional[str] = None):
    if obj is None:
        return
    if timestamp is None:
        timestamp = datetime.now().isoformat()
    items = obj.setdefault("items", [])
    for item in items:
        if isinstance(item, dict):
            item["status"] = status
            item["updated_at"] = timestamp


def update_receipts_for_request(uid: int, project: str, files: List[str], status: str, request: dict):
    prof = load_user(uid) or {}
    recmap = prof.get("receipts", {})
    recs = recmap.get(project, [])
    if not isinstance(recs, list) or not recs:
        return
    now_iso = datetime.now().isoformat()
    req_id = request.get("id")
    req_code = request.get("code") or req_id
    changed = False
    for entry in recs:
        if entry.get("file") not in files:
            continue
        history = entry.get("payout_history")
        if not isinstance(history, list):
            history = []
            entry["payout_history"] = history
        try:
            amount_value = float(entry.get("sum", 0.0))
        except (TypeError, ValueError):
            amount_value = 0.0
        history.append({
            "status": status,
            "timestamp": now_iso,
            "request_id": req_id,
            "code": req_code,
            "project": project,
            "amount": amount_value
        })
        payout = entry.get("payout") if isinstance(entry.get("payout"), dict) else {}
        if status in ("pending", "approved"):
            payout.update({
                "request_id": req_id,
                "code": req_code,
                "status": status,
                "updated_at": now_iso
            })
            if status == "pending":
                payout.setdefault("assigned_at", now_iso)
            if status == "approved":
                payout["approved_at"] = now_iso
            entry["payout"] = payout
        elif status == "confirmed":
            payout.update({
                "request_id": req_id,
                "code": req_code,
                "status": "confirmed",
                "updated_at": now_iso,
                "confirmed_at": now_iso,
                "assigned_at": payout.get("assigned_at", now_iso),
                "approved_at": payout.get("approved_at")
            })
            entry["payout"] = payout
            entry["paid"] = True
            entry["paid_at"] = now_iso
            entry["paid_request_id"] = req_id
            entry["paid_request_code"] = req_code
        elif status == "closed":
            if entry.get("paid") is not True:
                entry.pop("paid_request_id", None)
                entry.pop("paid_request_code", None)
                entry.pop("paid_at", None)
            entry["payout"] = None
        changed = True
    if changed:
        recmap[project] = recs
        prof["receipts"] = recmap
        save_user(prof)


# ========================== RECEIPT SAVE (FILE + EXCEL) ==========================
def ensure_user_receipts_dir(project: str, uid: int):
    os.makedirs(proj_receipts_dir(project, uid), exist_ok=True)

def _excel_ensure_sheet_user(wb: Workbook, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Дата", "Время", "Пользователь", "BSU", "Номер чека", "Сумма", "Файл", "Описание", "Оплачен (1/0/None)"])
        _autosize(ws)

def _excel_append_row(project: str, username: str, bsu: str, receipt_no: str, amount: float, filename: str, desc: str, paid):
    path = proj_ledger(project)
    if not os.path.exists(path):
        wb = Workbook(); ws = wb.active; ws.title = "Ledger"
        ws.append(["Дата", "Время", "Пользователь", "BSU", "Номер чека", "Сумма", "Файл", "Описание", "Оплачен (1/0/None)"])
        _autosize(ws)
        wb.save(path)
    wb = load_workbook(path)
    if "Ledger" not in wb.sheetnames:
        ws = wb.create_sheet("Ledger", 0)
        ws.append(["Дата", "Время", "Пользователь", "BSU", "Номер чека", "Сумма", "Файл", "Описание", "Оплачен (1/0/None)"])

    ws = wb["Ledger"]
    now = datetime.now()
    paid_cell = None if paid is None else (1 if paid else 0)
    ws.append([now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), username, bsu, receipt_no, float(amount), filename, desc or "", paid_cell])
    _autosize(ws)

    # Лист пользователя
    user_sheet = f"{_safe_name(username)}_{bsu}"
    _excel_ensure_sheet_user(wb, user_sheet)
    wsu = wb[user_sheet]
    wsu.append([now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), username, bsu, receipt_no, float(amount), filename, desc or "", paid_cell])
    _autosize(wsu)

    wb.save(path)

def save_receipt(project: str, uid: int, amount: float, tmp_img: str, desc: str, paid) -> Tuple[str, str, datetime, str]:
    ensure_user_receipts_dir(project, uid)
    now = datetime.now()
    prof = load_user(uid) or {}
    username = prof.get("fullname") or f"User{uid}"
    bsu = prof.get("bsu", f"BSU-{uid%10000:04d}")
    rid = next_receipt_no(prof)  # RID-XXXX

    base = f"{_safe_name(username)}_BSU{bsu.replace('BSU-','')}_{rid}_{now.strftime('%Y-%m-%d_%H-%M-%S')}_amt-{amount:.2f}.jpg"
    dst = os.path.join(proj_receipts_dir(project, uid), base)
    # перестраховка на уникальность
    i = 1
    while os.path.exists(dst):
        base = f"{_safe_name(username)}_BSU{bsu.replace('BSU-','')}_{rid}_{now.strftime('%Y-%m-%d_%H-%M-%S')}_{i}_amt-{amount:.2f}.jpg"
        dst = os.path.join(proj_receipts_dir(project, uid), base)
        i += 1
    if tmp_img and os.path.exists(tmp_img):
        os.replace(tmp_img, dst)

    _excel_append_row(project, username, bsu, rid, amount, base, desc, paid)
    user_append_receipt(uid, project, now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), amount, base, desc, paid, rid)
    return base, dst, now, rid


# ========================== HELPERS & MENУС ==========================
def fmt_money(x: float) -> str: return f"{x:.2f}"


def h(value: Any) -> str:
    if value is None:
        return ""
    return html_escape(str(value), quote=False)


def format_datetime_short(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(value)


def format_day_month(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value).strftime("%d.%m")
    except Exception:
        return ""


NP_FIELD_LABELS = {
    "uk": {
        "ttn": "ТТН",
        "status": "Статус",
        "last_update": "Оновлено",
        "delivery_date": "Планова доставка",
        "estimated_date": "Орієнтовно",
        "recipient": "Ім’я",
        "recipient_city": "Місто",
        "recipient_warehouse": "Відділення",
        "sender": "Ім’я",
        "sender_city": "Місто",
        "sender_warehouse": "Відділення",
        "service_type": "Сервіс",
        "weight": "Вага",
        "cost": "Оціночна вартість",
        "section_summary": "СВОДКА",
        "section_recipient": "ОТРИМУВАЧ",
        "section_sender": "ВІДПРАВНИК",
        "section_parcel": "ПОСИЛКА",
    },
    "en": {
        "ttn": "TTN",
        "status": "Status",
        "last_update": "Updated",
        "delivery_date": "Planned delivery",
        "estimated_date": "Estimated",
        "recipient": "Name",
        "recipient_city": "City",
        "recipient_warehouse": "Branch",
        "sender": "Name",
        "sender_city": "City",
        "sender_warehouse": "Branch",
        "service_type": "Service",
        "weight": "Weight",
        "cost": "Declared value",
        "section_summary": "Summary",
        "section_recipient": "Recipient",
        "section_sender": "Sender",
        "section_parcel": "Parcel",
    },
    "de": {
        "ttn": "TTN",
        "status": "Status",
        "last_update": "Aktualisiert",
        "delivery_date": "Geplante Zustellung",
        "estimated_date": "Voraussichtlich",
        "recipient": "Name",
        "recipient_city": "Stadt",
        "recipient_warehouse": "Filiale",
        "sender": "Name",
        "sender_city": "Stadt",
        "sender_warehouse": "Filiale",
        "service_type": "Service",
        "weight": "Gewicht",
        "cost": "Deklarierter Wert",
        "section_summary": "Übersicht",
        "section_recipient": "Empfänger",
        "section_sender": "Absender",
        "section_parcel": "Sendung",
    },
    "pl": {
        "ttn": "TTN",
        "status": "Status",
        "last_update": "Aktualizacja",
        "delivery_date": "Planowana dostawa",
        "estimated_date": "Szacunkowo",
        "recipient": "Imię",
        "recipient_city": "Miasto",
        "recipient_warehouse": "Oddział",
        "sender": "Imię",
        "sender_city": "Miasto",
        "sender_warehouse": "Oddział",
        "service_type": "Usługa",
        "weight": "Waga",
        "cost": "Deklarowana wartość",
        "section_summary": "Podsumowanie",
        "section_recipient": "Odbiorca",
        "section_sender": "Nadawca",
        "section_parcel": "Przesyłka",
    },
    "ru": {
        "ttn": "ТТН",
        "status": "Статус",
        "last_update": "Обновлено",
        "delivery_date": "Плановая доставка",
        "estimated_date": "Ориентировочно",
        "recipient": "Имя",
        "recipient_city": "Город",
        "recipient_warehouse": "Отделение",
        "sender": "Имя",
        "sender_city": "Город",
        "sender_warehouse": "Отделение",
        "service_type": "Сервис",
        "weight": "Вес",
        "cost": "Оценочная стоимость",
        "section_summary": "СВОДКА",
        "section_recipient": "ПОЛУЧАТЕЛЬ",
        "section_sender": "ОТПРАВИТЕЛЬ",
        "section_parcel": "ПОСЫЛКА",
    },
}

NP_WEIGHT_SUFFIX = {
    "uk": " кг",
    "en": " kg",
    "de": " kg",
    "pl": " kg",
    "ru": " кг",
}

NP_COST_SUFFIX = {
    "uk": " грн",
    "en": " UAH",
    "de": " UAH",
    "pl": " UAH",
    "ru": " грн",
}

NP_SECTION_ICONS = {
    "section_summary": "📌",
    "section_recipient": "🎯",
    "section_sender": "🚚",
    "section_parcel": "📦",
}

NP_TTN_TITLE = {
    "uk": "🧾 <b>Nova Poshta — квитанція</b>\n🔖 ТТН: <code>{ttn}</code>",
    "en": "🧾 <b>Nova Poshta — receipt</b>\n🔖 TTN: <code>{ttn}</code>",
    "de": "🧾 <b>Nova Poshta — Beleg</b>\n🔖 TTN: <code>{ttn}</code>",
    "pl": "🧾 <b>Nova Poshta — potwierdzenie</b>\n🔖 TTN: <code>{ttn}</code>",
    "ru": "🧾 <b>Nova Poshta — квитанция</b>\n🔖 ТТН: <code>{ttn}</code>",
}

NP_ASSIGN_INFO_LINE = {
    "uk": "🏢 Передав адміністратор: {name} • {time}",
    "en": "🏢 Assigned by admin {name} • {time}",
    "de": "🏢 Zugewiesen durch Admin {name} • {time}",
    "pl": "🏢 Przypisane przez admina {name} • {time}",
    "ru": "🏢 Передал администратор: {name} • {time}",
}

NP_ASSIGN_DELIVERED_LINE = {
    "uk": "✅ Отримання підтверджено: {time}",
    "en": "✅ Delivery confirmed: {time}",
    "de": "✅ Empfang bestätigt: {time}",
    "pl": "✅ Odbiór potwierdzony: {time}",
    "ru": "✅ Получение подтверждено: {time}",
}

NP_ADMIN_NOTE_PREFIX = {
    "uk": "💬 Коментар адміністратора: {note}",
    "en": "💬 Admin note: {note}",
    "de": "💬 Hinweis des Admins: {note}",
    "pl": "💬 Notatka administratora: {note}",
    "ru": "💬 Комментарий администратора: {note}",
}

NP_COMMENT_SECTION_TITLE = {
    "uk": "💬 Коментарі ({count})",
    "en": "💬 Comments ({count})",
    "de": "💬 Kommentare ({count})",
    "pl": "💬 Komentarze ({count})",
    "ru": "💬 Комментарии ({count})",
}

NP_DELIVERY_RECEIPT_TITLE = {
    "uk": "📦 Посилка отримана",
    "en": "📦 Parcel received",
    "de": "📦 Sendung erhalten",
    "pl": "📦 Przesyłka odebrana",
    "ru": "📦 Посылка получена",
}

NP_DELIVERY_STATUS_CONFIRMED = {
    "uk": "Підтверджено",
    "en": "Confirmed",
    "de": "Bestätigt",
    "pl": "Potwierdzono",
    "ru": "Подтверждено",
}

NP_DELIVERY_RECEIPT_LABELS = {
    "uk": {"ttn": "ТТН", "recipient": "Отримувач", "date": "Дата", "status": "Статус"},
    "en": {"ttn": "TTN", "recipient": "Recipient", "date": "Date", "status": "Status"},
    "de": {"ttn": "TTN", "recipient": "Empfänger", "date": "Datum", "status": "Status"},
    "pl": {"ttn": "TTN", "recipient": "Odbiorca", "date": "Data", "status": "Status"},
    "ru": {"ttn": "ТТН", "recipient": "Получатель", "date": "Дата", "status": "Статус"},
}

NP_DATETIME_CARD_FORMATS = {
    "uk": "%d.%m.%Y • %H:%M",
    "en": "%d.%m.%Y • %H:%M",
    "de": "%d.%m.%Y • %H:%M",
    "pl": "%d.%m.%Y • %H:%M",
    "ru": "%d.%m.%Y • %H:%M",
}

NP_REFRESH_BUTTON_LABEL = {
    "uk": "🔄 Оновити",
    "en": "🔄 Refresh",
    "de": "🔄 Aktualisieren",
    "pl": "🔄 Odśwież",
    "ru": "🔄 Обновить",
}

NP_NOTE_BUTTON_LABEL = {
    "uk": "💬 Додати коментар",
    "en": "💬 Add comment",
    "de": "💬 Kommentar hinzufügen",
    "pl": "💬 Dodaj komentarz",
    "ru": "💬 Добавить комментарий",
}

NP_CLOSE_BUTTON_LABEL = {
    "uk": "❌ Закрити",
    "en": "❌ Close",
    "de": "❌ Schließen",
    "pl": "❌ Zamknij",
    "ru": "❌ Закрыть",
}

NP_BOOKMARK_ADD_BUTTON = {
    "uk": "⭐ Додати в обране",
    "en": "⭐ Bookmark",
    "de": "⭐ Merken",
    "pl": "⭐ Oznacz",
    "ru": "⭐ Отметить",
}

NP_BOOKMARK_REMOVE_BUTTON = {
    "uk": "⭐ Прибрати з обраного",
    "en": "⭐ Remove bookmark",
    "de": "⭐ Entfernen",
    "pl": "⭐ Usuń oznaczenie",
    "ru": "⭐ Удалить отметку",
}

NP_MARK_RECEIVED_LABEL = {
    "uk": "✅ Посилку отримано",
    "en": "✅ Parcel received",
    "de": "✅ Sendung erhalten",
    "pl": "✅ Przesyłka odebrana",
    "ru": "✅ Посылка получена",
}

NP_CANCEL_BUTTON_LABEL = {
    "uk": "❌ Скасувати",
    "en": "❌ Cancel",
    "de": "❌ Abbrechen",
    "pl": "❌ Anuluj",
    "ru": "❌ Отменить",
}

NP_ASSIGN_SKIP_BUTTON_LABEL = {
    "uk": "⏭ Пропустити",
    "en": "⏭ Skip",
    "de": "⏭ Überspringen",
    "pl": "⏭ Pomiń",
    "ru": "⏭ Пропустить",
}

NP_CANCEL_WORDS = {"отмена", "cancel", "скасувати", "відміна", "anuluj", "abbrechen", "stop"}


def _np_pick(lang: str, mapping: Dict[str, str]) -> str:
    return mapping.get(lang) or mapping.get(DEFAULT_LANG) or next(iter(mapping.values()))


def _np_format_weight(lang: str, value: Any) -> str:
    try:
        num = float(str(value).replace(",", "."))
        if abs(num - round(num)) < 0.01:
            num_disp = str(int(round(num)))
        else:
            num_disp = f"{num:.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value)
    return f"{num_disp}{_np_pick(lang, NP_WEIGHT_SUFFIX)}"


def _np_format_cost(lang: str, value: Any) -> str:
    try:
        num = float(str(value).replace(",", "."))
    except Exception:
        return str(value)
    return f"{fmt_money(num)}{_np_pick(lang, NP_COST_SUFFIX)}"


def format_np_short_entry(payload: Optional[dict]) -> str:
    if not payload:
        return ""
    status = str(payload.get("Status") or payload.get("StatusCode") or payload.get("StatusDescription") or "").strip()
    city = str(payload.get("CityRecipient") or payload.get("CitySender") or "").strip()
    if status and city:
        return f"{status} · {city}"
    return status or city


def _np_extract_value(payload: Optional[dict], *keys: str) -> str:
    if not payload:
        return ""
    for key in keys:
        if key is None:
            continue
        raw = payload.get(key)
        if raw is None:
            continue
        if isinstance(raw, (int, float)):
            value = f"{raw}"
        else:
            value = str(raw)
        value = value.strip()
        if value:
            return value
    return ""


def _np_render_receipt_block(entries: List[Tuple[str, ...]]) -> str:
    items: List[Dict[str, Any]] = []
    for entry in entries:
        if not entry:
            continue
        kind = entry[0]
        if kind == "sep":
            if items and items[-1]["type"] != "sep":
                items.append({"type": "sep"})
            continue
        if kind == "section":
            title = str(entry[1]).strip()
            if title:
                items.append({"type": "section", "text": title})
            continue
        label = str(entry[1]).strip()
        value = ""
        if len(entry) > 2 and entry[2] is not None:
            value = str(entry[2]).strip()
        if not value and kind == "kv_opt":
            continue
        if not value:
            value = "—"
        items.append({"type": "kv", "label": label or "—", "value": value})

    while items and items[-1]["type"] == "sep":
        items.pop()

    if not items:
        return "—"

    kv_items = [item for item in items if item["type"] == "kv"]
    label_width = max((len(item["label"]) for item in kv_items), default=0)
    value_column = max(20, min(40, label_width + 4))

    lines: List[str] = []
    for item in items:
        if item["type"] == "sep":
            if lines and lines[-1] != "":
                lines.append("")
            continue
        if item["type"] == "section":
            if lines and lines[-1] != "":
                lines.append("")
            lines.append(item["text"])
            continue
        label = item["label"].strip()
        prefix = f"{label}:" if label else ""
        gap = value_column - len(prefix)
        if gap < 2:
            gap = 2
        spaces = " " * gap
        raw_value = item["value"]
        value_lines = [line.strip() for line in raw_value.splitlines()] or ["—"]
        first_line = value_lines[0] or "—"
        line_prefix = prefix + spaces
        lines.append(line_prefix + first_line)
        indent = " " * len(line_prefix)
        for extra in value_lines[1:]:
            extra_line = extra or "—"
            lines.append(indent + extra_line)

    while lines and lines[-1] == "":
        lines.pop()

    return "\n".join(lines)


def np_format_delivery_timestamp(value: Optional[str], lang: str) -> str:
    if not value:
        return ""
    raw = str(value)
    try:
        dt = datetime.fromisoformat(raw)
    except Exception:
        try:
            dt = datetime.strptime(raw, "%Y-%m-%d %H:%M")
        except Exception:
            return raw
    fmt = NP_DATETIME_CARD_FORMATS.get(lang) or NP_DATETIME_CARD_FORMATS.get(DEFAULT_LANG) or "%d.%m.%Y • %H:%M"
    try:
        return dt.strftime(fmt)
    except Exception:
        return raw


def np_render_delivery_receipt(lang: str, ttn: Any, recipient: Any, delivered_at: Optional[str]) -> str:
    labels = (
        NP_DELIVERY_RECEIPT_LABELS.get(lang)
        or NP_DELIVERY_RECEIPT_LABELS.get(DEFAULT_LANG)
        or next(iter(NP_DELIVERY_RECEIPT_LABELS.values()))
    )
    entries: List[Tuple[str, ...]] = [
        ("kv", labels.get("ttn", "TTN"), str(ttn) if ttn is not None else ""),
        ("kv", labels.get("recipient", "Recipient"), str(recipient) if recipient is not None else ""),
        ("kv", labels.get("date", "Date"), np_format_delivery_timestamp(delivered_at, lang)),
        ("kv", labels.get("status", "Status"), _np_pick(lang, NP_DELIVERY_STATUS_CONFIRMED)),
    ]
    block_plain = _np_render_receipt_block(entries)
    header = _np_pick(lang, NP_DELIVERY_RECEIPT_TITLE)
    return f"{header}\n━━━━━━━━━━━━━━━━━━\n<pre>{html_escape(block_plain)}</pre>"


def format_np_status(uid: int, ttn: str, payload: Optional[dict],
                     note_entries: Optional[List[Dict[str, Any]]] = None,
                     assignment: Optional[dict] = None) -> str:
    lang = resolve_lang(uid)
    labels = NP_FIELD_LABELS.get(lang) or NP_FIELD_LABELS[DEFAULT_LANG]
    header = _np_pick(lang, NP_TTN_TITLE).format(ttn=h(ttn))

    def field_label(key: str) -> str:
        return labels.get(key, key)

    def section_title(key: str) -> str:
        base = labels.get(key, key)
        icon = NP_SECTION_ICONS.get(key)
        title = base.strip().upper()
        if icon and not title.startswith(icon):
            return f"{icon} {title}"
        return title

    summary_rows: List[Tuple[str, ...]] = [
        ("section", section_title("section_summary")),
        ("kv", field_label("ttn"), str(ttn)),
        ("kv", field_label("status"), _np_extract_value(payload, "Status", "StatusDescription", "StatusCode") or "—"),
    ]

    last_update = _np_extract_value(payload, "LastUpdatedDate")
    if last_update:
        summary_rows.append(("kv_opt", field_label("last_update"), last_update))
    delivery_date = _np_extract_value(payload, "ScheduledDeliveryDate")
    if delivery_date:
        summary_rows.append(("kv_opt", field_label("delivery_date"), delivery_date))
    estimated_date = _np_extract_value(payload, "EstimatedDeliveryDate")
    if estimated_date:
        summary_rows.append(("kv_opt", field_label("estimated_date"), estimated_date))

    recipient_section: List[Tuple[str, ...]] = []
    recipient_name = _np_extract_value(payload, "RecipientFullName", "RecipientDescription", "RecipientName")
    if recipient_name:
        recipient_section.append(("kv", field_label("recipient"), recipient_name))
    recipient_city = _np_extract_value(payload, "CityRecipient")
    if recipient_city:
        recipient_section.append(("kv_opt", field_label("recipient_city"), recipient_city))
    recipient_branch = _np_extract_value(payload, "WarehouseRecipient")
    if recipient_branch:
        recipient_section.append(("kv_opt", field_label("recipient_warehouse"), recipient_branch))

    sender_section: List[Tuple[str, ...]] = []
    sender_name = _np_extract_value(payload, "SenderFullNameEW", "SenderFullName", "SenderName")
    if sender_name:
        sender_section.append(("kv", field_label("sender"), sender_name))
    sender_city = _np_extract_value(payload, "CitySender")
    if sender_city:
        sender_section.append(("kv_opt", field_label("sender_city"), sender_city))
    sender_branch = _np_extract_value(payload, "WarehouseSender")
    if sender_branch:
        sender_section.append(("kv_opt", field_label("sender_warehouse"), sender_branch))

    parcel_section: List[Tuple[str, ...]] = []
    service_type = _np_extract_value(payload, "ServiceType")
    if service_type:
        parcel_section.append(("kv_opt", field_label("service_type"), service_type))
    weight_raw = _np_extract_value(payload, "DocumentWeight", "FactualWeight")
    if weight_raw:
        parcel_section.append(("kv", field_label("weight"), _np_format_weight(lang, weight_raw)))
    cost_raw = _np_extract_value(payload, "DocumentCost", "EstimatedDeliveryCost")
    if cost_raw:
        parcel_section.append(("kv", field_label("cost"), _np_format_cost(lang, cost_raw)))

    receipt_entries: List[Tuple[str, ...]] = list(summary_rows)

    def push_section(title_key: str, rows: List[Tuple[str, ...]]):
        if not rows:
            return
        if receipt_entries:
            receipt_entries.append(("sep",))
        receipt_entries.append(("section", section_title(title_key)))
        receipt_entries.extend(rows)

    push_section("section_recipient", recipient_section)
    push_section("section_sender", sender_section)
    push_section("section_parcel", parcel_section)

    block_plain = _np_render_receipt_block(receipt_entries)
    block_html = f"<pre>{html_escape(block_plain)}</pre>"

    note_entries = list(note_entries or [])

    parts: List[str] = [header, block_html]

    footer_lines: List[str] = []
    comment_lines: List[str] = []
    if note_entries:
        comment_lines.append(_np_pick(lang, NP_COMMENT_SECTION_TITLE).format(count=len(note_entries)))
        for note in note_entries[:3]:
            timestamp_raw = note.get("timestamp") if isinstance(note, dict) else None
            timestamp = format_datetime_short(timestamp_raw) or (timestamp_raw or "")
            timestamp_disp = h(timestamp) if timestamp else "—"
            note_text = (note.get("text") if isinstance(note, dict) else "") or ""
            snippet = _np_trim_label(note_text.strip(), 220) if note_text else "—"
            comment_lines.append(f"• {timestamp_disp} — {h(snippet)}")
        if len(note_entries) > 3:
            comment_lines.append("…")
    if comment_lines:
        footer_lines.extend(comment_lines)

    assignment_lines: List[str] = []
    if assignment:
        admin_id = assignment.get("assigned_by")
        admin_name = None
        if admin_id:
            prof = load_user(admin_id) or {}
            admin_name = prof.get("fullname") or prof.get("tg", {}).get("first_name")
        admin_display = admin_name or (f"ID {admin_id}" if admin_id else "—")
        assigned_time = format_datetime_short(assignment.get("created_at")) or assignment.get("created_at") or "—"
        assignment_lines.append(
            _np_pick(lang, NP_ASSIGN_INFO_LINE).format(name=h(admin_display), time=h(assigned_time))
        )
        note_text = assignment.get("note")
        if note_text:
            assignment_lines.append(_np_pick(lang, NP_ADMIN_NOTE_PREFIX).format(note=h(note_text)))
        delivered_at = assignment.get("delivered_at")
        if delivered_at:
            delivered_time = format_datetime_short(delivered_at) or delivered_at or "—"
            assignment_lines.append(
                _np_pick(lang, NP_ASSIGN_DELIVERED_LINE).format(time=h(delivered_time))
            )

    if assignment_lines:
        if footer_lines:
            footer_lines.append("")
        footer_lines.extend(assignment_lines)

    if footer_lines:
        parts.append("\n".join(footer_lines))

    return "\n\n".join(part for part in parts if part)


async def np_send_card(uid: int, chat_id: int, text: str,
                       kb: Optional[InlineKeyboardMarkup] = None) -> types.Message:
    runtime = users_runtime.setdefault(uid, {})
    previous = runtime.get("np_last_card")
    if isinstance(previous, (list, tuple)) and len(previous) == 2:
        prev_chat, prev_mid = previous
        try:
            await bot.delete_message(prev_chat, prev_mid)
        except Exception:
            pass
        flow_items = runtime.get("flow_msgs", [])
        runtime["flow_msgs"] = [item for item in flow_items if not (item[0] == prev_chat and item[1] == prev_mid)]
    msg = await bot.send_message(chat_id, text, reply_markup=kb, disable_web_page_preview=True)
    flow_track(uid, msg)
    runtime["np_last_card"] = (msg.chat.id, msg.message_id)
    return msg


def np_prepare_view(uid: int, ttn: str, payload: Optional[dict] = None,
                    force_fetch: bool = False,
                    allow_bookmark: bool = True) -> Tuple[Optional[str], Optional[InlineKeyboardMarkup], Optional[dict], Optional[dict], Optional[str]]:
    actual_payload = payload or np_get_cached_status(uid, ttn)
    error_message = None
    if force_fetch or actual_payload is None:
        success, fetched_payload, error_message = np_fetch_tracking(ttn)
        if not success:
            return None, None, None, None, error_message
        actual_payload = fetched_payload
        np_remember_search(uid, ttn, actual_payload)

    assignment = np_get_assignment(ttn)
    if actual_payload and assignment:
        np_refresh_assignment_status(ttn, actual_payload)

    if assignment and not (assignment.get("assigned_to") == uid or uid in admins):
        assignment_display = None
    else:
        assignment_display = assignment

    notes_map = np_list_notes(uid, ttn)
    note_entries = notes_map.get(ttn, [])
    text = format_np_status(uid, ttn, actual_payload, note_entries=note_entries, assignment=assignment_display)
    kb = kb_np_result(
        uid,
        ttn,
        bookmarked=np_has_bookmark(uid, ttn),
        allow_assign=(uid in admins),
        assignment=assignment_display,
        allow_bookmark=allow_bookmark,
    )
    return text, kb, actual_payload, assignment_display, None


def receipt_status_text(paid, target: Any = DEFAULT_LANG) -> str:
    if paid is True:
        return tr(target, "STATUS_PAID")
    if paid is False:
        return tr(target, "STATUS_UNPAID")
    return tr(target, "STATUS_UNKNOWN")


def format_receipt_caption(receipt: dict, project: Optional[str] = None) -> str:
    date_part = h(receipt.get("date", "—")) or "—"
    time_raw = receipt.get("time")
    date_line = f"📅 {date_part} {h(time_raw)}".strip() if time_raw else f"📅 {date_part}"
    try:
        amount = float(receipt.get("sum", 0.0))
    except (TypeError, ValueError):
        amount = 0.0
    desc = receipt.get("desc")
    desc_text = h(desc) if desc else "—"
    file_name = receipt.get("file")
    file_text = h(file_name) if file_name else "—"
    lines = [f"🆔 Номер: <b>{h(receipt.get('receipt_no', '—'))}</b>"]
    if project:
        lines.append(f"📂 Проект: <b>{h(project)}</b>")
    lines.append(date_line)
    lines.append(f"💰 {fmt_money(amount)} грн")
    lines.append(f"📝 {desc_text}")
    lines.append(f"🔖 {receipt_status_text(receipt.get('paid'))}")
    lines.append(f"📄 {file_text}")
    payout = receipt.get("payout") if isinstance(receipt.get("payout"), dict) else None
    if payout and payout.get("status"):
        code_raw = payout.get("code") or payout.get("request_id")
        code_txt = h(code_raw) if code_raw else "—"
        status = payout.get("status")
        if status == "pending":
            ts = format_datetime_short(payout.get("assigned_at") or payout.get("updated_at"))
            tail = f" ({ts})" if ts else ""
            lines.append(f"⏳ Запрос: {code_txt}{tail}")
        elif status == "approved":
            ts = format_datetime_short(payout.get("updated_at"))
            tail = f" ({ts})" if ts else ""
            lines.append(f"💶 Запрос одобрен: {code_txt}{tail}")
        elif status == "confirmed":
            ts = format_datetime_short(payout.get("confirmed_at"))
            tail = f" ({ts})" if ts else ""
            lines.append(f"💸 Выплата подтверждена: {code_txt}{tail}")
        elif status == "closed":
            ts = format_datetime_short(payout.get("updated_at"))
            tail = f" ({ts})" if ts else ""
            lines.append(f"📭 Запрос закрыт: {code_txt}{tail}")
    elif receipt.get("paid") is True:
        ts = format_datetime_short(receipt.get("paid_at"))
        code_txt = h(receipt.get("paid_request_code")) if receipt.get("paid_request_code") else None
        if code_txt:
            tail = f" ({ts})" if ts else ""
            lines.append(f"💸 Подтверждено по {code_txt}{tail}")
        elif ts:
            lines.append(f"💸 Оплачено {ts}")
    return "\n".join(lines)


def format_photo_caption(project_info: dict, entry: dict) -> str:
    project_name = h(project_info.get("name", "—")) or "—"
    project_code = h(project_info.get("code") or "—")
    original = h(entry.get("original") or entry.get("file") or "—")
    stored = h(entry.get("file") or "—")
    uploader_name = h(entry.get("uploader_name") or "—")
    uploader_bsu = h(entry.get("uploader_bsu") or "—")
    uploader_id = entry.get("uploader_id")
    uploader_id_text = h(str(uploader_id)) if uploader_id is not None else "—"
    uploaded_at = format_datetime_short(entry.get("uploaded_at")) or "—"

    lines = [
        "🖼 <b>Карточка фотографии</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"🏗 Объект: <b>{project_name}</b> ({project_code})",
        f"📛 Оригинал: {original}",
        f"📂 Файл архива: {stored}",
        f"👤 Автор: {uploader_name} (ID {uploader_id_text}, {uploader_bsu})",
        f"🕒 Загружено: {h(uploaded_at)}",
    ]

    meta = entry.get("meta") if isinstance(entry.get("meta"), dict) else {}
    meta_lines: List[str] = []
    captured = meta.get("captured_at")
    if captured:
        meta_lines.append(f"📸 Дата съёмки: {h(captured)}")
    gps = meta.get("gps") if isinstance(meta, dict) else None
    if isinstance(gps, dict) and gps.get("lat") is not None and gps.get("lon") is not None:
        meta_lines.append(f"🌐 Координаты: {gps['lat']:.6f}, {gps['lon']:.6f}")
    address = meta.get("address")
    if address:
        meta_lines.append(f"🏙 Локация (EXIF): {h(address)}")
    camera = meta.get("camera")
    if camera:
        meta_lines.append(f"📷 Камера: {h(camera)}")

    if meta_lines:
        lines.append("")
        lines.extend(meta_lines)

    lines.extend(["", "ℹ️ Используйте кнопки под сообщением, чтобы получить оригинал или закрыть карточку."])
    return "\n".join(lines)


def photo_entry_keyboard(project: str, entry: dict, viewer_id: int, *, file_exists: bool = True) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    entry_id = entry.get("id") or entry.get("file")
    project_id = project_token(project)
    buttons: List[InlineKeyboardButton] = []
    if entry_id and file_exists:
        buttons.append(InlineKeyboardButton("📤 Оригинал", callback_data=f"photo_original:{project_id}:{entry_id}"))
    if entry_id and (viewer_id in admins or viewer_id == entry.get("uploader_id")):
        buttons.append(InlineKeyboardButton("🗑 Удалить", callback_data=f"photo_delete:{project_id}:{entry_id}"))
    if buttons:
        kb.row(*buttons)
    kb.add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
    return kb


def format_receipt_stat_entry(index: int, receipt: dict) -> str:
    indent = "&nbsp;&nbsp;&nbsp;"
    rid = h(receipt.get("receipt_no", "—")) or "—"
    date_text = h(receipt.get("date") or "—") or "—"
    time_raw = receipt.get("time")
    if time_raw:
        date_text = f"{date_text} {h(time_raw)}"
    try:
        amount = float(receipt.get("sum", 0.0))
    except (TypeError, ValueError):
        amount = 0.0
    desc = receipt.get("desc")
    desc_text = h(desc) if desc else "—"
    file_name = receipt.get("file")
    file_text = h(file_name) if file_name else "—"
    status_text = receipt_status_text(receipt.get("paid"))
    lines = [
        f"{index:02d}. <b>{rid}</b>",
        f"{indent}📅 {date_text}",
        f"{indent}💰 {fmt_money(amount)} грн — {status_text}",
        f"{indent}📝 {desc_text}",
        f"{indent}📄 {file_text}",
    ]
    payout_note = ""
    payout_obj = receipt.get("payout") if isinstance(receipt.get("payout"), dict) else None
    if payout_obj and payout_obj.get("status"):
        status = payout_obj.get("status")
        icon_map = {
            "pending": "⏳",
            "approved": "💶",
            "confirmed": "💸",
            "closed": "📭",
        }
        label_map = {
            "pending": "Запрос отправлен администратору",
            "approved": "Выплата одобрена",
            "confirmed": "Выплата подтверждена",
            "closed": "Запрос закрыт",
        }
        icon = icon_map.get(status, "ℹ️")
        label = label_map.get(status, status)
        extras: List[str] = []
        code = payout_obj.get("code") or payout_obj.get("request_id")
        if code:
            extras.append(f"ID {h(code)}")
        stamp = format_datetime_short(
            payout_obj.get("updated_at")
            or payout_obj.get("confirmed_at")
            or payout_obj.get("assigned_at")
        )
        if stamp:
            extras.append(stamp)
        if extras:
            payout_note = f"{icon} {label} — {' — '.join(extras)}"
        else:
            payout_note = f"{icon} {label}"
    elif receipt.get("paid") is True:
        extras: List[str] = []
        code = receipt.get("paid_request_code")
        if code:
            extras.append(f"ID {h(code)}")
        stamp = format_datetime_short(receipt.get("paid_at"))
        if stamp:
            extras.append(stamp)
        if extras:
            payout_note = f"💸 Выплата подтверждена — {' — '.join(extras)}"
    elif isinstance(receipt.get("payout_history"), list) and receipt["payout_history"]:
        last_event = receipt["payout_history"][-1]
        if isinstance(last_event, dict) and last_event.get("status"):
            status = last_event.get("status")
            if status == "closed":
                extras: List[str] = []
                code = last_event.get("code") or last_event.get("request_id")
                if code:
                    extras.append(f"ID {h(code)}")
                stamp = format_datetime_short(last_event.get("timestamp"))
                if stamp:
                    extras.append(stamp)
                if extras:
                    payout_note = f"📭 Запрос закрыт вручную — {' — '.join(extras)}"
                else:
                    payout_note = "📭 Запрос закрыт вручную"
    if payout_note:
        lines.append(f"{indent}{payout_note}")
    return "\n".join(lines)


def extract_receipt_prefix(message: Optional[types.Message]) -> Optional[str]:
    if not message:
        return None
    raw = (message.caption or message.text or "").strip()
    if not raw:
        return None
    first_line = raw.split("\n", 1)[0].strip()
    if first_line.startswith("🧾"):
        return first_line
    return None


async def send_receipt_card(chat_id: int, project: str, owner_uid: int, receipt: dict,
                            kb: Optional[InlineKeyboardMarkup] = None,
                            include_project: bool = False,
                            prefix: Optional[str] = None) -> types.Message:
    caption = format_receipt_caption(receipt, project if include_project else None)
    if prefix:
        caption = f"{prefix}\n{caption}"
    file_name = receipt.get("file") or ""
    path = os.path.join(proj_receipts_dir(project, owner_uid), file_name) if file_name else ""
    if path and os.path.exists(path):
        try:
            return await bot.send_photo(chat_id, InputFile(path), caption=caption, reply_markup=kb)
        except Exception:
            pass
    notices: List[str] = []
    if file_name:
        if path and not os.path.exists(path):
            notices.append("⚠️ Фото чека не найдено.")
        else:
            notices.append("⚠️ Не удалось отправить фото чека.")
    else:
        notices.append("⚠️ Фото чека отсутствует.")
    body = "\n".join(notices + [caption])
    return await bot.send_message(chat_id, body, reply_markup=kb)


def project_status_text(uid: int) -> str:
    if not active_project["name"]:
        return tr(uid, "ANCHOR_NO_PROJECT", bot=h(BOT_NAME))
    info = load_project_info(active_project["name"])
    photo_total = project_photo_count(active_project["name"])
    assignments = np_list_assignments(uid)
    total_assigned = len(assignments)
    pending_assigned = sum(1 for item in assignments if not item.get("delivered_at"))
    delivered_count = max(0, total_assigned - pending_assigned)
    bsg_section = tr(
        uid,
        "ANCHOR_PROJECT_BSG_SUMMARY",
        total=total_assigned,
        pending=pending_assigned,
        delivered=delivered_count,
    )
    alerts_section = alerts_anchor_section(uid)
    name = h(info.get("name", "—")) or "—"
    region = h(info.get("region") or "—")
    location = h(info.get("location", "—")) or "—"
    start = h(info.get("start_date", "—")) or "—"
    end = h(info.get("end_date", "—")) or "—"
    code = h(info.get("code") or "—")
    return tr(
        uid,
        "ANCHOR_PROJECT",
        bot=h(BOT_NAME),
        name=name,
        code=code,
        region=region,
        location=location,
        photos=photo_total,
        start=start,
        end=end,
        bsg_section=bsg_section,
        alerts_section=alerts_section,
    )




def kb_root(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(
        InlineKeyboardButton(tr(uid, "BTN_CHECKS"), callback_data="menu_checks"),
        InlineKeyboardButton(tr(uid, "BTN_DOCUMENTS"), callback_data="menu_docs"),
    )
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PHOTO_TIMELINE"), callback_data="menu_photos"))
    kb.row(
        InlineKeyboardButton(tr(uid, "BTN_FINANCE"), callback_data="menu_finance"),
        InlineKeyboardButton(tr(uid, "BTN_ALERTS"), callback_data="menu_alerts"),
    )
    kb.row(
        InlineKeyboardButton(tr(uid, "BTN_SOS"), callback_data="menu_sos"),
        InlineKeyboardButton(tr(uid, "BTN_NOVA_POSHTA"), callback_data="menu_np"),
    )
    kb.add(InlineKeyboardButton(tr(uid, "BTN_SETTINGS"), callback_data="menu_settings"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE"), callback_data="menu_profile"))
    if uid in admins:
        kb.add(InlineKeyboardButton(tr(uid, "BTN_ADMIN"), callback_data="menu_admin"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_ABOUT"), callback_data="menu_about"))
    return kb


def kb_profile_cancel(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_CANCEL"), callback_data="profile_cancel"))
    return kb


def kb_admin_edit_cancel(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_CANCEL"), callback_data="adm_edit_cancel"))
    return kb


def profile_has_photo(profile: dict) -> bool:
    photo = profile.get("photo") or {}
    if not isinstance(photo, dict):
        return False
    path = user_profile_photo_path(profile.get("user_id")) if profile.get("user_id") else None
    return bool(path and os.path.exists(path) and photo.get("status") != "skipped")


def profile_photo_status_label(uid: int, profile: dict) -> str:
    photo = profile.get("photo") or {}
    if not photo:
        return tr(uid, "PROFILE_PHOTO_STATUS_MISSING")
    status = photo.get("status")
    if status == "skipped":
        return tr(uid, "PROFILE_PHOTO_STATUS_SKIPPED")
    if profile_has_photo(profile):
        return tr(uid, "PROFILE_PHOTO_STATUS_OK")
    return tr(uid, "PROFILE_PHOTO_STATUS_MISSING")


def profile_summary_text(uid: int, profile: dict, edit_mode: bool = False) -> str:
    missing = tr(uid, "PROFILE_VALUE_MISSING")
    last_name = h(profile.get("last_name") or missing)
    first_name = h(profile.get("first_name") or missing)
    middle_name = h(profile.get("middle_name") or missing)
    birthdate = format_birthdate_display(profile.get("birthdate"), resolve_lang(uid))
    region = h(profile.get("region") or missing)
    phone = h(profile.get("phone") or missing)
    tg = profile.get("tg") or {}
    tg_id = str(profile.get("user_id", "—"))
    tg_username = tg.get("username")
    username_disp = h(f"@{tg_username}" if tg_username else missing)
    bsu = h(profile.get("bsu") or missing)
    photo_status = h(profile_photo_status_label(uid, profile))
    lines = [
        tr(uid, "PROFILE_HEADER"),
        "━━━━━━━━━━━━━━━━━━",
        f"{tr(uid, 'PROFILE_FIELD_LAST_NAME')}: <b>{last_name}</b>",
        f"{tr(uid, 'PROFILE_FIELD_FIRST_NAME')}: <b>{first_name}</b>",
        f"{tr(uid, 'PROFILE_FIELD_MIDDLE_NAME')}: <b>{middle_name}</b>",
        f"{tr(uid, 'PROFILE_FIELD_BIRTHDATE')}: <b>{h(birthdate)}</b>",
        f"{tr(uid, 'PROFILE_FIELD_REGION')}: <b>{region}</b>",
        f"{tr(uid, 'PROFILE_FIELD_PHONE')}: <b>{phone}</b>",
        f"{tr(uid, 'PROFILE_FIELD_BSU')}: <b>{bsu}</b>",
        f"{tr(uid, 'PROFILE_FIELD_TG_USERNAME')}: <b>{username_disp}</b>",
        f"{tr(uid, 'PROFILE_FIELD_TG_ID')}: <code>{tg_id}</code>",
        f"{tr(uid, 'PROFILE_FIELD_PHOTO')}: <b>{photo_status}</b>",
    ]
    if edit_mode:
        lines.append("")
        lines.append(tr(uid, "PROFILE_EDIT_HINT"))
    return "\n".join(lines)


def kb_profile_menu(uid: int, profile: dict, edit_mode: bool = False, show_photo: bool = False) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    has_photo = profile_has_photo(profile)
    if edit_mode:
        kb.row(
            InlineKeyboardButton(tr(uid, "BTN_PROFILE_FIELD_LAST"), callback_data="profile_edit_last"),
            InlineKeyboardButton(tr(uid, "BTN_PROFILE_FIELD_FIRST"), callback_data="profile_edit_first"),
        )
        kb.row(
            InlineKeyboardButton(tr(uid, "BTN_PROFILE_FIELD_MIDDLE"), callback_data="profile_edit_middle"),
            InlineKeyboardButton(tr(uid, "BTN_PROFILE_FIELD_BIRTHDATE"), callback_data="profile_edit_birthdate"),
        )
        kb.row(
            InlineKeyboardButton(tr(uid, "BTN_PROFILE_FIELD_REGION"), callback_data="profile_edit_region"),
            InlineKeyboardButton(tr(uid, "BTN_PROFILE_FIELD_PHONE"), callback_data="profile_edit_phone"),
        )
        kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_UPDATE_PHOTO"), callback_data="profile_edit_photo"))
        if has_photo:
            kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_REMOVE_PHOTO"), callback_data="profile_remove_photo"))
        kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_DONE"), callback_data="profile_done"))
    else:
        kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_EDIT"), callback_data="profile_edit"))
        if has_photo:
            if show_photo:
                kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_HIDE_PHOTO"), callback_data="profile_hide_photo"))
            else:
                kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_VIEW_PHOTO"), callback_data="profile_view_photo"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_ROOT"), callback_data="back_root"))
    return kb


def kb_profile_region_prompt(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "REGISTER_REGION_BUTTON"), callback_data="profile_region_open"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_CANCEL"), callback_data="profile_cancel"))
    return kb


def kb_profile_region_picker(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for idx, region in enumerate(UKRAINE_REGIONS):
        kb.insert(InlineKeyboardButton(region, callback_data=f"profile_region_pick:{idx}"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_CANCEL"), callback_data="profile_cancel"))
    return kb


def kb_profile_phone_keyboard(uid: int) -> ReplyKeyboardMarkup:
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(KeyboardButton(tr(uid, "BTN_SEND_PHONE"), request_contact=True))
    kb.add(KeyboardButton(tr(uid, "BTN_PROFILE_CANCEL")))
    return kb


async def show_profile(uid: int, *, edit_mode: Optional[bool] = None, show_photo: Optional[bool] = None):
    runtime = profile_runtime(uid)
    if edit_mode is None or show_photo is None:
        current_edit, current_photo = profile_get_flags(uid)
        if edit_mode is None:
            edit_mode = current_edit
        if show_photo is None:
            show_photo = current_photo
    profile = load_user(uid) or ensure_user(uid, runtime.get("tg", {}))
    profile.setdefault("user_id", uid)
    has_photo = profile_has_photo(profile)
    if show_photo and not has_photo:
        await profile_send_notification(uid, tr(uid, "PROFILE_NO_PHOTO"))
        show_photo = False
    profile_set_flags(uid, edit_mode=edit_mode, show_photo=show_photo and has_photo)
    caption = profile_summary_text(uid, profile, edit_mode=edit_mode)
    kb = kb_profile_menu(uid, profile, edit_mode=edit_mode, show_photo=show_photo and has_photo)
    if show_photo and has_photo:
        await anchor_replace_with_photo(uid, user_profile_photo_path(uid), caption, kb)
    else:
        profile_set_flags(uid, show_photo=False)
        chat = profile_chat_id(uid)
        if chat:
            await anchor_upsert(uid, chat, caption, kb)


def kb_alerts(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_BTN_OVERVIEW"), callback_data="alerts_overview"))
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_BTN_ACTIVE"), callback_data="alerts_active"))
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_BTN_HISTORY"), callback_data="alerts_history"))
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_BTN_SUBSCRIPTIONS"), callback_data="alerts_subscriptions"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_ROOT"), callback_data="back_root"))
    return kb


def kb_checks(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "BTN_ADD_RECEIPT"), callback_data="check_add"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_MY_STATS"), callback_data="check_stats"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_MY_RECEIPTS"), callback_data="check_list"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_RECEIPT_HISTORY"), callback_data="check_history"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_ROOT"), callback_data="back_root"))
    return kb


def kb_photos(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "BTN_UPLOAD_PHOTO"), callback_data="photo_upload"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_VIEW_OBJECT_PHOTOS"), callback_data="photo_view"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_ROOT"), callback_data="back_root"))
    return kb


def kb_photo_session_controls(has_uploads: bool) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    if has_uploads:
        kb.row(
            InlineKeyboardButton("🗂 Посмотреть загруженное", callback_data="photo_session_preview"),
            InlineKeyboardButton("✅ Завершить загрузку", callback_data="photo_finish"),
        )
    kb.add(InlineKeyboardButton("❌ Отменить", callback_data="photo_cancel"))
    return kb


def kb_photo_view_actions() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(
        InlineKeyboardButton("❌ Закрыть", callback_data="photo_view_close"),
        InlineKeyboardButton("🏠 На главную", callback_data="photo_view_root"),
    )
    return kb


def _format_photo_session_entry(idx: int, entry: dict) -> str:
    original = entry.get("original") or entry.get("file") or "—"
    uploaded_at = entry.get("uploaded_at")
    if isinstance(uploaded_at, str):
        try:
            dt = datetime.fromisoformat(uploaded_at)
            uploaded_at = dt.strftime("%d.%m.%Y %H:%M")
        except ValueError:
            uploaded_at = uploaded_at.replace("T", " ")
    return f"{idx}. {h(original)} — {h(uploaded_at or '—')}"


def _build_photo_session_text(info: dict, uploaded: List[dict], last_entry: Optional[dict] = None) -> str:
    name = h(info.get("name", "—"))
    code = h(info.get("code") or "—")
    lines = [
        "📤 <b>Загрузка фотографий объекта</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"📂 Проект: <b>{name}</b> ({code})",
        "",
        "Отправляйте одно или несколько изображений: можно прикреплять фото напрямую или документом без сжатия.",
        "Каждый файл сохраняется в хронологию вместе с вашим именем и временем загрузки.",
        "",
    ]
    if uploaded:
        lines.append(f"📸 Уже загружено: <b>{len(uploaded)}</b>")
        lines.append("Список файлов:")
        lines.extend(_format_photo_session_entry(idx + 1, entry) for idx, entry in enumerate(uploaded))
        if last_entry:
            marker = last_entry.get("original") or last_entry.get("file")
            if marker:
                lines.append("")
                lines.append(f"🆕 Последнее добавление: <b>{h(marker)}</b>")
        lines.append("")
        lines.append("Используйте кнопку «🗂 Посмотреть загруженное», чтобы пересмотреть файлы без повторной загрузки.")
    else:
        lines.append("Пока нет загруженных файлов. Отправьте первое фото, чтобы начать список.")
    lines.append("")
    lines.append("Продолжайте отправлять фото или используйте кнопки ниже для завершения или отмены сессии.")
    return "\n".join(lines)


async def _photo_refresh_session_message(chat_id: int, uid: int, state: FSMContext, info: dict,
                                         uploaded: List[dict], last_entry: Optional[dict] = None):
    data = await state.get_data()
    target = data.get("photo_session_message")
    text = _build_photo_session_text(info, uploaded, last_entry)
    kb = kb_photo_session_controls(bool(uploaded))
    if isinstance(target, (list, tuple)) and len(target) == 2:
        tgt_chat, tgt_id = target
        try:
            await bot.edit_message_text(text, tgt_chat, tgt_id, reply_markup=kb)
            return
        except MessageNotModified:
            return
        except Exception:
            await _delete_message_safe(tgt_chat, tgt_id)
    msg = await bot.send_message(chat_id, text, reply_markup=kb)
    flow_track(uid, msg)
    await state.update_data(photo_session_message=(msg.chat.id, msg.message_id))


def kb_finance_root(user_has_pending_confirm: bool=False) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    if user_has_pending_confirm:
        kb.add(InlineKeyboardButton("✅ Подтвердить получение выплат", callback_data="fin_confirm_list"))
    kb.add(InlineKeyboardButton("⏳ Неоплаченные чеки", callback_data="fin_unpaid_list"))
    kb.add(InlineKeyboardButton("📨 Запросить выплату", callback_data="fin_request_payout"))
    kb.add(InlineKeyboardButton("📚 История выплат", callback_data="fin_history"))
    kb.add(InlineKeyboardButton("⬅️ На главную", callback_data="back_root"))
    return kb


def kb_novaposhta(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_INTERFACE"), callback_data="np_interface"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_SEARCH"), callback_data="np_search"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_HISTORY"), callback_data="np_history"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_BOOKMARKS"), callback_data="np_bookmarks"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_ASSIGNED"), callback_data="np_assigned"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_RECEIVED"), callback_data="np_received"))
    if uid in admins:
        kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_ASSIGN_SEND"), callback_data="np_assign_start"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_ROOT"), callback_data="back_root"))
    return kb


def kb_np_cancel(uid: int) -> InlineKeyboardMarkup:
    lang = resolve_lang(uid)
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(_np_pick(lang, NP_CANCEL_BUTTON_LABEL), callback_data="np_cancel"))
    return kb


def kb_np_assign_note(uid: int) -> InlineKeyboardMarkup:
    lang = resolve_lang(uid)
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(_np_pick(lang, NP_ASSIGN_SKIP_BUTTON_LABEL), callback_data="np_assign_skip"))
    kb.add(InlineKeyboardButton(_np_pick(lang, NP_CANCEL_BUTTON_LABEL), callback_data="np_cancel"))
    return kb


def kb_np_result(uid: int, ttn: str, *, bookmarked: bool,
                 allow_assign: bool = False,
                 assignment: Optional[dict] = None,
                 allow_bookmark: bool = True) -> InlineKeyboardMarkup:
    lang = resolve_lang(uid)
    kb = InlineKeyboardMarkup()

    refresh_btn = InlineKeyboardButton(
        _np_pick(lang, NP_REFRESH_BUTTON_LABEL),
        callback_data=f"np_refresh:{ttn}"
    )

    if allow_bookmark:
        bookmark_label = _np_pick(
            lang,
            NP_BOOKMARK_REMOVE_BUTTON if bookmarked else NP_BOOKMARK_ADD_BUTTON
        )
        bookmark_btn = InlineKeyboardButton(bookmark_label, callback_data=f"np_bookmark:{ttn}")
        kb.row(refresh_btn, bookmark_btn)
    else:
        kb.add(refresh_btn)

    kb.add(InlineKeyboardButton(_np_pick(lang, NP_NOTE_BUTTON_LABEL), callback_data=f"np_note:{ttn}"))

    if assignment and not assignment.get("delivered_at"):
        kb.add(InlineKeyboardButton(_np_pick(lang, NP_MARK_RECEIVED_LABEL), callback_data=f"np_assigned_received:{ttn}"))

    if allow_assign:
        kb.add(InlineKeyboardButton(tr(uid, "BTN_NP_ASSIGN_SEND"), callback_data=f"np_assign_quick:{ttn}"))

    kb.add(InlineKeyboardButton(_np_pick(lang, NP_CLOSE_BUTTON_LABEL), callback_data="np_close"))
    return kb


def np_build_list_keyboard(uid: int, options: List[Tuple[str, str]], prefix: str,
                           back_callback: str = "menu_np") -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    for value, label in options:
        kb.add(InlineKeyboardButton(label, callback_data=f"{prefix}:{value}"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_SETTINGS"), callback_data=back_callback))
    return kb


def kb_admin_root() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("👥 Пользователи", callback_data="adm_users"))
    kb.add(InlineKeyboardButton("📂 Проекты", callback_data="adm_projects"))
    kb.add(InlineKeyboardButton("💵 Финансы", callback_data="adm_finance"))
    kb.add(InlineKeyboardButton("⬅️ На главную", callback_data="back_root"))
    return kb


def kb_admin_projects() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("📋 Список", callback_data="proj_list"))
    kb.add(InlineKeyboardButton("➕ Создать", callback_data="proj_create"))
    kb.add(InlineKeyboardButton("🔄 Активировать", callback_data="proj_activate"))
    kb.add(InlineKeyboardButton("✅ Завершить", callback_data="proj_finish"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="menu_admin"))
    return kb


def admin_collect_user_stats(profile: dict) -> dict:
    receipts = profile.get("receipts") or {}
    projects = sorted(receipts.keys()) if isinstance(receipts, dict) else []
    total_count = 0
    total_sum = 0.0
    paid_sum = 0.0
    unpaid_sum = 0.0
    for recs in receipts.values():
        if not isinstance(recs, list):
            continue
        for entry in recs:
            try:
                amount = float(entry.get("sum") or 0.0)
            except (TypeError, ValueError):
                amount = 0.0
            total_count += 1
            total_sum += amount
            if entry.get("paid") is True:
                paid_sum += amount
            elif entry.get("paid") is False:
                unpaid_sum += amount
    payouts = profile.get("payouts") or []
    pending_payouts = [p for p in payouts if (p.get("status") not in {"confirmed", "closed"})]
    confirmed_payouts = [p for p in payouts if p.get("status") in {"confirmed", "closed"}]
    return {
        "projects": projects,
        "total_count": total_count,
        "total_sum": total_sum,
        "paid_sum": paid_sum,
        "unpaid_sum": unpaid_sum,
        "pending_payouts": pending_payouts,
        "confirmed_payouts": confirmed_payouts,
    }


def admin_user_card_text(viewer_uid: int, profile: dict, *, edit_mode: bool = False) -> str:
    stats = admin_collect_user_stats(profile)
    base = profile_summary_text(viewer_uid, profile, edit_mode=False)
    lines = [base, "", "💼 <b>Активність</b>"]
    lines.append(f"📂 Проєктів: <b>{len(stats['projects'])}</b>")
    lines.append(f"🧾 Чеків: <b>{stats['total_count']}</b>")
    lines.append(f"💰 Сума чеків: <b>{fmt_money(stats['total_sum'])} грн</b>")
    lines.append(f"✅ Оплачено: <b>{fmt_money(stats['paid_sum'])} грн</b>")
    lines.append(f"⏳ Очікує виплат: <b>{fmt_money(stats['unpaid_sum'])} грн</b>")
    lines.append("")
    lines.append("💵 <b>Запити на виплати</b>")
    lines.append(f"⌛ В роботі: <b>{len(stats['pending_payouts'])}</b>")
    lines.append(f"📗 Завершено: <b>{len(stats['confirmed_payouts'])}</b>")
    if edit_mode:
        lines.append("")
        lines.append(tr(viewer_uid, "PROFILE_EDIT_HINT"))
    return "\n".join(lines)


def kb_admin_user(viewer_uid: int, profile: dict, *, show_photo: bool = False, edit_mode: bool = False) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    if profile_has_photo(profile):
        label = "📝 Повернути текст" if show_photo else "👁 Показати фото"
        kb.add(InlineKeyboardButton(label, callback_data="adm_user_photo_toggle"))
    if edit_mode:
        kb.row(
            InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_FIELD_LAST"), callback_data="adm_edit_last"),
            InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_FIELD_FIRST"), callback_data="adm_edit_first"),
        )
        kb.row(
            InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_FIELD_MIDDLE"), callback_data="adm_edit_middle"),
            InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_FIELD_BIRTHDATE"), callback_data="adm_edit_birthdate"),
        )
        kb.row(
            InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_FIELD_REGION"), callback_data="adm_edit_region"),
            InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_FIELD_PHONE"), callback_data="adm_edit_phone"),
        )
        kb.add(InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_UPDATE_PHOTO"), callback_data="adm_edit_photo"))
        if profile_has_photo(profile):
            kb.add(InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_REMOVE_PHOTO"), callback_data="adm_edit_remove_photo"))
        kb.add(InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_DONE"), callback_data="adm_user_edit_done"))
    else:
        kb.add(InlineKeyboardButton(tr(viewer_uid, "BTN_PROFILE_EDIT"), callback_data="adm_user_edit"))
        kb.row(
            InlineKeyboardButton("📊 Статистика", callback_data="adm_stat_choose"),
            InlineKeyboardButton("🧾 Чеки", callback_data="adm_recs_choose"),
        )
        kb.add(InlineKeyboardButton("💵 Фінанси", callback_data="adm_user_finance"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_users"))
    return kb


def kb_region_select() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for idx, region in enumerate(UKRAINE_REGIONS):
        kb.insert(InlineKeyboardButton(region, callback_data=f"proj_region_{idx}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="proj_create_cancel"))
    return kb


def kb_pdf_upload() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("➕ Ещё файл", callback_data="pdf_more"))
    kb.add(InlineKeyboardButton("✅ Завершить", callback_data="pdf_finish"))
    return kb


def kb_broadcast_close() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("❌ Закрыть сообщение", callback_data="broadcast_close"))
    return kb


def kb_preview() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("✏️ Изменить сумму", callback_data="edit_amount"))
    kb.add(InlineKeyboardButton("🖼 Заменить фото", callback_data="edit_photo"))
    kb.add(InlineKeyboardButton("📝 Изменить описание", callback_data="edit_desc"))
    kb.add(InlineKeyboardButton("🔁 Изменить статус оплаты", callback_data="edit_paid"))
    kb.add(InlineKeyboardButton("✅ Сохранить", callback_data="save_receipt"))
    kb.add(InlineKeyboardButton("❌ Отменить", callback_data="cancel_receipt"))
    return kb


def kb_receipt_cancel() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("❌ Отменить", callback_data="cancel_receipt"))
    return kb


def kb_desc_prompt() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("Пропустить", callback_data="desc_skip"))
    kb.add(InlineKeyboardButton("❌ Отменить", callback_data="cancel_receipt"))
    return kb


def kb_choose_paid(ask_later: bool=True, allow_cancel: bool=False, flow_cancel: bool=False) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(
        InlineKeyboardButton("✅ Оплачено", callback_data="paid_yes"),
        InlineKeyboardButton("❌ Не оплачено", callback_data="paid_no")
    )
    if ask_later:
        kb.add(InlineKeyboardButton("⏭ Указать позже", callback_data="paid_later"))
    if allow_cancel:
        kb.add(InlineKeyboardButton("❌ Отменить изменение", callback_data="edit_cancel"))
    if flow_cancel:
        kb.add(InlineKeyboardButton("❌ Отменить", callback_data="cancel_receipt"))
    return kb


def kb_edit_cancel() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("❌ Отменить изменение", callback_data="edit_cancel"))
    return kb


def kb_saved_receipt() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("❌ Закрыть", callback_data="close_saved_receipt"))
    return kb


def kb_next_step(target: Any, callback_data: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(target, "BTN_NEXT"), callback_data=callback_data))
    return kb


def kb_settings(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "SETTINGS_LANGUAGE"), callback_data="settings_language"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_ROOT"), callback_data="back_root"))
    return kb


def kb_language_settings(uid: int) -> InlineKeyboardMarkup:
    kb = kb_language_picker("settings_lang")
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_SETTINGS"), callback_data="settings_back"))
    return kb


def inline_kb_signature(kb: Optional[InlineKeyboardMarkup]) -> Any:
    if not kb or not kb.inline_keyboard: return None
    sign = []
    for row in kb.inline_keyboard:
        sign.append(tuple((btn.text, btn.callback_data or btn.url or "") for btn in row))
    return tuple(sign)


# ========================== ANCHOR ==========================
async def anchor_upsert(uid: int, chat_id: int, text: Optional[str] = None, kb: Optional[InlineKeyboardMarkup] = None):
    if text is None: text = project_status_text(uid)
    if kb is None: kb = kb_root(uid)
    text = str(text); kb_sign = inline_kb_signature(kb)

    ur = users_runtime.setdefault(uid, {})
    last_text = ur.get("last_anchor_text"); last_kb = ur.get("last_anchor_kb")
    anchor = ur.get("anchor")

    if anchor and last_text == text and last_kb == kb_sign:
        return

    if anchor:
        try:
            await bot.edit_message_text(text, chat_id, anchor, reply_markup=kb)
            ur["last_anchor_text"] = text; ur["last_anchor_kb"] = kb_sign
            ur["anchor_mode"] = "text"
            return
        except MessageNotModified:
            return
        except Exception:
            try:
                await bot.delete_message(chat_id, anchor)
            except Exception:
                pass

    msg = await bot.send_message(chat_id, text, reply_markup=kb)
    ur["anchor"] = msg.message_id
    ur["last_anchor_text"] = text; ur["last_anchor_kb"] = kb_sign
    ur["anchor_mode"] = "text"


async def anchor_show_root(uid: int):
    chat = users_runtime.get(uid, {}).get("tg", {}).get("chat_id")
    if chat: await anchor_upsert(uid, chat, project_status_text(uid), kb_root(uid))


async def anchor_replace_with_photo(uid: int, photo_path: str, caption: str, kb: InlineKeyboardMarkup):
    runtime = users_runtime.setdefault(uid, {})
    chat = runtime.get("tg", {}).get("chat_id")
    if not chat:
        return
    if not os.path.exists(photo_path):
        await anchor_upsert(uid, chat, caption, kb)
        return
    anchor = runtime.get("anchor")
    kb_sign = inline_kb_signature(kb)
    media = types.InputMediaPhoto(InputFile(photo_path), caption=caption, parse_mode="HTML")
    try:
        if anchor:
            await bot.edit_message_media(chat_id=chat, message_id=anchor, media=media, reply_markup=kb)
        else:
            msg = await bot.send_photo(chat, InputFile(photo_path), caption=caption, reply_markup=kb)
            runtime["anchor"] = msg.message_id
    except MessageNotModified:
        pass
    except Exception:
        try:
            if anchor:
                await bot.delete_message(chat, anchor)
        except Exception:
            pass
        msg = await bot.send_photo(chat, InputFile(photo_path), caption=caption, reply_markup=kb)
        runtime["anchor"] = msg.message_id
    runtime["last_anchor_text"] = caption
    runtime["last_anchor_kb"] = kb_sign
    runtime["anchor_mode"] = "photo"


async def anchor_show_text(uid: int, text: str, kb: InlineKeyboardMarkup):
    chat = users_runtime.get(uid, {}).get("tg", {}).get("chat_id")
    if chat: await anchor_upsert(uid, chat, text, kb)


async def update_all_anchors():
    for uid in list(users_runtime.keys()):
        await anchor_show_root(uid)


def profile_runtime(uid: int) -> dict:
    return users_runtime.setdefault(uid, {})


def profile_chat_id(uid: int) -> Optional[int]:
    return profile_runtime(uid).get("tg", {}).get("chat_id")


async def profile_clear_prompt(uid: int):
    runtime = profile_runtime(uid)
    prompt = runtime.pop("profile_prompt", None)
    if isinstance(prompt, (list, tuple)) and len(prompt) == 2:
        await _delete_message_safe(prompt[0], prompt[1])


async def profile_send_prompt(uid: int, text: str, reply_markup: Optional[Union[InlineKeyboardMarkup, ReplyKeyboardMarkup]] = None):
    chat = profile_chat_id(uid)
    if not chat:
        return None
    await profile_clear_prompt(uid)
    runtime = profile_runtime(uid)
    if isinstance(reply_markup, ReplyKeyboardMarkup):
        runtime["profile_reply_keyboard"] = True
    else:
        runtime.pop("profile_reply_keyboard", None)
    msg = await bot.send_message(chat, text, reply_markup=reply_markup)
    runtime["profile_prompt"] = (msg.chat.id, msg.message_id)
    flow_track(uid, msg)
    return msg


def admin_edit_runtime(uid: int) -> dict:
    runtime = users_runtime.setdefault(uid, {})
    return runtime.setdefault("admin_edit", {})


async def admin_edit_clear_prompt(uid: int):
    runtime = admin_edit_runtime(uid)
    prompt = runtime.pop("prompt", None)
    if isinstance(prompt, (list, tuple)) and len(prompt) == 2:
        await _delete_message_safe(prompt[0], prompt[1])


async def admin_edit_send_prompt(
    uid: int,
    text: str,
    reply_markup: Optional[Union[InlineKeyboardMarkup, ReplyKeyboardMarkup]] = None,
):
    chat = users_runtime.get(uid, {}).get("tg", {}).get("chat_id")
    if not chat:
        return None
    await admin_edit_clear_prompt(uid)
    runtime = admin_edit_runtime(uid)
    if isinstance(reply_markup, ReplyKeyboardMarkup):
        runtime["reply_keyboard"] = True
    else:
        runtime.pop("reply_keyboard", None)
    msg = await bot.send_message(chat, text, reply_markup=reply_markup)
    runtime["prompt"] = (msg.chat.id, msg.message_id)
    flow_track(uid, msg)
    return msg


async def admin_edit_notify(uid: int, text: str, *, remove_keyboard: bool = False):
    chat = users_runtime.get(uid, {}).get("tg", {}).get("chat_id")
    if not chat:
        return
    markup = ReplyKeyboardRemove() if remove_keyboard else None
    msg = await bot.send_message(chat, text, reply_markup=markup)
    flow_track(uid, msg)
    schedule_auto_delete(msg.chat.id, msg.message_id, delay=8)


async def profile_send_notification(uid: int, text: str, *, remove_keyboard: bool = False):
    chat = profile_chat_id(uid)
    if not chat:
        return
    markup = ReplyKeyboardRemove() if remove_keyboard else None
    msg = await bot.send_message(chat, text, reply_markup=markup)
    flow_track(uid, msg)
    schedule_auto_delete(msg.chat.id, msg.message_id, delay=8)


def profile_set_flags(uid: int, *, edit_mode: Optional[bool] = None, show_photo: Optional[bool] = None):
    runtime = profile_runtime(uid)
    if edit_mode is not None:
        runtime["profile_edit_mode"] = bool(edit_mode)
    if show_photo is not None:
        runtime["profile_show_photo"] = bool(show_photo)


def profile_get_flags(uid: int) -> Tuple[bool, bool]:
    runtime = profile_runtime(uid)
    return bool(runtime.get("profile_edit_mode")), bool(runtime.get("profile_show_photo"))


async def profile_abort(uid: int, state: FSMContext, *, remove_keyboard: bool = False):
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_CANCELLED"), remove_keyboard=remove_keyboard)
    await show_profile(uid, edit_mode=True, show_photo=False)


# ========================== FLOW CLEANER ==========================
def flow_track(uid: int, msg: Optional[types.Message], bucket: str = "flow_msgs"):
    if not msg:
        return
    runtime = users_runtime.setdefault(uid, {})
    runtime.setdefault(bucket, []).append((msg.chat.id, msg.message_id))


def flow_track_warning(uid: int, msg: Optional[types.Message]):
    flow_track(uid, msg, bucket="flow_warns")


async def _flow_clear_bucket(uid: int, bucket: str):
    runtime = users_runtime.setdefault(uid, {})
    tracked = list(runtime.get(bucket, []))
    runtime[bucket] = []
    if not tracked:
        return
    await asyncio.gather(*[
        _delete_message_safe(chat_id, mid)
        for chat_id, mid in tracked
        if chat_id and mid
    ], return_exceptions=True)


async def flow_clear_warnings(uid: int):
    await _flow_clear_bucket(uid, "flow_warns")


async def flow_clear(uid: int):
    runtime = users_runtime.setdefault(uid, {})
    await _flow_clear_bucket(uid, "flow_msgs")
    await _flow_clear_bucket(uid, "flow_warns")
    last_card = runtime.pop("np_last_card", None)
    if isinstance(last_card, (list, tuple)) and len(last_card) == 2:
        await _delete_message_safe(last_card[0], last_card[1])
    runtime.pop("alerts_cards", None)


async def flow_delete_message(uid: int, message: Optional[types.Message]):
    if not message:
        return
    if is_anchor_message(uid, message.message_id):
        return
    await _delete_message_safe(message.chat.id, message.message_id)


def schedule_auto_delete(chat_id: int, message_id: int, delay: int = 10):
    async def _delayed():
        try:
            await asyncio.sleep(delay)
        except Exception:
            return
        await _delete_message_safe(chat_id, message_id)

    asyncio.create_task(_delayed())


async def clear_then_anchor(uid: int, text: str, kb: InlineKeyboardMarkup):
    await flow_clear(uid)
    await anchor_show_text(uid, text, kb)



async def _delete_message_safe(chat_id: Optional[int], message_id: Optional[int]):
    if not chat_id or not message_id:
        return
    try:
        await bot.delete_message(chat_id, message_id)
    except Exception:
        pass


async def remove_preview_message(state: FSMContext):
    data = await state.get_data()
    info = data.get("preview_message")
    if not info:
        return
    chat_id = message_id = None
    if isinstance(info, (list, tuple)) and len(info) == 2:
        chat_id, message_id = info
    elif isinstance(info, dict):
        chat_id = info.get("chat_id")
        message_id = info.get("message_id")
    await _delete_message_safe(chat_id, message_id)
    await state.update_data(preview_message=None)


async def clear_edit_prompt(state: FSMContext):
    data = await state.get_data()
    info = data.get("edit_prompt")
    if not info:
        return
    chat_id = message_id = None
    if isinstance(info, (list, tuple)) and len(info) == 2:
        chat_id, message_id = info
    elif isinstance(info, dict):
        chat_id = info.get("chat_id")
        message_id = info.get("message_id")
    await _delete_message_safe(chat_id, message_id)
    await state.update_data(edit_prompt=None)


async def clear_step_prompt(state: FSMContext):
    if state is None:
        return
    data = await state.get_data()
    info = data.get("step_prompt")
    if not info:
        return
    chat_id = message_id = None
    if isinstance(info, (list, tuple)) and len(info) == 2:
        chat_id, message_id = info
    elif isinstance(info, dict):
        chat_id = info.get("chat_id")
        message_id = info.get("message_id")
    await _delete_message_safe(chat_id, message_id)
    await state.update_data(step_prompt=None)


async def remember_step_prompt(state: FSMContext, msg: types.Message):
    if state is None or msg is None:
        return msg
    await state.update_data(step_prompt=(msg.chat.id, msg.message_id))
    return msg


def is_anchor_message(uid: int, message_id: Optional[int]) -> bool:
    if message_id is None:
        return False
    return users_runtime.get(uid, {}).get("anchor") == message_id


async def delete_if_not_anchor(uid: int, chat_id: Optional[int], message_id: Optional[int]):
    if not chat_id or not message_id:
        return
    if is_anchor_message(uid, message_id):
        return
    await _delete_message_safe(chat_id, message_id)


async def send_receipt_preview(uid: int, chat_id: int, state: FSMContext):
    data = await state.get_data()
    tmp = data.get("tmp_img")
    amount = data.get("amount")
    if not tmp or amount is None:
        return None
    desc_raw = data.get("desc", "")
    desc_text = h(desc_raw) if desc_raw else "—"
    paid = data.get("paid")
    await clear_edit_prompt(state)
    await remove_preview_message(state)
    status_txt = "✅ Оплачен" if paid is True else ("❌ Не оплачен" if paid is False else "⏳ Не указан")
    caption = (
        "🧾 <b>Предпросмотр чека</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"💰 Сумма: {fmt_money(float(amount))} грн\n"
        f"📝 Описание: {desc_text}\n"
        f"🔖 Статус: {status_txt}\n\n"
        "Используйте кнопки под сообщением, чтобы скорректировать данные или сохранить чек."
    )
    msg = await bot.send_photo(chat_id, InputFile(tmp), caption=caption, reply_markup=kb_preview())
    flow_track(uid, msg)
    await state.update_data(preview_message=(chat_id, msg.message_id))
    return msg


async def admin_send_receipt_photos(admin_uid: int, chat_id: int, target_uid: int, project: str, files: List[str]):
    recs = user_project_receipts(target_uid, project)
    by_file = {r.get("file"): r for r in recs}
    base_dir = proj_receipts_dir(project, target_uid)
    for fname in files:
        path = os.path.join(base_dir, fname)
        if not os.path.exists(path):
            continue
        r = by_file.get(fname)
        desc = r.get("desc") if r else None
        status = r.get("paid") if r else None
        status_txt = "✅ Оплачен" if status is True else ("❌ Не оплачен" if status is False else "⏳ Не указан")
        caption_parts = []
        if r:
            caption_parts.append(f"🆔 Номер: <b>{h(r.get('receipt_no','—'))}</b>")
            caption_parts.append(f"📅 {h(r.get('date','—'))} {h(r.get('time',''))}")
            amount = float(r.get('sum') or 0.0)
            caption_parts.append(f"💰 {fmt_money(amount)} грн")
            caption_parts.append(f"📝 {h(desc) if desc else '—'}")
            caption_parts.append(f"🔖 {status_txt}")
        caption_parts.append(f"📄 {h(fname)}")
        caption = "\n".join(caption_parts)
        kb = InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
        try:
            msg = await bot.send_photo(chat_id, InputFile(path), caption=caption, reply_markup=kb)
            flow_track(admin_uid, msg)
        except Exception:
            continue


# ========================== START / ONBOARD ==========================
def kb_language_picker(prefix: str = "lang_select") -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    for code, label in LANG_ORDER:
        kb.add(InlineKeyboardButton(label, callback_data=f"{prefix}:{code}"))
    return kb


def registration_button_label(target: Any) -> str:
    lang = resolve_lang(target)
    labels = {
        "uk": "🤖 Далі",
        "en": "🤖 DALL·E",
        "de": "🤖 Weiter",
        "pl": "🤖 Dalej",
        "ru": "🤖 Дальше",
    }
    return labels.get(lang, labels.get(DEFAULT_LANG, "▶️ Next"))


def kb_registration_next(target: Any, callback_data: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(registration_button_label(target), callback_data=callback_data))
    return kb


def kb_region_prompt(target: Any) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(target, "REGISTER_REGION_BUTTON"), callback_data="reg_region_open"))
    return kb


def kb_admin_edit_region_prompt(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(uid, "REGISTER_REGION_BUTTON"), callback_data="adm_reg_open"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_CANCEL"), callback_data="adm_edit_cancel"))
    return kb


def kb_region_picker(target: Any) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for idx, region in enumerate(UKRAINE_REGIONS):
        kb.insert(InlineKeyboardButton(region, callback_data=f"reg_region_pick:{idx}"))
    return kb


def kb_admin_edit_region_picker(uid: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for idx, region in enumerate(UKRAINE_REGIONS):
        kb.insert(InlineKeyboardButton(region, callback_data=f"adm_reg_pick:{idx}"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_CANCEL"), callback_data="adm_edit_cancel"))
    return kb


def kb_phone_keyboard(target: Any) -> ReplyKeyboardMarkup:
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(KeyboardButton(tr(target, "BTN_SEND_PHONE"), request_contact=True))
    return kb


def kb_photo_keyboard(target: Any) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(tr(target, "BTN_SKIP"), callback_data="reg_photo_skip"))
    return kb


def kb_admin_edit_next(uid: int, callback_data: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton(registration_button_label(uid), callback_data=callback_data))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_PROFILE_CANCEL"), callback_data="adm_edit_cancel"))
    return kb


@dp.message_handler(commands=["start"], state="*")
async def start_cmd(m: types.Message, state: FSMContext):
    ensure_dirs()
    sync_state()
    uid = m.from_user.id
    runtime = users_runtime.setdefault(uid, {})

    await state.finish()
    await flow_clear(uid)

    runtime["tg"] = {
        "user_id": uid,
        "chat_id": m.chat.id,
        "username": m.from_user.username,
        "first_name": m.from_user.first_name,
        "last_name": m.from_user.last_name,
        "last_seen": datetime.now(timezone.utc).isoformat(),
    }

    ensure_user(uid, runtime["tg"], lang=m.from_user.language_code)
    profile = load_user(uid) or {}
    runtime["onboard_registered"] = bool(
        profile.get("profile_completed") or (
            profile.get("first_name") and profile.get("last_name") and profile.get("phone")
        )
    )

    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass

    prompt = await bot.send_message(m.chat.id, tr(uid, "LANGUAGE_PROMPT"), reply_markup=kb_language_picker())
    runtime["language_prompt"] = (prompt.chat.id, prompt.message_id)
    await OnboardFSM.language.set()


@dp.callback_query_handler(lambda c: c.data.startswith("lang_select:"), state=OnboardFSM.language)
async def onboard_language_selected(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    code = c.data.split(":", 1)[1]
    if code not in LANG_CODES:
        await c.answer()
        return

    runtime = users_runtime.setdefault(uid, {})
    prompt = runtime.pop("language_prompt", None)
    if prompt:
        await _delete_message_safe(prompt[0], prompt[1])

    set_user_lang(uid, code, confirmed=True)
    profile = load_user(uid) or {}
    runtime["onboard_registered"] = bool(
        profile.get("profile_completed") or (
            profile.get("first_name") and profile.get("last_name") and profile.get("phone")
        )
    )

    chat_id = runtime.get("tg", {}).get("chat_id") or c.message.chat.id
    confirm = await bot.send_message(
        chat_id,
        tr(uid, "ONBOARD_LANGUAGE_CONFIRMED", language=LANG_LABELS.get(code, code)),
        reply_markup=kb_registration_next(uid, "onboard_stage:welcome"),
    )
    runtime["onboard_intro"] = {"chat_id": confirm.chat.id, "message_id": confirm.message_id}
    await state.set_state(OnboardFSM.welcome.state)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("onboard_stage:"), state=[OnboardFSM.welcome, OnboardFSM.briefing, OnboardFSM.instructions])
async def onboard_stage_step(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    runtime = users_runtime.setdefault(uid, {})
    info = runtime.get("onboard_intro") or {}
    chat_id = info.get("chat_id") or c.message.chat.id
    message_id = info.get("message_id") or c.message.message_id
    profile = load_user(uid) or {}
    stage = c.data.split(":", 1)[1]

    if stage == "welcome":
        display_name = profile.get("first_name") or runtime.get("tg", {}).get("first_name") or profile.get("fullname") or runtime.get("tg", {}).get("username") or f"ID {uid}"
        text = tr(uid, "ONBOARD_WELCOME", name=h(display_name), bot=h(BOT_NAME))
        kb = kb_registration_next(uid, "onboard_stage:briefing")
        try:
            await bot.edit_message_text(text, chat_id, message_id, reply_markup=kb)
            runtime["onboard_intro"] = {"chat_id": chat_id, "message_id": message_id}
        except MessageNotModified:
            pass
        except Exception:
            msg = await bot.send_message(chat_id, text, reply_markup=kb)
            runtime["onboard_intro"] = {"chat_id": msg.chat.id, "message_id": msg.message_id}
        await state.set_state(OnboardFSM.briefing.state)
        await c.answer()
        return

    if stage == "briefing":
        text = tr(uid, "ONBOARD_BRIEFING")
        if runtime.get("onboard_registered"):
            text = f"{text}\n\n{tr(uid, 'ONBOARD_RETURNING_SHORTCUT')}"
        kb = kb_registration_next(uid, "onboard_stage:instructions")
        try:
            await bot.edit_message_text(text, chat_id, message_id, reply_markup=kb)
            runtime["onboard_intro"] = {"chat_id": chat_id, "message_id": message_id}
        except MessageNotModified:
            pass
        except Exception:
            msg = await bot.send_message(chat_id, text, reply_markup=kb)
            runtime["onboard_intro"] = {"chat_id": msg.chat.id, "message_id": msg.message_id}
        await state.set_state(OnboardFSM.instructions.state)
        await c.answer()
        return

    if stage == "instructions":
        try:
            await bot.delete_message(chat_id, message_id)
        except Exception:
            pass
        runtime.pop("onboard_intro", None)
        await flow_clear(uid)
        intro = await bot.send_message(chat_id, tr(uid, "REGISTER_INTRO_PROMPT"))
        flow_track(uid, intro)
        await onboard_prompt_last_name(uid, chat_id, state)
        await c.answer()
        return

    await c.answer()


async def onboard_prompt_last_name(uid: int, chat_id: int, state: FSMContext):
    await flow_clear_warnings(uid)
    msg = await bot.send_message(chat_id, tr(uid, "REGISTER_LAST_NAME_PROMPT"))
    flow_track(uid, msg)
    await state.set_state(OnboardFSM.last_name.state)


async def onboard_prompt_first_name(uid: int, chat_id: int, state: FSMContext):
    await flow_clear_warnings(uid)
    msg = await bot.send_message(chat_id, tr(uid, "REGISTER_FIRST_NAME_PROMPT"))
    flow_track(uid, msg)
    await state.set_state(OnboardFSM.first_name.state)


async def onboard_prompt_middle_name(uid: int, chat_id: int, state: FSMContext):
    await flow_clear_warnings(uid)
    msg = await bot.send_message(chat_id, tr(uid, "REGISTER_MIDDLE_NAME_PROMPT"))
    flow_track(uid, msg)
    await state.set_state(OnboardFSM.middle_name.state)


async def onboard_prompt_birthdate(uid: int, chat_id: int, state: FSMContext):
    await flow_clear_warnings(uid)
    msg = await bot.send_message(chat_id, tr(uid, "REGISTER_BIRTHDATE_PROMPT"))
    flow_track(uid, msg)
    await state.set_state(OnboardFSM.birthdate.state)


async def onboard_prompt_region(uid: int, chat_id: int, state: FSMContext):
    await flow_clear_warnings(uid)
    msg = await bot.send_message(chat_id, tr(uid, "REGISTER_REGION_PROMPT"), reply_markup=kb_region_prompt(uid))
    flow_track(uid, msg)
    await state.set_state(OnboardFSM.region.state)


async def onboard_prompt_phone(uid: int, chat_id: int, state: FSMContext):
    await flow_clear_warnings(uid)
    msg = await bot.send_message(chat_id, tr(uid, "REGISTER_PHONE_PROMPT_NEW"), reply_markup=kb_phone_keyboard(uid))
    flow_track(uid, msg)
    await state.set_state(OnboardFSM.phone.state)


async def onboard_prompt_photo(uid: int, chat_id: int, state: FSMContext):
    await flow_clear_warnings(uid)
    msg = await bot.send_message(chat_id, tr(uid, "REGISTER_PHOTO_PROMPT"), reply_markup=kb_photo_keyboard(uid))
    flow_track(uid, msg)
    await state.set_state(OnboardFSM.photo.state)


@dp.message_handler(state=OnboardFSM.last_name, content_types=ContentType.TEXT)
async def onboard_last_name(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    raw = normalize_person_name(m.text)
    await flow_delete_message(uid, m)
    if not validate_name(raw):
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_LAST_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_clear_warnings(uid)
    await state.update_data(last_name=beautify_name(raw))
    await onboard_prompt_first_name(uid, m.chat.id, state)


@dp.message_handler(state=OnboardFSM.first_name, content_types=ContentType.TEXT)
async def onboard_first_name(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    raw = normalize_person_name(m.text)
    await flow_delete_message(uid, m)
    if not validate_name(raw):
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_FIRST_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_clear_warnings(uid)
    await state.update_data(first_name=beautify_name(raw))
    await onboard_prompt_middle_name(uid, m.chat.id, state)


@dp.message_handler(state=OnboardFSM.middle_name, content_types=ContentType.TEXT)
async def onboard_middle_name(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    raw = normalize_person_name(m.text)
    await flow_delete_message(uid, m)
    if raw.lower() in SKIP_KEYWORDS:
        await flow_clear_warnings(uid)
        await state.update_data(middle_name="")
        await onboard_prompt_birthdate(uid, m.chat.id, state)
        return
    if not validate_name(raw):
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_MIDDLE_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_clear_warnings(uid)
    await state.update_data(middle_name=beautify_name(raw))
    await onboard_prompt_birthdate(uid, m.chat.id, state)


@dp.message_handler(state=OnboardFSM.birthdate, content_types=ContentType.TEXT)
async def onboard_birthdate(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    await flow_delete_message(uid, m)
    parsed = parse_birthdate_text(m.text)
    if not parsed:
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_BIRTHDATE_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_clear_warnings(uid)
    await state.update_data(birthdate=parsed.strftime("%Y-%m-%d"))
    await onboard_prompt_region(uid, m.chat.id, state)


@dp.message_handler(state=OnboardFSM.region, content_types=ContentType.ANY)
async def onboard_region_text(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_REGION_REMIND"), reply_markup=kb_region_prompt(uid))
    flow_track_warning(uid, warn)


@dp.callback_query_handler(lambda c: c.data == "reg_region_open", state=OnboardFSM.region)
async def onboard_region_open(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    runtime = users_runtime.setdefault(uid, {})
    chat_id = runtime.get("tg", {}).get("chat_id") or c.message.chat.id
    picker = await bot.send_message(chat_id, tr(uid, "REGISTER_REGION_PICK"), reply_markup=kb_region_picker(uid))
    flow_track(uid, picker)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("reg_region_pick:"), state=OnboardFSM.region)
async def onboard_region_pick(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    parts = c.data.split(":", 1)
    try:
        idx = int(parts[1])
        region = UKRAINE_REGIONS[idx]
    except Exception:
        await c.answer()
        return
    await state.update_data(region=region)
    await flow_clear_warnings(uid)
    await delete_if_not_anchor(uid, c.message.chat.id, c.message.message_id)
    runtime = users_runtime.setdefault(uid, {})
    chat_id = runtime.get("tg", {}).get("chat_id") or c.message.chat.id
    confirm = await bot.send_message(chat_id, tr(uid, "REGISTER_REGION_SELECTED", region=h(region)), reply_markup=kb_registration_next(uid, "reg_region_confirm"))
    flow_track(uid, confirm)
    await state.set_state(OnboardFSM.region_confirm.state)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "reg_region_confirm", state=OnboardFSM.region_confirm)
async def onboard_region_confirm(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await delete_if_not_anchor(uid, c.message.chat.id, c.message.message_id)
    runtime = users_runtime.setdefault(uid, {})
    chat_id = runtime.get("tg", {}).get("chat_id") or c.message.chat.id
    await onboard_prompt_phone(uid, chat_id, state)
    await c.answer()


@dp.message_handler(content_types=ContentType.CONTACT, state=OnboardFSM.phone)
async def onboard_phone_contact(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    phone = (m.contact.phone_number if m.contact else "").strip()
    await flow_delete_message(uid, m)
    if not phone:
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHONE_WARN"), reply_markup=kb_phone_keyboard(uid))
        flow_track_warning(uid, warn)
        return
    await flow_clear_warnings(uid)
    await state.update_data(phone=phone)
    ack = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHONE_OK"), reply_markup=ReplyKeyboardRemove())
    flow_track(uid, ack)
    await onboard_prompt_photo(uid, m.chat.id, state)


@dp.message_handler(state=OnboardFSM.phone, content_types=ContentType.TEXT)
async def onboard_phone_text(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHONE_TEXT_WARN"), reply_markup=kb_phone_keyboard(uid))
    flow_track_warning(uid, warn)


@dp.callback_query_handler(lambda c: c.data == "reg_photo_skip", state=OnboardFSM.photo)
async def onboard_photo_skip(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    runtime = users_runtime.setdefault(uid, {})
    chat_id = runtime.get("tg", {}).get("chat_id") or c.message.chat.id
    note = await bot.send_message(chat_id, tr(uid, "REGISTER_PHOTO_SKIP_CONFIRM"))
    flow_track(uid, note)
    await finalize_registration(uid, chat_id, state, photo_meta=None, skipped=True)
    await c.answer()


@dp.message_handler(content_types=ContentType.PHOTO, state=OnboardFSM.photo)
async def onboard_photo_received(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    photo = m.photo[-1] if m.photo else None
    if not photo:
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"), reply_markup=kb_photo_keyboard(uid))
        flow_track_warning(uid, warn)
        return
    try:
        meta = await store_profile_photo(uid, photo)
    except Exception:
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"), reply_markup=kb_photo_keyboard(uid))
        flow_track_warning(uid, warn)
        return
    await flow_clear_warnings(uid)
    ack = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_RECEIVED"))
    flow_track(uid, ack)
    runtime = users_runtime.setdefault(uid, {})
    chat_id = runtime.get("tg", {}).get("chat_id") or m.chat.id
    await finalize_registration(uid, chat_id, state, photo_meta=meta, skipped=False)


@dp.message_handler(state=OnboardFSM.photo, content_types=ContentType.ANY)
async def onboard_photo_other(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if m.content_type == ContentType.PHOTO:
        return
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"), reply_markup=kb_photo_keyboard(uid))
    flow_track_warning(uid, warn)


async def finalize_registration(uid: int, chat_id: int, state: FSMContext, photo_meta: Optional[dict], skipped: bool):
    data = await state.get_data()
    last_name = data.get("last_name", "")
    first_name = data.get("first_name", "")
    middle_name = data.get("middle_name", "")
    birthdate = data.get("birthdate", "")
    region = data.get("region", "")
    phone = data.get("phone", "")

    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["last_name"] = last_name
    profile["first_name"] = first_name
    profile["middle_name"] = middle_name
    profile["fullname"] = compose_fullname(last_name, first_name, middle_name) or profile.get("fullname")
    profile["birthdate"] = birthdate
    profile["region"] = region
    profile["phone"] = phone
    if photo_meta is not None:
        profile["photo"] = photo_meta
    elif skipped and not profile.get("photo"):
        profile["photo"] = {"status": "skipped", "updated_at": datetime.now(timezone.utc).isoformat()}
    profile["profile_completed"] = bool(last_name and first_name and phone and region)
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)

    await state.finish()
    await flow_clear(uid)
    runtime = users_runtime.setdefault(uid, {})
    runtime["onboard_registered"] = True
    runtime.pop("onboard_intro", None)

    confirm = await bot.send_message(chat_id, tr(uid, "REGISTER_FINISH_CONFIRM", code=h(profile.get("bsu", "—"))))
    schedule_auto_delete(confirm.chat.id, confirm.message_id, delay=20)
    await anchor_show_root(uid)


# ========================== PROFILE EDIT HANDLERS ==========================
@dp.message_handler(state=ProfileEditFSM.waiting_last_name, content_types=ContentType.TEXT)
async def profile_last_name_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    raw = normalize_person_name(m.text)
    if not raw or not NAME_VALID_RE.match(raw):
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_LAST_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_delete_message(uid, m)
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["last_name"] = beautify_name(raw)
    profile["fullname"] = compose_fullname(profile["last_name"], profile.get("first_name", ""), profile.get("middle_name"))
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await show_profile(uid, edit_mode=True, show_photo=False)


@dp.message_handler(state=ProfileEditFSM.waiting_first_name, content_types=ContentType.TEXT)
async def profile_first_name_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    raw = normalize_person_name(m.text)
    if not raw or not NAME_VALID_RE.match(raw):
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_FIRST_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_delete_message(uid, m)
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["first_name"] = beautify_name(raw)
    profile["fullname"] = compose_fullname(profile.get("last_name", ""), profile["first_name"], profile.get("middle_name"))
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await show_profile(uid, edit_mode=True, show_photo=False)


@dp.message_handler(state=ProfileEditFSM.waiting_middle_name, content_types=ContentType.TEXT)
async def profile_middle_name_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    raw = normalize_person_name(m.text)
    if raw and raw.lower() in SKIP_KEYWORDS:
        cleaned = ""
    else:
        if raw and NAME_VALID_RE.match(raw):
            cleaned = beautify_name(raw)
        else:
            await flow_delete_message(uid, m)
            warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_MIDDLE_NAME_WARN"))
            flow_track_warning(uid, warn)
            return
    await flow_delete_message(uid, m)
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["middle_name"] = cleaned
    profile["fullname"] = compose_fullname(profile.get("last_name", ""), profile.get("first_name", ""), cleaned)
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await show_profile(uid, edit_mode=True, show_photo=False)


@dp.message_handler(state=ProfileEditFSM.waiting_birthdate, content_types=ContentType.TEXT)
async def profile_birthdate_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    dt = parse_birthdate_text(m.text)
    if not dt:
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_BIRTHDATE_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_delete_message(uid, m)
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["birthdate"] = dt.strftime("%Y-%m-%d")
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await show_profile(uid, edit_mode=True, show_photo=False)


@dp.message_handler(state=ProfileEditFSM.waiting_region, content_types=ContentType.TEXT)
async def profile_region_text(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_REGION_REMIND"), reply_markup=kb_profile_region_prompt(uid))
    flow_track_warning(uid, warn)


@dp.callback_query_handler(lambda c: c.data == "profile_region_open", state=ProfileEditFSM.waiting_region)
async def profile_region_open(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await profile_send_prompt(uid, tr(uid, "REGISTER_REGION_PICK"), reply_markup=kb_profile_region_picker(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("profile_region_pick:"), state=ProfileEditFSM.waiting_region)
async def profile_region_pick(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    try:
        idx = int(c.data.split(":", 1)[1])
        region = UKRAINE_REGIONS[idx]
    except Exception:
        await c.answer(tr(uid, "REGISTER_REGION_REMIND"), show_alert=True)
        return
    await state.update_data(profile_region=region)
    await profile_send_prompt(uid, tr(uid, "REGISTER_REGION_SELECTED", region=h(region)), reply_markup=kb_registration_next(uid, "profile_region_confirm"))
    await ProfileEditFSM.region_confirm.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_region_confirm", state=ProfileEditFSM.region_confirm)
async def profile_region_confirm(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    data = await state.get_data()
    region = data.get("profile_region")
    if not region:
        await c.answer(tr(uid, "REGISTER_REGION_REMIND"), show_alert=True)
        return
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["region"] = region
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await show_profile(uid, edit_mode=True, show_photo=False)
    await c.answer()


@dp.message_handler(content_types=ContentType.CONTACT, state=ProfileEditFSM.waiting_phone)
async def profile_phone_contact(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    contact = m.contact
    if not contact or not contact.phone_number:
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHONE_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_delete_message(uid, m)
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["phone"] = contact.phone_number
    profile.setdefault("tg", {})["contact_user_id"] = contact.user_id
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    runtime = profile_runtime(uid)
    runtime.pop("profile_reply_keyboard", None)
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_PHONE_SAVED"), remove_keyboard=True)
    await show_profile(uid, edit_mode=True, show_photo=False)


@dp.message_handler(state=ProfileEditFSM.waiting_phone, content_types=ContentType.TEXT)
async def profile_phone_text(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    text = (m.text or "").strip()
    cancel_label = tr(uid, "BTN_PROFILE_CANCEL")
    if text == cancel_label:
        runtime = profile_runtime(uid)
        runtime.pop("profile_reply_keyboard", None)
        await flow_delete_message(uid, m)
        await profile_abort(uid, state, remove_keyboard=True)
        return
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHONE_TEXT_WARN"))
    flow_track_warning(uid, warn)


@dp.message_handler(state=ProfileEditFSM.waiting_photo, content_types=ContentType.PHOTO)
async def profile_photo_received(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    await flow_delete_message(uid, m)
    try:
        meta = await store_profile_photo(uid, m.photo[-1])
    except Exception:
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"))
        flow_track_warning(uid, warn)
        return
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    profile["photo"] = meta or {"status": "uploaded", "updated_at": datetime.now(timezone.utc).isoformat()}
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.finish()
    await profile_send_notification(uid, tr(uid, "PROFILE_PHOTO_UPDATED"))
    await show_profile(uid, edit_mode=True, show_photo=False)


@dp.message_handler(state=ProfileEditFSM.waiting_photo, content_types=ContentType.ANY)
async def profile_photo_invalid(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"))
    flow_track_warning(uid, warn)


@dp.callback_query_handler(lambda c: c.data == "profile_remove_photo")
async def profile_remove_photo_cb(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    profile = ensure_user(uid, users_runtime.get(uid, {}).get("tg", {}))
    remove_profile_photo(uid)
    profile["photo"] = {}
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await flow_clear_warnings(uid)
    await profile_send_notification(uid, tr(uid, "PROFILE_PHOTO_REMOVED"))
    await show_profile(uid, edit_mode=True, show_photo=False)
    await c.answer()


# ========================== ADMIN PROMOTE ==========================
# ========================== ADMIN PROMOTE ==========================
@dp.message_handler(lambda m: m.text and m.text.strip() == ADMIN_CODE)
async def become_admin(m: types.Message):
    uid = m.from_user.id
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    admins.add(uid); persist_state()
    x = await m.answer(
        "✅ Доступ администратора активирован.\n"
        "Используйте кнопку «🧑‍💼 Админ-панель» в главном меню, чтобы управлять пользователями, проектами и финансами."
    )
    flow_track(uid, x)
    await anchor_show_root(uid)


@dp.message_handler(content_types=ContentType.ANY, state=None)
async def fallback_message(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    text = m.text or ""
    if text.startswith("/"):
        return
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    warn = await bot.send_message(m.chat.id, tr(uid, "INVALID_COMMAND"))
    flow_track(uid, warn)


# ========================== NAVIGATION ==========================
@dp.callback_query_handler(lambda c: c.data == "menu_about")
async def menu_about(c: types.CallbackQuery):
    uid = c.from_user.id
    prof = load_user(uid) or {}
    about_text = (
        f"🤖 <b>{h(BOT_NAME)}</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"Версия: <code>{h(BOT_VERSION)}</code>\n"
        f"Ревизия: <code>{h(BOT_REVISION)}</code>\n"
        f"User ID: <code>{uid}</code>\n"
        f"BSU код: <b>{h(prof.get('bsu', '—'))}</b>\n\n"
        "📌 <b>Что умеет бот</b>\n"
        "• хранить и структурировать чеки;\n"
        "• помогать с запросами на выплаты;\n"
        "• распространять документы и новости по проекту.\n\n"
        "Используйте кнопки ниже, чтобы вернуться в главное меню и продолжить работу."
    )
    await clear_then_anchor(
        uid,
        about_text,
        kb=InlineKeyboardMarkup().add(InlineKeyboardButton("⬅️ На главную", callback_data="back_root"))
    )
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "menu_alerts")
async def menu_alerts(c: types.CallbackQuery):
    uid = c.from_user.id
    intro = tr(uid, "ALERTS_MENU_INTRO")
    count = alerts_active_oblast_count()
    if count:
        intro = f"{intro}\n\n{alerts_active_summary_line(uid)}"
    await clear_then_anchor(uid, intro, kb_alerts(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "alerts_active")
async def alerts_active_view(c: types.CallbackQuery):
    uid = c.from_user.id
    regions = alerts_user_regions(uid)
    if not regions:
        await clear_then_anchor(uid, tr(uid, "ALERTS_NO_REGIONS"), kb_alerts(uid))
        await c.answer()
        return
    events = alerts_collect_active_for_user(uid)
    if not events:
        await clear_then_anchor(uid, tr(uid, "ALERTS_NO_ACTIVE"), kb_alerts(uid))
        await c.answer()
        return
    lang = resolve_lang(uid)
    labels = alerts_field_labels(lang)
    status_labels = ALERTS_STATUS_TEXT.get(lang) or ALERTS_STATUS_TEXT[DEFAULT_LANG]
    divider = tr(uid, "ALERTS_ACTIVE_DIVIDER")
    lines = [
        tr(uid, "ALERTS_ACTIVE_HEADER", count=len(events)),
        divider,
    ]
    for event in events[:10]:
        region_display = alerts_display_region_name(
            event.get("region") or event.get("region_display") or "",
            lang,
            short=True,
        )
        type_text = alerts_type_label(event, lang)
        severity_text = alerts_severity_label(event, lang)
        summary_text = type_text or status_labels.get("alert", "")
        if severity_text:
            summary_text = f"{summary_text} • {severity_text}" if summary_text else severity_text
        lines.append(f"🔴 {h(region_display)} — {h(summary_text)}")
        started_display = alerts_format_datetime_display(event.get("started_at"))
        if not started_display:
            started_display = alerts_format_timestamp(event.get("started_at")) or labels["status_unknown"]
        lines.append(f"⏱ {h(started_display)}")
    lines.append(divider)
    lines.append("")
    lines.append(tr(uid, "ALERTS_ACTIVE_SUMMARY_TOTAL", count=len(events)))
    lines.append(tr(uid, "ALERTS_ACTIVE_SUMMARY_USER"))
    lines.append(tr(uid, "ALERTS_ACTIVE_SUMMARY_PROJECT"))
    await clear_then_anchor(uid, "\n".join(lines), kb_alerts(uid))
    await alerts_send_card(uid, c.message.chat.id, events, "active", index=0)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "alerts_overview")
async def alerts_overview_view(c: types.CallbackQuery):
    uid = c.from_user.id
    text = alerts_regions_overview_text(uid)
    await clear_then_anchor(uid, text, kb_alerts(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "alerts_history")
async def alerts_history_view(c: types.CallbackQuery):
    uid = c.from_user.id
    regions = alerts_user_regions(uid)
    if not regions:
        await clear_then_anchor(uid, tr(uid, "ALERTS_NO_REGIONS"), kb_alerts(uid))
        await c.answer()
        return
    events = alerts_collect_history_for_user(uid)
    if not events:
        await clear_then_anchor(uid, tr(uid, "ALERTS_NO_HISTORY"), kb_alerts(uid))
        await c.answer()
        return
    lang = resolve_lang(uid)
    status_labels = ALERTS_STATUS_TEXT.get(lang) or ALERTS_STATUS_TEXT[DEFAULT_LANG]
    labels = alerts_field_labels(lang)
    indent = "&nbsp;&nbsp;&nbsp;"
    lines = [tr(uid, "ALERTS_HISTORY_HEADER", count=len(events))]
    for idx, event in enumerate(events[:10], start=1):
        region_display = alerts_display_region_name(event.get("region") or event.get("region_display") or "", lang)
        start_text = alerts_format_timestamp(event.get("started_at")) or labels["status_unknown"]
        end_text = alerts_format_timestamp(event.get("ended_at")) if event.get("ended_at") else labels["status_active"]
        type_text = alerts_type_label(event, lang)
        severity_text = alerts_severity_label(event, lang)
        ended = bool(event.get("ended_at"))
        status_key = "standdown" if ended else "alert"
        status_icon = "🟡" if ended else "🔴"
        summary_parts = [status_labels[status_key], type_text]
        if severity_text:
            summary_parts.append(severity_text)
        lines.append(f"{idx}. {status_icon} <b>{h(region_display)}</b> — {h(' • '.join(summary_parts))}")
        lines.append(f"{indent}⏱ {h(labels['started'])}: {h(start_text)}")
        if ended:
            lines.append(f"{indent}🛑 {h(labels['ended'])}: {h(end_text)}")
    await clear_then_anchor(uid, "\n".join(lines), kb_alerts(uid))
    await alerts_send_card(uid, c.message.chat.id, events, "history", index=0)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "alerts_subscriptions")
async def alerts_subscriptions_menu(c: types.CallbackQuery):
    uid = c.from_user.id
    text, kb = alerts_subscription_view(uid, page=0)
    await clear_then_anchor(uid, text, kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("alerts_sub_page:"))
async def alerts_subscriptions_page(c: types.CallbackQuery):
    uid = c.from_user.id
    try:
        page = int(c.data.split(":", 1)[1])
    except ValueError:
        page = 0
    text, kb = alerts_subscription_view(uid, page=page)
    await anchor_show_text(uid, text, kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("alerts_toggle:"))
async def alerts_toggle_subscription(c: types.CallbackQuery):
    uid = c.from_user.id
    try:
        _, page_raw, idx_raw = c.data.split(":", 2)
        page = int(page_raw)
        region_index = int(idx_raw)
    except Exception:
        await c.answer("", show_alert=False)
        return
    profile = load_user(uid) or {"user_id": uid}
    alerts = alerts_profile_block(profile)
    region = alerts_canonical_region(UKRAINE_REGIONS[region_index]) or UKRAINE_REGIONS[region_index]
    items = alerts.get("regions", [])
    add = region not in items
    alerts_update_subscription(uid, region_index, add)
    text, kb = alerts_subscription_view(uid, page=page)
    await anchor_show_text(uid, text, kb)
    key = "ALERTS_SUBS_ADDED" if add else "ALERTS_SUBS_REMOVED"
    await c.answer(tr(uid, key, region=h(region)), show_alert=False)


@dp.callback_query_handler(lambda c: c.data == "alerts_locked")
async def alerts_locked_info(c: types.CallbackQuery):
    uid = c.from_user.id
    await c.answer(tr(uid, "ALERTS_SUBS_LOCKED"), show_alert=True)


@dp.callback_query_handler(lambda c: c.data.startswith("alerts_index:"))
async def alerts_card_index_stub(c: types.CallbackQuery):
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("alerts_nav:"))
async def alerts_card_nav(c: types.CallbackQuery):
    uid = c.from_user.id
    parts = c.data.split(":", 2)
    if len(parts) != 3:
        await c.answer()
        return
    context = parts[1]
    try:
        target_index = int(parts[2])
    except ValueError:
        await c.answer()
        return
    runtime = users_runtime.setdefault(uid, {})
    cards = runtime.get("alerts_cards", {})
    card = cards.get(context)
    if not card:
        await c.answer()
        return
    event_ids: List[str] = card.get("events", [])
    events: List[Dict[str, Any]] = []
    for event_id in event_ids:
        event = _alerts_get_event(event_id)
        if event:
            events.append(event)
    if not events:
        await c.answer(tr(uid, "ALERTS_NO_ACTIVE"), show_alert=True)
        return
    target_index = max(0, min(target_index, len(events) - 1))
    current_index = max(0, min(int(card.get("index", 0)), len(events) - 1))
    if target_index == current_index:
        await c.answer()
        return
    card["index"] = target_index
    lang = resolve_lang(uid)
    text = alerts_format_card(events[target_index], lang, index=target_index, total=len(events))
    kb = alerts_card_keyboard(uid, context, len(events), target_index)
    try:
        await bot.edit_message_text(text, c.message.chat.id, c.message.message_id, reply_markup=kb, disable_web_page_preview=True)
    except MessageNotModified:
        pass
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("alerts_close:"))
async def alerts_close_card(c: types.CallbackQuery):
    uid = c.from_user.id
    context = c.data.split(":", 1)[1]
    runtime = users_runtime.setdefault(uid, {})
    cards = runtime.setdefault("alerts_cards", {})
    cards.pop(context, None)
    try:
        await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception:
        pass
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("alerts_push:"))
async def alerts_push_actions(c: types.CallbackQuery):
    uid = c.from_user.id
    parts = c.data.split(":", 2)
    if len(parts) != 3:
        await c.answer()
        return
    action, token = parts[1], parts[2]
    entry = alerts_push_get(uid, token)
    if not entry:
        await c.answer(tr(uid, "ALERTS_NO_ACTIVE"), show_alert=True)
        return
    event_id = entry.get("event_id")
    event = _alerts_get_event(event_id) if event_id else None
    if action in {"expand", "collapse"}:
        if not event:
            await c.answer(tr(uid, "ALERTS_NO_ACTIVE"), show_alert=True)
            return
        expanded = action == "expand"
        kind = entry.get("kind") or ("end" if event.get("ended_at") else "start")
        entry["kind"] = kind
        entry["expanded"] = expanded
        text = alerts_push_render(uid, event, kind, expanded=expanded)
        kb = alerts_push_keyboard(uid, token, expanded)
        try:
            await bot.edit_message_text(
                text,
                c.message.chat.id,
                c.message.message_id,
                reply_markup=kb,
                disable_web_page_preview=True,
            )
        except MessageNotModified:
            pass
        except MessageCantBeEdited:
            pass
        alerts_push_store(uid, token, entry)
        await c.answer()
        return
    if action == "delete":
        alerts_push_remove(uid, token)
        try:
            await bot.delete_message(c.message.chat.id, c.message.message_id)
        except Exception:
            pass
        await c.answer()
        return
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "back_root")
async def back_root(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, project_status_text(uid), kb_root(uid))
    await c.answer()


# ========================== ALERTS STORAGE ==========================
ALERTS_STORAGE_BASE = os.path.join("data", ALERTS_DIRNAME)
_alerts_state_cache: Dict[str, Dict[str, Any]] = {}
_alerts_user_cache: Dict[str, Dict[str, Any]] = {}


def _alerts_resolve_project(project: Optional[str] = None) -> Optional[str]:
    if project:
        return project
    return None


def _alerts_context_key(project: Optional[str] = None) -> str:
    resolved = _alerts_resolve_project(project)
    return resolved or "__global__"


def alerts_storage_root(project: Optional[str] = None) -> str:
    base = ALERTS_STORAGE_BASE
    os.makedirs(base, exist_ok=True)

    def _ensure_flattened() -> None:
        legacy_global = os.path.join(base, "__global__")
        if os.path.isdir(legacy_global):
            for name in (ALERTS_STATE_FILENAME, ALERTS_USERS_FILENAME):
                src = os.path.join(legacy_global, name)
                dst = os.path.join(base, name)
                if os.path.exists(src) and not os.path.exists(dst):
                    os.replace(src, dst)
            legacy_history = os.path.join(legacy_global, ALERTS_HISTORY_DIRNAME)
            if os.path.isdir(legacy_history):
                target_history = os.path.join(base, ALERTS_HISTORY_DIRNAME)
                os.makedirs(target_history, exist_ok=True)
                for entry in os.listdir(legacy_history):
                    src = os.path.join(legacy_history, entry)
                    dst = os.path.join(target_history, entry)
                    if os.path.exists(src) and not os.path.exists(dst):
                        os.replace(src, dst)
            try:
                if not os.listdir(legacy_global):
                    os.rmdir(legacy_global)
            except Exception:
                pass

    _ensure_flattened()

    if project and project not in {"", "__global__"}:
        legacy_path = os.path.join(base, project)
        if os.path.isdir(legacy_path):
            return legacy_path
    return base


def alerts_state_file(project: Optional[str] = None) -> str:
    return os.path.join(alerts_storage_root(project), ALERTS_STATE_FILENAME)


def alerts_history_dir(project: Optional[str] = None) -> str:
    path = os.path.join(alerts_storage_root(project), ALERTS_HISTORY_DIRNAME)
    os.makedirs(path, exist_ok=True)
    return path


def alerts_history_file(project: Optional[str] = None, day_key: Optional[str] = None) -> str:
    day = day_key or alerts_today_key()
    return os.path.join(alerts_history_dir(project), f"{day}.json")


def alerts_users_file(project: Optional[str] = None) -> str:
    return os.path.join(alerts_storage_root(project), ALERTS_USERS_FILENAME)


def _alerts_save_timeline(project: Optional[str], timeline: List[Dict[str, Any]], day_key: Optional[str] = None) -> None:
    path = alerts_history_file(project, day_key)
    tmp_file = f"{path}.tmp"
    with open(tmp_file, "w", encoding="utf-8") as fh:
        json.dump(timeline, fh, ensure_ascii=False, indent=2)
    os.replace(tmp_file, path)


def _alerts_load_timeline(project: Optional[str], day_key: Optional[str] = None) -> List[Dict[str, Any]]:
    path = alerts_history_file(project, day_key)
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        if isinstance(payload, list):
            return payload
    except Exception:
        pass
    return []


def _alerts_write_state_payload(project: Optional[str], payload: Dict[str, Any]) -> None:
    timeline = list(payload.get(ALERTS_TIMELINE_KEY) or [])
    day_key = payload.get("_timeline_day") or alerts_today_key()
    state_copy = dict(payload)
    state_copy.pop(ALERTS_TIMELINE_KEY, None)
    state_copy.pop("_timeline_day", None)
    state_copy.pop("_context_project", None)
    path = alerts_state_file(project)
    tmp_file = f"{path}.tmp"
    with open(tmp_file, "w", encoding="utf-8") as fh:
        json.dump(state_copy, fh, ensure_ascii=False, indent=2)
    os.replace(tmp_file, path)
    _alerts_save_timeline(project, timeline, day_key)


def _alerts_migrate_legacy(project: Optional[str]) -> None:
    if not project:
        return
    legacy_root = os.path.join(proj_path(project), ALERTS_DIRNAME)
    if not os.path.isdir(legacy_root):
        return
    state_path = alerts_state_file(project)
    subs_path = alerts_users_file(project)
    migrated = False
    if not os.path.exists(state_path):
        legacy_history = os.path.join(legacy_root, ALERTS_LEGACY_HISTORY_FILENAME)
        if os.path.exists(legacy_history):
            try:
                with open(legacy_history, "r", encoding="utf-8") as fh:
                    legacy_payload = json.load(fh)
                if not isinstance(legacy_payload, dict):
                    raise ValueError("Invalid legacy alerts state")
            except Exception:
                legacy_payload = _alerts_blank_state()
            payload = _alerts_blank_state()
            payload.update({
                "events": legacy_payload.get("events", {}),
                "regions": legacy_payload.get("regions", {}),
                "last_fetch": legacy_payload.get("last_fetch"),
            })
            timeline = legacy_payload.get(ALERTS_TIMELINE_KEY)
            if isinstance(timeline, list):
                payload[ALERTS_TIMELINE_KEY] = timeline
            payload["_timeline_day"] = alerts_today_key()
            payload["_context_project"] = project
            _alerts_write_state_payload(project, payload)
            migrated = True
    if not os.path.exists(subs_path):
        legacy_subs = os.path.join(legacy_root, ALERTS_USERS_FILENAME)
        if os.path.exists(legacy_subs):
            try:
                with open(legacy_subs, "r", encoding="utf-8") as fh:
                    subs_payload = json.load(fh)
                if isinstance(subs_payload, dict):
                    tmp_file = f"{subs_path}.tmp"
                    with open(tmp_file, "w", encoding="utf-8") as fh:
                        json.dump(subs_payload, fh, ensure_ascii=False, indent=2)
                    os.replace(tmp_file, subs_path)
                    migrated = True
            except Exception:
                pass
    if migrated:
        try:
            legacy_marker = os.path.join(legacy_root, "migrated.txt")
            with open(legacy_marker, "w", encoding="utf-8") as fh:
                fh.write(alerts_now().isoformat())
        except Exception:
            pass


if ZoneInfo:
    try:
        ALERTS_TIMEZONE = ZoneInfo("Europe/Kiev")
    except Exception:
        ALERTS_TIMEZONE = timezone.utc
else:
    ALERTS_TIMEZONE = timezone.utc

ALERTS_REGION_EQUIVALENTS: Dict[str, List[str]] = {
    "Винницкая область": ["Вінницька область", "Vinnytska oblast", "Vinnytsia region"],
    "Волынская область": ["Волинська область", "Volynska oblast", "Volyn region"],
    "Днепропетровская область": ["Дніпропетровська область", "Dnipropetrovska oblast", "Dnipropetrovsk region"],
    "Донецкая область": ["Донецька область", "Donetska oblast", "Donetsk region"],
    "Житомирская область": ["Житомирська область", "Zhytomyrska oblast", "Zhytomyr region"],
    "Закарпатская область": ["Закарпатська область", "Zakarpatska oblast", "Zakarpattia region"],
    "Запорожская область": ["Запорізька область", "Zaporizka oblast", "Zaporizhzhia region"],
    "Ивано-Франковская область": ["Івано-Франківська область", "Ivano-Frankivska oblast", "Ivano-Frankivsk region"],
    "Киевская область": [
        "Київська область",
        "Kyivska oblast",
        "Kyiv region",
        "м. Київ",
        "Київ",
        "Kyiv",
        "Kiev",
        "Kyiv City",
    ],
    "Кировоградская область": ["Кіровоградська область", "Kirovohradska oblast", "Kirovohrad region"],
    "Луганская область": ["Луганська область", "Luhanska oblast", "Luhansk region"],
    "Львовская область": ["Львівська область", "Lvivska oblast", "Lviv region"],
    "Николаевская область": ["Миколаївська область", "Mykolaivska oblast", "Mykolaiv region"],
    "Одесская область": ["Одеська область", "Odeska oblast", "Odesa region"],
    "Полтавская область": ["Полтавська область", "Poltavska oblast", "Poltava region"],
    "Ровенская область": ["Рівненська область", "Rivnenska oblast", "Rivne region"],
    "Сумская область": ["Сумська область", "Sumska oblast", "Sumy region"],
    "Тернопольская область": ["Тернопільська область", "Ternopilska oblast", "Ternopil region"],
    "Харьковская область": ["Харківська область", "Kharkivska oblast", "Kharkiv region"],
    "Херсонская область": ["Херсонська область", "Khersonska oblast", "Kherson region"],
    "Хмельницкая область": ["Хмельницька область", "Khmelnytska oblast", "Khmelnytskyi region"],
    "Черкасская область": ["Черкаська область", "Cherkaska oblast", "Cherkasy region"],
    "Черниговская область": ["Чернігівська область", "Chernihivska oblast", "Chernihiv region"],
    "Черновицкая область": ["Чернівецька область", "Chernivetska oblast", "Chernivtsi region"],
}

ALERTS_REGION_SHORT_NAMES: Dict[str, Dict[str, str]] = {
    "Винницкая область": {
        "uk": "Вінниця",
        "ru": "Винница",
        "en": "Vinnytsia",
        "de": "Winnyzja",
        "pl": "Winnica",
    },
    "Волынская область": {
        "uk": "Волинь",
        "ru": "Волынь",
        "en": "Volyn",
        "de": "Wolhynien",
        "pl": "Wołyń",
    },
    "Днепропетровская область": {
        "uk": "Дніпро",
        "ru": "Днепр",
        "en": "Dnipro",
        "de": "Dnipro",
        "pl": "Dnipro",
    },
    "Донецкая область": {
        "uk": "Донецьк",
        "ru": "Донецк",
        "en": "Donetsk",
        "de": "Donezk",
        "pl": "Donieck",
    },
    "Житомирская область": {
        "uk": "Житомир",
        "ru": "Житомир",
        "en": "Zhytomyr",
        "de": "Schytomyr",
        "pl": "Żytomierz",
    },
    "Закарпатская область": {
        "uk": "Закарпаття",
        "ru": "Закарпатье",
        "en": "Zakarpattia",
        "de": "Transkarpatien",
        "pl": "Zakarpacie",
    },
    "Запорожская область": {
        "uk": "Запоріжжя",
        "ru": "Запорожье",
        "en": "Zaporizhzhia",
        "de": "Saporischschja",
        "pl": "Zaporoże",
    },
    "Ивано-Франковская область": {
        "uk": "Івано-Франківськ",
        "ru": "Ивано-Франковск",
        "en": "Ivano-Frankivsk",
        "de": "Iwano-Frankiwsk",
        "pl": "Iwano-Frankiwsk",
    },
    "Киевская область": {
        "uk": "Київ",
        "ru": "Киев",
        "en": "Kyiv",
        "de": "Kyjiw",
        "pl": "Kijów",
    },
    "Кировоградская область": {
        "uk": "Кропивницький",
        "ru": "Кропивницкий",
        "en": "Kropyvnytskyi",
        "de": "Kropywnyzkyj",
        "pl": "Kropywnycki",
    },
    "Луганская область": {
        "uk": "Луганськ",
        "ru": "Луганск",
        "en": "Luhansk",
        "de": "Luhansk",
        "pl": "Ługańsk",
    },
    "Львовская область": {
        "uk": "Львів",
        "ru": "Львов",
        "en": "Lviv",
        "de": "Lwiw",
        "pl": "Lwów",
    },
    "Николаевская область": {
        "uk": "Миколаїв",
        "ru": "Николаев",
        "en": "Mykolaiv",
        "de": "Mykolajiw",
        "pl": "Mikołajów",
    },
    "Одесская область": {
        "uk": "Одеса",
        "ru": "Одесса",
        "en": "Odesa",
        "de": "Odessa",
        "pl": "Odessa",
    },
    "Полтавская область": {
        "uk": "Полтава",
        "ru": "Полтава",
        "en": "Poltava",
        "de": "Poltawa",
        "pl": "Połtawa",
    },
    "Ровенская область": {
        "uk": "Рівне",
        "ru": "Ровно",
        "en": "Rivne",
        "de": "Riwne",
        "pl": "Równe",
    },
    "Сумская область": {
        "uk": "Суми",
        "ru": "Сумы",
        "en": "Sumy",
        "de": "Sumy",
        "pl": "Sumy",
    },
    "Тернопольская область": {
        "uk": "Тернопіль",
        "ru": "Тернополь",
        "en": "Ternopil",
        "de": "Ternopil",
        "pl": "Tarnopol",
    },
    "Харьковская область": {
        "uk": "Харків",
        "ru": "Харьков",
        "en": "Kharkiv",
        "de": "Charkiw",
        "pl": "Charków",
    },
    "Херсонская область": {
        "uk": "Херсон",
        "ru": "Херсон",
        "en": "Kherson",
        "de": "Cherson",
        "pl": "Chersoń",
    },
    "Хмельницкая область": {
        "uk": "Хмельницький",
        "ru": "Хмельницкий",
        "en": "Khmelnytskyi",
        "de": "Chmelnyzkyj",
        "pl": "Chmielnicki",
    },
    "Черкасская область": {
        "uk": "Черкаси",
        "ru": "Черкассы",
        "en": "Cherkasy",
        "de": "Tscherkassy",
        "pl": "Czerkasy",
    },
    "Черниговская область": {
        "uk": "Чернігів",
        "ru": "Чернигов",
        "en": "Chernihiv",
        "de": "Tschernihiw",
        "pl": "Czernihów",
    },
    "Черновицкая область": {
        "uk": "Чернівці",
        "ru": "Черновцы",
        "en": "Chernivtsi",
        "de": "Tscherniwzi",
        "pl": "Czerniowce",
    },
}

ALERTS_TYPE_ALIASES: Dict[str, str] = {
    "air_raid": "air_raid",
    "air-raid": "air_raid",
    "airalert": "air_raid",
    "air alert": "air_raid",
    "air_raid_alert": "air_raid",
    "artillery": "artillery_shelling",
    "artillery_shelling": "artillery_shelling",
    "shelling": "artillery_shelling",
    "missile": "missile_strike",
    "missile_attack": "missile_strike",
    "missile_strike": "missile_strike",
    "rocket": "missile_strike",
    "rocket_attack": "missile_strike",
    "ballistic": "missile_strike",
    "ballistic_missile": "missile_strike",
    "drone": "drone_attack",
    "drone_attack": "drone_attack",
    "uav": "drone_attack",
    "nuclear": "nuclear",
    "nuclear_threat": "nuclear",
    "chemical": "chemical",
    "chemical_threat": "chemical",
    "urban_fights": "urban_fights",
    "urban": "urban_fights",
    "ground_assault": "urban_fights",
    "unknown": "unknown",
}

ALERTS_DEFAULT_SEVERITY: Dict[str, str] = {
    "air_raid": "high",
    "artillery_shelling": "high",
    "missile_strike": "critical",
    "drone_attack": "medium",
    "nuclear": "critical",
    "chemical": "critical",
    "urban_fights": "high",
    "unknown": "high",
}

ALERTS_SEVERITY_KEYWORDS: Dict[str, str] = {
    "critical": "critical",
    "extreme": "critical",
    "highest": "critical",
    "max": "critical",
    "максим": "critical",
    "крит": "critical",
    "violet": "critical",
    "purple": "critical",
    "фіолет": "critical",
    "червон": "high",
    "red": "high",
    "висок": "high",
    "high": "high",
    "повітряна": "high",
    "жовт": "medium",
    "yellow": "medium",
    "orange": "medium",
    "помаран": "medium",
    "середн": "medium",
    "medium": "medium",
    "elevated": "medium",
    "зел": "low",
    "green": "low",
    "низ": "low",
    "low": "low",
    "мінім": "low",
    "none": "low",
    "відсут": "low",
    "без": "low",
}

ALERTS_TYPE_LABELS: Dict[str, Dict[str, str]] = {
    "air_raid": {
        "uk": "Повітряна тривога",
        "en": "Air raid alert",
        "de": "Luftalarm",
        "pl": "Alarm lotniczy",
        "ru": "Воздушная тревога",
    },
    "artillery_shelling": {
        "uk": "Артилерійський обстріл",
        "en": "Artillery shelling",
        "de": "Artilleriebeschuss",
        "pl": "Ostrzał artyleryjski",
        "ru": "Артиллерийский обстрел",
    },
    "missile_strike": {
        "uk": "Ракетна загроза",
        "en": "Missile threat",
        "de": "Raketenbedrohung",
        "pl": "Zagrożenie rakietowe",
        "ru": "Ракетная угроза",
    },
    "drone_attack": {
        "uk": "Атака дронів",
        "en": "Drone attack",
        "de": "Drohnenangriff",
        "pl": "Atak dronów",
        "ru": "Атака дронов",
    },
    "nuclear": {
        "uk": "Ядерна загроза",
        "en": "Nuclear threat",
        "de": "Atomare Gefahr",
        "pl": "Zagrożenie nuklearne",
        "ru": "Ядерная угроза",
    },
    "chemical": {
        "uk": "Хімічна загроза",
        "en": "Chemical threat",
        "de": "Chemische Gefahr",
        "pl": "Zagrożenie chemiczne",
        "ru": "Химическая угроза",
    },
    "urban_fights": {
        "uk": "Бої в місті",
        "en": "Urban fights",
        "de": "Straßenkämpfe",
        "pl": "Walki w mieście",
        "ru": "Бои в городе",
    },
    "unknown": {
        "uk": "Тривога",
        "en": "Alert",
        "de": "Alarm",
        "pl": "Alarm",
        "ru": "Тревога",
    },
}

ALERTS_TYPE_ICONS: Dict[str, str] = {
    "air_raid": "📡",
    "missile_strike": "🚀",
    "drone_attack": "🛩",
    "artillery_shelling": "💥",
    "urban_fights": "⚔️",
    "nuclear": "☢",
    "chemical": "☣",
    "unknown": "🚨",
}

ALERTS_SEVERITY_LABELS: Dict[str, Dict[str, str]] = {
    "low": {
        "icon": "🟢",
        "uk": "Низький",
        "en": "Low",
        "de": "Niedrig",
        "pl": "Niski",
        "ru": "Низкий",
    },
    "medium": {
        "icon": "🟡",
        "uk": "Серйозний",
        "en": "Serious",
        "de": "Ernst",
        "pl": "Poważny",
        "ru": "Серьёзный",
    },
    "high": {
        "icon": "🟠",
        "uk": "Високий",
        "en": "High",
        "de": "Hoch",
        "pl": "Wysoki",
        "ru": "Высокий",
    },
    "critical": {
        "icon": "🔴",
        "uk": "Небезпечний",
        "en": "Critical",
        "de": "Kritisch",
        "pl": "Krytyczny",
        "ru": "Опасный",
    },
}

ALERTS_STATUS_TEXT: Dict[str, Dict[str, str]] = {
    "uk": {
        "alert": "Тривога",
        "standdown": "Відбій тривоги",
        "calm": "Відбій тривоги",
    },
    "en": {
        "alert": "Alert",
        "standdown": "Alert cleared",
        "calm": "Alert cleared",
    },
    "de": {
        "alert": "Alarm",
        "standdown": "Alarm beendet",
        "calm": "Alarm aufgehoben",
    },
    "pl": {
        "alert": "Alarm",
        "standdown": "Alarm odwołano",
        "calm": "Alarm odwołano",
    },
    "ru": {
        "alert": "Тревога",
        "standdown": "Отбой тревоги",
        "calm": "Отбой тревоги",
    },
}

ALERTS_DURATION_FORMS: Dict[str, Dict[str, Tuple[str, ...]]] = {
    "uk": {
        "hour": ("год", "год", "год"),
        "minute": ("хв", "хв", "хв"),
        "alarm": ("тривога", "тривоги", "тривог"),
    },
    "ru": {
        "hour": ("ч", "ч", "ч"),
        "minute": ("мин", "мин", "мин"),
        "alarm": ("тревога", "тревоги", "тревог"),
    },
    "pl": {
        "hour": ("godz", "godz", "godz"),
        "minute": ("min", "min", "min"),
        "alarm": ("alarm", "alarmy", "alarmów"),
    },
    "de": {
        "hour": ("Std", "Std"),
        "minute": ("Min", "Min"),
        "alarm": ("Alarm", "Alarme"),
    },
    "en": {
        "hour": ("hr", "hrs"),
        "minute": ("min", "min"),
        "alarm": ("alert", "alerts"),
    },
}

ALERTS_RECOMMENDATIONS: Dict[str, Dict[str, List[str]]] = {
    "default": {
        "uk": [
            "— Прямуйте в укриття",
            "— Зачиніть двері та вікна",
            "— Тримайтеся подалі від вікон і вітрин",
        ],
        "en": [
            "— Move to shelter",
            "— Close doors and windows",
            "— Stay away from windows and glass",
        ],
        "de": [
            "— Begeben Sie sich in einen Schutzraum",
            "— Schließen Sie Türen und Fenster",
            "— Halten Sie Abstand von Fenstern und Glasflächen",
        ],
        "pl": [
            "— Udać się do schronu",
            "— Zamknij drzwi i okna",
            "— Trzymaj się z dala od okien i witryn",
        ],
        "ru": [
            "— Следуйте в укрытие",
            "— Закройте двери и окна",
            "— Избегайте окон и витрин",
        ],
    }
}

ALERTS_OVERVIEW_STATUS_TEXT: Dict[str, Dict[str, str]] = {
    "uk": {
        "alert": "Тривога",
        "standdown": "Відбій",
        "calm": "Відбій",
    },
    "en": {
        "alert": "Alert",
        "standdown": "Cleared",
        "calm": "Cleared",
    },
    "de": {
        "alert": "Alarm",
        "standdown": "Entwarnung",
        "calm": "Entwarnung",
    },
    "pl": {
        "alert": "Alarm",
        "standdown": "Odwołано",
        "calm": "Odwołано",
    },
    "ru": {
        "alert": "Тревога",
        "standdown": "Отбой",
        "calm": "Отбой",
    },
}

ALERTS_LOCATION_TYPE_LABELS: Dict[str, Dict[str, str]] = {
    "oblast": {
        "uk": "Область",
        "en": "Oblast",
        "de": "Oblast",
        "pl": "Obwód",
        "ru": "Область",
    },
    "raion": {
        "uk": "Район",
        "en": "District",
        "de": "Rajon",
        "pl": "Rejon",
        "ru": "Район",
    },
    "hromada": {
        "uk": "Громада",
        "en": "Community",
        "de": "Gemeinde",
        "pl": "Hromada",
        "ru": "Громада",
    },
    "community": {
        "uk": "Громада",
        "en": "Community",
        "de": "Gemeinschaft",
        "pl": "Wspólnota",
        "ru": "Сообщество",
    },
    "city": {
        "uk": "Місто",
        "en": "City",
        "de": "Stadt",
        "pl": "Miasto",
        "ru": "Город",
    },
    "settlement": {
        "uk": "Населений пункт",
        "en": "Settlement",
        "de": "Siedlung",
        "pl": "Osada",
        "ru": "Населённый пункт",
    },
    "village": {
        "uk": "Село",
        "en": "Village",
        "de": "Dorf",
        "pl": "Wieś",
        "ru": "Деревня",
    },
}

ALERTS_FIELD_LABELS: Dict[str, Dict[str, str]] = {
    "uk": {
        "header_active": "🚨 УВАГА! ТРИВОГА 🚨",
        "header_ended": "🟢 ВІДБІЙ ТРИВОГИ",
        "type": "Тип",
        "region": "Область",
        "location": "Локація",
        "location_type": "Тип локації",
        "coordinates": "Координати",
        "severity": "Рівень",
        "cause": "Причина",
        "details": "Деталі",
        "started": "Початок",
        "ended": "Відбій",
        "duration": "Тривалість",
        "message": "Повідомлення",
        "source": "Джерело",
        "status_active": "ще триває",
        "status_unknown": "—",
    },
    "en": {
        "header_active": "🚨 ALERT IN PROGRESS 🚨",
        "header_ended": "🟢 ALERT ENDED",
        "type": "Type",
        "region": "Oblast",
        "location": "Location",
        "location_type": "Location type",
        "coordinates": "Coordinates",
        "severity": "Severity",
        "cause": "Cause",
        "details": "Details",
        "started": "Start",
        "ended": "End",
        "duration": "Duration",
        "message": "Message",
        "source": "Source",
        "status_active": "still active",
        "status_unknown": "—",
    },
    "de": {
        "header_active": "🚨 ALARM AKTIV 🚨",
        "header_ended": "🟢 ALARM BEENDET",
        "type": "Art",
        "region": "Oblast",
        "location": "Ort",
        "location_type": "Ortstyp",
        "coordinates": "Koordinaten",
        "severity": "Stufe",
        "cause": "Ursache",
        "details": "Details",
        "started": "Beginn",
        "ended": "Ende",
        "duration": "Dauer",
        "message": "Meldung",
        "source": "Quelle",
        "status_active": "läuft noch",
        "status_unknown": "—",
    },
    "pl": {
        "header_active": "🚨 TRWA ALARM 🚨",
        "header_ended": "🟢 ALARM ODWOŁANY",
        "type": "Typ",
        "region": "Obwód",
        "location": "Lokalizacja",
        "location_type": "Typ lokalizacji",
        "coordinates": "Współrzędne",
        "severity": "Poziom",
        "cause": "Przyczyna",
        "details": "Szczegóły",
        "started": "Początek",
        "ended": "Zakończenie",
        "duration": "Czas trwania",
        "message": "Komunikat",
        "source": "Źródło",
        "status_active": "wciąż trwa",
        "status_unknown": "—",
    },
    "ru": {
        "header_active": "🚨 ТРЕВОГА! 🚨",
        "header_ended": "🟢 ОТБОЙ ТРЕВОГИ",
        "type": "Тип",
        "region": "Область",
        "location": "Локация",
        "location_type": "Тип локации",
        "coordinates": "Координаты",
        "severity": "Уровень",
        "cause": "Причина",
        "details": "Детали",
        "started": "Начало",
        "ended": "Отбой",
        "duration": "Длительность",
        "message": "Сообщение",
        "source": "Источник",
        "status_active": "ещё продолжается",
        "status_unknown": "—",
    },
}


def _alerts_ensure_storage(project: Optional[str] = None) -> None:
    alerts_storage_root(project)
    _alerts_migrate_legacy(project)


def _alerts_blank_state() -> Dict[str, Any]:
    return {
        "events": {},
        "regions": {},
        "last_fetch": None,
        ALERTS_TIMELINE_KEY: [],
        "_timeline_day": alerts_today_key(),
        "_context_project": None,
    }


def _alerts_refresh_timeline_day(state: Dict[str, Any], project: Optional[str]) -> None:
    expected_day = alerts_today_key()
    current_day = state.get("_timeline_day")
    if current_day == expected_day:
        return
    context_project = project if project is not None else state.get("_context_project")
    timeline = _alerts_load_timeline(context_project, expected_day)
    state[ALERTS_TIMELINE_KEY] = timeline
    state["_timeline_day"] = expected_day


def _alerts_load_state(project: Optional[str] = None) -> Dict[str, Any]:
    key = _alerts_context_key(project)
    cached = _alerts_state_cache.get(key)
    resolved = _alerts_resolve_project(project)
    if cached is not None:
        _alerts_refresh_timeline_day(cached, resolved)
        return cached
    _alerts_ensure_storage(resolved)
    path = alerts_state_file(resolved)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                payload = json.load(fh)
            if not isinstance(payload, dict):
                raise ValueError("Invalid alerts state")
        except Exception:
            payload = _alerts_blank_state()
    else:
        payload = _alerts_blank_state()
    payload.setdefault("events", {})
    payload.setdefault("regions", {})
    payload.setdefault("last_fetch", None)
    payload.setdefault("_timeline_day", alerts_today_key())
    timeline_day = payload.get("_timeline_day") or alerts_today_key()
    payload[ALERTS_TIMELINE_KEY] = _alerts_load_timeline(resolved, timeline_day)
    _alerts_refresh_timeline_day(payload, resolved)
    payload["_context_project"] = resolved
    _alerts_state_cache[key] = payload
    return payload


def _alerts_save_state(project: Optional[str] = None) -> None:
    def _write(target_key: str, payload: Dict[str, Any]) -> None:
        if payload is None:
            return
        resolved = None if target_key == "__global__" else target_key
        _alerts_refresh_timeline_day(payload, resolved)
        _alerts_ensure_storage(resolved)
        _alerts_write_state_payload(resolved, payload)

    if project is None:
        for key, payload in list(_alerts_state_cache.items()):
            if payload is not None:
                _write(key, payload)
        return
    key = _alerts_context_key(project)
    payload = _alerts_state_cache.get(key)
    if payload is None:
        return
    _write(key, payload)


def _alerts_blank_user_state() -> Dict[str, Any]:
    return {}


def _alerts_load_users(project: Optional[str] = None) -> Dict[str, Any]:
    key = _alerts_context_key(project)
    cached = _alerts_user_cache.get(key)
    if cached is not None:
        return cached
    resolved = _alerts_resolve_project(project)
    _alerts_ensure_storage(resolved)
    path = alerts_users_file(resolved)
    if not os.path.exists(path):
        payload = _alerts_blank_user_state()
        _alerts_user_cache[key] = payload
        _alerts_save_users(project)
        return payload
    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        if not isinstance(payload, dict):
            raise ValueError("Invalid alerts user state")
    except Exception:
        payload = _alerts_blank_user_state()
    _alerts_user_cache[key] = payload
    return payload


def _alerts_save_users(project: Optional[str] = None) -> None:
    def _write(target_key: str, payload: Dict[str, Any]) -> None:
        resolved = None if target_key == "__global__" else target_key
        _alerts_ensure_storage(resolved)
        path = alerts_users_file(resolved)
        tmp_file = f"{path}.tmp"
        with open(tmp_file, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2)
        os.replace(tmp_file, path)

    if project is None:
        for key, payload in list(_alerts_user_cache.items()):
            if payload is not None:
                _write(key, payload)
        return
    key = _alerts_context_key(project)
    payload = _alerts_user_cache.get(key)
    if payload is None:
        return
    _write(key, payload)


def _alerts_user_entry(uid: int, project: Optional[str] = None) -> Dict[str, Any]:
    store = _alerts_load_users(project)
    key = str(uid)
    created = key not in store
    entry = store.setdefault(key, {"regions": [], "last_seen": {}})
    if not isinstance(entry.get("regions"), list):
        entry["regions"] = []
    if not isinstance(entry.get("last_seen"), dict):
        entry["last_seen"] = {}
    if created:
        _alerts_save_users(project)
    return entry


def _alerts_region_state(region: str) -> Dict[str, Any]:
    state = _alerts_load_state()
    bucket = state.setdefault("regions", {}).setdefault(region, {})
    bucket.setdefault("active", [])
    bucket.setdefault("history", [])
    return bucket


def alerts_canonical_region(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    cleaned = str(name).strip()
    if not cleaned:
        return None
    lower = cleaned.lower()
    for canonical, aliases in ALERTS_REGION_EQUIVALENTS.items():
        if lower == canonical.lower():
            return canonical
        for alias in aliases:
            if lower == alias.lower():
                return canonical
    return cleaned


def alerts_region_storage_keys(*candidates: Optional[str]) -> Set[str]:
    keys: Set[str] = set()
    for candidate in candidates:
        if not candidate:
            continue
        cleaned = str(candidate).strip()
        if not cleaned:
            continue
        keys.add(cleaned)
        canonical = alerts_canonical_region(cleaned)
        if canonical:
            keys.add(canonical)
    return keys


def alerts_sanitize_notes(notes: Any) -> List[Dict[str, str]]:
    sanitized: List[Dict[str, str]] = []
    if isinstance(notes, list):
        for entry in notes:
            if isinstance(entry, dict):
                note_type = str(entry.get("type") or "").strip()
                title = str(entry.get("title") or "").strip()
                text = str(entry.get("text") or entry.get("value") or entry.get("note") or "").strip()
                cleaned: Dict[str, str] = {}
                if note_type:
                    cleaned["type"] = note_type
                if title:
                    cleaned["title"] = title
                if text:
                    cleaned["text"] = text
                if cleaned:
                    sanitized.append(cleaned)
            elif isinstance(entry, str):
                text = entry.strip()
                if text:
                    sanitized.append({"text": text})
    return sanitized


def alerts_extract_note_fields(payload: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    severity_raw = ""
    cause = ""
    details_parts: List[str] = []
    message = ""
    source = ""

    for note in alerts_sanitize_notes(payload.get("notes")):
        text = str(note.get("text") or "").strip()
        if not text:
            continue
        note_type = str(note.get("type") or "").lower()
        title_lower = str(note.get("title") or "").lower()

        def matches(keyword: str) -> bool:
            return keyword in note_type or keyword in title_lower

        if not severity_raw and (matches("severity") or matches("level") or matches("рів") or matches("статус")):
            severity_raw = text
            continue
        if not cause and (matches("cause") or matches("reason") or matches("прич")):
            cause = text
            continue
        if not source and (matches("source") or matches("issuer") or matches("джерел")):
            source = text
            continue
        if not message and (matches("message") or matches("status") or matches("опис") or matches("повідом") or note_type in ("message", "status")):
            message = text
            continue
        if matches("detail") or matches("детал") or matches("info") or note_type in ("details", "description"):
            details_parts.append(text)
            continue
        if not message:
            message = text
        else:
            details_parts.append(text)

    details = " • ".join(part for part in details_parts if part)
    return severity_raw, cause, details, message, source


def alerts_normalize_type_code(raw_type: str) -> str:
    base = (raw_type or "").strip().lower()
    if not base:
        return "unknown"
    mapped = ALERTS_TYPE_ALIASES.get(base, base)
    if mapped not in ALERTS_TYPE_LABELS:
        return "unknown"
    return mapped


def alerts_normalize_severity(raw_severity: Optional[str], type_code: str) -> str:
    candidate = str(raw_severity or "").strip()
    lowered = candidate.lower()
    if lowered:
        for keyword, mapped in ALERTS_SEVERITY_KEYWORDS.items():
            if keyword in lowered:
                return mapped
        numeric_map = {"4": "critical", "3": "high", "2": "medium", "1": "low", "0": "low"}
        roman_map = {"iv": "critical", "iii": "high", "ii": "medium", "i": "low"}
        if lowered in numeric_map:
            return numeric_map[lowered]
        if lowered in roman_map:
            return roman_map[lowered]
    return ""


def alerts_normalize_event(raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not isinstance(raw, dict):
        return None
    payload = dict(raw)
    event_id = payload.get("id") or payload.get("alert_id") or ""
    if not event_id:
        region_hint = str(
            payload.get("location_oblast")
            or payload.get("region")
            or payload.get("area")
            or payload.get("location_title")
            or ""
        ).strip()
        started_hint = str(payload.get("started_at") or payload.get("start") or payload.get("timestamp") or "")
        event_id = f"{region_hint}|{payload.get('alert_type') or payload.get('type') or 'alert'}|{started_hint}"
    event_id = str(event_id)

    raw_type = str(payload.get("alert_type") or payload.get("type") or "").strip()
    type_code = alerts_normalize_type_code(raw_type)

    oblast_title = str(payload.get("location_oblast") or payload.get("region") or payload.get("area") or "").strip()
    location_title = str(payload.get("location_title") or payload.get("location") or payload.get("city") or "").strip()
    if not oblast_title and location_title:
        oblast_title = location_title
    region_original = oblast_title or location_title or str(payload.get("region") or payload.get("area") or "").strip()
    region_canonical = alerts_canonical_region(region_original)

    started_at = payload.get("started_at") or payload.get("start") or payload.get("timestamp") or ""
    ended_at = payload.get("finished_at") or payload.get("ended_at") or payload.get("end") or ""
    updated_at = payload.get("updated_at") or payload.get("last_updated_at") or datetime.now(timezone.utc).isoformat()

    severity_note, cause, details, message_note, source_note = alerts_extract_note_fields(payload)
    message = str(payload.get("message") or payload.get("text") or message_note or cause or details or "").strip()
    source = str(payload.get("source") or payload.get("issuer") or source_note or "alerts.in.ua").strip()

    severity_code = alerts_normalize_severity(payload.get("severity") or severity_note, type_code)

    notes_clean = alerts_sanitize_notes(payload.get("notes"))
    extra_payload = {
        "severity": severity_code,
        "cause": cause,
        "details": details,
        "severity_note": severity_note,
        "type_raw": raw_type,
        "oblast_uid": payload.get("location_oblast_uid") or payload.get("oblast_uid"),
        "oblast_title": oblast_title or region_original,
        "notes": notes_clean,
    }

    clean_extra: Dict[str, Any] = {}
    for key, value in extra_payload.items():
        if key == "notes":
            if value:
                clean_extra[key] = value
            continue
        if isinstance(value, str):
            value = value.strip()
        if value in (None, "", []):
            continue
        clean_extra[key] = value
    if severity_code:
        clean_extra.setdefault("severity", severity_code)

    return {
        "id": event_id,
        "type": type_code or "unknown",
        "type_raw": raw_type or "unknown",
        "region": region_canonical or region_original,
        "region_display": region_original or region_canonical or "",
        "started_at": str(started_at) if started_at else "",
        "ended_at": str(ended_at) if ended_at else "",
        "message": message,
        "source": source or "alerts.in.ua",
        "extra": clean_extra,
        "updated_at": str(updated_at) if updated_at else datetime.now(timezone.utc).isoformat(),
    }


def _alerts_user_agent() -> str:
    base = "Bot.BSG-alerts/1.0 (+https://alerts.in.ua)"
    token = ALERTS_API_TOKEN or ""
    if len(token) >= 5:
        return f"{base} token:{token[:5]}"
    return base


def _alerts_request_headers() -> Dict[str, str]:
    if not ALERTS_API_TOKEN:
        return {}
    return {
        "Authorization": f"Bearer {ALERTS_API_TOKEN}",
        "X-API-Key": ALERTS_API_TOKEN,
        "Accept": "application/json",
        "Accept-Language": "uk-UA",
        "User-Agent": _alerts_user_agent(),
    }


def _alerts_api_get(endpoint: str, params: Optional[Dict[str, Any]] = None) -> Tuple[bool, str, Any]:
    if not ALERTS_API_TOKEN:
        return False, "API token is empty", None
    endpoint = endpoint if endpoint.startswith("/") else f"/{endpoint}"
    url = f"{ALERTS_API_BASE_URL}{endpoint}"
    headers = _alerts_request_headers()
    try:
        response = requests.get(url, headers=headers, params=params, timeout=ALERTS_API_TIMEOUT)
    except requests.RequestException as exc:
        return False, str(exc), None
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        return False, str(exc), None
    if response.status_code == 204:
        return True, "", {}
    try:
        data = response.json()
    except ValueError:
        return False, "Некорректный ответ тревог", None
    return True, "", data


def alerts_fetch_remote() -> Tuple[bool, str, List[Dict[str, Any]]]:
    ok, error, data = _alerts_api_get(ALERTS_API_ACTIVE_ENDPOINT)
    if not ok:
        return False, f"Запрос тревог не удался: {error}", []
    if not isinstance(data, dict):
        return True, "", []
    items = data.get("alerts")
    if not isinstance(items, list):
        items = data.get("data") if isinstance(data.get("data"), list) else []
    events: List[Dict[str, Any]] = []
    for item in items:
        normalized = alerts_normalize_event(item)
        if normalized:
            events.append(normalized)
    return True, "", events


def alerts_fetch_history_by_oblast(oblast_uid: Union[int, str], period: str = ALERTS_DEFAULT_HISTORY_PERIOD) -> Tuple[bool, str, List[Dict[str, Any]]]:
    if not oblast_uid:
        return False, "Порожній ідентифікатор області", []
    endpoint = ALERTS_API_HISTORY_TEMPLATE.format(uid=oblast_uid, period=period or ALERTS_DEFAULT_HISTORY_PERIOD)
    ok, error, data = _alerts_api_get(endpoint)
    if not ok:
        return False, error, []
    if isinstance(data, dict):
        items = data.get("alerts") or data.get("data") or []
    else:
        items = data if isinstance(data, list) else []
    events: List[Dict[str, Any]] = []
    for item in items:
        normalized = alerts_normalize_event(item)
        if normalized:
            events.append(normalized)
    return True, "", events


def alerts_history_events(oblast_uid: Union[int, str]) -> Tuple[bool, str, List[Dict[str, Any]]]:
    key = str(oblast_uid)
    now = datetime.now(timezone.utc)
    cached = alerts_history_cache.get(key)
    if cached:
        fetched_at = cached.get("fetched_at")
        if isinstance(fetched_at, datetime) and (now - fetched_at).total_seconds() < ALERTS_HISTORY_CACHE_TTL:
            return True, "", cached.get("events", [])
    ok, error, events = alerts_fetch_history_by_oblast(oblast_uid)
    if ok:
        alerts_history_cache[key] = {"fetched_at": now, "events": events}
    return ok, error, events


def alerts_merge_extra(base: Optional[Dict[str, Any]], update: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    merged: Dict[str, Any] = dict(base or {})
    for key, value in (update or {}).items():
        if key == "notes":
            if value:
                merged[key] = value
            continue
        if isinstance(value, str):
            value = value.strip()
        if value in (None, "", []):
            continue
        merged[key] = value
    return merged


def alerts_enrich_from_history(event: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    extra = event.get("extra") or {}
    oblast_uid = extra.get("oblast_uid")
    if not oblast_uid:
        return None
    ok, error, history_events = alerts_history_events(oblast_uid)
    if not ok:
        if error:
            print(f"[alerts] history fetch failed for oblast {oblast_uid}: {error}")
        return None
    event_id = str(event.get("id"))
    for hist_event in history_events:
        if str(hist_event.get("id")) != event_id:
            continue
        merged_extra = alerts_merge_extra(event.get("extra"), hist_event.get("extra"))
        return {
            "ended_at": hist_event.get("ended_at") or event.get("ended_at"),
            "message": hist_event.get("message") or event.get("message"),
            "source": hist_event.get("source") or event.get("source"),
            "extra": merged_extra,
            "region": hist_event.get("region") or event.get("region"),
            "region_display": hist_event.get("region_display") or event.get("region_display"),
            "started_at": hist_event.get("started_at") or event.get("started_at"),
            "updated_at": hist_event.get("updated_at") or event.get("updated_at"),
        }
    return None


def alerts_refresh_once() -> Tuple[List[str], List[str]]:
    ok, error, events = alerts_fetch_remote()
    if not ok:
        print(f"[alerts] {error}")
        return [], []
    state = _alerts_load_state()
    events_map = state.setdefault("events", {})
    regions_map = state.setdefault("regions", {})

    previous_active_ids: Set[str] = {str(eid) for eid, payload in events_map.items() if not payload.get("ended_at")}
    start_notify: List[str] = []
    end_notify: List[str] = []
    seen_ids: Set[str] = set()

    for event in events:
        event_id = str(event["id"])
        seen_ids.add(event_id)
        stored = events_map.get(event_id)
        ended_now = bool(event.get("ended_at"))
        if stored:
            previously_ended = bool(stored.get("ended_at"))
            merged_extra = alerts_merge_extra(stored.get("extra"), event.get("extra"))
            stored.update(event)
            if merged_extra:
                stored["extra"] = merged_extra
            if not previously_ended and ended_now:
                stored.setdefault("notified_end", False)
                if event_id not in end_notify:
                    end_notify.append(event_id)
        else:
            event["notified_start"] = bool(ended_now)
            event["notified_end"] = False
            events_map[event_id] = event
            if ended_now:
                if event_id not in end_notify:
                    end_notify.append(event_id)
            else:
                if event_id not in start_notify:
                    start_notify.append(event_id)

        region_keys = alerts_region_storage_keys(event.get("region"), event.get("region_display"))
        if not region_keys:
            region_keys = {""}
        for region_key in region_keys:
            bucket = regions_map.setdefault(region_key, {"active": [], "history": []})
            history = bucket.setdefault("history", [])
            active = bucket.setdefault("active", [])
            if event_id not in history:
                history.insert(0, event_id)
            if ended_now:
                if event_id in active:
                    active.remove(event_id)
            else:
                if event_id not in active:
                    active.append(event_id)

    missing_active = previous_active_ids - seen_ids
    if missing_active:
        now_iso = datetime.now(timezone.utc).isoformat()
        for event_id in list(missing_active):
            stored = events_map.get(event_id)
            if not stored or stored.get("ended_at"):
                continue
            enriched = alerts_enrich_from_history(stored)
            if enriched:
                merged_extra = alerts_merge_extra(stored.get("extra"), enriched.get("extra"))
                stored.update({k: v for k, v in enriched.items() if k != "extra"})
                if merged_extra:
                    stored["extra"] = merged_extra
            else:
                stored["ended_at"] = now_iso
            stored.setdefault("notified_end", False)
            if event_id not in end_notify:
                end_notify.append(event_id)
            region_keys = alerts_region_storage_keys(stored.get("region"), stored.get("region_display")) or {""}
            for region_key in region_keys:
                bucket = regions_map.setdefault(region_key, {"active": [], "history": []})
                active = bucket.setdefault("active", [])
                if event_id in active:
                    active.remove(event_id)

    for region_key, bucket in regions_map.items():
        active = bucket.get("active", [])
        bucket["active"] = [eid for eid in active if not events_map.get(eid, {}).get("ended_at")]

    alerts_record_timeline(state, start_notify, "start")
    alerts_record_timeline(state, end_notify, "end")
    state["last_fetch"] = datetime.now(timezone.utc).isoformat()
    _alerts_save_state()
    return start_notify, end_notify


def _alerts_get_event(event_id: str) -> Optional[Dict[str, Any]]:
    state = _alerts_load_state()
    payload = state.get("events", {}).get(event_id)
    if payload:
        return dict(payload)
    return None


def _alerts_mark_notified(event_id: str, kind: str) -> None:
    state = _alerts_load_state()
    payload = state.get("events", {}).get(event_id)
    if not payload:
        return
    if kind == "start":
        payload["notified_start"] = True
    elif kind == "end":
        payload["notified_end"] = True
    _alerts_save_state()


def _alerts_timeline_bucket(state: Dict[str, Any]) -> List[Dict[str, Any]]:
    timeline = state.setdefault(ALERTS_TIMELINE_KEY, [])
    if isinstance(timeline, list):
        return timeline
    timeline = []
    state[ALERTS_TIMELINE_KEY] = timeline
    return timeline


def alerts_record_timeline(state: Dict[str, Any], event_ids: List[str], kind: str) -> None:
    if not event_ids:
        return
    _alerts_refresh_timeline_day(state, state.get("_context_project"))
    events_map = state.get("events", {})
    timeline = _alerts_timeline_bucket(state)
    recorded_at = alerts_now().isoformat()
    for event_id in event_ids:
        event = events_map.get(event_id)
        if not event:
            continue
        started_at = event.get("started_at")
        ended_at = event.get("ended_at")
        if kind == "start" and not started_at:
            started_at = recorded_at
            event["started_at"] = started_at
        if kind == "end":
            if not started_at:
                started_at = recorded_at
                event.setdefault("started_at", started_at)
            if not ended_at:
                ended_at = recorded_at
                event["ended_at"] = ended_at
        canonical = alerts_canonical_region(event.get("region") or event.get("region_display"))
        region_value = canonical or event.get("region") or event.get("region_display") or ""
        extra = event.get("extra") or {}
        entry = {
            "event_id": event_id,
            "kind": kind,
            "region": region_value,
            "type": event.get("type") or "",
            "severity": extra.get("severity") or "",
            "started_at": started_at,
            "ended_at": ended_at,
            "cause": extra.get("cause") or "",
            "details": extra.get("details") or "",
            "message": event.get("message") or "",
            "recorded_at": recorded_at,
        }
        timeline.append(entry)


def alerts_parse_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                dt = datetime.strptime(raw, fmt)
                dt = dt.replace(tzinfo=timezone.utc)
                break
            except ValueError:
                continue
        else:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    try:
        return dt.astimezone(ALERTS_TIMEZONE)
    except Exception:
        return dt


def alerts_now() -> datetime:
    try:
        tz = ALERTS_TIMEZONE  # type: ignore[name-defined]
    except Exception:
        tz = timezone.utc
    return datetime.now(tz)


def alerts_today_key() -> str:
    return alerts_now().strftime("%Y-%m-%d")


def alerts_format_timestamp(value: Optional[str]) -> str:
    dt = alerts_parse_datetime(value)
    if not dt:
        return value or ""
    return dt.strftime("%Y-%m-%d %H:%M")


def alerts_format_datetime_display(value: Optional[str]) -> str:
    date, clock = alerts_format_push_date_pair(value)
    if date and clock:
        return f"{date} • {clock}"
    if date:
        return date
    if clock:
        return clock
    return ""


def alerts_format_clock(value: Optional[str]) -> str:
    dt = alerts_parse_datetime(value)
    if not dt:
        return ""
    return dt.strftime("%H:%M")


def alerts_type_icon(event: Dict[str, Any]) -> str:
    type_code = event.get("type") or ""
    return ALERTS_TYPE_ICONS.get(type_code, ALERTS_TYPE_ICONS["unknown"])


def alerts_select_form(value: int, forms: Tuple[str, ...]) -> str:
    if len(forms) == 2:
        return forms[0] if abs(value) == 1 else forms[1]
    n = abs(value) % 100
    if 11 <= n <= 14:
        return forms[2]
    n = abs(value) % 10
    if n == 1:
        return forms[0]
    if 2 <= n <= 4:
        return forms[1]
    return forms[2]


def alerts_unit_form(lang: str, unit: str, value: int) -> str:
    forms_map = ALERTS_DURATION_FORMS.get(lang) or ALERTS_DURATION_FORMS.get(DEFAULT_LANG) or {}
    forms = forms_map.get(unit)
    if not forms:
        # fallback to english plural rules
        forms = (unit, f"{unit}s")
    return alerts_select_form(value, forms)


def alerts_format_duration_value(seconds: int, lang: str) -> str:
    if seconds <= 0:
        return tr(lang, "ALERTS_DURATION_LESS_MINUTE")
    total_minutes = max(1, seconds // 60)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    parts: List[str] = []
    if hours:
        unit = alerts_unit_form(lang, "hour", hours)
        parts.append(f"{hours} {unit}")
    if minutes:
        unit = alerts_unit_form(lang, "minute", minutes)
        parts.append(f"{minutes} {unit}")
    if not parts:
        unit = alerts_unit_form(lang, "minute", 1)
        parts.append(f"1 {unit}")
    return " ".join(parts)


def alerts_duration_seconds(start: Optional[str], end: Optional[str] = None) -> int:
    start_dt = alerts_parse_datetime(start)
    if not start_dt:
        return 0
    end_dt = alerts_parse_datetime(end) if end else alerts_now()
    if not end_dt:
        end_dt = alerts_now()
    delta = end_dt - start_dt
    return max(0, int(delta.total_seconds()))


def alerts_duration_phrase(start: Optional[str], end: Optional[str], lang: str, ongoing: bool) -> str:
    seconds = alerts_duration_seconds(start, end)
    value = alerts_format_duration_value(seconds, lang)
    key = "ALERTS_DURATION_RUNNING" if ongoing else "ALERTS_DURATION_COMPLETED"
    return tr(lang, key, duration=value)


def alerts_format_push_timestamp(value: Optional[str]) -> str:
    dt = alerts_parse_datetime(value)
    if not dt:
        return ""
    return dt.strftime("%d.%m.%Y %H:%M")


def alerts_format_push_date_pair(value: Optional[str]) -> Tuple[str, str]:
    dt = alerts_parse_datetime(value)
    if not dt:
        return "", ""
    return dt.strftime("%d.%m.%Y"), dt.strftime("%H:%M")


def alerts_country_count_label(count: int, lang: str) -> str:
    unit = alerts_unit_form(lang, "alarm", count)
    return f"{count} {unit}"


def alerts_region_active_value(count: int, lang: str) -> str:
    if count <= 0:
        return ""
    if lang == "uk":
        return f"активна {count}" if count == 1 else f"активні {count}"
    if lang == "ru":
        return f"активна {count}" if count == 1 else f"активны {count}"
    if lang == "pl":
        return f"aktywna {count}" if count == 1 else f"aktywne {count}"
    if lang == "de":
        return f"aktiv {count}"
    return f"active {count}"


def alerts_recommendation_block(event: Dict[str, Any], lang: str) -> str:
    mapping = ALERTS_RECOMMENDATIONS.get(event.get("type"))
    if not mapping:
        mapping = ALERTS_RECOMMENDATIONS.get("default", {})
    lines = mapping.get(lang) or mapping.get(DEFAULT_LANG) or []
    return "\n".join(lines)


def alerts_type_label(event: Dict[str, Any], lang: str) -> str:
    mapping = ALERTS_TYPE_LABELS.get(event.get("type")) or ALERTS_TYPE_LABELS.get("unknown")
    return mapping.get(lang) or mapping.get(DEFAULT_LANG) or event.get("type") or "Alert"


def alerts_severity_label(event: Dict[str, Any], lang: str) -> str:
    severity = (event.get("extra") or {}).get("severity") or ""
    if not severity:
        return ""
    mapping = ALERTS_SEVERITY_LABELS.get(severity)
    if not mapping:
        return severity.capitalize()
    icon = mapping.get("icon", "")
    text = mapping.get(lang) or mapping.get(DEFAULT_LANG) or severity
    return f"{icon} {text}" if icon else text


def alerts_field_labels(lang: str) -> Dict[str, str]:
    return ALERTS_FIELD_LABELS.get(lang) or ALERTS_FIELD_LABELS[DEFAULT_LANG]


def alerts_format_row(icon: str, label: str, value: str) -> List[str]:
    if not value:
        return []
    text = str(value)
    prefix = f"{icon} {label}: "
    indent = " " * len(prefix)
    rows = [prefix + text.splitlines()[0]]
    for part in text.splitlines()[1:]:
        rows.append(indent + part)
    return rows


def alerts_format_card(event: Dict[str, Any], lang: str, index: Optional[int] = None, total: Optional[int] = None) -> str:
    labels = alerts_field_labels(lang)
    status_labels = ALERTS_STATUS_TEXT.get(lang) or ALERTS_STATUS_TEXT[DEFAULT_LANG]
    ended = bool(event.get("ended_at"))
    region_display = alerts_display_region_name(event.get("region_display") or event.get("region") or "", lang)
    status_label = status_labels["standdown" if ended else "alert"].upper()
    header_icon = "🟢" if ended else "🚨"
    lines: List[str] = [f"{header_icon} {status_label} — {region_display}", "━━━━━━━━━━━━━━━━━━"]
    type_label = alerts_type_label(event, lang)
    type_icon = alerts_type_icon(event)
    lines.extend(alerts_format_row(type_icon, labels["type"], type_label))
    severity_value = alerts_severity_label(event, lang)
    if severity_value:
        lines.extend(alerts_format_row("⚠️", labels["severity"], severity_value))
    extra = event.get("extra") or {}
    cause = extra.get("cause") or ""
    lines.extend(alerts_format_row("🎯", labels["cause"], cause))
    details = extra.get("details") or ""
    lines.extend(alerts_format_row("🔎", labels["details"], details))
    started_display = alerts_format_datetime_display(event.get("started_at")) or labels["status_unknown"]
    lines.extend(alerts_format_row("🕒", labels["started"], started_display))
    if ended:
        end_display = alerts_format_datetime_display(event.get("ended_at")) or labels["status_unknown"]
    else:
        end_display = labels["status_active"]
    lines.extend(alerts_format_row("✅", labels["ended"], end_display))
    duration_seconds = alerts_duration_seconds(event.get("started_at"), event.get("ended_at") if ended else None)
    duration_value = alerts_format_duration_value(duration_seconds, lang)
    lines.extend(alerts_format_row("⏱️", labels["duration"], duration_value))
    lines.extend(alerts_format_row("📢", labels["message"], event.get("message") or ""))
    if index is not None and total:
        lines.append("━━━━━━━━━━━━━━━━━━")
        lines.append(tr(lang, "ALERTS_CARD_INDEX", index=index + 1, total=total))
    return "\n".join(line for line in lines if line)


def alerts_profile_block(profile: dict) -> dict:
    uid = profile.get("user_id")
    if not uid:
        return {"regions": [], "last_seen": {}}
    entry = _alerts_user_entry(uid)
    legacy = profile.get("alerts")
    migrated = False
    if isinstance(legacy, dict):
        legacy_regions = legacy.get("regions", [])
        if isinstance(legacy_regions, list):
            for region in legacy_regions:
                canonical = alerts_canonical_region(region) or region
                if canonical and canonical not in entry["regions"]:
                    entry["regions"].append(canonical)
                    migrated = True
        legacy_seen = legacy.get("last_seen")
        if isinstance(legacy_seen, dict):
            entry["last_seen"].update({str(k): v for k, v in legacy_seen.items()})
            migrated = True
        if migrated:
            _alerts_save_users()
        profile.pop("alerts", None)
        try:
            save_user(profile)
        except Exception:
            pass
    return entry


def alerts_user_regions(uid: int) -> List[str]:
    regions: List[str] = []
    if active_project.get("name"):
        info = load_project_info(active_project["name"])
        project_region = info.get("region")
        canonical = alerts_canonical_region(project_region)
        if canonical:
            regions.append(canonical)
        elif project_region:
            regions.append(project_region)
    profile = load_user(uid) or {"user_id": uid}
    alerts = alerts_profile_block(profile)
    for region in alerts.get("regions", []):
        canonical = alerts_canonical_region(region)
        if canonical and canonical not in regions:
            regions.append(canonical)
        elif region not in regions:
            regions.append(region)
    return regions


def alerts_shorten_region_label(name: str, lang: str) -> str:
    text = str(name or "").strip()
    lowered = text.lower()
    suffix_map = [
        ("область", "обл."),
        ("oblast", "obl."),
        ("region", "reg."),
        ("obwód", "obw."),
        ("obwod", "obw."),
    ]
    for suffix, replacement in suffix_map:
        if lowered.endswith(suffix):
            base = text[: -len(suffix)].rstrip(" -")
            if not base:
                return text
            return f"{base} {replacement}".strip()
    return text


def alerts_display_region_name(region: str, lang: str, short: bool = False) -> str:
    canonical = alerts_canonical_region(region) or region
    aliases = ALERTS_REGION_EQUIVALENTS.get(canonical)
    if not aliases:
        result = canonical
    elif lang == "ru":
        result = canonical
    elif lang == "en":
        for alias in aliases:
            if re.search(r"[A-Za-z]", alias):
                result = alias
                break
        else:
            result = aliases[-1]
    else:
        result = aliases[0]
    if short:
        return alerts_shorten_region_label(result, lang)
    return result


def alerts_trim_region_suffix(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return ""
    lowered = text.lower()
    suffixes = (
        " область",
        "область",
        " обл.",
        " oblast",
        " region",
        " obwód",
        " obwod",
        " obw.",
        " reg.",
    )
    for suffix in suffixes:
        suffix_clean = suffix.strip()
        if not suffix_clean:
            continue
        if lowered.endswith(suffix_clean):
            candidate = text[: -len(suffix_clean)].rstrip(" -–—,.;:")
            if candidate:
                return candidate
    return text


def alerts_overview_region_label(region: str, lang: str) -> str:
    canonical = alerts_canonical_region(region) or region
    mapping = ALERTS_REGION_SHORT_NAMES.get(canonical)
    if mapping:
        label = mapping.get(lang) or mapping.get(DEFAULT_LANG)
        if label:
            return label
    display = alerts_display_region_name(canonical, lang)
    trimmed = alerts_trim_region_suffix(display)
    return trimmed or display


def alerts_display_width(text: str) -> int:
    """Estimate visual column width for aligning overview rows."""
    if not text:
        return 0
    width = 0
    for char in str(text):
        if unicodedata.combining(char):
            continue
        if char in ("\u00A0", "\u202F", "\u2007"):
            width += 1
            continue
        east_asian = unicodedata.east_asian_width(char)
        if east_asian in ("F", "W"):
            width += 2
        else:
            width += 1
    return width


def alerts_regions_overview_text(uid: int) -> str:
    lang = resolve_lang(uid)
    status_labels = ALERTS_OVERVIEW_STATUS_TEXT.get(lang) or {}
    default_status_labels = ALERTS_OVERVIEW_STATUS_TEXT.get(DEFAULT_LANG) or {}

    def get_status_label(key: str) -> str:
        return status_labels.get(key) or default_status_labels.get(key) or ""

    header = tr(uid, "ALERTS_OVERVIEW_HEADER")
    entries: List[Dict[str, Any]] = []
    max_name_width = 0
    for index, raw_region in enumerate(UKRAINE_REGIONS, start=1):
        canonical, active_event, last_event = alerts_region_snapshot(raw_region)
        display_name = alerts_overview_region_label(canonical, lang)
        name_width = alerts_display_width(display_name)
        max_name_width = max(max_name_width, name_width)
        if active_event:
            status_text = get_status_label("alert")
            time_text = alerts_format_clock(active_event.get("started_at")) or "--:--"
            icon = "🔴"
        else:
            icon = "🟢"
            status_text = get_status_label("standdown") or get_status_label("calm")
            end_clock = ""
            if last_event and last_event.get("ended_at"):
                end_clock = alerts_format_clock(last_event.get("ended_at"))
            time_text = end_clock or "--:--"
        entries.append(
            {
                "index": index,
                "icon": icon,
                "name": display_name,
                "name_width": name_width,
                "status": status_text,
                "time": time_text,
            }
        )

    lines: List[str] = [h(header), ""]
    for entry in entries:
        stored_width = entry.get("name_width")
        if stored_width is None:
            stored_width = alerts_display_width(entry["name"])
        name_padding = max_name_width - stored_width
        padded_name = f"{entry['name']}{' ' * max(name_padding, 0)}"
        number = f"{entry['index']:2d}"
        status_text = h(entry["status"])
        time_text = h(entry["time"])
        lines.append(
            f"{number}. {entry['icon']} {h(padded_name)} — {status_text} • {time_text}"
        )
        lines.append("")

    while len(lines) > 2 and lines[-1] == "":
        lines.pop()

    updated_clock = alerts_now().strftime("%H:%M")
    lines.append("")
    lines.append("━━━━━━━━━━━━━━━━━━")
    updated_line = tr(uid, "ALERTS_OVERVIEW_UPDATED").format(time=updated_clock)
    lines.append(h(updated_line))
    lines.append("")
    lines.append(h(tr(uid, "ALERTS_OVERVIEW_GUIDE")))
    body = "\n".join(lines)
    return f"<pre>{body}</pre>"


def alerts_collect_active_for_user(uid: int) -> List[Dict[str, Any]]:
    state = _alerts_load_state()
    events_map = state.get("events", {})
    lang = resolve_lang(uid)
    aggregated: Dict[str, Dict[str, Any]] = {}
    for region in alerts_user_regions(uid):
        bucket = state.get("regions", {}).get(region) or {}
        for event_id in bucket.get("active", []):
            event = events_map.get(event_id)
            if not event or event.get("ended_at"):
                continue
            canonical = alerts_canonical_region(event.get("region") or event.get("region_display") or region) or region
            stored = aggregated.get(canonical)
            started_at = event.get("started_at") or ""
            if not stored or (started_at > (stored.get("started_at") or "")):
                copy = dict(event)
                copy["region"] = canonical
                copy["region_display"] = alerts_display_region_name(canonical, lang)
                aggregated[canonical] = copy
    events = list(aggregated.values())
    events.sort(key=lambda item: alerts_display_region_name(item.get("region") or item.get("region_display") or "", lang))
    return events


def alerts_collect_history_for_user(uid: int, limit: int = 40) -> List[Dict[str, Any]]:
    state = _alerts_load_state()
    events_map = state.get("events", {})
    regions_selected = {alerts_canonical_region(r) or r for r in alerts_user_regions(uid)}
    seen: Set[str] = set()
    collected: List[Dict[str, Any]] = []
    timeline = list(_alerts_timeline_bucket(state))
    if timeline:
        for entry in reversed(timeline):
            event_id = str(entry.get("event_id") or "")
            if not event_id or event_id in seen:
                continue
            region_value = alerts_canonical_region(entry.get("region")) or entry.get("region") or ""
            if regions_selected and region_value and region_value not in regions_selected:
                continue
            event = events_map.get(event_id)
            if not event:
                continue
            copy = dict(event)
            if region_value:
                copy.setdefault("region", region_value)
            collected.append(copy)
            seen.add(event_id)
            if len(collected) >= limit:
                break
    if not collected:
        fallback_regions = alerts_user_regions(uid)
        for region in fallback_regions:
            canonical = alerts_canonical_region(region) or region
            bucket = state.get("regions", {}).get(canonical) or state.get("regions", {}).get(region) or {}
            for event_id in bucket.get("history", []):
                if event_id in seen:
                    continue
                event = events_map.get(event_id)
                if event:
                    collected.append(dict(event))
                    seen.add(event_id)
                if len(collected) >= limit:
                    break
            if len(collected) >= limit:
                break
    collected.sort(key=lambda item: item.get("started_at") or "", reverse=True)
    return collected[:limit]


def alerts_subscription_view(uid: int, page: int = 0) -> Tuple[str, InlineKeyboardMarkup]:
    profile = load_user(uid) or {}
    alerts = alerts_profile_block(profile)
    project_region = None
    if active_project.get("name"):
        info = load_project_info(active_project["name"])
        project_region = info.get("region") or ""
    canonical_project = alerts_canonical_region(project_region)
    selected = alerts_user_regions(uid)
    lang = resolve_lang(uid)
    lines = [tr(uid, "ALERTS_SUBS_HEADER"), tr(uid, "ALERTS_SUBS_DIVIDER")]
    if canonical_project:
        project_label = alerts_display_region_name(canonical_project, lang, short=False)
        lines.append(tr(uid, "ALERTS_SUBS_NOTE_HAS_PROJECT", region=h(project_label)))
    else:
        lines.append(tr(uid, "ALERTS_SUBS_NOTE_NO_PROJECT"))
    lines.append("")
    lines.append(tr(uid, "ALERTS_SUBS_LIST_TITLE"))
    if selected:
        labels = []
        for name in selected:
            full_label = alerts_display_region_name(name, lang, short=False)
            trimmed = alerts_trim_region_suffix(full_label)
            labels.append(h(trimmed or full_label))
        chunk_size = 4
        for idx in range(0, len(labels), chunk_size):
            lines.append(" • ".join(labels[idx:idx + chunk_size]))
    else:
        lines.append(tr(uid, "ALERTS_SUBS_LIST_EMPTY"))
    lines.append("")
    lines.append(tr(uid, "ALERTS_SUBS_DIVIDER"))
    lines.append(tr(uid, "ALERTS_SUBS_MANAGE"))
    kb = alerts_build_subscription_keyboard(uid, page, canonical_project, alerts)
    return "\n".join(lines), kb


def alerts_build_subscription_keyboard(uid: int, page: int, project_region: Optional[str], alerts: dict) -> InlineKeyboardMarkup:
    per_page = 6
    total = len(UKRAINE_REGIONS)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(0, min(page, total_pages - 1))
    start = page * per_page
    chunk = UKRAINE_REGIONS[start:start + per_page]
    selected = {alerts_canonical_region(x) or x for x in alerts.get("regions", [])}
    kb = InlineKeyboardMarkup(row_width=2)
    lang = resolve_lang(uid)
    for idx, region in enumerate(chunk):
        canonical = alerts_canonical_region(region) or region
        label_text = alerts_display_region_name(canonical, lang, short=True)
        if project_region and canonical == alerts_canonical_region(project_region):
            label = f"🔒 {label_text}"
            callback = "alerts_locked"
        else:
            is_selected = canonical in selected
            prefix = "✅" if is_selected else "➕"
            label = f"{prefix} {label_text}"
            callback = f"alerts_toggle:{page}:{start + idx}"
        kb.insert(InlineKeyboardButton(label, callback_data=callback))
    if total_pages > 1:
        nav: List[InlineKeyboardButton] = []
        if page > 0:
            nav.append(InlineKeyboardButton("◀️", callback_data=f"alerts_sub_page:{page - 1}"))
        nav.append(InlineKeyboardButton(tr(uid, "ALERTS_SUBS_PAGE", current=page + 1, total=total_pages), callback_data=f"alerts_sub_page:{page}"))
        if page < total_pages - 1:
            nav.append(InlineKeyboardButton("▶️", callback_data=f"alerts_sub_page:{page + 1}"))
        kb.row(*nav)
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_BACK_TO_MENU"), callback_data="menu_alerts"))
    kb.add(InlineKeyboardButton(tr(uid, "BTN_BACK_ROOT"), callback_data="back_root"))
    return kb


def alerts_update_subscription(uid: int, region_index: int, add: bool) -> bool:
    if region_index < 0 or region_index >= len(UKRAINE_REGIONS):
        return False
    region = alerts_canonical_region(UKRAINE_REGIONS[region_index]) or UKRAINE_REGIONS[region_index]
    profile = load_user(uid) or {"user_id": uid}
    alerts = alerts_profile_block(profile)
    items = alerts.setdefault("regions", [])
    changed = False
    if add:
        if region not in items:
            items.append(region)
            changed = True
    else:
        if region in items:
            items.remove(region)
            changed = True
    if changed:
        _alerts_save_users()
    return changed


def alerts_card_keyboard(uid: int, context: str, total: int, index: int) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    if total > 1:
        row: List[InlineKeyboardButton] = []
        if index > 0:
            row.append(InlineKeyboardButton("◀️", callback_data=f"alerts_nav:{context}:{index - 1}"))
        row.append(
            InlineKeyboardButton(
                tr(uid, "ALERTS_CARD_INDEX", index=index + 1, total=total),
                callback_data=f"alerts_index:{context}:{index}",
            )
        )
        if index < total - 1:
            row.append(InlineKeyboardButton("▶️", callback_data=f"alerts_nav:{context}:{index + 1}"))
        kb.row(*row)
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_CLOSE_CARD"), callback_data=f"alerts_close:{context}"))
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_BACK_TO_MENU"), callback_data="menu_alerts"))
    return kb


def alerts_push_keyboard(uid: int, token: str, expanded: bool) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=1)
    if expanded:
        kb.add(InlineKeyboardButton(tr(uid, "ALERTS_PUSH_COLLAPSE"), callback_data=f"alerts_push:collapse:{token}"))
    else:
        kb.add(InlineKeyboardButton(tr(uid, "ALERTS_PUSH_OPEN"), callback_data=f"alerts_push:expand:{token}"))
    kb.add(InlineKeyboardButton(tr(uid, "ALERTS_PUSH_DELETE"), callback_data=f"alerts_push:delete:{token}"))
    return kb


async def alerts_send_card(uid: int, chat_id: int, events: List[Dict[str, Any]], context: str, index: int = 0) -> Optional[types.Message]:
    if not events:
        return None
    runtime = users_runtime.setdefault(uid, {})
    cards = runtime.setdefault("alerts_cards", {})
    previous = cards.get(context, {}).get("message")
    if isinstance(previous, (list, tuple)) and len(previous) == 2:
        await _delete_message_safe(previous[0], previous[1])
    index = max(0, min(index, len(events) - 1))
    lang = resolve_lang(uid)
    text = alerts_format_card(events[index], lang, index=index, total=len(events))
    kb = alerts_card_keyboard(uid, context, len(events), index)
    msg = await bot.send_message(chat_id, text, reply_markup=kb, disable_web_page_preview=True)
    flow_track(uid, msg)
    cards[context] = {
        "events": [event["id"] for event in events],
        "index": index,
        "message": (msg.chat.id, msg.message_id),
    }
    return msg


def alerts_active_oblast_count() -> int:
    state = _alerts_load_state()
    events_map = state.get("events", {})
    oblasts: Set[str] = set()
    for payload in events_map.values():
        if not isinstance(payload, dict):
            continue
        if payload.get("ended_at"):
            continue
        region_name = payload.get("region") or payload.get("region_display") or ""
        canonical = alerts_canonical_region(region_name)
        normalized = (canonical or region_name or "").strip()
        if not normalized:
            continue
        lower = normalized.lower()
        if any(token in lower for token in ("област", "oblast")):
            oblasts.add(normalized)
    return len(oblasts)


def alerts_active_summary_line(uid: int) -> str:
    count = alerts_active_oblast_count()
    return tr(uid, "ANCHOR_ALERT_SUMMARY", count=count)


def alerts_region_snapshot(region_key: str) -> Tuple[str, Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    state = _alerts_load_state()
    events_map = state.get("events", {})
    regions_map = state.get("regions", {})
    canonical = alerts_canonical_region(region_key) or region_key
    bucket = regions_map.get(canonical) or regions_map.get(region_key) or {}
    active_event: Optional[Dict[str, Any]] = None
    for event_id in bucket.get("active", []):
        payload = events_map.get(event_id)
        if payload and not payload.get("ended_at"):
            if not active_event or (payload.get("started_at") or "") > (active_event.get("started_at") or ""):
                active_event = payload
    last_event: Optional[Dict[str, Any]] = None
    for event_id in bucket.get("history", []):
        payload = events_map.get(event_id)
        if payload:
            if not last_event or (payload.get("ended_at") or "") > (last_event.get("ended_at") or ""):
                last_event = payload
    if not last_event or not last_event.get("ended_at"):
        timeline = list(_alerts_timeline_bucket(state))
        if timeline:
            for entry in reversed(timeline):
                region_value = alerts_canonical_region(entry.get("region")) or entry.get("region") or ""
                if (region_value or "") != canonical:
                    continue
                if entry.get("kind") != "end":
                    continue
                event_id = str(entry.get("event_id") or "")
                payload = dict(events_map.get(event_id) or {})
                if not payload:
                    payload = {"id": event_id or f"timeline|{canonical}|{entry.get('recorded_at') or ''}"}
                payload.setdefault("region", canonical)
                payload.setdefault("region_display", entry.get("region") or canonical)
                if entry.get("started_at") and not payload.get("started_at"):
                    payload["started_at"] = entry["started_at"]
                end_value = entry.get("ended_at") or entry.get("recorded_at")
                if end_value:
                    payload["ended_at"] = end_value
                last_event = payload
                break
    return canonical, active_event, last_event


def alerts_region_active_total(region_key: str) -> int:
    state = _alerts_load_state()
    events_map = state.get("events", {})
    regions_map = state.get("regions", {})
    canonical = alerts_canonical_region(region_key) or region_key
    bucket = regions_map.get(canonical) or {}
    count = 0
    for event_id in bucket.get("active", []):
        payload = events_map.get(event_id)
        if payload and not payload.get("ended_at"):
            count += 1
    return count


def alerts_anchor_region_block(uid: int, region_key: str) -> Optional[str]:
    lang = resolve_lang(uid)
    canonical, active_event, last_event = alerts_region_snapshot(region_key)
    display_region = alerts_display_region_name(canonical, lang)
    status_labels = ALERTS_STATUS_TEXT.get(lang) or ALERTS_STATUS_TEXT[DEFAULT_LANG]
    if active_event:
        type_text = alerts_type_label(active_event, lang)
        severity_text = alerts_severity_label(active_event, lang)
        start_clock = alerts_format_clock(active_event.get("started_at"))
        extra = active_event.get("extra") or {}
        cause_text = extra.get("cause") or ""
        details: List[str] = []
        if type_text:
            details.append(type_text)
        if cause_text:
            details.append(cause_text)
        if severity_text:
            details.append(severity_text)
        if start_clock:
            details.append(start_clock)
        line = f"🔴 <b>{h(display_region)}</b> — {h(status_labels['alert'])}"
        if details:
            line += " • " + " • ".join(h(part) for part in details if part)
        return line
    if last_event and last_event.get("ended_at"):
        ended_dt = alerts_parse_datetime(last_event.get("ended_at"))
        now_dt = alerts_now()
        if ended_dt and (now_dt - ended_dt) <= timedelta(seconds=ALERTS_STANDDOWN_DISPLAY_WINDOW):
            type_text = alerts_type_label(last_event, lang)
            severity_text = alerts_severity_label(last_event, lang)
            start_clock = alerts_format_clock(last_event.get("started_at"))
            end_clock = alerts_format_clock(last_event.get("ended_at"))
            extra = last_event.get("extra") or {}
            cause_text = extra.get("cause") or ""
            details: List[str] = []
            if type_text:
                details.append(type_text)
            if cause_text:
                details.append(cause_text)
            if severity_text:
                details.append(severity_text)
            time_segment = ""
            if start_clock and end_clock:
                time_segment = f"{start_clock} → {end_clock}"
            elif start_clock:
                time_segment = start_clock
            elif end_clock:
                time_segment = end_clock
            if time_segment:
                details.append(time_segment)
            line = f"🟡 <b>{h(display_region)}</b> — {h(status_labels['standdown'])}"
            if details:
                line += " • " + " • ".join(h(part) for part in details if part)
            return line
    return ""


def alerts_anchor_section(uid: int) -> str:
    summary = alerts_active_summary_line(uid)
    regions: List[str] = []
    for region in alerts_user_regions(uid):
        canonical = alerts_canonical_region(region) or region
        if canonical and canonical not in regions:
            regions.append(canonical)
    lines: List[str] = [summary] if summary else []
    for region in regions:
        block = alerts_anchor_region_block(uid, region)
        if block:
            lines.append(block)
    if not lines:
        return ""
    head = lines[0]
    tail = lines[1:4]
    return "\n".join([head] + tail)


def alerts_recipients_for_event(event: Dict[str, Any]) -> List[Tuple[int, Dict[str, Any]]]:
    recipients: List[Tuple[int, Dict[str, Any]]] = []
    target_region = alerts_canonical_region(event.get("region") or event.get("region_display")) or event.get("region")
    if not target_region:
        return recipients
    for profile in load_all_users():
        uid = profile.get("user_id")
        if not uid:
            continue
        regions = alerts_user_regions(uid)
        canonical_regions = {alerts_canonical_region(r) or r for r in regions}
        if target_region not in canonical_regions:
            continue
        recipients.append((uid, profile))
    return recipients


def alerts_push_summary_text(uid: int, event: Dict[str, Any], kind: str) -> str:
    lang = resolve_lang(uid)
    ended = kind == "end" or bool(event.get("ended_at"))
    region_display = alerts_display_region_name(event.get("region_display") or event.get("region") or "", lang)
    header_key = "ALERTS_PUSH_HEADER_STANDDOWN" if ended else "ALERTS_PUSH_HEADER_ALERT"
    header = tr(uid, header_key, region=region_display)
    lead_key = "ALERTS_PUSH_SUMMARY_LEAD_STANDDOWN" if ended else "ALERTS_PUSH_SUMMARY_LEAD_ALERT"
    lead_line = tr(uid, lead_key)
    type_label = alerts_type_label(event, lang)
    type_icon = alerts_type_icon(event)
    start_display = alerts_format_push_timestamp(event.get("started_at")) or "--:--"
    if ended:
        end_source = event.get("ended_at") or event.get("updated_at")
        end_display = alerts_format_push_timestamp(end_source) or "—"
        body = tr(
            uid,
            "ALERTS_PUSH_SUMMARY_ENDED",
            icon=type_icon,
            type=type_label,
            start=start_display,
            ended=end_display,
        )
    else:
        progress = alerts_duration_phrase(event.get("started_at"), None, lang, True)
        body = tr(
            uid,
            "ALERTS_PUSH_SUMMARY_RUNNING",
            icon=type_icon,
            type=type_label,
            start=start_display,
            progress=progress,
        )
    lines: List[str] = []
    if ended and lead_line:
        lines.append(lead_line)
        lines.append("")
    lines.append(header)
    lines.append(body)
    return "\n".join(lines).strip()


def alerts_push_detail_text(uid: int, event: Dict[str, Any], kind: str) -> str:
    lang = resolve_lang(uid)
    ended = kind == "end" or bool(event.get("ended_at"))
    region_display = alerts_display_region_name(event.get("region_display") or event.get("region") or "", lang)
    title_key = "ALERTS_PUSH_DETAIL_TITLE_STANDDOWN" if ended else "ALERTS_PUSH_DETAIL_TITLE_ALERT"
    title = tr(uid, title_key, region=region_display)
    type_label = alerts_type_label(event, lang)
    type_icon = alerts_type_icon(event)
    start_date, start_time = alerts_format_push_date_pair(event.get("started_at"))
    end_source = event.get("ended_at") or event.get("updated_at")
    end_date, end_time = alerts_format_push_date_pair(end_source) if ended else ("", "")
    duration_seconds = alerts_duration_seconds(event.get("started_at"), end_source if ended else None)
    duration_value = alerts_format_duration_value(duration_seconds, lang)
    stats_country = alerts_country_count_label(alerts_active_oblast_count(), lang)
    region_key = alerts_canonical_region(event.get("region") or event.get("region_display") or "") or event.get("region") or ""
    region_active = alerts_region_active_total(region_key) if region_key else 0
    if region_active:
        value = alerts_region_active_value(region_active, lang)
        stats_region = tr(uid, "ALERTS_PUSH_DETAIL_STATS_REGION_ACTIVE", value=value)
    else:
        stats_region = tr(uid, "ALERTS_PUSH_DETAIL_STATS_REGION_CLEAR")
    recommendations = alerts_recommendation_block(event, lang)

    lines: List[str] = [
        title,
        "━━━━━━━━━━━━━━━━━━",
        "",
        tr(uid, "ALERTS_PUSH_DETAIL_TYPE", icon=type_icon, value=type_label),
        "",
        tr(uid, "ALERTS_PUSH_DETAIL_START", date=start_date or "—", time=start_time or "--:--"),
    ]
    if ended:
        lines.append(tr(uid, "ALERTS_PUSH_DETAIL_END_STANDDOWN", date=end_date or "—", time=end_time or "--:--"))
    lines.append(tr(uid, "ALERTS_PUSH_DETAIL_DURATION", duration=duration_value))
    lines.extend(
        [
            "",
            "━━━━━━━━━━━━━━━━━━",
            tr(uid, "ALERTS_PUSH_DETAIL_STATS_HEADER"),
            tr(uid, "ALERTS_PUSH_DETAIL_STATS_COUNTRY", value=stats_country),
            stats_region,
        ]
    )
    if ended:
        lines.extend(
            [
                "",
                "━━━━━━━━━━━━━━━━━━",
                tr(uid, "ALERTS_PUSH_DETAIL_STANDDOWN_HEADER"),
                tr(uid, "ALERTS_PUSH_DETAIL_STANDDOWN_NOTE"),
            ]
        )
    else:
        lines.extend(
            [
                "",
                "━━━━━━━━━━━━━━━━━━",
                tr(uid, "ALERTS_PUSH_DETAIL_RECOMMENDATIONS_HEADER"),
            ]
        )
        if recommendations:
            lines.append(recommendations)
        lines.extend([
            "",
            "━━━━━━━━━━━━━━━━━━",
            tr(uid, "ALERTS_PUSH_DETAIL_FOOTER"),
        ])
    return "\n".join(lines)


def alerts_push_render(uid: int, event: Dict[str, Any], kind: str, expanded: bool = False) -> str:
    if expanded:
        return alerts_push_detail_text(uid, event, kind)
    return alerts_push_summary_text(uid, event, kind)


def alerts_push_store(uid: int, token: str, payload: Dict[str, Any]) -> None:
    runtime = users_runtime.setdefault(uid, {})
    registry = runtime.setdefault("alerts_pushes", {})
    registry[token] = payload


def alerts_push_get(uid: int, token: str) -> Optional[Dict[str, Any]]:
    runtime = users_runtime.get(uid, {})
    registry = runtime.get("alerts_pushes", {})
    entry = registry.get(token)
    if not isinstance(entry, dict):
        return None
    return entry


def alerts_push_remove(uid: int, token: str) -> Optional[Dict[str, Any]]:
    runtime = users_runtime.get(uid, {})
    registry = runtime.get("alerts_pushes")
    if isinstance(registry, dict):
        return registry.pop(token, None)
    return None


def alerts_notification_text(uid: int, event: Dict[str, Any], kind: str) -> str:
    return alerts_push_summary_text(uid, event, kind)


async def alerts_broadcast(event_id: str, kind: str) -> None:
    event = _alerts_get_event(event_id)
    if not event:
        return
    if kind == "start" and event.get("notified_start"):
        return
    if kind == "end" and event.get("notified_end"):
        return
    recipients = alerts_recipients_for_event(event)
    if not recipients:
        _alerts_mark_notified(event_id, kind)
        return
    for uid, profile in recipients:
        chat_id = users_runtime.get(uid, {}).get("tg", {}).get("chat_id")
        if not chat_id:
            chat_id = (profile.get("tg") or {}).get("chat_id")
        if not chat_id:
            continue
        try:
            text = alerts_push_summary_text(uid, event, kind)
            token = secrets.token_hex(4)
            while alerts_push_get(uid, token):
                token = secrets.token_hex(4)
            kb = alerts_push_keyboard(uid, token, expanded=False)
            message = await bot.send_message(chat_id, text, reply_markup=kb, disable_web_page_preview=True)
            alerts_push_store(
                uid,
                token,
                {
                    "event_id": event_id,
                    "kind": kind,
                    "message_id": message.message_id,
                    "chat_id": message.chat.id,
                    "expanded": False,
                },
            )
        except Exception:
            continue
    _alerts_mark_notified(event_id, kind)


async def alerts_dispatch_updates(start_ids: List[str], end_ids: List[str]) -> None:
    changed = False
    for event_id in start_ids:
        event = _alerts_get_event(event_id)
        if not event:
            continue
        await alerts_broadcast(event_id, "start")
        changed = True
    for event_id in end_ids:
        event = _alerts_get_event(event_id)
        if not event:
            continue
        await alerts_broadcast(event_id, "end")
        changed = True
    if changed:
        await update_all_anchors()


async def alerts_poll_loop() -> None:
    global alerts_poll_task
    try:
        await asyncio.sleep(ALERTS_POLL_INTERVAL)
        while True:
            try:
                start_ids, end_ids = await asyncio.to_thread(alerts_refresh_once)
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                print(f"[alerts] refresh error: {exc}")
                start_ids, end_ids = [], []
            if start_ids or end_ids:
                await alerts_dispatch_updates(start_ids, end_ids)
            await asyncio.sleep(ALERTS_POLL_INTERVAL)
    except asyncio.CancelledError:
        pass
    finally:
        alerts_poll_task = None


async def alerts_bootstrap() -> None:
    try:
        start_ids, end_ids = await asyncio.to_thread(alerts_refresh_once)
    except Exception as exc:
        print(f"[alerts] initial refresh error: {exc}")
        await update_all_anchors()
        return
    if start_ids or end_ids:
        await alerts_dispatch_updates(start_ids, end_ids)
    else:
        await update_all_anchors()


async def alerts_start_polling() -> None:
    global alerts_poll_task
    if alerts_poll_task and not alerts_poll_task.done():
        return
    alerts_poll_task = asyncio.create_task(alerts_poll_loop())


# ========================== NOVA POSHTA STORAGE ==========================

NP_API_URL = "https://api.novaposhta.ua/v2.0/json/"
NOVA_POSHTA_API_KEY = "2b7d39d126d56e60cfc61d00cd0b452c"
NP_DATA_FILE = os.path.join("data", "nova_poshta.json")

_np_state_cache: Optional[Dict[str, Any]] = None


def _np_utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def _np_ensure_storage() -> None:
    os.makedirs(os.path.dirname(NP_DATA_FILE), exist_ok=True)


def _np_blank_state() -> Dict[str, Any]:
    return {"users": {}, "assignments": {}}


def _np_load_state() -> Dict[str, Any]:
    global _np_state_cache
    if _np_state_cache is not None:
        return _np_state_cache
    _np_ensure_storage()
    if not os.path.exists(NP_DATA_FILE):
        _np_state_cache = _np_blank_state()
        _np_save_state()
        return _np_state_cache
    try:
        with open(NP_DATA_FILE, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        if not isinstance(payload, dict):
            raise ValueError("Invalid payload")
        payload.setdefault("users", {})
        payload.setdefault("assignments", {})
        _np_state_cache = payload
    except Exception:
        _np_state_cache = _np_blank_state()
        _np_save_state()
    return _np_state_cache


def _np_save_state() -> None:
    if _np_state_cache is None:
        return
    _np_ensure_storage()
    tmp_file = f"{NP_DATA_FILE}.tmp"
    with open(tmp_file, "w", encoding="utf-8") as fh:
        json.dump(_np_state_cache, fh, ensure_ascii=False, indent=2)
    os.replace(tmp_file, NP_DATA_FILE)


def _np_user_bucket(uid: int) -> Dict[str, Any]:
    state = _np_load_state()
    user = state["users"].setdefault(str(uid), {})
    user.setdefault("history", [])
    user.setdefault("bookmarks", {})
    user.setdefault("notes", {})
    user.setdefault("assigned", {})
    return user


def np_fetch_tracking(ttn: str) -> Tuple[bool, Optional[Dict[str, Any]], str]:
    number = (ttn or "").strip()
    if not number:
        return False, None, "Пустой номер ТТН"
    if not re.fullmatch(r"[0-9A-Za-z-]{5,40}", number):
        return False, None, "Некорректный формат ТТН"

    payload = {
        "apiKey": NOVA_POSHTA_API_KEY,
        "modelName": "TrackingDocument",
        "calledMethod": "getStatusDocuments",
        "methodProperties": {
            "Documents": [{"DocumentNumber": number}],
        },
    }
    try:
        response = requests.post(NP_API_URL, json=payload, timeout=15)
        response.raise_for_status()
    except requests.RequestException as exc:
        return False, None, f"Ошибка запроса: {exc}"

    try:
        data = response.json()
    except ValueError:
        return False, None, "Не удалось разобрать ответ API"

    docs = data.get("data") or []
    if docs:
        doc = docs[0]
        if isinstance(doc, dict):
            doc = dict(doc)
            doc.setdefault("Number", number)
            doc.setdefault("DocumentNumber", number)
            return True, doc, ""

    errors = data.get("errors") or data.get("message") or data.get("error")
    if isinstance(errors, list):
        msg = "; ".join(str(e) for e in errors if e)
    else:
        msg = str(errors or "Номер не найден")
    if not msg:
        msg = "Номер не найден"
    return False, None, msg


def np_remember_search(uid: int, ttn: str, status_payload: Optional[Dict[str, Any]]) -> None:
    user = _np_user_bucket(uid)
    status_payload = status_payload or {}
    history = [entry for entry in user["history"] if entry.get("ttn") != ttn]
    history.insert(0, {
        "ttn": ttn,
        "timestamp": _np_utcnow(),
        "status_payload": status_payload,
    })
    user["history"] = history[:25]
    _np_save_state()


def np_get_history(uid: int) -> List[Dict[str, Any]]:
    user = _np_user_bucket(uid)
    return list(user["history"])


def np_get_cached_status(uid: int, ttn: str) -> Optional[Dict[str, Any]]:
    user = _np_user_bucket(uid)
    for entry in user["history"]:
        if entry.get("ttn") == ttn:
            return entry.get("status_payload")
    bookmark = user["bookmarks"].get(ttn)
    if isinstance(bookmark, dict):
        return bookmark.get("status_payload")
    assignment = _np_load_state()["assignments"].get(ttn)
    if isinstance(assignment, dict) and (
        assignment.get("assigned_to") == uid or assignment.get("assigned_by") == uid
    ):
        return assignment.get("status_payload")
    return None


def _np_set_bookmark(uid: int, ttn: str, status_payload: Optional[Dict[str, Any]]) -> None:
    user = _np_user_bucket(uid)
    user["bookmarks"][ttn] = {
        "added_at": _np_utcnow(),
        "status_payload": status_payload or {},
    }
    _np_save_state()


def np_remove_bookmark(uid: int, ttn: str) -> None:
    user = _np_user_bucket(uid)
    user["bookmarks"].pop(ttn, None)
    _np_save_state()


def np_toggle_bookmark(uid: int, ttn: str, status_payload: Optional[Dict[str, Any]] = None) -> bool:
    user = _np_user_bucket(uid)
    if ttn in user["bookmarks"]:
        np_remove_bookmark(uid, ttn)
        return False
    if status_payload is None:
        status_payload = np_get_cached_status(uid, ttn) or {}
    _np_set_bookmark(uid, ttn, status_payload)
    return True


def np_list_bookmarks(uid: int) -> List[Tuple[str, Dict[str, Any]]]:
    user = _np_user_bucket(uid)
    items = []
    for ttn, payload in user["bookmarks"].items():
        entry = dict(payload)
        entry["ttn"] = ttn
        items.append((ttn, entry))
    items.sort(key=lambda x: x[1].get("added_at", ""), reverse=True)
    return items


def np_has_bookmark(uid: int, ttn: str) -> bool:
    return ttn in _np_user_bucket(uid)["bookmarks"]


def np_add_note(uid: int, ttn: str, text: str) -> Dict[str, Any]:
    user = _np_user_bucket(uid)
    bucket = user["notes"].setdefault(ttn, [])
    note = {
        "note_id": secrets.token_hex(6),
        "ttn": ttn,
        "text": text,
        "timestamp": _np_utcnow(),
    }
    bucket.insert(0, note)
    user["notes"][ttn] = bucket[:20]
    _np_save_state()
    return note


def np_list_notes(uid: int, ttn: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
    user = _np_user_bucket(uid)
    notes = user["notes"]
    if ttn is not None:
        return {ttn: list(notes.get(ttn, []))}
    return {key: list(value) for key, value in notes.items() if value}


def np_assign_parcel(admin_uid: int, target_uid: int, ttn: str,
                     status_payload: Optional[Dict[str, Any]], note: Optional[str] = None) -> Dict[str, Any]:
    state = _np_load_state()
    now = _np_utcnow()
    assignment = state["assignments"].get(ttn, {})
    assignment.update({
        "ttn": ttn,
        "assigned_to": target_uid,
        "assigned_by": admin_uid,
        "note": note or "",
        "created_at": assignment.get("created_at") or now,
        "updated_at": now,
        "status_payload": status_payload or assignment.get("status_payload") or {},
        "delivered_at": assignment.get("delivered_at"),
        "delivery_note": assignment.get("delivery_note", ""),
    })
    state["assignments"][ttn] = assignment
    user = _np_user_bucket(target_uid)
    user["assigned"][ttn] = {
        "assigned_at": assignment.get("created_at") or now,
        "assigned_by": admin_uid,
    }
    _np_save_state()
    return dict(assignment)


def np_get_assignment(ttn: str) -> Optional[Dict[str, Any]]:
    state = _np_load_state()
    assignment = state["assignments"].get(ttn)
    if assignment:
        return dict(assignment)
    return None


def np_list_assignments(uid: int) -> List[Dict[str, Any]]:
    state = _np_load_state()
    result: List[Dict[str, Any]] = []
    for ttn, assignment in state["assignments"].items():
        if assignment.get("assigned_to") == uid:
            entry = dict(assignment)
            entry["ttn"] = ttn
            result.append(entry)
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result


def np_list_admin_assignments(admin_uid: int) -> List[Dict[str, Any]]:
    state = _np_load_state()
    result: List[Dict[str, Any]] = []
    for ttn, assignment in state["assignments"].items():
        if assignment.get("assigned_by") == admin_uid:
            entry = dict(assignment)
            entry["ttn"] = ttn
            result.append(entry)
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result


def np_mark_assignment_received(uid: int, ttn: str, delivery_note: str = "") -> Optional[Dict[str, Any]]:
    state = _np_load_state()
    assignment = state["assignments"].get(ttn)
    if not assignment or assignment.get("assigned_to") != uid:
        return None
    assignment["delivered_at"] = _np_utcnow()
    assignment["delivery_note"] = delivery_note or ""
    assignment["updated_at"] = assignment["delivered_at"]
    user = _np_user_bucket(uid)
    bucket = user["assigned"].setdefault(ttn, {})
    bucket["delivered_at"] = assignment["delivered_at"]
    _np_save_state()
    return dict(assignment)


def np_refresh_assignment_status(ttn: str, status_payload: Optional[Dict[str, Any]]) -> None:
    state = _np_load_state()
    assignment = state["assignments"].get(ttn)
    if not assignment:
        return
    assignment["status_payload"] = status_payload or {}
    assignment["updated_at"] = _np_utcnow()
    _np_save_state()


# ========================== NOVA POSHTA ==========================

def _np_clean_ttn(raw: str) -> str:
    return re.sub(r"[^0-9A-Za-z-]", "", (raw or "").strip())


def _np_trim_label(text: str, limit: int = 48) -> str:
    text = (text or "").strip()
    if len(text) <= limit:
        return text
    return text[: max(1, limit - 1)] + "…"


@dp.callback_query_handler(lambda c: c.data == "menu_np")
async def menu_np(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "np_interface")
async def np_interface_menu(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, tr(uid, "NP_INTERFACE_TEXT"), kb_novaposhta(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "np_search")
async def np_search_start(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear(uid)
    await state.set_state(NovaPoshtaFSM.waiting_ttn.state)
    prompt = await bot.send_message(c.message.chat.id, tr(uid, "NP_PROMPT_TTN"), reply_markup=kb_np_cancel(uid))
    flow_track(uid, prompt)
    await c.answer()


@dp.message_handler(state=NovaPoshtaFSM.waiting_ttn, content_types=ContentType.TEXT)
async def np_receive_ttn(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    text = (m.text or "").strip()
    if not text:
        return
    if text.lower() in NP_CANCEL_WORDS:
        await state.finish()
        await flow_clear(uid)
        await anchor_show_text(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))
        try:
            await bot.delete_message(m.chat.id, m.message_id)
        except Exception:
            pass
        return

    ttn = _np_clean_ttn(text) or text
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass

    progress = await bot.send_message(m.chat.id, tr(uid, "NP_SEARCH_PROGRESS"))
    flow_track(uid, progress)

    text_body, kb, payload, assignment_display, error_message = np_prepare_view(uid, ttn, force_fetch=True)
    await flow_clear(uid)

    if not text_body:
        warn_text = tr(uid, "NP_SEARCH_ERROR", error=error_message or "—")
        lowered = (error_message or "").lower()
        if any(token in lowered for token in ("не найден", "не знайден", "not found", "невірн", "неверн")):
            warn_text = tr(uid, "NP_SEARCH_NOT_FOUND", ttn=h(ttn))
        warn = await bot.send_message(m.chat.id, warn_text)
        flow_track(uid, warn)
        prompt = await bot.send_message(m.chat.id, tr(uid, "NP_PROMPT_TTN"), reply_markup=kb_np_cancel(uid))
        flow_track(uid, prompt)
        return

    await state.finish()
    await np_send_card(uid, m.chat.id, text_body, kb)


@dp.callback_query_handler(lambda c: c.data.startswith("np_refresh:"))
async def np_refresh_detail(c: types.CallbackQuery):
    uid = c.from_user.id
    ttn = c.data.split(":", 1)[1]
    text_body, kb, _, _, error_message = np_prepare_view(uid, ttn, force_fetch=True)
    if not text_body:
        await c.answer(tr(uid, "NP_SEARCH_ERROR", error=error_message or "—"), show_alert=True)
        return
    try:
        await bot.edit_message_text(
            text_body,
            c.message.chat.id,
            c.message.message_id,
            reply_markup=kb,
            disable_web_page_preview=True,
        )
    except MessageNotModified:
        pass
    except MessageCantBeEdited:
        await c.answer(tr(uid, "NP_REFRESH_NOT_POSSIBLE"), show_alert=True)
        return
    except Exception:
        await c.answer(tr(uid, "NP_REFRESH_NOT_POSSIBLE"), show_alert=True)
        return
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_bookmark:"))
async def np_toggle_bookmark_cb(c: types.CallbackQuery):
    uid = c.from_user.id
    ttn = c.data.split(":", 1)[1]
    payload = np_get_cached_status(uid, ttn)
    if payload is None:
        success, payload, error_message = np_fetch_tracking(ttn)
        if not success:
            await c.answer(tr(uid, "NP_SEARCH_ERROR", error=error_message or "—"), show_alert=True)
            return
        np_remember_search(uid, ttn, payload)
    added = np_toggle_bookmark(uid, ttn, status_payload=payload)
    text_body, kb, _, _, _ = np_prepare_view(uid, ttn, payload=payload)
    if text_body and kb:
        try:
            await bot.edit_message_text(text_body, c.message.chat.id, c.message.message_id, reply_markup=kb)
        except Exception:
            pass
    await c.answer(tr(uid, "NP_BOOKMARK_ADDED" if added else "NP_BOOKMARK_REMOVED"))


@dp.callback_query_handler(lambda c: c.data.startswith("np_note:"))
async def np_note_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    ttn = c.data.split(":", 1)[1]
    await state.set_state(NovaPoshtaFSM.waiting_note.state)
    await state.update_data(note_ttn=ttn, note_message=(c.message.chat.id, c.message.message_id))
    prompt = await bot.send_message(c.message.chat.id, tr(uid, "NP_NOTE_PROMPT", ttn=h(ttn)), reply_markup=kb_np_cancel(uid))
    flow_track(uid, prompt)
    await c.answer()


@dp.message_handler(state=NovaPoshtaFSM.waiting_note, content_types=ContentType.TEXT)
async def np_note_receive(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    data = await state.get_data()
    ttn = data.get("note_ttn")
    if not ttn:
        await state.finish()
        return
    text = (m.text or "").strip()
    if not text:
        return
    if text.lower() in NP_CANCEL_WORDS:
        await state.finish()
        await flow_clear(uid)
        notice = await bot.send_message(m.chat.id, tr(uid, "NP_NOTE_CANCELLED"))
        flow_track(uid, notice)
        return
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    np_add_note(uid, ttn, text)
    await state.finish()
    await flow_clear(uid)
    chat_id, message_id = data.get("note_message", (None, None))
    text_body, kb, _, _, _ = np_prepare_view(uid, ttn)
    if text_body and kb and chat_id and message_id:
        try:
            await bot.edit_message_text(text_body, chat_id, message_id, reply_markup=kb)
        except Exception:
            await bot.send_message(chat_id, text_body, reply_markup=kb)
    ack = await bot.send_message(m.chat.id, tr(uid, "NP_NOTE_SAVED"))
    flow_track(uid, ack)


@dp.callback_query_handler(lambda c: c.data == "np_history")
async def np_history_menu(c: types.CallbackQuery):
    uid = c.from_user.id
    history = np_get_history(uid)
    if not history:
        await clear_then_anchor(uid, tr(uid, "NP_HISTORY_EMPTY"), kb_novaposhta(uid))
        await c.answer()
        return
    lines = [tr(uid, "NP_HISTORY_HEADER")]
    options: List[Tuple[str, str]] = []
    for idx, entry in enumerate(history[:10], start=1):
        ttn = entry.get("ttn")
        if not ttn:
            continue
        timestamp = format_datetime_short(entry.get("timestamp")) or entry.get("timestamp") or ""
        summary = format_np_short_entry(entry.get("status_payload"))
        line = f"{idx}. <b>{h(ttn)}</b>"
        if summary:
            line += f" — {h(summary)}"
        if timestamp:
            line += f" ({h(timestamp)})"
        lines.append(line)
        label_parts = [ttn]
        if timestamp:
            label_parts.append(timestamp)
        options.append((ttn, _np_trim_label(" • ".join(label_parts))))
    if not options:
        await clear_then_anchor(uid, tr(uid, "NP_HISTORY_EMPTY"), kb_novaposhta(uid))
    else:
        kb = np_build_list_keyboard(uid, options, "np_history_show")
        await clear_then_anchor(uid, "\n".join(lines), kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_history_show:"))
async def np_history_show_cb(c: types.CallbackQuery):
    uid = c.from_user.id
    ttn = c.data.split(":", 1)[1]
    text_body, kb, _, _, error_message = np_prepare_view(uid, ttn)
    if not text_body:
        await c.answer(tr(uid, "NP_SEARCH_ERROR", error=error_message or "—"), show_alert=True)
        return
    await np_send_card(uid, c.message.chat.id, text_body, kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "np_bookmarks")
async def np_bookmarks_menu(c: types.CallbackQuery):
    uid = c.from_user.id
    bookmarks = np_list_bookmarks(uid)
    if not bookmarks:
        await clear_then_anchor(uid, tr(uid, "NP_BOOKMARKS_EMPTY"), kb_novaposhta(uid))
        await c.answer()
        return
    lines = [tr(uid, "NP_BOOKMARKS_HEADER")]
    options: List[Tuple[str, str]] = []
    for idx, (ttn, entry) in enumerate(bookmarks[:10], start=1):
        timestamp = format_datetime_short(entry.get("added_at")) or entry.get("added_at") or ""
        summary = format_np_short_entry(entry.get("status_payload"))
        line = f"{idx}. <b>{h(ttn)}</b>"
        if summary:
            line += f" — {h(summary)}"
        if timestamp:
            line += f" ({h(timestamp)})"
        lines.append(line)
        label_parts = [ttn]
        if timestamp:
            label_parts.append(timestamp)
        options.append((ttn, _np_trim_label(" • ".join(label_parts))))
    kb = np_build_list_keyboard(uid, options, "np_bookmark_show")
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_bookmark_show:"))
async def np_bookmark_show_cb(c: types.CallbackQuery):
    uid = c.from_user.id
    ttn = c.data.split(":", 1)[1]
    text_body, kb, _, _, error_message = np_prepare_view(uid, ttn)
    if not text_body:
        await c.answer(tr(uid, "NP_SEARCH_ERROR", error=error_message or "—"), show_alert=True)
        return
    await np_send_card(uid, c.message.chat.id, text_body, kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "np_assigned")
async def np_assigned_menu(c: types.CallbackQuery):
    uid = c.from_user.id
    assignments = np_list_assignments(uid)
    if not assignments:
        await clear_then_anchor(uid, tr(uid, "NP_ASSIGNED_EMPTY"), kb_novaposhta(uid))
        await c.answer()
        return
    lines = [tr(uid, "NP_ASSIGNED_HEADER")]
    options: List[Tuple[str, str]] = []
    for idx, assignment in enumerate(assignments[:10], start=1):
        ttn = assignment.get("ttn")
        status_short = format_np_short_entry(assignment.get("status_payload"))
        assigned_time = format_datetime_short(assignment.get("created_at")) or assignment.get("created_at") or ""
        delivered_time = format_datetime_short(assignment.get("delivered_at")) if assignment.get("delivered_at") else ""
        line = f"{idx}. <b>{h(ttn)}</b>"
        if status_short:
            line += f" — {h(status_short)}"
        if assigned_time:
            line += f" ({h(assigned_time)})"
        if delivered_time:
            line += f" ✔️ {h(delivered_time)}"
        lines.append(line)
        label_parts = [ttn]
        if delivered_time:
            label_parts.append("✔")
        elif assigned_time:
            label_parts.append(assigned_time)
        options.append((ttn, _np_trim_label(" • ".join(label_parts))))
    kb = np_build_list_keyboard(uid, options, "np_assigned_detail")
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "np_received")
async def np_received_menu(c: types.CallbackQuery):
    uid = c.from_user.id
    assignments = [a for a in np_list_assignments(uid) if a.get("delivered_at")]
    if not assignments:
        await clear_then_anchor(uid, tr(uid, "NP_RECEIVED_EMPTY"), kb_novaposhta(uid))
        await c.answer()
        return
    lines = [tr(uid, "NP_RECEIVED_HEADER")]
    options: List[Tuple[str, str]] = []
    for idx, assignment in enumerate(assignments[:10], start=1):
        ttn = assignment.get("ttn")
        status_short = format_np_short_entry(assignment.get("status_payload"))
        delivered_time = format_datetime_short(assignment.get("delivered_at")) or assignment.get("delivered_at") or ""
        line = f"{idx}. <b>{h(ttn)}</b>"
        if status_short:
            line += f" — {h(status_short)}"
        if delivered_time:
            line += f" ✔️ {h(delivered_time)}"
        lines.append(line)
        label_parts = [ttn]
        if delivered_time:
            label_parts.append(delivered_time)
        options.append((ttn, _np_trim_label(" • ".join(label_parts))))
    kb = np_build_list_keyboard(uid, options, "np_assigned_detail")
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_assigned_detail:"))
async def np_assigned_detail_cb(c: types.CallbackQuery):
    uid = c.from_user.id
    ttn = c.data.split(":", 1)[1]
    assignment = np_get_assignment(ttn)
    payload = assignment.get("status_payload") if assignment else None
    text_body, kb, _, _, error_message = np_prepare_view(uid, ttn, payload=payload)
    if not text_body:
        await c.answer(tr(uid, "NP_SEARCH_ERROR", error=error_message or "—"), show_alert=True)
        return
    await np_send_card(uid, c.message.chat.id, text_body, kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_assigned_received:"))
async def np_assigned_received_cb(c: types.CallbackQuery):
    uid = c.from_user.id
    ttn = c.data.split(":", 1)[1]
    assignment = np_mark_assignment_received(uid, ttn)
    if not assignment:
        await c.answer(tr(uid, "NP_ASSIGNMENT_ALREADY_DONE"), show_alert=True)
        return
    text_body, kb, _, _, error_message = np_prepare_view(uid, ttn, payload=assignment.get("status_payload"))
    if text_body and kb:
        try:
            await bot.edit_message_text(text_body, c.message.chat.id, c.message.message_id, reply_markup=kb)
        except Exception:
            pass
    await c.answer(tr(uid, "NP_DELIVERY_ACK_RECORDED"))

    user_profile = load_user(uid) or {"user_id": uid}
    user_name = user_profile.get("fullname") or (user_profile.get("tg") or {}).get("first_name") or f"User {uid}"
    delivered_at = assignment.get("delivered_at")

    await anchor_show_root(uid)
    assigned_by = assignment.get("assigned_by")
    if assigned_by and assigned_by != uid:
        await anchor_show_root(assigned_by)

    user_lang = resolve_lang(uid)
    user_receipt = np_render_delivery_receipt(user_lang, ttn, user_name, delivered_at)
    user_kb = InlineKeyboardMarkup().add(
        InlineKeyboardButton(_np_pick(user_lang, NP_CLOSE_BUTTON_LABEL), callback_data="np_close")
    )
    try:
        receipt_msg = await bot.send_message(
            c.message.chat.id,
            user_receipt,
            reply_markup=user_kb,
            disable_web_page_preview=True,
        )
        flow_track(uid, receipt_msg)
    except Exception:
        pass

    for admin_id in admins:
        chat_id = users_runtime.get(admin_id, {}).get("tg", {}).get("chat_id") or (load_user(admin_id) or {}).get("tg", {}).get("chat_id")
        if not chat_id:
            continue
        admin_lang = resolve_lang(admin_id)
        alert = np_render_delivery_receipt(admin_lang, ttn, user_name, delivered_at)
        kb_admin = InlineKeyboardMarkup().add(
            InlineKeyboardButton(_np_pick(admin_lang, NP_CLOSE_BUTTON_LABEL), callback_data="np_close")
        )
        try:
            await bot.send_message(chat_id, alert, reply_markup=kb_admin, disable_web_page_preview=True)
        except Exception:
            continue


def np_assign_candidate_profiles() -> List[dict]:
    profiles = load_all_users()
    def sort_key(profile: dict) -> Tuple[str, int]:
        name = (profile.get("fullname") or (profile.get("tg") or {}).get("first_name") or "").strip().lower()
        return name, profile.get("user_id", 0)
    return sorted(profiles, key=sort_key)


def np_assign_format_label(profile: dict) -> str:
    name = profile.get("fullname") or (profile.get("tg") or {}).get("first_name") or f"ID {profile.get('user_id')}"
    bsu = profile.get("bsu") or f"ID {profile.get('user_id')}"
    label = f"{name} • {bsu}"
    return _np_trim_label(label, 36)


async def np_assign_show_picker(uid: int, state: FSMContext, chat_id: int, page: int = 0) -> None:
    profiles = np_assign_candidate_profiles()
    per_page = 6
    total = len(profiles)
    if total == 0:
        warn = await bot.send_message(chat_id, tr(uid, "NP_ASSIGN_USER_NOT_FOUND"))
        flow_track(uid, warn)
        return
    max_page = max(0, (total - 1) // per_page)
    page = max(0, min(page, max_page))
    start = page * per_page
    chunk = profiles[start:start + per_page]
    kb = InlineKeyboardMarkup()
    for profile in chunk:
        user_id = profile.get("user_id")
        if user_id is None:
            continue
        kb.add(InlineKeyboardButton(np_assign_format_label(profile), callback_data=f"np_assign_pick:{user_id}"))
    nav_buttons: List[InlineKeyboardButton] = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton("⬅️", callback_data=f"np_assign_page:{page-1}"))
    if page < max_page:
        nav_buttons.append(InlineKeyboardButton("➡️", callback_data=f"np_assign_page:{page+1}"))
    if nav_buttons:
        kb.row(*nav_buttons)
    kb.add(InlineKeyboardButton(_np_pick(resolve_lang(uid), NP_CANCEL_BUTTON_LABEL), callback_data="np_cancel"))
    text = tr(uid, "NP_ASSIGN_PROMPT_USER")
    if max_page:
        text += f"\n\n{page + 1}/{max_page + 1}"
    data = await state.get_data()
    picker_info = data.get("assign_picker")
    message_id = None
    if isinstance(picker_info, (list, tuple)) and len(picker_info) >= 2:
        stored_chat, stored_mid = picker_info[0], picker_info[1]
        if stored_chat == chat_id:
            message_id = stored_mid
    if message_id:
        try:
            await bot.edit_message_text(text, chat_id, message_id, reply_markup=kb)
        except Exception:
            msg = await bot.send_message(chat_id, text, reply_markup=kb)
            flow_track(uid, msg)
            await state.update_data(assign_picker=(msg.chat.id, msg.message_id, page))
        else:
            await state.update_data(assign_picker=(chat_id, message_id, page))
    else:
        msg = await bot.send_message(chat_id, text, reply_markup=kb)
        flow_track(uid, msg)
        await state.update_data(assign_picker=(msg.chat.id, msg.message_id, page))


async def np_assign_clear_picker(state: FSMContext):
    data = await state.get_data()
    picker_info = data.get("assign_picker")
    if isinstance(picker_info, (list, tuple)) and len(picker_info) >= 2:
        chat_id, message_id = picker_info[0], picker_info[1]
        try:
            await bot.delete_message(chat_id, message_id)
        except Exception:
            pass
    await state.update_data(assign_picker=None)


async def np_assign_user_selected(uid: int, profile: dict, state: FSMContext, chat_id: int):
    await np_assign_clear_picker(state)
    await flow_clear(uid)
    fullname = profile.get("fullname") or (profile.get("tg") or {}).get("first_name") or f"User {profile.get('user_id')}"
    bsu = profile.get("bsu", "—")
    summary = await bot.send_message(chat_id, f"👤 <b>{h(fullname)}</b> — BSU {h(bsu)}")
    flow_track(uid, summary)
    prompt = await bot.send_message(chat_id, tr(uid, "NP_ASSIGN_PROMPT_NOTE"), reply_markup=kb_np_assign_note(uid))
    flow_track(uid, prompt)
    await state.set_state(NovaPoshtaFSM.waiting_assign_note.state)


async def np_assign_finalize(uid: int, state: FSMContext, chat_id: int, note_text: str) -> None:
    data = await state.get_data()
    ttn = data.get("assign_ttn")
    payload = data.get("assign_payload")
    target_id = data.get("assign_user_id")
    note_text = (note_text or "").strip()

    await np_assign_clear_picker(state)
    await state.finish()
    await flow_clear(uid)

    if not target_id or not ttn or not payload:
        warn = await bot.send_message(chat_id, tr(uid, "NP_ASSIGN_CANCELLED"))
        flow_track(uid, warn)
        await anchor_show_text(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))
        return

    target_profile = load_user(target_id) or {"user_id": target_id}
    assignment = np_assign_parcel(uid, target_id, ttn, payload, note=note_text)
    await anchor_show_root(target_id)
    admin_profile = load_user(uid) or {"user_id": uid}
    admin_name = admin_profile.get("fullname") or (admin_profile.get("tg") or {}).get("first_name") or f"ID {uid}"
    target_name = target_profile.get("fullname") or (target_profile.get("tg") or {}).get("first_name") or f"User {target_id}"
    lang = resolve_lang(uid)
    assigned_time = format_datetime_short(assignment.get("updated_at")) or assignment.get("updated_at") or "—"

    confirm_text = tr(uid, "NP_ASSIGN_DONE", ttn=h(ttn), user=h(target_name), time=h(assigned_time))
    note_display = (assignment.get("note") or "").strip()
    if note_display:
        confirm_text = f"{confirm_text}\n\n{_np_pick(lang, NP_ASSIGN_DONE_NOTE_LABEL).format(note=h(note_display))}"

    confirm_kb = InlineKeyboardMarkup().add(
        InlineKeyboardButton(_np_pick(lang, NP_CLOSE_BUTTON_LABEL), callback_data="np_close")
    )
    confirm = await bot.send_message(chat_id, confirm_text, reply_markup=confirm_kb)
    flow_track(uid, confirm)
    await anchor_show_text(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))

    target_chat = users_runtime.get(target_id, {}).get("tg", {}).get("chat_id") or (target_profile.get("tg") or {}).get("chat_id")
    if target_chat:
        target_text, target_kb, _, _, _ = np_prepare_view(target_id, ttn, payload=payload)
        notify_prefix = tr(target_id, "NP_ASSIGN_NOTIFY_USER", admin=h(admin_name), ttn=h(ttn))
        body = f"{notify_prefix}\n\n{target_text}" if target_text else notify_prefix
        try:
            await bot.send_message(target_chat, body, reply_markup=target_kb)
        except Exception:
            pass


@dp.callback_query_handler(lambda c: c.data == "np_assign_start")
async def np_assign_start_cb(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        await c.answer("⛔", show_alert=True)
        return
    await flow_clear(uid)
    await state.set_state(NovaPoshtaFSM.waiting_assign_ttn.state)
    prompt = await bot.send_message(c.message.chat.id, tr(uid, "NP_ASSIGN_PROMPT_TTN"), reply_markup=kb_np_cancel(uid))
    flow_track(uid, prompt)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_assign_quick:"), state="*")
async def np_assign_quick_cb(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        await c.answer("⛔", show_alert=True)
        return
    ttn = c.data.split(":", 1)[1]
    payload = np_get_cached_status(uid, ttn)
    if payload is None:
        success, payload, error_message = np_fetch_tracking(ttn)
        if not success:
            await c.answer(tr(uid, "NP_SEARCH_ERROR", error=error_message or "—"), show_alert=True)
            return
        np_remember_search(uid, ttn, payload)
    await flow_clear(uid)
    await state.set_state(NovaPoshtaFSM.waiting_assign_user.state)
    await state.update_data(assign_ttn=ttn, assign_payload=payload)
    await np_assign_show_picker(uid, state, c.message.chat.id, page=0)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_assign_page:"), state="*")
async def np_assign_page_cb(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        await c.answer("⛔", show_alert=True)
        return
    current = await state.get_state()
    if current != NovaPoshtaFSM.waiting_assign_user.state:
        await c.answer()
        return
    try:
        page = int(c.data.split(":", 1)[1])
    except ValueError:
        await c.answer()
        return
    await np_assign_show_picker(uid, state, c.message.chat.id, page=page)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("np_assign_pick:"), state="*")
async def np_assign_pick_cb(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        await c.answer("⛔", show_alert=True)
        return
    current = await state.get_state()
    if current != NovaPoshtaFSM.waiting_assign_user.state:
        await c.answer()
        return
    try:
        target_id = int(c.data.split(":", 1)[1])
    except ValueError:
        await c.answer()
        return
    profile = load_user(target_id)
    if not profile:
        await c.answer(tr(uid, "NP_ASSIGN_USER_NOT_FOUND"), show_alert=True)
        return
    await state.update_data(assign_user_id=target_id)
    await np_assign_user_selected(uid, profile, state, c.message.chat.id)
    await c.answer()


@dp.message_handler(state=NovaPoshtaFSM.waiting_assign_ttn, content_types=ContentType.TEXT)
async def np_assign_receive_ttn(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        await state.finish()
        return
    text = (m.text or "").strip()
    if not text:
        return
    if text.lower() in NP_CANCEL_WORDS:
        await np_assign_clear_picker(state)
        await state.finish()
        await flow_clear(uid)
        try:
            await bot.delete_message(m.chat.id, m.message_id)
        except Exception:
            pass
        await anchor_show_text(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))
        return
    ttn = _np_clean_ttn(text) or text
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    text_body, kb, payload, _, error_message = np_prepare_view(uid, ttn, force_fetch=True)
    await flow_clear(uid)
    if not text_body:
        warn = await bot.send_message(m.chat.id, tr(uid, "NP_SEARCH_ERROR", error=error_message or "—"))
        flow_track(uid, warn)
        prompt = await bot.send_message(m.chat.id, tr(uid, "NP_ASSIGN_PROMPT_TTN"), reply_markup=kb_np_cancel(uid))
        flow_track(uid, prompt)
        return
    await state.set_state(NovaPoshtaFSM.waiting_assign_user.state)
    await state.update_data(assign_ttn=ttn, assign_payload=payload)
    preview = await bot.send_message(m.chat.id, text_body, reply_markup=kb)
    flow_track(uid, preview)
    await np_assign_show_picker(uid, state, m.chat.id, page=0)


@dp.message_handler(state=NovaPoshtaFSM.waiting_assign_user, content_types=ContentType.ANY)
async def np_assign_receive_user(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        await state.finish()
        return
    if m.text and m.text.strip().lower() in NP_CANCEL_WORDS:
        await np_assign_clear_picker(state)
        await state.finish()
        await flow_clear(uid)
        try:
            await bot.delete_message(m.chat.id, m.message_id)
        except Exception:
            pass
        await anchor_show_text(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))
        return
    profile = resolve_user_reference(m)
    if not profile:
        warn = await bot.send_message(m.chat.id, tr(uid, "NP_ASSIGN_USER_NOT_FOUND"))
        flow_track(uid, warn)
        return
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    await state.update_data(assign_user_id=profile.get("user_id"))
    await np_assign_user_selected(uid, profile, state, m.chat.id)


@dp.message_handler(state=NovaPoshtaFSM.waiting_assign_note, content_types=ContentType.TEXT)
async def np_assign_receive_note(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        await state.finish()
        return
    note_text = (m.text or "").strip()
    if note_text.lower() in NP_CANCEL_WORDS:
        await np_assign_clear_picker(state)
        await state.finish()
        await flow_clear(uid)
        try:
            await bot.delete_message(m.chat.id, m.message_id)
        except Exception:
            pass
        await anchor_show_text(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))
        return
    if note_text == "-":
        note_text = ""
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    await np_assign_finalize(uid, state, m.chat.id, note_text)


@dp.callback_query_handler(lambda c: c.data == "np_assign_skip", state="*")
async def np_assign_skip_cb(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        await c.answer("⛔", show_alert=True)
        return
    current = await state.get_state()
    if current != NovaPoshtaFSM.waiting_assign_note.state:
        await c.answer()
        return
    await np_assign_finalize(uid, state, c.message.chat.id, "")
    await c.answer(tr(uid, "NP_ASSIGN_SKIP_TOAST"))


@dp.callback_query_handler(lambda c: c.data == "np_close", state="*")
async def np_close_message(c: types.CallbackQuery):
    try:
        await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception:
        pass
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "np_cancel", state="*")
async def np_cancel_flow(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    current = await state.get_state()
    if current and current.startswith("NovaPoshtaFSM"):
        await np_assign_clear_picker(state)
        await state.finish()
    await flow_clear(uid)
    try:
        await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception:
        pass
    await anchor_show_text(uid, tr(uid, "NP_MENU_TITLE"), kb_novaposhta(uid))
    await c.answer(tr(uid, "NP_CANCELLED_TOAST"))

# ========================== CHECKS ==========================
@dp.callback_query_handler(lambda c: c.data == "menu_profile")
async def menu_profile(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await state.finish()
    await flow_clear(uid)
    profile_set_flags(uid, edit_mode=False, show_photo=False)
    await show_profile(uid, edit_mode=False, show_photo=False)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit")
async def profile_enter_edit(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await state.finish()
    await flow_clear(uid)
    profile_set_flags(uid, edit_mode=True, show_photo=False)
    await show_profile(uid, edit_mode=True, show_photo=False)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_done", state="*")
async def profile_exit_edit(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await profile_clear_prompt(uid)
    await flow_clear_warnings(uid)
    runtime = profile_runtime(uid)
    remove_keyboard = bool(runtime.pop("profile_reply_keyboard", False))
    await state.finish()
    profile_set_flags(uid, edit_mode=False, show_photo=False)
    if remove_keyboard:
        await profile_send_notification(uid, tr(uid, "PROFILE_CANCELLED"), remove_keyboard=True)
    await show_profile(uid, edit_mode=False, show_photo=False)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_view_photo")
async def profile_view_photo(c: types.CallbackQuery):
    uid = c.from_user.id
    profile_set_flags(uid, show_photo=True)
    await show_profile(uid, show_photo=True)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_hide_photo")
async def profile_hide_photo(c: types.CallbackQuery):
    uid = c.from_user.id
    profile_set_flags(uid, show_photo=False)
    await show_profile(uid, show_photo=False)
    await c.answer()


@dp.callback_query_handler(
    lambda c: c.data == "profile_cancel",
    state=[
        ProfileEditFSM.waiting_last_name,
        ProfileEditFSM.waiting_first_name,
        ProfileEditFSM.waiting_middle_name,
        ProfileEditFSM.waiting_birthdate,
        ProfileEditFSM.waiting_region,
        ProfileEditFSM.region_confirm,
        ProfileEditFSM.waiting_phone,
        ProfileEditFSM.waiting_photo,
    ]
)
async def profile_cancel_edit(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    runtime = profile_runtime(uid)
    remove_keyboard = bool(runtime.pop("profile_reply_keyboard", False))
    await profile_abort(uid, state, remove_keyboard=remove_keyboard)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit_last")
async def profile_prompt_last(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear_warnings(uid)
    await profile_send_prompt(uid, tr(uid, "PROFILE_PROMPT_LAST_NAME"), reply_markup=kb_profile_cancel(uid))
    await ProfileEditFSM.waiting_last_name.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit_first")
async def profile_prompt_first(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear_warnings(uid)
    await profile_send_prompt(uid, tr(uid, "PROFILE_PROMPT_FIRST_NAME"), reply_markup=kb_profile_cancel(uid))
    await ProfileEditFSM.waiting_first_name.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit_middle")
async def profile_prompt_middle(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear_warnings(uid)
    await profile_send_prompt(uid, tr(uid, "PROFILE_PROMPT_MIDDLE_NAME"), reply_markup=kb_profile_cancel(uid))
    await ProfileEditFSM.waiting_middle_name.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit_birthdate")
async def profile_prompt_birthdate(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear_warnings(uid)
    await profile_send_prompt(uid, tr(uid, "PROFILE_PROMPT_BIRTHDATE"), reply_markup=kb_profile_cancel(uid))
    await ProfileEditFSM.waiting_birthdate.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit_region")
async def profile_prompt_region(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear_warnings(uid)
    await profile_send_prompt(uid, tr(uid, "PROFILE_PROMPT_REGION"), reply_markup=kb_profile_region_prompt(uid))
    await ProfileEditFSM.waiting_region.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit_phone")
async def profile_prompt_phone(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear_warnings(uid)
    kb = kb_profile_phone_keyboard(uid)
    await profile_send_prompt(uid, tr(uid, "PROFILE_PROMPT_PHONE"), reply_markup=kb)
    await ProfileEditFSM.waiting_phone.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "profile_edit_photo")
async def profile_prompt_photo(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear_warnings(uid)
    await profile_send_prompt(uid, tr(uid, "PROFILE_PROMPT_PHOTO"), reply_markup=kb_profile_cancel(uid))
    await ProfileEditFSM.waiting_photo.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "menu_checks")
async def menu_checks(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, tr(uid, "CHECKS_MENU_INTRO"), kb_checks(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "menu_settings")
async def menu_settings(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, tr(uid, "SETTINGS_TITLE"), kb_settings(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "settings_language")
async def settings_language(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, tr(uid, "LANGUAGE_PROMPT"), kb_language_settings(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "settings_back")
async def settings_back(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, tr(uid, "SETTINGS_TITLE"), kb_settings(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("settings_lang:"))
async def settings_lang_change(c: types.CallbackQuery):
    uid = c.from_user.id
    code = c.data.split(":", 1)[1]
    if code not in LANG_CODES:
        await c.answer(tr(uid, "INVALID_COMMAND"), show_alert=True)
        return
    set_user_lang(uid, code, confirmed=True)
    await c.answer()
    await clear_then_anchor(uid, tr(uid, "LANGUAGE_SELECTED", language=LANG_LABELS[code]), kb_settings(uid))
    await anchor_show_root(uid)


@dp.callback_query_handler(lambda c: c.data == "check_stats")
async def check_stats(c: types.CallbackQuery):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    proj = active_project["name"]
    recs = user_project_receipts(uid, proj)
    cnt = len(recs)
    total = round(sum(float(r.get("sum") or 0.0) for r in recs), 2)
    paid_recs = [r for r in recs if r.get("paid") is True]
    unpaid_recs = [r for r in recs if r.get("paid") is False]
    pending_recs = [r for r in recs if r.get("paid") is None]
    paid_sum = round(sum(float(r.get("sum") or 0.0) for r in paid_recs), 2)
    unpaid_sum = round(sum(float(r.get("sum") or 0.0) for r in unpaid_recs), 2)
    pending_sum = round(sum(float(r.get("sum") or 0.0) for r in pending_recs), 2)
    summary_lines = [
        "📊 <b>Личная статистика по чекам</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"📂 Проект: <b>{h(proj)}</b>",
        f"🧾 Всего чеков: <b>{cnt}</b>",
        f"💰 Сумма чеков: <b>{fmt_money(total)} грн</b>",
        f"💸 Оплачено фирмой: <b>{fmt_money(paid_sum)} грн</b> ({len(paid_recs)} шт.)",
        f"⏳ Ожидает оплаты: <b>{fmt_money(unpaid_sum)} грн</b> ({len(unpaid_recs)} шт.)",
    ]
    if pending_recs:
        summary_lines.append(f"❔ Статус не указан: <b>{fmt_money(pending_sum)} грн</b> ({len(pending_recs)} шт.)")
    if cnt == 0:
        summary_lines.append("")
        summary_lines.append("Добавьте первый чек через кнопку «📷 Добавить чек», и здесь появится расширенная таблица с деталями.")
        await clear_then_anchor(uid, "\n".join(summary_lines), kb_checks(uid))
        await c.answer()
        return
    summary_lines.append("")
    summary_lines.append("Полная расшифровка всех чеков приведена ниже в порядке от новых к старым.")
    await clear_then_anchor(uid, "\n".join(summary_lines), kb_checks(uid))

    sorted_recs = sorted(
        recs,
        key=lambda r: (
            r.get("date") or "",
            r.get("time") or "",
            r.get("receipt_no") or "",
        ),
        reverse=True,
    )
    chat_id = c.message.chat.id
    chunk_size = 12
    total_recs = len(sorted_recs)
    for start in range(0, total_recs, chunk_size):
        chunk = sorted_recs[start:start + chunk_size]
        lines: List[str] = [
            "🗂 <b>Детализация чеков</b>",
            f"Строки {start + 1}–{start + len(chunk)} из {total_recs}",
            "━━━━━━━━━━━━━━━━━━",
        ]
        for idx, receipt in enumerate(chunk, start=start + 1):
            lines.append(format_receipt_stat_entry(idx, receipt))
            lines.append("")
        message_text = "\n".join(line for line in lines if line is not None).strip()
        msg = await bot.send_message(chat_id, message_text, disable_web_page_preview=True)
        flow_track(uid, msg)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "check_list")
async def check_list(c: types.CallbackQuery):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    proj = active_project["name"]
    recs = user_project_receipts(uid, proj)
    if not recs:
        await c.answer("У вас ещё нет сохранённых чеков. Добавьте первый через кнопку «📷 Добавить чек».", show_alert=True); return
    total = len(recs)
    header = (
        "📁 <b>Все загруженные чеки</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"📂 Проект: <b>{h(proj)}</b>\n"
        f"🧾 Всего чеков: <b>{total}</b>\n\n"
        "Просматривайте карточки ниже. Для чеков без статуса оплаты появятся кнопки, чтобы указать «Фирма оплатила» или «Фирма не оплатила»."
    )
    await clear_then_anchor(uid, header, kb_checks(uid))
    chat_id = c.message.chat.id
    for idx, r in enumerate(recs, start=1):
        token = r.get("receipt_no") or r.get("file") or ""
        rows: List[List[InlineKeyboardButton]] = []
        if r.get("paid") is None and token:
            rows.append([
                InlineKeyboardButton("✅ Фирма оплатила", callback_data=f"userpaid_yes:{proj}:{token}"),
                InlineKeyboardButton("❌ Фирма не оплатила", callback_data=f"userpaid_no:{proj}:{token}")
            ])
        rows.append([InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close")])
        kb: Optional[InlineKeyboardMarkup] = None
        if rows:
            kb = InlineKeyboardMarkup()
            for row in rows:
                kb.row(*row)
        prefix = f"🧾 Чек {idx} из {total}"
        try:
            msg = await send_receipt_card(chat_id, proj, uid, r, kb=kb, prefix=prefix)
        except Exception as exc:
            print(f"[check_list] failed to send receipt card: {exc}")
            fallback = (
                f"{prefix}\n"
                f"🆔 Номер: <b>{h(r.get('receipt_no', '—'))}</b>\n"
                f"💰 {fmt_money(float(r.get('sum') or 0.0))} грн\n"
                f"🔖 {receipt_status_text(r.get('paid'))}"
            )
            msg = await bot.send_message(chat_id, fallback, reply_markup=kb)
        flow_track(uid, msg)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "check_history")
async def check_history(c: types.CallbackQuery):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    proj = active_project["name"]
    recs = user_project_receipts(uid, proj)
    if not recs:
        await clear_then_anchor(
            uid,
            f"🗂 История чеков пока пуста.\n📂 Проект: <b>{h(proj)}</b>\n\nДобавьте первый чек через кнопку «📷 Добавить чек», и здесь появится полный журнал операций.",
            kb_checks(uid)
        )
        return await c.answer()
    sorted_recs = sorted(
        recs,
        key=lambda r: (
            r.get("date") or "",
            r.get("time") or "",
            r.get("receipt_no") or ""
        )
    )
    display_recs = sorted_recs[-30:]
    lines = [
        "🗂 <b>История чеков и оплат</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"📂 Проект: <b>{h(proj)}</b>",
        f"🗂 Записей в подборке: <b>{len(display_recs)}</b>",
        "",
        "Последние операции перечислены ниже в хронологическом порядке. Значки справа подсказывают текущий статус выплат.",
        "",
    ]
    for r in display_recs:
        try:
            amount = float(r.get("sum", 0.0))
        except (TypeError, ValueError):
            amount = 0.0
        base = f"• {h(r.get('receipt_no', '—'))} — {fmt_money(amount)} грн — {receipt_status_text(r.get('paid'))}"
        extra = ""
        payout = r.get("payout") if isinstance(r.get("payout"), dict) else None
        if payout and payout.get("status"):
            status = payout.get("status")
            code = payout.get("code") or payout.get("request_id")
            code_disp = h(code) if code else "—"
            if status == "pending":
                ts = format_datetime_short(payout.get("assigned_at") or payout.get("updated_at"))
                extra = f" ⏳ {code_disp}{f' — {ts}' if ts else ''}"
            elif status == "approved":
                ts = format_datetime_short(payout.get("updated_at"))
                extra = f" 💶 {code_disp}{f' — {ts}' if ts else ''}"
            elif status == "confirmed":
                ts = format_datetime_short(payout.get("confirmed_at"))
                extra = f" 💸 {code_disp}{f' — {ts}' if ts else ''}"
            elif status == "closed":
                ts = format_datetime_short(payout.get("updated_at"))
                extra = f" 📭 {code_disp}{f' — {ts}' if ts else ''}"
        elif r.get("paid") is True:
            ts = format_datetime_short(r.get("paid_at"))
            code = r.get("paid_request_code")
            if code:
                extra = f" 💸 {h(code)}{f' — {ts}' if ts else ''}"
            elif ts:
                extra = f" 💸 {ts}"
        elif isinstance(r.get("payout_history"), list) and r["payout_history"]:
            last_event = r["payout_history"][-1]
            if isinstance(last_event, dict) and last_event.get("status") == "closed":
                code = last_event.get("code") or last_event.get("request_id")
                ts = format_datetime_short(last_event.get("timestamp"))
                extra = f" 📭 {h(code) if code else 'Запрос'}{f' — {ts}' if ts else ''}"
        lines.append(base + extra)
    await clear_then_anchor(uid, "\n".join(lines), kb_checks(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("userpaid_"))
async def userpaid_set(c: types.CallbackQuery):
    uid = c.from_user.id
    try:
        action, proj, token = c.data.split(":", 2)
    except:
        return await c.answer("Ошибка")
    prof = load_user(uid) or {}
    recs = prof.get("receipts", {}).get(proj, [])
    changed = False
    for r in recs:
        rid = r.get("receipt_no") or ""
        fname = r.get("file") or ""
        if token and token not in (rid, fname):
            continue
        payout_status = (r.get("payout") or {}).get("status") if isinstance(r.get("payout"), dict) else None
        if payout_status in ("pending", "approved"):
            return await c.answer("Запрос по этому чеку уже в обработке.", show_alert=True)
        new_value = True if action.endswith("yes") else False
        r["paid"] = new_value
        now_iso = datetime.now().isoformat()
        if new_value:
            r["paid_at"] = now_iso
        else:
            r.pop("paid_at", None)
        r.pop("paid_request_id", None)
        r.pop("paid_request_code", None)
        if isinstance(r.get("payout_history"), list):
            r["payout_history"].append({
                "status": "manual_paid" if new_value else "manual_unpaid",
                "timestamp": now_iso,
                "project": proj,
                "amount": float(r.get("sum") or 0.0)
            })
        r["payout"] = None if r.get("payout") else None
        changed = True
        break
    if changed:
        prof.setdefault("receipts", {})[proj] = recs
        save_user(prof)
        close_kb = InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
        caption_body = format_receipt_caption(r)
        prefix = extract_receipt_prefix(c.message)
        new_caption = f"{prefix}\n{caption_body}" if prefix else caption_body
        if c.message:
            try:
                await bot.edit_message_caption(
                    c.message.chat.id,
                    c.message.message_id,
                    caption=new_caption,
                    reply_markup=close_kb
                )
            except Exception:
                try:
                    await bot.edit_message_text(
                        new_caption,
                        c.message.chat.id,
                        c.message.message_id,
                        reply_markup=close_kb,
                        disable_web_page_preview=True
                    )
                except Exception:
                    pass
        await update_all_anchors()
        await c.answer("Статус обновлён")
    else:
        await c.answer("Не найден чек или статус уже установлен", show_alert=True)


@dp.callback_query_handler(lambda c: c.data == "check_add")
async def check_add(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    await flow_clear(uid)
    tip = await bot.send_message(
        c.message.chat.id,
        (
            "📸 <b>Шаг 1 из 4.</b>\n"
            "Пришлите один чёткий снимок чека. После загрузки мы попросим указать сумму, описание и статус оплаты.\n\n"
            "Если передумали — нажмите «Отменить»."
        ),
        reply_markup=kb_receipt_cancel()
    )
    flow_track(uid, tip)
    await state.update_data(tmp_img=None, amount=None, photo_set=False, replace_photo=False, desc="", paid=None)
    await remember_step_prompt(state, tip)
    await ReceiptFSM.waiting_photo.set()
    await c.answer()


@dp.message_handler(commands=["cancel"], state="*")
async def cancel_any(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    current_state = await state.get_state()
    if current_state and current_state.startswith(ProfileEditFSM.__name__):
        runtime = profile_runtime(uid)
        remove_keyboard = bool(runtime.pop("profile_reply_keyboard", False))
        await profile_abort(uid, state, remove_keyboard=remove_keyboard)
        try:
            await bot.delete_message(m.chat.id, m.message_id)
        except Exception:
            pass
        return
    if current_state and current_state.startswith(ReceiptFSM.__name__):
        await remove_preview_message(state)
        await clear_edit_prompt(state)
        await clear_step_prompt(state)
    await state.finish()
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    await flow_clear(uid)
    if current_state and current_state.startswith(PhotoFSM.__name__):
        text = (
            "🚫 <b>Загрузка фотографий прервана.</b>\n"
            "Вы всегда можете вернуться в раздел «🖼 Фотохронология», чтобы продолжить." 
        )
        await anchor_show_text(uid, text, kb_photos(uid))
    else:
        text = (
            "❌ <b>Действие отменено.</b>\n"
            "Чтобы возобновить работу с чеками, откройте раздел «🧾 Чеки» заново."
        )
        await anchor_show_text(uid, text, kb_checks(uid))


@dp.callback_query_handler(lambda c: c.data == "cancel_receipt", state="*")
async def cancel_receipt_btn(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await remove_preview_message(state)
    await clear_edit_prompt(state)
    await clear_step_prompt(state)
    await state.finish(); await flow_clear(uid)
    await anchor_show_text(uid, tr(uid, "CHECKS_SECTION_TITLE"), kb_checks(uid))
    await c.answer("Отменено.")


@dp.message_handler(content_types=ContentType.PHOTO, state=ReceiptFSM.waiting_photo)
async def rcp_photo(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    data = await state.get_data()
    replacing = bool(data.get("replace_photo"))
    already = bool(data.get("photo_set"))
    editing = data.get("editing")

    if already and not replacing:
        warn = await bot.send_message(
            m.chat.id,
            "⚠️ Для одного чека можно прикрепить только один снимок. Используйте «🖼 Заменить фото» или команду /cancel, чтобы начать заново."
        )
        flow_track(uid, warn)
        try: await bot.delete_message(m.chat.id, m.message_id)
        except: pass
        return

    await clear_step_prompt(state)
    tmp = f"tmp_{uid}.jpg"
    await m.photo[-1].download(destination_file=tmp)
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass

    await state.update_data(tmp_img=tmp, photo_set=True, replace_photo=False)
    if editing == "photo":
        await state.update_data(editing=None)
        await send_receipt_preview(uid, m.chat.id, state)
        await ReceiptFSM.preview.set()
        return

    if data.get("amount") is None:
        ask = await bot.send_message(
            m.chat.id,
            "💰 <b>Шаг 2 из 4.</b> Укажите сумму чека в гривнах (пример: 123.45). Используйте точку в качестве разделителя копеек.",
            reply_markup=kb_receipt_cancel(),
        )
        flow_track(uid, ask)
        await remember_step_prompt(state, ask)
        await ReceiptFSM.waiting_amount.set()
    else:
        tip = await bot.send_message(
            m.chat.id,
            "📝 Хотите добавить описание к чеку? Отправьте текст или нажмите «Пропустить».",
            reply_markup=kb_desc_prompt(),
        )
        flow_track(uid, tip)
        await remember_step_prompt(state, tip)
        await ReceiptFSM.waiting_description.set()


@dp.message_handler(lambda m: m.content_type != ContentType.PHOTO, state=ReceiptFSM.waiting_photo, content_types=ContentType.ANY)
async def rcp_photo_reject(m: types.Message, state: FSMContext):
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass


@dp.message_handler(state=ReceiptFSM.waiting_amount, content_types=ContentType.TEXT)
async def rcp_amount(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    txt = (m.text or "").strip().replace(",", ".")
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    try:
        val = float(txt)
    except Exception:
        warn = await bot.send_message(
            m.chat.id,
            "❗ Не удалось распознать сумму. Введите число в формате 123.45 — точка отделяет гривны от копеек."
        )
        flow_track(uid, warn); return

    data = await state.get_data()
    editing = data.get("editing")
    tmp = data.get("tmp_img")
    if not tmp:
        warn = await bot.send_message(
            m.chat.id,
            "⚠️ Сначала загрузите фотографию чека — без неё мы не сможем сохранить запись."
        )
        flow_track(uid, warn); return
    await state.update_data(amount=val)
    await clear_step_prompt(state)
    if editing == "amount":
        await state.update_data(editing=None)
        await send_receipt_preview(uid, m.chat.id, state)
        await ReceiptFSM.preview.set()
        return
    tip = await bot.send_message(
        m.chat.id,
        "📝 <b>Шаг 3 из 4.</b> Добавьте краткое описание (например, цель покупки) или нажмите «Пропустить».",
        reply_markup=kb_desc_prompt(),
    )
    flow_track(uid, tip)
    await remember_step_prompt(state, tip)
    await ReceiptFSM.waiting_description.set()


@dp.callback_query_handler(lambda c: c.data == "desc_skip", state=ReceiptFSM.waiting_description)
async def desc_skip(c: types.CallbackQuery, state: FSMContext):
    await clear_step_prompt(state)
    await state.update_data(desc="")
    kb = kb_choose_paid(ask_later=True, flow_cancel=True)
    msg = await bot.send_message(
        c.message.chat.id,
        "🔖 <b>Шаг 4 из 4.</b> Укажите статус оплаты для этого чека.",
        reply_markup=kb
    )
    flow_track(c.from_user.id, msg)
    await remember_step_prompt(state, msg)
    await ReceiptFSM.waiting_paid_choice.set()
    await c.answer()


@dp.message_handler(state=ReceiptFSM.waiting_description, content_types=ContentType.TEXT)
async def rcp_desc(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    desc = (m.text or "").strip()
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    data = await state.get_data()
    editing = data.get("editing")
    await state.update_data(desc=desc)
    await clear_step_prompt(state)
    if editing == "desc":
        await state.update_data(editing=None)
        await send_receipt_preview(uid, m.chat.id, state)
        await ReceiptFSM.preview.set()
        return
    kb = kb_choose_paid(ask_later=True, flow_cancel=True)
    msg = await bot.send_message(
        m.chat.id,
        "🔖 <b>Шаг 4 из 4.</b> Выберите статус оплаты для этого чека.",
        reply_markup=kb
    )
    flow_track(uid, msg)
    await remember_step_prompt(state, msg)
    await ReceiptFSM.waiting_paid_choice.set()


@dp.callback_query_handler(lambda c: c.data == "edit_cancel", state=[ReceiptFSM.waiting_amount, ReceiptFSM.waiting_description, ReceiptFSM.waiting_photo])
async def edit_cancel_action(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await clear_edit_prompt(state)
    await state.update_data(editing=None, replace_photo=False)
    await send_receipt_preview(uid, c.message.chat.id, state)
    await ReceiptFSM.preview.set()
    await c.answer("Отменено")


@dp.callback_query_handler(lambda c: c.data in ("paid_yes","paid_no","paid_later"), state=ReceiptFSM.waiting_paid_choice)
async def paid_choice(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    data = await state.get_data()
    paid = None
    if c.data == "paid_yes": paid = True
    elif c.data == "paid_no": paid = False
    await clear_step_prompt(state)
    await state.update_data(paid=paid)
    await send_receipt_preview(uid, c.message.chat.id, state)
    await ReceiptFSM.preview.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data in ("edit_amount", "edit_photo", "edit_desc", "edit_paid", "save_receipt", "cancel_receipt"), state=ReceiptFSM.preview)
async def rcp_preview_actions(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    data = await state.get_data()

    if c.data == "edit_amount":
        await clear_edit_prompt(state)
        await state.update_data(editing="amount")
        msg = await bot.send_message(
            c.message.chat.id,
            "✏️ Укажите новую сумму для чека (пример: 123.45).",
            reply_markup=kb_edit_cancel()
        )
        flow_track(uid, msg)
        await state.update_data(edit_prompt=(c.message.chat.id, msg.message_id))
        await ReceiptFSM.waiting_amount.set()
        return await c.answer()

    if c.data == "edit_photo":
        await clear_edit_prompt(state)
        await state.update_data(replace_photo=True, photo_set=False, editing="photo")
        msg = await bot.send_message(
            c.message.chat.id,
            "🖼 Пришлите обновлённое фото чека.",
            reply_markup=kb_edit_cancel()
        )
        flow_track(uid, msg)
        await state.update_data(edit_prompt=(c.message.chat.id, msg.message_id))
        await ReceiptFSM.waiting_photo.set()
        return await c.answer()

    if c.data == "edit_desc":
        await clear_edit_prompt(state)
        await state.update_data(editing="desc")
        msg = await bot.send_message(
            c.message.chat.id,
            "📝 Введите новое описание для этого чека.",
            reply_markup=kb_edit_cancel()
        )
        flow_track(uid, msg)
        await state.update_data(edit_prompt=(c.message.chat.id, msg.message_id))
        await ReceiptFSM.waiting_description.set()
        return await c.answer()

    if c.data == "edit_paid":
        await clear_edit_prompt(state)
        kb = kb_choose_paid(ask_later=True, allow_cancel=True)
        msg = await bot.send_message(
            c.message.chat.id,
            "🔖 Обновите статус оплаты для чека с помощью кнопок ниже.",
            reply_markup=kb
        )
        flow_track(uid, msg)
        await remember_step_prompt(state, msg)
        await ReceiptFSM.waiting_paid_choice.set()
        return await c.answer()

    if c.data == "cancel_receipt":
        await remove_preview_message(state)
        await clear_edit_prompt(state)
        await flow_clear(uid); await state.finish()
        await anchor_show_text(uid, tr(uid, "CHECKS_SECTION_TITLE"), kb_checks(uid))
        return await c.answer("Отменено.")

    if c.data == "save_receipt":
        proj = active_project["name"]
        if not proj:
            await c.answer("❗ Нет активного проекта", show_alert=True); return
        if not data.get("tmp_img") or data.get("amount") is None:
            await c.answer("⚠️ Не хватает данных (фото/сумма).", show_alert=True); return
        fname, path, now, rid = save_receipt(proj, uid, float(data["amount"]), data.get("tmp_img"), data.get("desc",""), data.get("paid"))
        await remove_preview_message(state)
        await clear_edit_prompt(state)
        await clear_step_prompt(state)
        await flow_clear(uid)
        await state.finish()
        status_txt = receipt_status_text(data.get('paid'))
        desc_value = data.get('desc')
        desc_text = h(desc_value) if desc_value else "—"
        caption = (
            "✅ Чек сохранён!\n"
            "━━━━━━━━━━━━━━━━━━\n"
            f"📂 Проект: <b>{h(proj)}</b>\n"
            f"🆔 Номер: <b>{h(rid)}</b>\n"
            f"📅 Дата сохранения: {now.strftime('%Y-%m-%d %H:%M')}\n"
            f"💰 Сумма: {fmt_money(float(data['amount']))} грн\n"
            f"📝 Описание: {desc_text}\n"
            f"🔖 Статус: {status_txt}\n"
            f"📄 Файл: {h(fname)}\n\n"
            "Карточка доступна в разделе «📁 Мои чеки». Нажмите «Закрыть», когда ознакомитесь."
        )
        await bot.send_photo(c.message.chat.id, InputFile(path), caption=caption, reply_markup=kb_saved_receipt())
        await anchor_show_text(uid, tr(uid, "CHECKS_SECTION_TITLE"), kb_checks(uid))
        return await c.answer("Сохранено.")


# ========================== PHOTO TIMELINE ==========================
@dp.callback_query_handler(lambda c: c.data == "menu_photos")
async def menu_photos(c: types.CallbackQuery):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    info = load_project_info(active_project["name"])
    text = (
        "🖼 <b>Фотохронология объекта</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"📂 Проект: <b>{h(info.get('name', '—'))}</b>\n"
        f"🆔 Код объекта: {h(info.get('code') or '—')}\n\n"
        "Здесь хранится визуальная история работ: загружайте новые фото или просматривайте архивные снимки.\n"
        "В архиве можно отправить оригинал файла и удалить свои снимки при необходимости."
    )
    await clear_then_anchor(uid, text, kb_photos(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "photo_upload")
async def photo_upload(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    project = active_project["name"]
    info = load_project_info(project)
    await flow_clear(uid)
    await state.finish()
    await state.update_data(photo_project=project, uploaded=[], photo_session_message=None)
    await _photo_refresh_session_message(c.message.chat.id, uid, state, info, [])
    await PhotoFSM.collecting.set()
    await c.answer()


def _detect_extension(name: str, fallback: str = ".jpg") -> str:
    ext = (os.path.splitext(name)[1] or "").lower()
    if not ext:
        return fallback
    return ext


def _should_send_as_photo(ext: str) -> bool:
    return ext.lower() in {".jpg", ".jpeg", ".png", ".webp"}


@dp.message_handler(state=PhotoFSM.collecting, content_types=[ContentType.PHOTO, ContentType.DOCUMENT])
async def photo_collect_media(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    data = await state.get_data()
    project = data.get("photo_project")
    if not project:
        await state.finish()
        return
    info = load_project_info(project)
    prof = load_user(uid) or {"fullname": f"User{uid}", "bsu": "—"}
    now = datetime.now()

    if m.content_type == ContentType.PHOTO:
        original_name = f"photo_{now.strftime('%Y%m%d_%H%M%S')}.jpg"
        ext = ".jpg"
        file_obj = m.photo[-1]
    else:
        doc = m.document
        original_name = doc.file_name or f"file_{now.strftime('%Y%m%d_%H%M%S')}"
        ext = _detect_extension(original_name, fallback=".jpg")
        if ext.lower() not in ALLOWED_IMAGE_EXTENSIONS:
            warn = await bot.send_message(
                m.chat.id,
                "⚠️ Можно загружать только изображения (JPG, PNG, WEBP, BMP, TIF, HEIC).",
                reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
            )
            flow_track(uid, warn)
            try:
                await bot.delete_message(m.chat.id, m.message_id)
            except Exception:
                pass
            return
        file_obj = doc

    base_name = os.path.splitext(original_name)[0]
    safe_original = _sanitize_filename(base_name) or f"photo_{now.strftime('%Y%m%d_%H%M%S')}"
    safe_user = _sanitize_filename(prof.get("fullname") or f"User{uid}")
    project_code = info.get("code") or project
    prefix = f"{project_code}_{safe_original}_UID{uid}"
    if safe_user:
        prefix += f"_{safe_user}"
    filename = f"{prefix}{ext.lower()}"
    target_dir = proj_photos_dir(project)
    os.makedirs(target_dir, exist_ok=True)
    counter = 1
    while os.path.exists(os.path.join(target_dir, filename)):
        filename = f"{prefix}_{counter}{ext.lower()}"
        counter += 1
    dst_path = os.path.join(target_dir, filename)

    try:
        if m.content_type == ContentType.PHOTO:
            await file_obj.download(destination_file=dst_path)
        else:
            await file_obj.download(destination_file=dst_path)
        try:
            await bot.delete_message(m.chat.id, m.message_id)
        except Exception:
            pass
    except Exception as exc:
        warn = await bot.send_message(
            m.chat.id,
            "⚠️ Не удалось сохранить файл. Попробуйте ещё раз или отправьте другой формат.",
            reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
        )
        flow_track(uid, warn)
        return

    metadata = extract_image_metadata(dst_path)
    entry = {
        "id": generate_photo_id(),
        "file": filename,
        "original": original_name,
        "uploaded_at": now.isoformat(),
        "uploader_id": uid,
        "uploader_name": prof.get("fullname", "—"),
        "uploader_bsu": prof.get("bsu", "—"),
        "meta": metadata,
    }
    photos = load_project_photos(project)
    photos.append(entry)
    save_project_photos(project, photos)

    uploaded = list(data.get("uploaded") or [])
    uploaded.append(entry)
    await state.update_data(uploaded=uploaded)
    await _photo_refresh_session_message(m.chat.id, uid, state, info, uploaded, entry)
    await update_all_anchors()


@dp.message_handler(state=PhotoFSM.collecting, content_types=ContentType.TEXT)
async def photo_collect_text(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    data = await state.get_data()
    uploaded = data.get("uploaded") or []
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    hint = await bot.send_message(
        m.chat.id,
        "ℹ️ Отправьте фотографию или изображение документом. Когда закончите, нажмите «✅ Завершить загрузку».",
        reply_markup=kb_photo_session_controls(bool(uploaded))
    )
    flow_track(uid, hint)


@dp.callback_query_handler(lambda c: c.data == "photo_session_preview", state=PhotoFSM.collecting)
async def photo_session_preview(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    data = await state.get_data()
    uploaded = data.get("uploaded") or []
    if not uploaded:
        return await c.answer("Пока нет загруженных файлов", show_alert=True)
    project = data.get("photo_project")
    if not project:
        return await c.answer("Проект не найден", show_alert=True)
    await c.answer("Отправляю файлы…")
    header = await bot.send_message(
        c.message.chat.id,
        f"🗂 <b>Загруженные в этой сессии</b>\nВсего файлов: <b>{len(uploaded)}</b>.",
        reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
    )
    flow_track(uid, header)
    base_dir = proj_photos_dir(project)
    for entry in uploaded:
        stored = entry.get("file") or ""
        original = entry.get("original") or stored
        if not stored:
            warn = await bot.send_message(
                c.message.chat.id,
                f"⚠️ Не удалось определить файл для {h(original)}.",
                reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
            )
            flow_track(uid, warn)
            continue
        path = os.path.join(base_dir, stored)
        if not os.path.exists(path):
            warn = await bot.send_message(
                c.message.chat.id,
                f"⚠️ Файл {h(stored)} не найден на диске.",
                reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
            )
            flow_track(uid, warn)
            continue
        caption = f"📁 {h(original)}"
        kb = InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
        try:
            ext = os.path.splitext(stored)[1].lower()
            file_input = InputFile(path)
            if _should_send_as_photo(ext):
                msg = await bot.send_photo(c.message.chat.id, file_input, caption=caption, reply_markup=kb)
            else:
                msg = await bot.send_document(c.message.chat.id, file_input, caption=caption, reply_markup=kb)
            flow_track(uid, msg)
        except Exception:
            warn = await bot.send_message(
                c.message.chat.id,
                f"⚠️ Не удалось отправить {h(original)}.",
                reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
            )
            flow_track(uid, warn)


@dp.callback_query_handler(lambda c: c.data == "photo_finish", state=PhotoFSM.collecting)
async def photo_finish(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    data = await state.get_data()
    uploaded = data.get("uploaded") or []
    await state.finish()
    summary = (
        "🖼 <b>Фотосессия завершена</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"Добавлено файлов: <b>{len(uploaded)}</b>.\n"
        "Архив можно открыть через «🖼 Просмотреть фотографии объекта».\n"
        "При просмотре вы можете удалить свои снимки или запросить оригиналы при необходимости."
    )
    await clear_then_anchor(uid, summary, kb_photos(uid))
    await c.answer("Готово")


@dp.callback_query_handler(lambda c: c.data == "photo_cancel", state=PhotoFSM.collecting)
async def photo_cancel(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    data = await state.get_data()
    uploaded = len(data.get("uploaded") or [])
    await state.finish()
    note = (
        "🚫 <b>Загрузка остановлена</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        "Новые файлы больше не принимаются."
    )
    if uploaded:
        note += f"\nУже сохранено: <b>{uploaded}</b>."
    await clear_then_anchor(uid, note, kb_photos(uid))
    await c.answer("Остановлено")


@dp.callback_query_handler(lambda c: c.data == "photo_view")
async def photo_view(c: types.CallbackQuery):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    project = active_project["name"]
    info = load_project_info(project)
    photos = load_project_photos(project)
    if not photos:
        text = (
            "🖼 <b>Фотохронология</b>\n"
            "━━━━━━━━━━━━━━━━━━\n"
            "Для этого проекта ещё не загружено ни одного изображения."
        )
        await clear_then_anchor(uid, text, kb_photos(uid))
        await c.answer()
        return

    header = (
        "🖼 <b>Архив фотографий</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"📂 Проект: <b>{h(info.get('name', '—'))}</b> ({h(info.get('code') or '—')})\n"
        f"📸 Всего снимков: <b>{len(photos)}</b>."
    )
    await clear_then_anchor(uid, header, kb_photos(uid))

    for entry in photos:
        stored = entry.get("file") or ""
        path = os.path.join(proj_photos_dir(project), stored) if stored else ""
        file_exists = bool(stored and os.path.exists(path))
        caption = format_photo_caption(info, entry)
        kb = photo_entry_keyboard(project, entry, uid, file_exists=file_exists)
        if file_exists:
            ext = os.path.splitext(stored)[1].lower()
            try:
                if _should_send_as_photo(ext):
                    msg = await bot.send_photo(c.message.chat.id, InputFile(path), caption=caption, reply_markup=kb)
                else:
                    msg = await bot.send_document(c.message.chat.id, InputFile(path), caption=caption, reply_markup=kb)
                flow_track(uid, msg)
                continue
            except Exception:
                file_exists = False
                kb = photo_entry_keyboard(project, entry, uid, file_exists=False)
        fallback_text = caption + "\n\n⚠️ Не удалось отобразить файл в Telegram, но запись остаётся в архиве."
        warn = await bot.send_message(c.message.chat.id, fallback_text, reply_markup=kb)
        flow_track(uid, warn)
    footer = (
        "📁 <b>Просмотр архива завершён</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        "Выберите дальнейшее действие с помощью кнопок ниже."
    )
    tail = await bot.send_message(c.message.chat.id, footer, reply_markup=kb_photo_view_actions())
    flow_track(uid, tail)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data in {"photo_view_close", "photo_view_root", "photo_view_menu"})
async def photo_view_controls(c: types.CallbackQuery):
    uid = c.from_user.id
    action = c.data
    if action == "photo_view_menu":
        await menu_photos(c)
        return
    if action == "photo_view_root":
        await c.answer("Главное меню открыто")
        await flow_clear(uid)
        await anchor_show_root(uid)
    else:
        await c.answer("Просмотр закрыт")
        await flow_clear(uid)


@dp.callback_query_handler(lambda c: c.data.startswith("photo_original:"))
async def photo_send_original(c: types.CallbackQuery):
    uid = c.from_user.id
    parts = c.data.split(":", 2)
    if len(parts) != 3:
        return await c.answer("Некорректный запрос", show_alert=True)
    _, token, entry_id = parts
    project = project_from_token(token)
    if not project:
        return await c.answer("Проект не найден", show_alert=True)
    _, entry, _ = find_photo_entry(project, entry_id)
    if not entry:
        return await c.answer("Фотография уже удалена.", show_alert=True)
    file_name = entry.get("file")
    if not file_name:
        return await c.answer("Оригинал отсутствует", show_alert=True)
    path = os.path.join(proj_photos_dir(project), file_name)
    if not os.path.exists(path):
        return await c.answer("Файл не найден", show_alert=True)
    original_name = entry.get("original") or os.path.basename(path)
    try:
        msg = await bot.send_document(
            c.message.chat.id,
            InputFile(path, filename=original_name),
            caption=f"📤 Оригинал: {h(original_name)}",
            reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
        )
        flow_track(uid, msg)
        await c.answer("Оригинал отправлен")
    except Exception:
        await c.answer("Не удалось отправить файл", show_alert=True)


@dp.callback_query_handler(lambda c: c.data.startswith("photo_delete:"))
async def photo_delete(c: types.CallbackQuery):
    uid = c.from_user.id
    parts = c.data.split(":", 2)
    if len(parts) != 3:
        return await c.answer("Некорректный запрос", show_alert=True)
    _, token, entry_id = parts
    project = project_from_token(token)
    if not project:
        return await c.answer("Проект не найден", show_alert=True)
    index, entry, photos = find_photo_entry(project, entry_id)
    if entry is None or index < 0:
        return await c.answer("Фотография уже удалена", show_alert=True)
    uploader_id = entry.get("uploader_id")
    if uid not in admins and uid != uploader_id:
        return await c.answer("Удаление недоступно", show_alert=True)

    file_name = entry.get("file")
    if file_name:
        path = os.path.join(proj_photos_dir(project), file_name)
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass

    photos.pop(index)
    save_project_photos(project, photos)

    try:
        await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception:
        pass

    await c.answer("Фотография удалена")
    await update_all_anchors()

    if uploader_id and uploader_id != uid:
        uploader_profile = load_user(uploader_id) or {}
        notify_chat = (
            users_runtime.get(uploader_id, {}).get("tg", {}).get("chat_id")
            or uploader_profile.get("tg", {}).get("chat_id")
        )
        if notify_chat:
            actor_profile = load_user(uid) or {}
            actor_name = actor_profile.get("fullname") or actor_profile.get("tg", {}).get("first_name") or f"ID {uid}"
            info = load_project_info(project)
            note = (
                "🗑 <b>Фото удалено администратором</b>\n"
                "━━━━━━━━━━━━━━━━━━\n"
                f"📂 Проект: <b>{h(info.get('name', '—'))}</b> ({h(info.get('code') or '—')})\n"
                f"📛 Оригинал: {h(entry.get('original') or entry.get('file') or '—')}\n"
                f"Удалил: {h(actor_name)} (ID {uid})"
            )
            kb = InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
            try:
                msg = await bot.send_message(notify_chat, note, reply_markup=kb)
                flow_track(uploader_id, msg)
            except Exception:
                pass


# ========================== DOCUMENTS ==========================
@dp.callback_query_handler(lambda c: c.data == "menu_docs")
async def menu_docs(c: types.CallbackQuery):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    proj = active_project["name"]
    folder = proj_pdf_dir(proj)
    pdfs = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")] if os.path.exists(folder) else []
    if not pdfs:
        await clear_then_anchor(
            uid,
            f"📑 Документы проекта <b>{h(proj)}</b>\n━━━━━━━━━━━━━━━━━━\nПока нет загруженных файлов. Обратитесь к администратору, если ожидаете документацию.",
            kb=InlineKeyboardMarkup().add(InlineKeyboardButton("⬅️ На главную", callback_data="back_root"))
        )
        return await c.answer()
    await clear_then_anchor(
        uid,
        f"📑 Документы проекта <b>{h(proj)}</b>\n━━━━━━━━━━━━━━━━━━\nДоступно файлов: <b>{len(pdfs)}</b>. Откройте нужный документ из списка ниже.",
        kb=InlineKeyboardMarkup().add(InlineKeyboardButton("⬅️ На главную", callback_data="back_root"))
    )
    for f in pdfs[:10]:
        p = os.path.join(folder, f)
        if os.path.exists(p):
            msg = await bot.send_document(c.message.chat.id, InputFile(p), caption=h(f))
            flow_track(uid, msg)
    await c.answer()


# ========================== FINANCE (USER) ==========================
def user_has_approved_not_confirmed(uid: int) -> bool:
    prof = load_user(uid) or {}
    for ref in iter_user_payout_refs(prof):
        obj = finance_load_request(ref.get("id"), ref.get("project"))
        if obj and obj.get("status") == "approved" and obj.get("user_id") == uid:
            return True
    return False


@dp.callback_query_handler(lambda c: c.data == "menu_finance")
async def finance_menu(c: types.CallbackQuery):
    uid = c.from_user.id
    if not active_project["name"]:
        return await c.answer("❗ Нет активного проекта", show_alert=True)
    proj = active_project["name"]
    stats = user_project_stats(uid, proj)
    lines = [
        "💵 <b>Финансовый раздел</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"📂 Проект: <b>{h(proj)}</b>",
        f"🧾 Загружено чеков: <b>{stats['count']}</b>",
        f"💰 Общая сумма: <b>{fmt_money(stats['total'])} грн</b>",
        f"✅ Оплачено фирмой: <b>{fmt_money(stats['paid'])} грн</b>",
        f"❌ Ожидает оплаты: <b>{fmt_money(stats['unpaid'])} грн</b>"
    ]
    if stats["pending"]:
        lines.append(f"⏳ Уже в запросах: <b>{fmt_money(stats['pending'])} грн</b>")
    if stats["unspecified"]:
        lines.append(f"❔ Без статуса оплаты: <b>{fmt_money(stats['unspecified'])} грн</b>")
    alerts: List[str] = []
    active_req = finance_active_request_for_user(uid, proj)
    if active_req:
        status = active_req.get("status")
        status_human = {"pending": "ожидает подтверждения", "approved": "одобрена"}.get(status, status or "в обработке")
        code = active_req.get("code", active_req.get("id"))
        alerts.append(f"📨 Активный запрос: <b>{h(code)}</b> — {h(status_human)}")
    if user_has_approved_not_confirmed(uid):
        alerts.insert(0, "⚠️ Есть одобренные выплаты. Подтвердите получение денег через соответствующий пункт меню.")
    if alerts:
        lines.append("")
        lines.extend(alerts)
    lines.append("")
    lines.append("Выберите действие ниже, чтобы посмотреть детали чеков, отправить запрос на выплату или подтвердить получение средств.")
    text = "\n".join(lines)
    await clear_then_anchor(uid, text, kb_finance_root(user_has_pending_confirm=user_has_approved_not_confirmed(uid)))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "fin_unpaid_list")
async def finance_unpaid_list(c: types.CallbackQuery):
    uid = c.from_user.id
    proj = active_project["name"]
    if not proj: return await c.answer("❗ Нет активного проекта", show_alert=True)
    recs = user_project_receipts(uid, proj)
    unpaid: List[dict] = []
    pending: List[dict] = []
    for r in recs:
        if r.get("paid") is False:
            payout_status = (r.get("payout") or {}).get("status") if isinstance(r.get("payout"), dict) else None
            if payout_status in ("pending", "approved"):
                pending.append(r)
            else:
                unpaid.append(r)
    if not unpaid and not pending:
        return await c.answer("Неоплаченных чеков нет 🎉", show_alert=True)
    lines = [
        "⏳ <b>Неоплаченные чеки и запросы</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"📂 Проект: <b>{h(proj)}</b>",
        ""
    ]
    if unpaid:
        lines.append(f"❌ Готовы к запросу ({len(unpaid)} шт.):")
        total_unpaid = 0.0
        for r in unpaid:
            try:
                amount = float(r.get("sum") or 0.0)
            except (TypeError, ValueError):
                amount = 0.0
            total_unpaid += amount
            moment = f"{h(r.get('date','—'))} {h(r.get('time',''))}".strip()
            desc = r.get('desc')
            desc_text = h(desc) if desc else "—"
            rid = h(r.get('receipt_no', '—'))
            lines.append(f"• {moment} — {fmt_money(amount)} грн — {desc_text} — #{rid}")
        lines.append(f"Итого к запросу: <b>{fmt_money(total_unpaid)} грн</b>")
        lines.append("")
    if pending:
        lines.append(f"⏳ Уже в запросах ({len(pending)} шт.):")
        total_pending = 0.0
        for r in pending:
            try:
                amount = float(r.get("sum") or 0.0)
            except (TypeError, ValueError):
                amount = 0.0
            total_pending += amount
            moment = f"{h(r.get('date','—'))} {h(r.get('time',''))}".strip()
            code = ((r.get("payout") or {}).get("code") or (r.get("payout") or {}).get("request_id")) if isinstance(r.get("payout"), dict) else None
            lines.append(f"• {moment} — {fmt_money(amount)} грн — запрос {h(code) if code else '—'}")
        lines.append(f"Всего в запросах: <b>{fmt_money(total_pending)} грн</b>")
    lines.append("")
    lines.append("Подайте новый запрос на выплату, чтобы закрыть чеки из раздела «Готовы к запросу».")
    await clear_then_anchor(uid, "\n".join(lines), kb_finance_root(user_has_pending_confirm=user_has_approved_not_confirmed(uid)))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "fin_request_payout")
async def finance_request_payout(c: types.CallbackQuery):
    uid = c.from_user.id
    proj = active_project["name"]
    if not proj: return await c.answer("❗ Нет активного проекта", show_alert=True)
    existing = finance_active_request_for_user(uid, proj)
    if existing:
        code = existing.get("code", existing.get("id"))
        status = existing.get("status")
        status_human = {"pending": "ожидает подтверждения", "approved": "одобрена"}.get(status, status or "в обработке")
        await clear_then_anchor(
            uid,
            (
                "📨 <b>Запрос уже в обработке</b>\n"
                f"Код: <b>{h(code)}</b>\n"
                f"Текущий статус: {h(status_human)}.\n\n"
                "Дождитесь ответа администратора или подтвердите получение средств в разделе подтверждений."
            ),
            kb_finance_root(user_has_pending_confirm=user_has_approved_not_confirmed(uid))
        )
        return await c.answer("Есть активный запрос на выплату.", show_alert=True)
    recs = user_project_receipts(uid, proj)
    eligible: List[dict] = []
    locked: List[dict] = []
    for r in recs:
        if r.get("paid") is False:
            payout_status = (r.get("payout") or {}).get("status") if isinstance(r.get("payout"), dict) else None
            if payout_status in ("pending", "approved"):
                locked.append(r)
            else:
                eligible.append(r)
    if not eligible:
        if locked:
            message_text = (
                "Все неоплаченные чеки уже добавлены в действующие запросы.\n"
                "Ожидайте подтверждения администратора или уведомления о выплате."
            )
        else:
            message_text = (
                "Фирма закрыла все ваши чеки — неоплаченных сумм не осталось.\n"
                "Добавьте новые чеки, чтобы сформировать следующий запрос."
            )
        await clear_then_anchor(
            uid,
            message_text,
            kb_finance_root(user_has_pending_confirm=user_has_approved_not_confirmed(uid))
        )
        return await c.answer("Нет чеков для запроса.", show_alert=True)
    req = finance_new_request(uid, proj, eligible)
    req_id = req["id"]
    req_code = req.get("code", req_id)
    total = float(req.get("sum") or 0.0)
    proj_info = load_project_info(proj)
    project_code_txt = h(proj_info.get('code') or '—')
    await c.answer("Запрос отправлен администратору.")

    prof = load_user(uid) or {}
    fullname = h(prof.get('fullname', '—'))
    bsu_code = h(prof.get('bsu', '—'))
    phone = h(prof.get('phone', '—'))
    username_raw = (prof.get('tg', {}) or {}).get('username')
    username_display = h(f"@{username_raw}" if username_raw else "—")
    receipts_line_parts = [h(r.get('receipt_no', '—')) for r in eligible[:10]]
    files_line = ", ".join(receipts_line_parts)
    if len(eligible) > 10:
        files_line += "…"
    region_txt = h(proj_info.get('region') or '—')
    location_txt = h(proj_info.get('location', '—'))
    req_code_disp = h(req_code)
    req_id_disp = h(req_id)
    text = (
        "📢 <b>Новый запрос на выплату</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"Код выплаты: <b>{req_code_disp}</b>\n"
        f"Файл: <code>{req_id_disp}</code>\n"
        f"👤 {fullname} (ID {uid}, {bsu_code})\n"
        f"📱 {phone}\n"
        f"🆔 {username_display}\n"
        f"📂 Проект: {h(proj)}\n"
        f"🆔 Код объекта: {project_code_txt}\n"
        f"🌍 Область: {region_txt}\n"
        f"📍 Локация: {location_txt}\n"
        f"❌ Неоплаченных чеков: {len(eligible)} шт.\n"
        f"💰 Сумма к выплате: <b>{fmt_money(total)} грн</b>\n"
        f"🧾 Номера чеков: {files_line}\n\n"
        "Используйте кнопки ниже, чтобы проверить карточки и подтвердить перечисление."
    )
    akb = InlineKeyboardMarkup()
    akb.add(InlineKeyboardButton("👀 Посмотреть чеки", callback_data=f"adm_req_view_checks:{req_id}"))
    akb.add(InlineKeyboardButton("✅ Выплатить", callback_data=f"adm_req_paid:{req_id}"))
    akb.add(InlineKeyboardButton("❌ Закрыть", callback_data=f"adm_req_close:{req_id}"))
    for ad in list(admins):
        chat_id = users_runtime.get(ad, {}).get("tg", {}).get("chat_id") or (load_user(ad) or {}).get("tg", {}).get("chat_id")
        if chat_id:
            try: await bot.send_message(chat_id, text, reply_markup=akb)
            except Exception: pass

    await clear_then_anchor(
        uid,
        (
            "📨 <b>Запрос на выплату отправлен</b>\n"
            f"Код: <b>{req_code_disp}</b>\n"
            f"Проект: {h(proj)} (код {project_code_txt})\n"
            f"Чеков в запросе: {len(eligible)} шт.\n"
            f"Сумма: <b>{fmt_money(total)} грн</b>\n\n"
            "Мы сообщим, когда администратор одобрит выплату или запросит уточнения."
        ),
        kb_finance_root(user_has_pending_confirm=user_has_approved_not_confirmed(uid))
    )


@dp.callback_query_handler(lambda c: c.data == "fin_history")
async def fin_history(c: types.CallbackQuery):
    uid = c.from_user.id
    prof = load_user(uid) or {}
    requests: List[dict] = []
    for ref in iter_user_payout_refs(prof):
        req = finance_load_request(ref.get("id"), ref.get("project"))
        if req and req.get("user_id") == uid:
            requests.append(req)
    if not requests:
        await clear_then_anchor(
            uid,
            "📚 История выплат пока пуста.\nОтправьте запрос на выплату, и здесь появятся все подтверждённые операции.",
            kb_finance_root(user_has_pending_confirm=user_has_approved_not_confirmed(uid))
        )
        return await c.answer()
    requests.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    status_map = {"pending": "В ожидании", "approved": "Одобрено", "confirmed": "Подтверждено", "closed": "Закрыто"}
    lines = [
        "📚 <b>История выплат</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"Записей найдено: <b>{len(requests)}</b>",
        "Нажмите на нужный код, чтобы открыть детали и посмотреть привязанные чеки.",
        ""
    ]
    kb = InlineKeyboardMarkup()
    for req in requests[:20]:
        code = req.get("code", req["id"])
        status_txt = status_map.get(req.get("status"), req.get("status", "—"))
        amount = fmt_money(float(req.get("sum") or 0.0))
        lines.append(f"• {h(code)} — {amount} грн — {h(status_txt)}")
        kb.add(InlineKeyboardButton(f"{code} • {status_txt}", callback_data=f"fin_hist_open:{req['id']}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="menu_finance"))
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("fin_hist_open:"))
async def fin_hist_open(c: types.CallbackQuery):
    uid = c.from_user.id
    req_id = c.data.split(":", 1)[1]
    obj = finance_load_request(req_id)
    if not obj or obj.get("user_id") != uid:
        return await c.answer("Запись не найдена", show_alert=True)
    code = obj.get("code", req_id)
    proj_info = load_project_info(obj.get("project")) if obj.get("project") else {}
    project_code_txt = h(proj_info.get("code") or "—")
    status_map = {"pending": "В ожидании", "approved": "Одобрено", "confirmed": "Подтверждено", "closed": "Закрыто"}
    status_disp = status_map.get(obj.get("status"), obj.get("status", "—"))
    lines = [
        f"💵 <b>Выплата {h(code)}</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"Статус: <b>{h(status_disp)}</b>",
        f"Сумма: <b>{fmt_money(float(obj.get('sum') or 0.0))} грн</b>",
        f"Проект: {h(obj.get('project', '—'))}",
        f"Связанных чеков: {len(obj.get('files', []))}",
        ""
    ]
    def fmt_ts(value: Optional[str]) -> str:
        if not value:
            return "—"
        return format_datetime_short(value) or "—"
    lines.extend([
        f"📅 Создано: {fmt_ts(obj.get('created_at'))}",
        f"✅ Одобрено администратором: {fmt_ts(obj.get('approved_at'))}",
        f"💸 Подтверждено пользователем: {fmt_ts(obj.get('confirmed_at'))}"
    ])
    lines.append("")
    lines.append("Используйте кнопки ниже, чтобы просмотреть привязанные чеки или вернуться к списку.")
    kb = InlineKeyboardMarkup()
    if obj.get("files"):
        kb.add(InlineKeyboardButton("🧾 Посмотреть чеки", callback_data=f"fin_hist_view:{req_id}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="fin_history"))
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("fin_hist_view:"))
async def fin_hist_view(c: types.CallbackQuery):
    uid = c.from_user.id
    req_id = c.data.split(":", 1)[1]
    obj = finance_load_request(req_id)
    if not obj or obj.get("user_id") != uid:
        return await c.answer("Запись не найдена", show_alert=True)
    files = obj.get("files", [])
    code = obj.get("code", req_id)
    recs = user_project_receipts(uid, obj.get("project"))
    by_file = {r.get("file"): r for r in recs}
    proj_info = load_project_info(obj.get("project")) if obj.get("project") else {}
    project_code_txt = h(proj_info.get("code") or "—")
    lines = [
        f"🧾 <b>Чеки выплаты {h(code)}</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"Проект: {h(obj.get('project','—'))} (код {project_code_txt})",
        f"Всего файлов: <b>{len(files)}</b>",
        ""
    ]
    for fname in files:
        r = by_file.get(fname)
        if r:
            amount = fmt_money(float(r.get("sum") or 0.0))
            desc = h(r.get("desc")) if r.get("desc") else "—"
            lines.append(f"• #{h(r.get('receipt_no','—'))} — {amount} грн — {desc}")
        else:
            lines.append(f"• {h(fname)}")
    kb = InlineKeyboardMarkup().add(InlineKeyboardButton("⬅️ Назад", callback_data=f"fin_hist_open:{req_id}"))
    await clear_then_anchor(uid, "\n".join(lines), kb)
    chat_id = c.message.chat.id
    for fname in files[:5]:
        r = by_file.get(fname)
        if r:
            msg = await send_receipt_card(chat_id, obj.get("project"), uid, r, include_project=False)
            flow_track(uid, msg)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "fin_confirm_list")
async def fin_confirm_list(c: types.CallbackQuery):
    uid = c.from_user.id
    prof = load_user(uid) or {}
    to_confirm = []
    for ref in iter_user_payout_refs(prof):
        obj = finance_load_request(ref.get("id"), ref.get("project"))
        if obj and obj.get("status") == "approved" and obj.get("user_id") == uid:
            to_confirm.append(obj)
    if not to_confirm:
        return await c.answer("Нет выплат, ожидающих подтверждения.")
    kb = InlineKeyboardMarkup()
    text_lines = [
        "✅ <b>Одобренные выплаты</b>",
        "━━━━━━━━━━━━━━━━━━",
        "Подтвердите получение средств по каждой записи, чтобы обновить статистику.",
        ""
    ]
    for o in to_confirm[:20]:
        code = o.get("code", o['id'])
        amount = float(o.get('sum') or 0.0)
        project_name_raw = o.get('project')
        if project_name_raw:
            proj_info = load_project_info(project_name_raw)
            project_name_disp = h(project_name_raw)
            project_code_txt = h(proj_info.get('code') or '—')
        else:
            project_name_disp = '—'
            project_code_txt = '—'
        text_lines.append(f"• {h(code)} — {fmt_money(amount)} грн — {project_name_disp} (код {project_code_txt})")
        kb.add(InlineKeyboardButton(f"Подтвердить {code}", callback_data=f"user_confirm_payout:{o['id']}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="menu_finance"))
    await clear_then_anchor(uid, "\n".join(text_lines), kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("user_confirm_payout:"))
async def user_confirm_payout(c: types.CallbackQuery):
    uid = c.from_user.id
    req_id = c.data.split(":",1)[1]
    prof = load_user(uid) or {}
    project_hint = None
    payout_entries = list(iter_user_payout_refs(prof))
    for ref in payout_entries:
        if ref.get("id") == req_id:
            project_hint = ref.get("project")
            break
    obj = finance_load_request(req_id, project_hint)
    if not obj or obj.get("user_id") != uid or obj.get("status") != "approved":
        return await c.answer("Запрос не найден/неправильный статус.", show_alert=True)
    now_iso = datetime.now().isoformat()
    obj["status"] = "confirmed"
    obj["confirmed_at"] = now_iso
    finance_update_items_status(obj, "confirmed", now_iso)
    finance_append_history(obj, "confirmed", {"by": uid})
    finance_save_request(obj)
    update_receipts_for_request(uid, obj.get("project"), obj.get("files", []), "confirmed", obj)
    fin_state_clear(obj.get("project", ""), uid)
    code = obj.get("code", obj["id"])
    proj_info = load_project_info(obj.get("project")) if obj.get("project") else {}
    amount = float(obj.get('sum') or 0.0)
    if c.message:
        await delete_if_not_anchor(uid, c.message.chat.id, c.message.message_id)
    # уведомим админа(ов)
    code_disp = h(code)
    project_disp = h(obj.get('project', '—'))
    proj_info = load_project_info(obj.get('project')) if obj.get('project') else {}
    project_code_txt = h(proj_info.get('code') or '—') if proj_info else '—'
    project_code_txt = h(proj_info.get('code') or '—') if proj_info else '—'
    fullname = h(prof.get('fullname', '—'))
    bsu_code = h(prof.get('bsu', '—'))
    phone = h(prof.get('phone', '—'))
    admin_note = (
        "💸 <b>Выплата подтверждена</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"Код: <b>{code_disp}</b>\n"
        f"Проект: {project_disp} (код {project_code_txt})\n"
        f"Сумма: <b>{fmt_money(amount)} грн</b>\n"
        f"Получатель: {fullname} (ID {uid}, {bsu_code})\n"
        f"Телефон: {phone}\n"
        f"Подтверждено: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    admin_kb = InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="admin_notice_close"))
    for ad in list(admins):
        chat_id = users_runtime.get(ad, {}).get("tg", {}).get("chat_id") or (load_user(ad) or {}).get("tg", {}).get("chat_id")
        if chat_id:
            try:
                await bot.send_message(chat_id, admin_note, reply_markup=admin_kb)
            except Exception:
                pass
    prof = load_user(uid) or {}
    remaining = []
    for ref in iter_user_payout_refs(prof):
        other = finance_load_request(ref.get("id"), ref.get("project"))
        if other and other.get("status") == "approved" and other.get("user_id") == uid:
            remaining.append(other)
    if remaining:
        kb = InlineKeyboardMarkup()
        text_lines = [
            "✅ <b>Одобренные выплаты</b>",
            "━━━━━━━━━━━━━━━━━━",
            "Подтвердите получение средств по каждой записи, чтобы закрыть запрос.",
            ""
        ]
        for o in remaining[:20]:
            oc = o.get("code", o["id"])
            amt = float(o.get("sum") or 0.0)
            text_lines.append(f"• {h(oc)} — {fmt_money(amt)} грн — {h(o.get('project','—'))}")
            kb.add(InlineKeyboardButton(f"Подтвердить {oc}", callback_data=f"user_confirm_payout:{o['id']}"))
        kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="menu_finance"))
        await anchor_show_text(uid, "\n".join(text_lines), kb)
    else:
        await anchor_show_text(
            uid,
            (
                "✅ <b>Выплата подтверждена</b>\n"
                f"Код: <b>{code_disp}</b>\n"
                f"Сумма: <b>{fmt_money(amount)} грн</b>\n\n"
                "Спасибо! Статистика обновлена, и запрос перенесён в историю выплат."
            ),
            kb_finance_root(user_has_pending_confirm=user_has_approved_not_confirmed(uid))
        )
    await c.answer("Подтверждено")


# ========================== ADMIN: FINANCE ==========================
@dp.callback_query_handler(lambda c: c.data == "adm_finance")
async def adm_finance(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("📬 Запросы (ожидают)", callback_data="adm_requests"))
    kb.add(InlineKeyboardButton("📚 История выплат", callback_data="adm_history"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="menu_admin"))
    await clear_then_anchor(
        uid,
        "💵 <b>Финансовый модуль администратора</b>\nВыберите раздел для просмотра запросов или истории выплат.",
        kb
    )
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_requests")
async def adm_requests(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    lst = finance_list("pending")
    if not lst:
        await clear_then_anchor(
            uid,
            "💵 <b>Запросов на выплату нет</b>\nВсе обращения сотрудников обработаны.",
            kb_admin_root()
        ); return await c.answer()
    kb = InlineKeyboardMarkup()
    for r in lst[:20]:
        code = r.get("code", r['id'])
        amount = float(r.get('sum') or 0.0)
        kb.add(InlineKeyboardButton(f"{code} • {fmt_money(amount)} грн • u{r['user_id']}", callback_data=f"adm_req_open:{r['id']}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_finance"))
    await clear_then_anchor(
        uid,
        "📬 <b>Запросы на выплату</b>\n━━━━━━━━━━━━━━━━━━\nВыберите обращение, чтобы изучить детали и принять решение.",
        kb
    )
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_history")
async def adm_history(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    lst = [x for x in finance_list() if x.get("status") in ("approved","confirmed","closed")]
    if not lst:
        await clear_then_anchor(
            uid,
            "📚 История выплат пуста.\nЗдесь появятся все одобренные и закрытые обращения.",
            kb_admin_root()
        ); return await c.answer()
    kb = InlineKeyboardMarkup()
    status_map = {"approved": "Одобрено", "confirmed": "Подтверждено", "closed": "Закрыто"}
    for o in lst[:30]:
        code = o.get("code", o['id'])
        status_txt = status_map.get(o.get("status"), o.get("status"))
        amount = float(o.get('sum') or 0.0)
        kb.add(InlineKeyboardButton(f"{code} • {fmt_money(amount)} грн • {status_txt}", callback_data=f"adm_hist_open:{o['id']}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_finance"))
    await clear_then_anchor(
        uid,
        "📚 <b>История выплат</b>\n━━━━━━━━━━━━━━━━━━\nВыберите запись, чтобы просмотреть подробности и связанные чеки.",
        kb
    )
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_hist_open:"))
async def adm_hist_open(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    req_id = c.data.split(":", 1)[1]
    obj = finance_load_request(req_id)
    if not obj:
        return await c.answer("Запрос не найден", show_alert=True)
    prof = load_user(obj.get("user_id")) or {}
    code = obj.get("code", req_id)

    def fmt_ts(value: Optional[str]) -> str:
        if not value:
            return "—"
        try:
            return datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return value

    status_map = {"pending": "В ожидании", "approved": "Одобрено", "confirmed": "Подтверждено", "closed": "Закрыто"}
    status_disp = status_map.get(obj.get('status'), obj.get('status'))
    project_disp = h(obj.get('project', '—'))
    code_disp = h(code)
    fullname = h(prof.get('fullname', '—'))
    bsu_code = h(prof.get('bsu', '—'))
    phone = h(prof.get('phone', '—'))
    username_raw = (prof.get('tg', {}) or {}).get('username')
    username_display = h(f"@{username_raw}" if username_raw else "—")
    text = (
        f"💵 <b>Выплата {code_disp}</b>\n\n"
        f"Статус: {h(status_disp)}\n"
        f"Сумма: <b>{fmt_money(float(obj.get('sum') or 0.0))} грн</b>\n"
        f"Проект: {project_disp} (код {project_code_txt})\n"
        f"Чеков: {len(obj.get('files', []))}\n\n"
        f"👤 {fullname} (ID {obj.get('user_id')}, {bsu_code})\n"
        f"📱 {phone}\n"
        f"🆔 {username_display}\n\n"
        f"Создано: {fmt_ts(obj.get('created_at'))}\n"
        f"Одобрено: {fmt_ts(obj.get('approved_at'))}\n"
        f"Подтверждено: {fmt_ts(obj.get('confirmed_at'))}"
    )
    kb = InlineKeyboardMarkup()
    if obj.get("files"):
        kb.add(InlineKeyboardButton("🧾 Посмотреть чеки", callback_data=f"adm_hist_view_checks:{req_id}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_history"))
    await clear_then_anchor(uid, text, kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_hist_view_checks:"))
async def adm_hist_view_checks(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    req_id = c.data.split(":", 1)[1]
    obj = finance_load_request(req_id)
    if not obj:
        return await c.answer("Запрос не найден", show_alert=True)
    files = obj.get("files", [])
    code = obj.get("code", req_id)
    lines = [f"🧾 Чеки выплаты <b>{h(code)}</b> ({len(files)})", ""]
    recs = user_project_receipts(obj.get("user_id"), obj.get("project"))
    by_file = {r.get("file"): r for r in recs}
    for fname in files:
        r = by_file.get(fname)
        if r:
            amount = float(r.get('sum') or 0.0)
            desc_text = h(r.get('desc')) if r.get('desc') else "—"
            lines.append(f"• {h(r.get('receipt_no','—'))} — {fmt_money(amount)} грн — {desc_text}")
        else:
            lines.append(f"• {h(fname)}")
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data=f"adm_hist_open:{req_id}"))
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await admin_send_receipt_photos(uid, c.message.chat.id, obj.get("user_id"), obj.get("project"), files)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_req_open:"))
async def adm_req_open(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    req_id = c.data.split(":",1)[1]
    obj = finance_load_request(req_id)
    if not obj: return await c.answer("Запрос не найден", show_alert=True)
    prof = load_user(obj["user_id"]) or {}
    code = obj.get("code", obj["id"])
    proj_info = load_project_info(obj.get("project")) if obj.get("project") else {"region": "", "location": ""}
    fullname = h(prof.get('fullname', '—'))
    bsu_code = h(prof.get('bsu', '—'))
    phone = h(prof.get('phone', '—'))
    username_raw = (prof.get('tg', {}) or {}).get('username')
    username_display = h(f"@{username_raw}" if username_raw else "—")
    code_disp = h(code)
    file_disp = h(obj['id'])
    project_name = h(obj.get('project', '—'))
    project_code_txt = h(proj_info.get('code') or '—')
    region_txt = h(proj_info.get('region') or '—')
    location_txt = h(proj_info.get('location', '—'))
    text = (
        "📢 <b>Запрос на выплату</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"Код выплаты: <b>{code_disp}</b>\n"
        f"Файл: <code>{file_disp}</code>\n"
        f"👤 {fullname} (ID {obj['user_id']}, {bsu_code})\n"
        f"📱 {phone}\n"
        f"🆔 {username_display}\n"
        f"📂 Проект: {project_name}\n"
        f"🆔 Код объекта: {project_code_txt}\n"
        f"🌍 Область: {region_txt}\n"
        f"📍 Локация: {location_txt}\n"
        f"❌ Неоплаченных чеков: {len(obj['files'])} шт.\n"
        f"💰 К выплате: <b>{fmt_money(float(obj.get('sum') or 0.0))} грн</b>\n\n"
        "Просмотрите прикреплённые чеки перед одобрением выплаты."
    )
    akb = InlineKeyboardMarkup()
    akb.add(InlineKeyboardButton("👀 Посмотреть чеки", callback_data=f"adm_req_view_checks:{req_id}"))
    akb.add(InlineKeyboardButton("✅ Выплатить", callback_data=f"adm_req_paid:{req_id}"))
    akb.add(InlineKeyboardButton("❌ Закрыть", callback_data=f"adm_req_close:{req_id}"))
    akb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_requests"))
    await clear_then_anchor(uid, text, akb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_req_view_checks:"))
async def adm_req_view_checks(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    req_id = c.data.split(":",1)[1]
    obj = finance_load_request(req_id)
    if not obj: return await c.answer("Запрос не найден", show_alert=True)
    # вытащим чеки пользователя по проекту с указанными файлами
    recs = user_project_receipts(obj["user_id"], obj["project"])
    by_file = {r["file"]: r for r in recs}
    code = obj.get("code", req_id)
    proj_info = load_project_info(obj.get("project")) if obj.get("project") else {}
    project_code_txt = h(proj_info.get("code") or "—")
    lines = [
        f"🧾 <b>Чеки выплаты {h(code)}</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"Проект: {h(obj.get('project','—'))} (код {project_code_txt})",
        f"Всего чеков: <b>{len(obj['files'])}</b>",
        ""
    ]
    for f in obj["files"]:
        r = by_file.get(f)
        if r:
            amount = float(r.get('sum') or 0.0)
            desc_text = h(r.get('desc')) if r.get('desc') else "—"
            lines.append(f"• {h(r.get('receipt_no','—'))} — {fmt_money(amount)} грн — {desc_text} — {h(r.get('date','—'))} {h(r.get('time',''))}")
        else:
            lines.append(f"• {h(f)}")
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("✅ Выплатить", callback_data=f"adm_req_paid:{req_id}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data=f"adm_req_open:{req_id}"))
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await admin_send_receipt_photos(uid, c.message.chat.id, obj["user_id"], obj["project"], obj.get("files", []))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_req_close:"))
async def adm_req_close(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    req_id = c.data.split(":",1)[1]
    obj = finance_load_request(req_id)
    if not obj or obj.get("status") != "pending":
        return await c.answer("Запрос не найден/обработан", show_alert=True)
    now_iso = datetime.now().isoformat()
    obj["status"] = "closed"
    obj["closed_at"] = now_iso
    finance_update_items_status(obj, "closed", now_iso)
    finance_append_history(obj, "closed", {"by": uid})
    finance_save_request(obj)
    update_receipts_for_request(obj.get("user_id"), obj.get("project"), obj.get("files", []), "closed", obj)
    fin_state_clear(obj.get("project", ""), obj.get("user_id"))
    code = obj.get("code", req_id)
    if c.message:
        await delete_if_not_anchor(uid, c.message.chat.id, c.message.message_id)
    user_id = obj.get("user_id")
    prof = load_user(user_id) or {}
    chat_id = users_runtime.get(user_id, {}).get("tg", {}).get("chat_id") or prof.get("tg", {}).get("chat_id")
    if chat_id:
        note = (
            "ℹ️ <b>Запрос на выплату закрыт</b>\n\n"
            f"Код: <b>{h(code)}</b>\n"
            f"Проект: {h(obj.get('project','—'))} (код {project_code_txt})\n"
            "Если нужна помощь — напишите администратору."
        )
        kb = InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
        try:
            await bot.send_message(chat_id, note, reply_markup=kb)
        except Exception:
            pass
    await clear_then_anchor(uid, f"🗂 Запрос {code} закрыт без выплат.", kb_admin_root())
    await c.answer("Закрыто")


@dp.callback_query_handler(lambda c: c.data.startswith("adm_req_paid:"))
async def adm_req_paid(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    req_id = c.data.split(":",1)[1]
    obj = finance_load_request(req_id)
    if not obj or obj.get("status") != "pending":
        return await c.answer("Запрос не найден/обработан", show_alert=True)

    now_iso = datetime.now().isoformat()
    obj["status"] = "approved"
    obj["approved_by"] = uid
    obj["approved_at"] = now_iso
    finance_update_items_status(obj, "approved", now_iso)
    finance_append_history(obj, "approved", {"by": uid})
    finance_save_request(obj)
    update_receipts_for_request(obj.get("user_id"), obj.get("project"), obj.get("files", []), "approved", obj)
    fin_state_set(obj.get("project", ""), obj.get("user_id"), req_id, "approved")

    user_id = obj["user_id"]
    prof = load_user(user_id) or {}
    code = obj.get("code", obj["id"])
    proj_info = load_project_info(obj.get("project")) if obj.get("project") else {}
    if c.message:
        await delete_if_not_anchor(uid, c.message.chat.id, c.message.message_id)

    chat_id = users_runtime.get(user_id, {}).get("tg", {}).get("chat_id") or prof.get("tg", {}).get("chat_id")
    recs = user_project_receipts(user_id, obj["project"])
    by_file = {r["file"]: r for r in recs}
    lines = []
    for f in obj["files"]:
        r = by_file.get(f)
        if r:
            amount = float(r.get("sum") or 0.0)
            lines.append(f"• {h(r.get('receipt_no','—'))} — {fmt_money(amount)} грн")
        else:
            lines.append(f"• {h(f)}")
    details = "\n".join(lines) if lines else "—"
    code_disp = h(code)
    project_disp = h(obj.get('project', '—'))
    project_code_txt = h(proj_info.get("code") or "—")
    user_text = (
        "💵 <b>Выплата согласована</b>\n\n"
        f"Код: <b>{code_disp}</b>\n"
        f"Проект: {project_disp} (код {project_code_txt})\n"
        f"💰 К выдаче: <b>{fmt_money(float(obj.get('sum') or 0.0))} грн</b>\n"
        "Вам должны передать указанную сумму. Как только получите деньги, подтвердите это внизу.\n\n"
        "Чеки в выплате:\n"
        f"{details}"
    )
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("✅ Деньги получены", callback_data=f"user_confirm_payout:{obj['id']}"))
    kb.add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close"))
    if chat_id:
        try:
            await bot.send_message(chat_id, user_text, reply_markup=kb)
        except Exception:
            pass

    await clear_then_anchor(uid, f"💸 Выплата {code_disp} по объекту {project_disp} (код {project_code_txt}) одобрена. Ожидаем подтверждения пользователя.", kb_admin_root())
    await c.answer("Выплата одобрена")


# ========================== SOS ==========================
@dp.callback_query_handler(lambda c: c.data == "menu_sos")
async def sos_start(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await state.finish()
    await flow_clear(uid)
    text = ("⚠️ Вы нажали кнопку <b>SOS</b>.\n\n"
            "Пожалуйста, подтвердите, что нажатие не случайно:")
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("✅ Подтвердить", callback_data="sos_confirm"))
    kb.add(InlineKeyboardButton("❌ Отменить", callback_data="sos_cancel"))
    msg = await bot.send_message(c.message.chat.id, text, reply_markup=kb)
    flow_track(uid, msg)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "sos_cancel")
async def sos_cancel(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await state.finish(); await flow_clear(uid); await anchor_show_root(uid)
    await c.answer("Отменено.")


@dp.callback_query_handler(lambda c: c.data == "sos_confirm")
async def sos_confirm(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await flow_clear(uid)
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.row(
        KeyboardButton("📍 Отправить местоположение", request_location=True),
        KeyboardButton("❌ Отменить")
    )
    msg = await bot.send_message(
        c.message.chat.id,
        "📍 Отправьте вашу геолокацию кнопкой ниже или отмените действие.",
        reply_markup=kb
    )
    flow_track(uid, msg)
    await SosFSM.waiting_location.set()
    await c.answer()


@dp.message_handler(state=SosFSM.waiting_location, content_types=ContentType.TEXT)
async def sos_location_text(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    text = (m.text or "").strip()
    normalized = text.replace("❌", "").replace("📍", "").strip().lower()
    if normalized in {"отменить", "cancel"} or normalized in NP_CANCEL_WORDS:
        try:
            await bot.delete_message(m.chat.id, m.message_id)
        except Exception:
            pass
        try:
            remove = await bot.send_message(m.chat.id, "⌨️", reply_markup=ReplyKeyboardRemove())
            await bot.delete_message(remove.chat.id, remove.message_id)
        except Exception:
            pass
        await flow_clear(uid)
        await state.finish()
        await anchor_show_root(uid)
        return
    try:
        await bot.delete_message(m.chat.id, m.message_id)
    except Exception:
        pass
    warn = await bot.send_message(
        m.chat.id,
        "📍 Используйте кнопку отправки геопозиции или «❌ Отменить», чтобы вернуться в меню."
    )
    flow_track(uid, warn)


@dp.message_handler(content_types=ContentType.LOCATION, state=SosFSM.waiting_location)
async def sos_location(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    lat, lon = m.location.latitude, m.location.longitude
    prof = load_user(uid) or {"user_id": uid}

    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass

    try:
        placeholder = await bot.send_message(m.chat.id, "⌨️", reply_markup=ReplyKeyboardRemove())
        await bot.delete_message(m.chat.id, placeholder.message_id)
    except Exception:
        pass

    sender_kb = InlineKeyboardMarkup().add(
        InlineKeyboardButton("❌ Закрыть уведомление", callback_data="sos_sender_close")
    )
    sender_msg = await bot.send_message(
        m.chat.id,
        "🆘 SOS включён. Сигнал принят.",
        reply_markup=sender_kb
    )
    users_runtime.setdefault(uid, {})["sos_sender_msg"] = sender_msg.message_id

    sos_text = (f"🚨 <b>SOS сигнал!</b>\n\n"
                f"👤 Имя: {prof.get('fullname','—')} ({prof.get('bsu','—')})\n"
                f"📱 Телефон: {prof.get('phone','—')}\n"
                f"🆔 Telegram: @{(prof.get('tg',{}) or {}).get('username','—')}\n"
                f"🕒 Время вызова: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                f"❗ Пользователь запросил помощь. Координаты ниже.")
    close_kb = InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть сообщение", callback_data="sos_close"))

    def get_chat_id_for_user(u_id: int) -> Optional[int]:
        ch = users_runtime.get(u_id, {}).get("tg", {}).get("chat_id")
        if ch: return ch
        up = load_user(u_id)
        return (up.get("tg", {}).get("chat_id")) if up else None

    for f in os.listdir(USERS_PATH):
        if not f.endswith(".json"): continue
        try:
            udata = json.load(open(os.path.join(USERS_PATH, f), "r", encoding="utf-8"))
        except:
            continue
        rec_uid = int(udata.get("user_id", 0))
        if rec_uid == uid:  # отправителю не шлём
            continue
        chat_id = get_chat_id_for_user(rec_uid)
        if not chat_id:
            continue
        try:
            msg_txt = await bot.send_message(chat_id, sos_text, reply_markup=close_kb)
            msg_loc = await bot.send_location(chat_id, latitude=lat, longitude=lon)
            users_runtime.setdefault(rec_uid, {})["last_sos_broadcast"] = (msg_txt.message_id, msg_loc.message_id)
        except Exception:
            pass

    await flow_clear(uid)
    await state.finish()
    await anchor_show_root(uid)


@dp.callback_query_handler(lambda c: c.data == "sos_close")
async def sos_close(c: types.CallbackQuery):
    uid = c.from_user.id
    try: await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception: pass
    pair = users_runtime.get(uid, {}).get("last_sos_broadcast")
    if pair:
        _, loc_id = pair
        try: await bot.delete_message(c.message.chat.id, loc_id)
        except: pass
        users_runtime[uid]["last_sos_broadcast"] = None
    await c.answer("Сообщение закрыто.")


@dp.callback_query_handler(lambda c: c.data == "sos_sender_close")
async def sos_sender_close(c: types.CallbackQuery):
    uid = c.from_user.id
    msg_id = users_runtime.get(uid, {}).pop("sos_sender_msg", None)
    if msg_id:
        try: await bot.delete_message(c.message.chat.id, msg_id)
        except: pass
    await c.answer("Уведомление закрыто.")


# ========================== ADMIN: USERS & PROJECTS ==========================
def paginate(lst: List[str], page: int, per_page: int=10) -> Tuple[List[str], int]:
    total = max(1, (len(lst) + per_page - 1) // per_page)
    page = max(1, min(page, total))
    start = (page-1)*per_page
    return lst[start:start+per_page], total


@dp.callback_query_handler(lambda c: c.data == "menu_admin")
async def menu_admin(c: types.CallbackQuery):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔ Доступ только для администраторов", show_alert=True)
    await clear_then_anchor(
        uid,
        "🧑‍💼 <b>Админ-панель</b>\nВыберите раздел для управления пользователями, проектами и финансами.",
        kb_admin_root()
    )
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_users")
async def adm_users(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    await state.reset_state(with_data=False)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    files = sorted([f for f in os.listdir(USERS_PATH) if f.endswith(".json")])
    page = 1
    slice_, total = paginate(files, page)
    kb = InlineKeyboardMarkup()
    for filename in slice_:
        try:
            target_uid = int(os.path.splitext(filename)[0])
        except ValueError:
            continue
        profile = load_user(target_uid) or {"user_id": target_uid}
        fullname = str(profile.get("fullname") or "—")
        code = str(profile.get("bsu") or "—")
        kb.add(
            InlineKeyboardButton(
                f"{fullname} ({code})",
                callback_data=f"adm_user_{target_uid}"
            )
        )
    if total > 1:
        kb.row(
            InlineKeyboardButton("⏮", callback_data=f"adm_users_page_1"),
            InlineKeyboardButton(f"{page}/{total}", callback_data="noop"),
            InlineKeyboardButton("⏭", callback_data=f"adm_users_page_{total}")
        )
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="menu_admin"))
    await clear_then_anchor(
        uid,
        "👥 <b>Пользователи</b>\n━━━━━━━━━━━━━━━━━━\nВыберите сотрудника, чтобы открыть карточку с подробностями.",
        kb
    )
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_users_page_"))
async def adm_users_page(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    await state.reset_state(with_data=False)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    files = sorted([f for f in os.listdir(USERS_PATH) if f.endswith(".json")])
    try:
        page = int(c.data.split("_")[-1])
    except Exception:
        page = 1
    slice_, total = paginate(files, page)
    kb = InlineKeyboardMarkup()
    for filename in slice_:
        try:
            target_uid = int(os.path.splitext(filename)[0])
        except ValueError:
            continue
        profile = load_user(target_uid) or {"user_id": target_uid}
        fullname = str(profile.get("fullname") or "—")
        code = str(profile.get("bsu") or "—")
        kb.add(
            InlineKeyboardButton(
                f"{fullname} ({code})",
                callback_data=f"adm_user_{target_uid}"
            )
        )
    if total > 1:
        prev_page = max(1, page-1); next_page = min(total, page+1)
        kb.row(
            InlineKeyboardButton("⏮", callback_data=f"adm_users_page_1"),
            InlineKeyboardButton(f"◀ {prev_page}", callback_data=f"adm_users_page_{prev_page}"),
            InlineKeyboardButton(f"{page}/{total}", callback_data="noop"),
            InlineKeyboardButton(f"{next_page} ▶", callback_data=f"adm_users_page_{next_page}"),
            InlineKeyboardButton("⏭", callback_data=f"adm_users_page_{total}")
        )
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="menu_admin"))
    await clear_then_anchor(
        uid,
        "👥 <b>Пользователи</b>\n━━━━━━━━━━━━━━━━━━\nВыберите сотрудника, чтобы открыть карточку с подробностями.",
        kb
    )
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_user_"))
async def adm_user_card(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    target = int(c.data.split("adm_user_",1)[1])
    await state.reset_state(with_data=False)
    await state.update_data(target_uid=target, admin_user_show_photo=False, admin_user_edit_mode=False)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=False)
    await c.answer()


async def admin_show_user(
    uid: int,
    target_uid: int,
    state: FSMContext,
    show_photo: Optional[bool] = None,
    edit_mode: Optional[bool] = None,
):
    data = await state.get_data()
    if show_photo is None:
        show_photo = bool(data.get("admin_user_show_photo"))
    if edit_mode is None:
        edit_mode = bool(data.get("admin_user_edit_mode"))
    await state.update_data(admin_user_show_photo=show_photo, admin_user_edit_mode=edit_mode)
    profile = load_user(target_uid) or {"user_id": target_uid}
    text = admin_user_card_text(uid, profile, edit_mode=edit_mode)
    kb = kb_admin_user(uid, profile, show_photo=show_photo, edit_mode=edit_mode)
    if show_photo and profile_has_photo(profile):
        await anchor_replace_with_photo(uid, user_profile_photo_path(target_uid), text, kb)
    else:
        chat = users_runtime.get(uid, {}).get("tg", {}).get("chat_id")
        if chat:
            await anchor_upsert(uid, chat, text, kb)


@dp.callback_query_handler(lambda c: c.data == "adm_user_photo_toggle")
async def adm_user_photo_toggle(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        return await c.answer("Користувач не знайдений", show_alert=True)
    current = bool(data.get("admin_user_show_photo"))
    await admin_show_user(uid, target, state, show_photo=not current)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_user_edit")
async def adm_user_edit(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.update_data(admin_user_edit_mode=True)
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_user_edit_done", state="*")
async def adm_user_edit_done(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        return await c.answer("Користувач не знайдений", show_alert=True)
    runtime = admin_edit_runtime(uid)
    runtime.pop("reply_keyboard", False)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await state.update_data(admin_user_edit_mode=False)
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=False)
    await c.answer()


@dp.callback_query_handler(
    lambda c: c.data == "adm_edit_cancel",
    state=[
        AdminProfileEditFSM.waiting_last_name,
        AdminProfileEditFSM.waiting_first_name,
        AdminProfileEditFSM.waiting_middle_name,
        AdminProfileEditFSM.waiting_birthdate,
        AdminProfileEditFSM.waiting_region,
        AdminProfileEditFSM.region_confirm,
        AdminProfileEditFSM.waiting_phone,
        AdminProfileEditFSM.waiting_photo,
    ],
)
async def adm_edit_cancel(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    data = await state.get_data()
    target = data.get("target_uid")
    runtime = admin_edit_runtime(uid)
    remove_keyboard = bool(runtime.pop("reply_keyboard", False))
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    if target:
        await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)
    if remove_keyboard:
        await admin_edit_notify(uid, tr(uid, "PROFILE_CANCELLED"), remove_keyboard=True)
    else:
        await c.answer(tr(uid, "PROFILE_CANCELLED"))


@dp.callback_query_handler(lambda c: c.data == "adm_edit_last")
async def adm_edit_last_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    if not (await state.get_data()).get("target_uid"):
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await AdminProfileEditFSM.waiting_last_name.set()
    await flow_clear_warnings(uid)
    await admin_edit_send_prompt(uid, tr(uid, "PROFILE_PROMPT_LAST_NAME"), reply_markup=kb_admin_edit_cancel(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_edit_first")
async def adm_edit_first_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    if not (await state.get_data()).get("target_uid"):
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await AdminProfileEditFSM.waiting_first_name.set()
    await flow_clear_warnings(uid)
    await admin_edit_send_prompt(uid, tr(uid, "PROFILE_PROMPT_FIRST_NAME"), reply_markup=kb_admin_edit_cancel(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_edit_middle")
async def adm_edit_middle_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    if not (await state.get_data()).get("target_uid"):
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await AdminProfileEditFSM.waiting_middle_name.set()
    await flow_clear_warnings(uid)
    await admin_edit_send_prompt(uid, tr(uid, "PROFILE_PROMPT_MIDDLE_NAME"), reply_markup=kb_admin_edit_cancel(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_edit_birthdate")
async def adm_edit_birthdate_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    if not (await state.get_data()).get("target_uid"):
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await AdminProfileEditFSM.waiting_birthdate.set()
    await flow_clear_warnings(uid)
    await admin_edit_send_prompt(uid, tr(uid, "PROFILE_PROMPT_BIRTHDATE"), reply_markup=kb_admin_edit_cancel(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_edit_region")
async def adm_edit_region_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    if not (await state.get_data()).get("target_uid"):
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await AdminProfileEditFSM.waiting_region.set()
    await flow_clear_warnings(uid)
    await admin_edit_send_prompt(uid, tr(uid, "PROFILE_PROMPT_REGION"), reply_markup=kb_admin_edit_region_prompt(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_reg_open", state=AdminProfileEditFSM.waiting_region)
async def adm_edit_region_open(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    await admin_edit_send_prompt(uid, tr(uid, "REGISTER_REGION_PICK"), reply_markup=kb_admin_edit_region_picker(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_reg_pick:"), state=AdminProfileEditFSM.waiting_region)
async def adm_edit_region_pick(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    try:
        idx = int(c.data.split(":", 1)[1])
        region = UKRAINE_REGIONS[idx]
    except Exception:
        return await c.answer(tr(uid, "REGISTER_REGION_REMIND"), show_alert=True)
    await state.update_data(admin_region=region)
    await admin_edit_send_prompt(
        uid,
        tr(uid, "REGISTER_REGION_SELECTED", region=h(region)),
        reply_markup=kb_admin_edit_next(uid, "adm_reg_confirm"),
    )
    await AdminProfileEditFSM.region_confirm.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_reg_confirm", state=AdminProfileEditFSM.region_confirm)
async def adm_edit_region_confirm(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    data = await state.get_data()
    target = data.get("target_uid")
    region = data.get("admin_region")
    if not target or not region:
        return await c.answer(tr(uid, "REGISTER_REGION_REMIND"), show_alert=True)
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    profile["region"] = region
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await admin_edit_notify(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)
    await c.answer()


@dp.message_handler(state=AdminProfileEditFSM.waiting_region, content_types=ContentType.TEXT)
async def adm_edit_region_text(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_REGION_REMIND"), reply_markup=kb_admin_edit_region_prompt(uid))
    flow_track_warning(uid, warn)


@dp.callback_query_handler(lambda c: c.data == "adm_edit_phone")
async def adm_edit_phone_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    if not (await state.get_data()).get("target_uid"):
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await AdminProfileEditFSM.waiting_phone.set()
    await flow_clear_warnings(uid)
    await admin_edit_send_prompt(uid, tr(uid, "PROFILE_PROMPT_PHONE"), reply_markup=kb_admin_edit_cancel(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_edit_photo")
async def adm_edit_photo_prompt(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    if not (await state.get_data()).get("target_uid"):
        return await c.answer("Користувач не знайдений", show_alert=True)
    await state.reset_state(with_data=False)
    await AdminProfileEditFSM.waiting_photo.set()
    await flow_clear_warnings(uid)
    await admin_edit_send_prompt(uid, tr(uid, "PROFILE_PROMPT_PHOTO"), reply_markup=kb_admin_edit_cancel(uid))
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_edit_remove_photo")
async def adm_edit_remove_photo(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        return await c.answer("Користувач не знайдений", show_alert=True)
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    remove_profile_photo(target)
    profile["photo"] = {}
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await flow_clear_warnings(uid)
    await admin_edit_notify(uid, tr(uid, "PROFILE_PHOTO_REMOVED"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)
    await c.answer()

@dp.message_handler(state=AdminProfileEditFSM.waiting_last_name, content_types=ContentType.TEXT)
async def adm_edit_last_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    raw = normalize_person_name(m.text)
    await flow_delete_message(uid, m)
    if not raw or not NAME_VALID_RE.match(raw):
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_LAST_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        warn = await bot.send_message(m.chat.id, "Користувач не знайдений")
        flow_track_warning(uid, warn)
        return
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    profile["last_name"] = beautify_name(raw)
    profile["fullname"] = compose_fullname(profile["last_name"], profile.get("first_name", ""), profile.get("middle_name"))
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await admin_edit_notify(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)


@dp.message_handler(state=AdminProfileEditFSM.waiting_first_name, content_types=ContentType.TEXT)
async def adm_edit_first_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    raw = normalize_person_name(m.text)
    await flow_delete_message(uid, m)
    if not raw or not NAME_VALID_RE.match(raw):
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_FIRST_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        warn = await bot.send_message(m.chat.id, "Користувач не знайдений")
        flow_track_warning(uid, warn)
        return
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    profile["first_name"] = beautify_name(raw)
    profile["fullname"] = compose_fullname(profile.get("last_name", ""), profile["first_name"], profile.get("middle_name"))
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await admin_edit_notify(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)


@dp.message_handler(state=AdminProfileEditFSM.waiting_middle_name, content_types=ContentType.TEXT)
async def adm_edit_middle_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    raw = normalize_person_name(m.text)
    await flow_delete_message(uid, m)
    if raw and raw.lower() in SKIP_KEYWORDS:
        cleaned = ""
    elif raw and NAME_VALID_RE.match(raw):
        cleaned = beautify_name(raw)
    else:
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_MIDDLE_NAME_WARN"))
        flow_track_warning(uid, warn)
        return
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        warn = await bot.send_message(m.chat.id, "Користувач не знайдений")
        flow_track_warning(uid, warn)
        return
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    profile["middle_name"] = cleaned
    profile["fullname"] = compose_fullname(profile.get("last_name", ""), profile.get("first_name", ""), cleaned)
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await admin_edit_notify(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)


@dp.message_handler(state=AdminProfileEditFSM.waiting_birthdate, content_types=ContentType.TEXT)
async def adm_edit_birthdate_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    dt = parse_birthdate_text(m.text)
    await flow_delete_message(uid, m)
    if not dt:
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_BIRTHDATE_WARN"))
        flow_track_warning(uid, warn)
        return
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        warn = await bot.send_message(m.chat.id, "Користувач не знайдений")
        flow_track_warning(uid, warn)
        return
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    profile["birthdate"] = dt.strftime("%Y-%m-%d")
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await admin_edit_notify(uid, tr(uid, "PROFILE_UPDATE_SUCCESS"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)


@dp.message_handler(state=AdminProfileEditFSM.waiting_phone, content_types=ContentType.TEXT)
async def adm_edit_phone_input(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    phone = sanitize_phone_input(m.text)
    await flow_delete_message(uid, m)
    if not phone:
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHONE_TEXT_WARN"))
        flow_track_warning(uid, warn)
        return
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        warn = await bot.send_message(m.chat.id, "Користувач не знайдений")
        flow_track_warning(uid, warn)
        return
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    profile["phone"] = phone
    profile.setdefault("tg", {})
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    runtime = admin_edit_runtime(uid)
    runtime.pop("reply_keyboard", None)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await admin_edit_notify(uid, tr(uid, "PROFILE_PHONE_SAVED"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)


@dp.message_handler(state=AdminProfileEditFSM.waiting_phone, content_types=ContentType.CONTACT)
async def adm_edit_phone_contact(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHONE_TEXT_WARN"))
    flow_track_warning(uid, warn)


@dp.message_handler(state=AdminProfileEditFSM.waiting_photo, content_types=ContentType.PHOTO)
async def adm_edit_photo_receive(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, "Користувач не знайдений")
        flow_track_warning(uid, warn)
        return
    photo = m.photo[-1] if m.photo else None
    if not photo:
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"))
        flow_track_warning(uid, warn)
        return
    try:
        meta = await store_profile_photo(target, photo)
    except Exception:
        await flow_delete_message(uid, m)
        warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"))
        flow_track_warning(uid, warn)
        return
    await flow_delete_message(uid, m)
    profile = load_user(target)
    if not profile:
        profile = ensure_user(target, {})
    profile["photo"] = meta or {"status": "uploaded", "updated_at": datetime.now(timezone.utc).isoformat()}
    profile["updated_at"] = datetime.now(timezone.utc).isoformat()
    save_user(profile)
    await admin_edit_clear_prompt(uid)
    await flow_clear_warnings(uid)
    await state.reset_state(with_data=False)
    await admin_edit_notify(uid, tr(uid, "PROFILE_PHOTO_UPDATED"))
    await admin_show_user(uid, target, state, show_photo=False, edit_mode=True)


@dp.message_handler(state=AdminProfileEditFSM.waiting_photo, content_types=ContentType.ANY)
async def adm_edit_photo_invalid(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    if uid not in admins:
        return
    if m.photo:
        return
    await flow_delete_message(uid, m)
    warn = await bot.send_message(m.chat.id, tr(uid, "REGISTER_PHOTO_WARN"))
    flow_track_warning(uid, warn)


@dp.callback_query_handler(lambda c: c.data == "adm_user_finance")
async def adm_user_finance_view(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins:
        return await c.answer("⛔", show_alert=True)
    data = await state.get_data()
    target = data.get("target_uid")
    if not target:
        return await c.answer("Користувач не знайдений", show_alert=True)
    profile = load_user(target) or {"user_id": target}
    stats = admin_collect_user_stats(profile)
    pending = stats["pending_payouts"]
    closed = stats["confirmed_payouts"]
    fullname = h(profile.get("fullname", "—"))
    bsu = h(profile.get("bsu", "—"))
    lines = [
        "💵 <b>Фінансовий огляд користувача</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"👤 {fullname} (BSU {bsu})",
        f"📂 Активних проєктів з чеками: <b>{len(stats['projects'])}</b>",
        f"🧾 Чеків у системі: <b>{stats['total_count']}</b>",
        f"💰 Загальна сума: <b>{fmt_money(stats['total_sum'])} грн</b>",
        f"✅ Оплачено: <b>{fmt_money(stats['paid_sum'])} грн</b>",
        f"⏳ Очікує: <b>{fmt_money(stats['unpaid_sum'])} грн</b>",
        "",
        f"📬 Запитів у роботі: <b>{len(pending)}</b>",
    ]
    if pending:
        for req in pending[:10]:
            code = req.get("code") or req.get("id")
            amount = fmt_money(float(req.get("sum") or 0.0))
            status = req.get("status", "—")
            lines.append(f"• {h(code)} — {amount} грн — {h(status)}")
    if closed:
        lines.append("")
        lines.append(f"📗 Підтверджено/закрито: <b>{len(closed)}</b>")
        for req in closed[-5:]:
            code = req.get("code") or req.get("id")
            amount = fmt_money(float(req.get("sum") or 0.0))
            status = req.get("status", "—")
            lines.append(f"• {h(code)} — {amount} грн — {h(status)}")
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data=f"adm_user_{target}"))
    await clear_then_anchor(uid, "\n".join(lines), kb)
    await c.answer()


# ======= Admin: per-project stats / receipts (reuse from previous version) =======
@dp.callback_query_handler(lambda c: c.data == "adm_stat_choose")
async def adm_stat_choose(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    target = (await state.get_data()).get("target_uid")
    projs = sorted(list((load_user(target) or {}).get("receipts", {}).keys()))
    kb = InlineKeyboardMarkup()
    for p in projs: kb.add(InlineKeyboardButton(p, callback_data=f"adm_stat_{p}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data=f"adm_user_{target}"))
    await clear_then_anchor(uid, "Выберите проект:", kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_stat_"))
async def adm_stat_show(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    proj = c.data.split("adm_stat_",1)[1]
    target = (await state.get_data()).get("target_uid")
    recs = user_project_receipts(target, proj)
    cnt = len(recs); total = round(sum(float(r.get("sum", 0.0)) for r in recs), 2)
    paid_sum = round(sum(float(r.get("sum", 0.0)) for r in recs if r.get("paid") is True), 2)
    unpaid_sum = round(sum(float(r.get("sum", 0.0)) for r in recs if r.get("paid") is False), 2)
    text = (f"📊 Статистика пользователя <b>{target}</b>\n"
            f"📂 Проект: <b>{h(proj)}</b>\n"
            f"• Чеков: <b>{cnt}</b>\n• Всего: <b>{fmt_money(total)} грн</b>\n"
            f"• ✅ Оплачено: <b>{fmt_money(paid_sum)} грн</b>\n"
            f"• ❌ Не оплачено: <b>{fmt_money(unpaid_sum)} грн</b>")
    kb = InlineKeyboardMarkup().add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_stat_choose"))
    await clear_then_anchor(uid, text, kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "adm_recs_choose")
async def adm_recs_choose(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    target = (await state.get_data()).get("target_uid")
    projs = sorted(list((load_user(target) or {}).get("receipts", {}).keys()))
    kb = InlineKeyboardMarkup()
    for p in projs: kb.add(InlineKeyboardButton(p, callback_data=f"adm_recs_{p}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_user_"+str(target)))
    await clear_then_anchor(uid, "Выберите проект:", kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("adm_recs_"))
async def adm_recs_show(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    proj = c.data.split("adm_recs_",1)[1]
    target = (await state.get_data()).get("target_uid")
    recs = user_project_receipts(target, proj)
    if not recs:
        return await c.answer("Чеков нет.", show_alert=True)
    lines = [f"📁 Чеки пользователя <b>{target}</b>", f"📂 Проект: <b>{h(proj)}</b>", ""]
    for r in recs[-30:]:
        status = "✅" if r.get("paid") is True else ("❌" if r.get("paid") is False else "⏳")
        amount = float(r.get('sum') or 0.0)
        desc_text = h(r.get('desc')) if r.get('desc') else "—"
        lines.append(
            f"• #{h(r.get('receipt_no',''))} — {h(r.get('date','—'))} {h(r.get('time',''))} — {fmt_money(amount)} грн {status} — {desc_text}"
        )
    kb = InlineKeyboardMarkup().add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_recs_choose"))
    await clear_then_anchor(uid, "\n".join(lines), kb)
    # последние 10 фото
    for r in recs[-10:]:
        msg = await send_receipt_card(c.message.chat.id, proj, target, r,
                                      kb=InlineKeyboardMarkup().add(InlineKeyboardButton("❌ Закрыть", callback_data="broadcast_close")),
                                      include_project=False)
        flow_track(uid, msg)
    await c.answer()


# ========================== PROJECTS (ADMIN) ==========================
@dp.callback_query_handler(lambda c: c.data == "adm_projects", state="*")
async def adm_projects(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if uid not in admins: return await c.answer("⛔", show_alert=True)
    await clear_step_prompt(state)
    await state.finish()
    await clear_then_anchor(uid, "📂 <b>Проекты</b>", kb_admin_projects())
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "proj_list")
async def proj_list_cb(c: types.CallbackQuery):
    uid = c.from_user.id
    projs = list_projects()
    if not projs:
        await clear_then_anchor(uid, "❗ Проектов нет.", kb_admin_projects()); return await c.answer()
    act = active_project["name"]
    lines = ["📋 <b>Список проектов</b>", ""]
    for p in projs:
        info = load_project_info(p)
        flag = " (★ активный)" if p == act else ""
        lines.append(
            f"• <b>{h(p)}</b>{flag}\n"
            f"  🆔 {h(info.get('code') or '—')}\n"
            f"  🌍 {h(info.get('region') or '—')}\n"
            f"  📍 {h(info.get('location') or '—')}\n"
            f"  📅 {h(info.get('start_date') or '?')} → {h(info.get('end_date') or '?')}"
        )
    await clear_then_anchor(uid, "\n".join(lines), kb_admin_projects())
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "proj_create")
async def proj_create_begin(c: types.CallbackQuery):
    uid = c.from_user.id
    await clear_then_anchor(uid, "✏️ Введите <b>название</b> проекта и отправьте сообщением.",
                            kb=InlineKeyboardMarkup().add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_projects")))
    await ProjectCreateFSM.enter_name.set()
    await c.answer()


@dp.callback_query_handler(lambda c: c.data == "proj_create_cancel", state="*")
async def proj_create_cancel(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    await clear_step_prompt(state)
    await state.finish()
    await flow_clear(uid)
    await anchor_show_text(uid, "📂 <b>Проекты</b>", kb_admin_projects())
    await c.answer("Отменено")


@dp.message_handler(state=ProjectCreateFSM.enter_name, content_types=ContentType.TEXT)
async def proj_enter_loc(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    name = (m.text or "").strip()
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    if not name:
        warn = await bot.send_message(m.chat.id, "❗ Некорректное название.")
        flow_track(uid, warn); return
    await clear_step_prompt(state)
    await state.update_data(name=name)
    msg = await bot.send_message(m.chat.id, "🌍 Выберите область проекта:", reply_markup=kb_region_select())
    flow_track(uid, msg)
    await remember_step_prompt(state, msg)
    await ProjectCreateFSM.enter_region.set()


@dp.callback_query_handler(lambda c: c.data.startswith("proj_region_"), state=ProjectCreateFSM.enter_region)
async def proj_select_region(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    try:
        idx = int(c.data.split("proj_region_", 1)[1])
    except Exception:
        return await c.answer("Неизвестная область", show_alert=True)
    if idx < 0 or idx >= len(UKRAINE_REGIONS):
        return await c.answer("Неизвестная область", show_alert=True)
    region = UKRAINE_REGIONS[idx]
    await clear_step_prompt(state)
    await state.update_data(region=region)
    msg = await bot.send_message(c.message.chat.id, "📍 Укажите <b>локацию</b> (город/адрес).")
    flow_track(uid, msg)
    await remember_step_prompt(state, msg)
    await ProjectCreateFSM.enter_location.set()
    await c.answer(region)


@dp.message_handler(state=ProjectCreateFSM.enter_location, content_types=ContentType.TEXT)
async def proj_enter_desc(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    loc = (m.text or "").strip()
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    if not loc:
        warn = await bot.send_message(m.chat.id, "❗ Укажите локацию.")
        flow_track(uid, warn); return
    await clear_step_prompt(state)
    await state.update_data(location=loc)
    msg = await bot.send_message(m.chat.id, "ℹ️ Краткое описание (необязательно). Если пропустить — отправьте «-».")
    flow_track(uid, msg)
    await remember_step_prompt(state, msg)
    await ProjectCreateFSM.enter_description.set()


@dp.message_handler(state=ProjectCreateFSM.enter_description, content_types=ContentType.TEXT)
async def proj_enter_start(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    desc = None if (m.text or "").strip() == "-" else (m.text or "").strip()
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    await clear_step_prompt(state)
    await state.update_data(description=desc)
    msg = await bot.send_message(m.chat.id, "📅 Введите <b>дату начала</b> (YYYY-MM-DD):")
    flow_track(uid, msg)
    await remember_step_prompt(state, msg)
    await ProjectCreateFSM.enter_start_date.set()


def _parse_date(text: str) -> Optional[str]:
    try: return datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
    except Exception: return None


@dp.message_handler(state=ProjectCreateFSM.enter_start_date, content_types=ContentType.TEXT)
async def proj_enter_end(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    sd = _parse_date((m.text or "").strip())
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    if not sd:
        warn = await bot.send_message(m.chat.id, "❗ Неверный формат. Пример: 2025-09-21")
        flow_track(uid, warn); return
    await clear_step_prompt(state)
    await state.update_data(start_date=sd)
    msg = await bot.send_message(m.chat.id, "📅 Введите <b>дату окончания</b> (YYYY-MM-DD):")
    flow_track(uid, msg)
    await remember_step_prompt(state, msg)
    await ProjectCreateFSM.enter_end_date.set()


@dp.message_handler(state=ProjectCreateFSM.enter_end_date, content_types=ContentType.TEXT)
async def proj_pdf_prompt(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    ed = _parse_date((m.text or "").strip())
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass
    if not ed:
        warn = await bot.send_message(m.chat.id, "❗ Неверный формат. Пример: 2025-12-31")
        flow_track(uid, warn); return
    await clear_step_prompt(state)
    await state.update_data(end_date=ed)
    data = await state.get_data()
    ensure_project_structure(data["name"])
    tip = await bot.send_message(m.chat.id, "📑 Загружайте PDF-файлы проекта (можно несколько). Когда закончите — нажмите «✅ Завершить».")
    flow_track(uid, tip)
    await state.update_data(step_prompt=None)
    await anchor_show_text(uid, "Загрузка PDF: пришлите документ(ы), затем «✅ Завершить».", kb_pdf_upload())
    await ProjectCreateFSM.upload_pdf.set()


@dp.message_handler(content_types=ContentType.DOCUMENT, state=ProjectCreateFSM.upload_pdf)
async def proj_pdf_upload(m: types.Message, state: FSMContext):
    uid = m.from_user.id
    data = await state.get_data(); name = data["name"]
    if not (m.document and (m.document.mime_type and "pdf" in m.document.mime_type.lower() or m.document.file_name.lower().endswith(".pdf"))):
        warn = await bot.send_message(m.chat.id, "⚠️ Допускаются только PDF.")
        flow_track(uid, warn)
        try: await bot.delete_message(m.chat.id, m.message_id)
        except: pass
        return
    dst = os.path.join(proj_pdf_dir(name), m.document.file_name)
    os.makedirs(proj_pdf_dir(name), exist_ok=True)
    await m.document.download(destination_file=dst)
    info = load_project_info(name)
    arr = info.get("pdf", [])
    if m.document.file_name not in arr:
        arr.append(m.document.file_name); info["pdf"] = arr; save_project_info(name, info)
    ok = await bot.send_message(m.chat.id, f"✅ PDF сохранён: {m.document.file_name}")
    flow_track(uid, ok)
    try: await bot.delete_message(m.chat.id, m.message_id)
    except: pass


@dp.callback_query_handler(lambda c: c.data in ("pdf_more", "pdf_finish"), state=ProjectCreateFSM.upload_pdf)
async def proj_pdf_buttons(c: types.CallbackQuery, state: FSMContext):
    uid = c.from_user.id
    if c.data == "pdf_more":
        await anchor_show_text(uid, "📑 Пришлите следующий PDF или «✅ Завершить».", kb_pdf_upload())
        return await c.answer("Жду файл")
    data = await state.get_data(); name = data["name"]
    info = load_project_info(name)
    info.update({
        "location": data["location"],
        "description": data.get("description") or "",
        "region": data.get("region") or "",
        "start_date": data["start_date"],
        "end_date": data["end_date"],
        "active": True
    })
    save_project_info(name, info); set_active_project(name)
    await state.finish()
    await clear_then_anchor(uid, f"✅ Проект «{h(name)}» (код {h(info.get('code') or '—')}) создан и активирован.", kb_admin_projects())
    await c.answer("Готово")
    await update_all_anchors()
    text = (
        f"🏗 <b>Старт нового проекта!</b>\n\n"
        f"📂 <b>{h(name)}</b>\n"
        f"🆔 Код объекта: {h(info.get('code') or '—')}\n"
        f"🌍 Область: {h(info.get('region') or '—')}\n"
        f"📍 Локация: {h(info.get('location') or '—')}\n"
        f"📅 Сроки: {h(info.get('start_date') or '—')} → {h(info.get('end_date') or '—')}\n"
        f"ℹ️ {h(info.get('description') or 'Документы — в разделе «📑 Документы»')}\n\n"
        f"Добавляйте чеки через «🧾 Чеки» → «📷 Добавить чек»."
    )
    for f in os.listdir(USERS_PATH):
        if not f.endswith(".json"): continue
        try:
            udata = json.load(open(os.path.join(USERS_PATH, f), "r", encoding="utf-8"))
        except:
            continue
        chat_id = users_runtime.get(udata["user_id"], {}).get("tg", {}).get("chat_id") or udata.get("tg", {}).get("chat_id")
        if chat_id:
            try: await bot.send_message(chat_id, text, reply_markup=kb_broadcast_close())
            except Exception: pass


@dp.callback_query_handler(lambda c: c.data == "proj_activate")
async def proj_activate(c: types.CallbackQuery):
    uid = c.from_user.id
    projs = list_projects()
    if not projs:
        await clear_then_anchor(uid, "❗ Нет проектов.", kb_admin_projects()); return await c.answer()
    kb = InlineKeyboardMarkup()
    for p in projs: kb.add(InlineKeyboardButton(p, callback_data=f"proj_act_{p}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_projects"))
    await clear_then_anchor(uid, "Выберите проект для активации:", kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("proj_act_"))
async def proj_activate_do(c: types.CallbackQuery):
    uid = c.from_user.id
    name = c.data.split("proj_act_",1)[1]
    ensure_project_structure(name)
    info = load_project_info(name); info["active"] = True; save_project_info(name, info)
    set_active_project(name)
    await clear_then_anchor(uid, f"✅ Активирован проект: <b>{h(name)}</b> (код {h(info.get('code') or '—')})", kb_admin_projects())
    await c.answer("Активирован")
    await update_all_anchors()


@dp.callback_query_handler(lambda c: c.data == "proj_finish")
async def proj_finish(c: types.CallbackQuery):
    uid = c.from_user.id
    projs = list_projects()
    if not projs:
        await clear_then_anchor(uid, "❗ Нет проектов.", kb_admin_projects()); return await c.answer()
    kb = InlineKeyboardMarkup()
    for p in projs: kb.add(InlineKeyboardButton(p, callback_data=f"proj_fin_{p}"))
    kb.add(InlineKeyboardButton("⬅️ Назад", callback_data="adm_projects"))
    await clear_then_anchor(uid, "Выберите проект для завершения:", kb)
    await c.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("proj_fin_"))
async def proj_finish_do(c: types.CallbackQuery):
    uid = c.from_user.id
    name = c.data.split("proj_fin_",1)[1]
    info = load_project_info(name); info["active"] = False; save_project_info(name, info)
    if active_project["name"] == name: set_active_project(None)
    await clear_then_anchor(uid, f"✅ Проект «{h(name)}» (код {h(info.get('code') or '—')}) помечен как завершён.", kb_admin_projects())
    await c.answer("Завершён")
    await update_all_anchors()
    code = info.get("code") or "—"
    variants = [
        f"🎉 <b>Проект «{h(name)}» (код {h(code)}) завершён!</b>\n\nСпасибо за отличную работу.",
        f"✅ <b>Объект «{h(name)}» (код {h(code)}) закрыт.</b>\n\nБлагодарим каждого!",
        f"✨ <b>Завершили «{h(name)}»!</b>\n\nКод объекта: {h(code)}. До встречи на новых проектах."
    ]
    text = random.choice(variants)
    for f in os.listdir(USERS_PATH):
        if not f.endswith(".json"): continue
        try:
            udata = json.load(open(os.path.join(USERS_PATH, f), "r", encoding="utf-8"))
        except:
            continue
        chat_id = users_runtime.get(udata["user_id"], {}).get("tg", {}).get("chat_id") or udata.get("tg", {}).get("chat_id")
        if chat_id:
            try: await bot.send_message(chat_id, text, reply_markup=kb_broadcast_close())
            except Exception: pass


@dp.callback_query_handler(lambda c: c.data == "broadcast_close")
async def broadcast_close(c: types.CallbackQuery):
    try: await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception: pass
    await c.answer("Сообщение закрыто.")


@dp.callback_query_handler(lambda c: c.data == "close_saved_receipt")
async def close_saved_receipt_cb(c: types.CallbackQuery):
    try: await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception: pass
    await c.answer("Сообщение закрыто.")


@dp.callback_query_handler(lambda c: c.data == "admin_notice_close")
async def admin_notice_close(c: types.CallbackQuery):
    try: await bot.delete_message(c.message.chat.id, c.message.message_id)
    except Exception: pass
    await c.answer("Сообщение закрыто.")


def _colorize_terminal(text: str, color: str) -> str:
    """Return text wrapped in ANSI color codes if the terminal supports it."""
    if not sys.stdout.isatty() or os.environ.get("NO_COLOR"):
        return text
    return f"\033[{color}m{text}\033[0m"


def print_startup_banner():
    """Print a vibrant startup banner for the SAARC Telegram bot."""
    launch_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"🤖 {BOT_NAME} v{BOT_VERSION}",
        f"🛠 Revision: {BOT_REVISION}",
        "🏢 Company: SAARC",
        f"⏱ Launch time: {launch_time}",
        "🚀 Telegram bot is warming up — have a wonderful session!",
    ]
    width = max(len(line) for line in lines) + 4
    border = "═" * (width - 2)
    left = _colorize_terminal("║", "95")
    right = _colorize_terminal("║", "95")
    print(_colorize_terminal(f"╔{border}╗", "95"))
    for line in lines:
        padded = line.ljust(width - 4)
        print(f"{left} {_colorize_terminal(padded, '96')} {right}")
    print(_colorize_terminal(f"╚{border}╝", "95"))
    ready_line = f"{BOT_NAME} v{BOT_VERSION} | {BOT_REVISION} | ready for SAARC 🚀"
    print(_colorize_terminal(ready_line, "92"))


async def on_startup(dispatcher):
    await alerts_bootstrap()
    await alerts_start_polling()


async def on_shutdown(dispatcher):
    global alerts_poll_task
    if alerts_poll_task:
        alerts_poll_task.cancel()
        try:
            await alerts_poll_task
        except asyncio.CancelledError:
            pass
        alerts_poll_task = None


# ========================== BOOT ==========================
if __name__ == "__main__":
    ensure_dirs()
    sync_state()
    print_startup_banner()
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup, on_shutdown=on_shutdown)
